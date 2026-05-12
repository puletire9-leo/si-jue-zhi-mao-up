# ===================== 【用户自定义配置区域】=====================
import os
import sys
import warnings

import pandas as pd
from openpyxl import load_workbook

# 文件名配置
CUSTOM_FILE_SUFFIX = ""  # 自定义文件名后缀
OUTPUT_FILE_BASE_NAME = "导入领星表"  # 文件名前缀

# 固定值配置（随时修改）
FIXED_VALUES = {
    '开发人': '',
    '产品负责人': '唐若,张亚芳,阳姣,尹心如,蒋舒,张奋奋,李杉',
    '采购员': '王亚成',
    '采购交期': 7,
    '辅料SKU': '2270356',
    '辅料比例_主料': 1,
    '辅料比例_辅料': 1,
}

# 源数据文件名（在读取文件信息文件夹中）
SOURCE_FILE_NAME = '产品汇总表-模版 .xlsx'

# 源数据文件路径（从命令行参数传入，优先使用）
SOURCE_FILE_PATH = None

# 输出路径文件（用于后端API传递生成的文件路径）
OUTPUT_PATH_FILE = None

# 开发人（从命令行参数传入，覆盖配置）
DEVELOPER = None
# ===================== 【配置区域结束】=====================

# 解析命令行参数
if '--output-path-file' in sys.argv:
    idx = sys.argv.index('--output-path-file')
    if idx + 1 < len(sys.argv):
        OUTPUT_PATH_FILE = sys.argv[idx + 1]

if '--developer' in sys.argv:
    idx = sys.argv.index('--developer')
    if idx + 1 < len(sys.argv):
        DEVELOPER = sys.argv[idx + 1]

if '--source-file' in sys.argv:
    idx = sys.argv.index('--source-file')
    if idx + 1 < len(sys.argv):
        SOURCE_FILE_PATH = sys.argv[idx + 1]

warnings.filterwarnings('ignore')



def find_column_in_dataframe(df, possible_names):
    """精准匹配源数据列名"""
    for name in possible_names:
        if name in df.columns:
            return name
    return None


def get_column_mappings():
    """源数据→模板字段映射"""
    return {
        'SKU': ['SKU'],
        '产品名称': ['产品名称'],
        '长cm': ['长cm'],
        '宽cm': ['宽cm'],
        '高cm': ['高cm'],
        '毛重（kg）': ['毛重（kg）'],
        '采购费用': ['采购费用'],
        '供应商链接': ['供应商链接'],
        '供应商': ['供应商'],
        '英国海关编码': ['英国海关编码'],
        '中文报关名': ['中文报关名'],
        '英文报关名': ['英文报关名'],
        '图片url': ['图片url'],
    }


def clear_column_content(ws, col_letter, start_row=2):
    """清空指定列从start_row开始的所有内容"""
    max_row = ws.max_row
    for row_idx in range(start_row, max_row + 1):
        ws[f'{col_letter}{row_idx}'] = None


def process_lingxing_import():
    # 1. 路径定义
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_dir = os.path.join(base_dir, '导入零星模版')
    source_dir = os.path.join(base_dir, '读取文件信息')
    
    template_path = os.path.join(template_dir, '领星导入模板.xlsx')
    
    # 优先使用传入的源文件路径，否则使用默认路径
    if SOURCE_FILE_PATH and os.path.exists(SOURCE_FILE_PATH):
        source_path = SOURCE_FILE_PATH
        print(f"使用传入的源文件: {source_path}")
    else:
        source_path = os.path.join(source_dir, SOURCE_FILE_NAME)
        print(f"使用默认源文件: {source_path}")
    
    # 使用临时文件作为输出，不保存到固定文件夹
    import tempfile
    output_filename = f'{OUTPUT_FILE_BASE_NAME}-{CUSTOM_FILE_SUFFIX}.xlsx' if CUSTOM_FILE_SUFFIX else f'{OUTPUT_FILE_BASE_NAME}.xlsx'
    output_path = os.path.join(tempfile.gettempdir(), output_filename)

    try:
        # 2. 读取源数据并校验
        print("=" * 60)
        print("1. 读取源数据文件...")
        print("=" * 60)
        print(f"源文件路径: {source_path}")
        source_df = pd.read_excel(source_path)
        print(f"DataFrame列: {source_df.columns.tolist()}")
        print(f"DataFrame形状: {source_df.shape}")
        print(f"DataFrame前3行:\n{source_df.head(3)}")
        
        required_columns = ['SKU', '产品名称', '长cm', '宽cm', '高cm', '毛重（kg）',
                            '采购费用', '供应商链接', '供应商', '英国海关编码',
                            '中文报关名', '英文报关名', '图片url']
        missing_cols = [col for col in required_columns if col not in source_df.columns]
        if missing_cols:
            raise ValueError(f"源文件缺失关键列：{', '.join(missing_cols)}")

        sku_count = len(source_df)
        print(f"[OK] 源数据读取成功，共 {sku_count} 条SKU记录")

        # 3. 列名映射
        print("\n" + "=" * 60)
        print("2. 匹配列名...")
        print("=" * 60)
        column_mappings = get_column_mappings()
        actual_columns = {}
        for target_name, possible_names in column_mappings.items():
            actual_name = find_column_in_dataframe(source_df, possible_names)
            if actual_name:
                actual_columns[target_name] = actual_name
                print(f"[OK] {target_name} -> {actual_name}（匹配成功）")
            else:
                raise ValueError(f"列 {target_name} 未在源文件中找到")

        # 4. 加载模板
        print("\n" + "=" * 60)
        print("3. 初始化模板...")
        print("=" * 60)
        wb = load_workbook(template_path)
        ws_product = wb['产品']

        # 先清空所有可能需要填充的列（从第2行开始）
        all_cols_to_clear = ['A', 'B', 'S', 'T', 'W', 'X', 'AG', 'AI', 'AJ', 'AK',
                             'BB', 'Y', 'BF', 'BG', 'BP', 'V']
        for col in all_cols_to_clear:
            clear_column_content(ws_product, col)
            print(f"[OK] 清空【产品】工作表{col}列所有内容")

        # ---------------------- 5.1 填充「产品」工作表（仅SKU行）----------------------
        print("\n" + "=" * 60)
        print("4. 填充领星模板数据...")
        print("=" * 60)
        print("[INFO] 正在填充【产品】工作表...")

        # 固定值定义（从配置区域读取，开发人使用传入的参数）
        developer_value = DEVELOPER if DEVELOPER else FIXED_VALUES['开发人']
        fixed_values = {
            'S': developer_value,
            'T': FIXED_VALUES['产品负责人'],
            'W': FIXED_VALUES['采购员'],
            'X': FIXED_VALUES['采购交期']
        }
        print(f"[OK] 开发人: {developer_value}")

        # 核心：仅填充SKU对应行
        start_row = 2
        for i in range(sku_count):
            row_idx = start_row + i
            row = source_df.iloc[i]

            # 获取SKU值
            sku = row[actual_columns['SKU']]

            # 只有当SKU不为空时才填充该行
            if pd.notna(sku) and str(sku).strip():
                # 填充数据列
                ws_product[f'A{row_idx}'] = sku  # SKU

                # 产品名称
                product_name = row[actual_columns['产品名称']]
                ws_product[f'B{row_idx}'] = product_name if pd.notna(product_name) else ''

                # 填充固定值
                for col_letter, fixed_value in fixed_values.items():
                    ws_product[f'{col_letter}{row_idx}'] = fixed_value

                # 毛重处理：kg→g
                gross_weight = row[actual_columns['毛重（kg）']]
                if pd.notna(gross_weight) and str(gross_weight).strip():
                    try:
                        ws_product[f'AG{row_idx}'] = float(gross_weight) * 1000
                    except:
                        ws_product[f'AG{row_idx}'] = ''
                else:
                    ws_product[f'AG{row_idx}'] = ''

                # 长、宽、高
                for col_letter, field in [('AI', '长cm'), ('AJ', '宽cm'), ('AK', '高cm')]:
                    value = row[actual_columns[field]]
                    ws_product[f'{col_letter}{row_idx}'] = value if pd.notna(value) else ''

                # 供应商链接
                supplier_link = row[actual_columns['供应商链接']]
                ws_product[f'BB{row_idx}'] = supplier_link if pd.notna(supplier_link) else ''

                # 采购费用
                purchase_cost = row[actual_columns['采购费用']]
                ws_product[f'Y{row_idx}'] = purchase_cost if pd.notna(purchase_cost) else ''

                # 中文报关名
                chinese_customs_name = row[actual_columns['中文报关名']]
                ws_product[f'BF{row_idx}'] = chinese_customs_name if pd.notna(chinese_customs_name) else ''

                # 英文报关名
                english_customs_name = row[actual_columns['英文报关名']]
                ws_product[f'BG{row_idx}'] = english_customs_name if pd.notna(english_customs_name) else ''

                # 英国海关编码
                uk_hs_code = row[actual_columns['英国海关编码']]
                ws_product[f'BP{row_idx}'] = uk_hs_code if pd.notna(uk_hs_code) else ''

                # 图片url
                image_url = row[actual_columns['图片url']]
                ws_product[f'V{row_idx}'] = image_url if pd.notna(image_url) else ''

        print(f"[OK] 【产品】工作表填充完成，仅填充 {sku_count} 行（SKU对应行）")

        # ---------------------- 5.2 填充「关联辅料」工作表----------------------
        print("[INFO] 正在填充【关联辅料】工作表...")
        ws_accessory = wb['关联辅料']

        # 清空所有列
        accessory_cols = ['A', 'C', 'E', 'F']
        for col in accessory_cols:
            clear_column_content(ws_accessory, col)

        # 仅填充SKU对应行
        for i in range(sku_count):
            row_idx = start_row + i
            row = source_df.iloc[i]

            # 获取SKU值
            sku = row[actual_columns['SKU']]

            # 只有当SKU不为空时才填充该行
            if pd.notna(sku) and str(sku).strip():
                ws_accessory[f'A{row_idx}'] = sku  # SKU
                ws_accessory[f'C{row_idx}'] = FIXED_VALUES['辅料SKU']  # 辅料SKU固定值
                ws_accessory[f'E{row_idx}'] = FIXED_VALUES['辅料比例_主料']  # 辅料比例(主料)固定值
                ws_accessory[f'F{row_idx}'] = FIXED_VALUES['辅料比例_辅料']  # 辅料比例(辅料)固定值

        print(f"[OK] 【关联辅料】工作表填充完成，仅填充 {sku_count} 行（SKU对应行）")

        # ---------------------- 5.3 填充「更多供应商报价」工作表----------------------
        print("[INFO] 正在填充【更多供应商报价】工作表...")
        ws_supplier = wb['更多供应商报价']

        # 清空所有列
        supplier_cols = ['A', 'B', 'C', 'K']
        for col in supplier_cols:
            clear_column_content(ws_supplier, col)

        # 仅填充SKU对应行
        for i in range(sku_count):
            row_idx = start_row + i
            row = source_df.iloc[i]

            # 获取SKU值
            sku = row[actual_columns['SKU']]

            # 只有当SKU不为空时才填充该行
            if pd.notna(sku) and str(sku).strip():
                ws_supplier[f'A{row_idx}'] = sku  # SKU

                # 产品名称
                product_name = row[actual_columns['产品名称']]
                ws_supplier[f'B{row_idx}'] = product_name if pd.notna(product_name) else ''

                # 供应商
                supplier = row[actual_columns['供应商']]
                ws_supplier[f'C{row_idx}'] = supplier if pd.notna(supplier) else ''

                # 供应商链接
                supplier_link = row[actual_columns['供应商链接']]
                ws_supplier[f'K{row_idx}'] = supplier_link if pd.notna(supplier_link) else ''

        print(f"[OK] 【更多供应商报价】工作表填充完成，仅填充 {sku_count} 行（SKU对应行）")

        # 6. 保存文件
        print("\n" + "=" * 60)
        print("5. 保存生成的文件...")
        print("=" * 60)
        wb.save(output_path)
        print(f"[OK] 文件已保存至：{output_path}")

        # 7. 输出报告
        print("\n" + "=" * 60)
        print("[REPORT] 处理完成报告")
        print("=" * 60)
        print(f"- 源数据SKU数量：{sku_count} 条")
        print(f"- 填充行数：{start_row} - {start_row + sku_count - 1} 行（仅SKU对应行）")
        print(f"- 非SKU行：所有列均为空白")
        print(f"- 输出文件：{output_filename}")
        print("=" * 60)

        # 如果指定了输出路径文件，将生成的文件路径写入
        if OUTPUT_PATH_FILE:
            try:
                with open(OUTPUT_PATH_FILE, 'w') as f:
                    f.write(output_path)
                print(f"[OK] 文件路径已写入: {OUTPUT_PATH_FILE}")
            except Exception as e:
                print(f"[ERROR] 写入文件路径失败: {e}")

        return output_path

    except Exception as e:
        print(f"\n[ERROR] 处理失败：{str(e)}")
        import traceback
        traceback.print_exc()
        return None


def main():
    """主程序入口（支持命令行调用）"""
    print("=" * 60)
    print("领星导入文件生成程序（仅填充SKU行版本）")
    print("=" * 60)

    # 前置检查
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, '导入零星模版', '领星导入模板.xlsx')
    source_path = os.path.join(base_dir, '读取文件信息', SOURCE_FILE_NAME)

    if not os.path.exists(template_path):
        print(f"[ERROR] 错误：未找到模板文件 -> {template_path}")
        if not OUTPUT_PATH_FILE:  # 只有在非API调用时才等待输入
            input("按Enter键退出...")
        exit(1)

    if not os.path.exists(source_path):
        print(f"[ERROR] 错误：未找到源数据文件 -> {source_path}")
        if not OUTPUT_PATH_FILE:  # 只有在非API调用时才等待输入
            input("按Enter键退出...")
        exit(1)

    # 执行处理
    result = process_lingxing_import()

    if result:
        print("\n[SUCCESS] 程序执行成功！")
        return 0
    else:
        print("\n[ERROR] 程序执行失败！")
        return 1


if __name__ == "__main__":
    exit_code = main()
    if not OUTPUT_PATH_FILE:  # 只有在非API调用时才等待输入
        input("\n按Enter键退出...")
    exit(exit_code)
