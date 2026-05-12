# ===================== 【匹配链接文件脚本】=====================
import os
import warnings

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings('ignore')


def match_urls_to_excel():
    """
    读取url.txt中的链接，按顺序匹配到产品汇总表的图片url列
    """
    # 1. 路径定义
    base_dir = os.path.dirname(os.path.abspath(__file__))
    source_dir = os.path.join(base_dir, '读取文件信息')
    
    url_file_path = os.path.join(source_dir, 'url.txt')
    excel_file_path = os.path.join(source_dir, '产品汇总表-刘淼 .xlsx')
    output_file_path = excel_file_path  # 直接修改源文件
    
    # 打印路径信息用于调试
    print(f"URL文件路径: {url_file_path}")
    print(f"Excel文件路径: {excel_file_path}")
    print(f"输出文件路径: {output_file_path}")
    print(f"文件是否存在: {os.path.exists(url_file_path)}")
    
    try:
        # 2. 读取url.txt文件
        print("=" * 60)
        print("1. 读取url.txt文件...")
        print("=" * 60)
        
        # 使用绝对路径读取文件
        print(f"尝试读取文件: {url_file_path}")
        
        # 尝试多次读取文件
        for i in range(3):
            try:
                with open(url_file_path, 'rb') as f:
                    content = f.read()
                print(f"第 {i+1} 次读取 - 文件大小: {len(content)} 字节")
                if len(content) > 0:
                    break
            except Exception as e:
                print(f"读取错误: {str(e)}")
                import time
                time.sleep(1)
        
        print(f"最终文件大小: {len(content)} 字节")
        if len(content) > 0:
            print(f"文件前50个字节: {content[:50]}")
        
        # 尝试不同编码解码
        encodings = ['utf-8', 'gbk', 'utf-8-sig']
        decoded_content = None
        
        for encoding in encodings:
            try:
                decoded_content = content.decode(encoding)
                print(f"使用 {encoding} 编码成功解码")
                print(f"解码后内容长度: {len(decoded_content)}")
                print(f"前100个字符: {decoded_content[:100]}...")
                break
            except Exception as e:
                print(f"{encoding} 编码解码失败: {str(e)}")
                continue
        
        if decoded_content is None:
            raise ValueError("无法解码文件内容")
        
        # 分割行并过滤空行
        urls = [line.strip() for line in decoded_content.split('\n') if line.strip()]
        
        # 打印前5个链接用于调试
        print(f"✅ 成功读取 {len(urls)} 个链接")
        if urls:
            print("前5个链接：")
            for i, url in enumerate(urls[:5]):
                print(f"  {i+1}. {url}")
        else:
            print("未找到任何链接")
            print("文件内容:")
            print(repr(decoded_content))
        
        # 3. 读取Excel文件
        print("\n" + "=" * 60)
        print("2. 读取Excel文件...")
        print("=" * 60)
        wb = load_workbook(excel_file_path)
        ws = wb.active  # 默认使用第一个工作表
        
        # 4. 查找图片url列
        print("\n" + "=" * 60)
        print("3. 查找图片url列...")
        print("=" * 60)
        image_url_column = None
        header_row = 1  # 假设第一行是表头
        
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=header_row, column=col).value
            if header_value and '图片url' in str(header_value):
                image_url_column = col
                print(f"✓ 找到图片url列：第 {col} 列")
                break
        
        if not image_url_column:
            raise ValueError("未找到图片url列")
        
        # 5. 按顺序填充链接
        print("\n" + "=" * 60)
        print("4. 按顺序填充链接...")
        print("=" * 60)
        
        # 从第二行开始填充（跳过表头）
        start_row = 2
        max_row = ws.max_row
        
        # 计算需要填充的行数
        fill_rows = min(max_row - start_row + 1, len(urls))
        
        for i in range(fill_rows):
            row = start_row + i
            ws.cell(row=row, column=image_url_column, value=urls[i])
            print(f"✓ 第 {row} 行：{urls[i]}")
        
        # 6. 保存文件
        print("\n" + "=" * 60)
        print("5. 保存文件...")
        print("=" * 60)
        wb.save(output_file_path)
        print(f"✅ 成功修改源文件：{output_file_path}")
        print(f"\n📊 统计信息：")
        print(f"- 读取的链接数：{len(urls)}")
        print(f"- 填充的行数：{fill_rows}")
        print(f"- 剩余未使用的链接数：{len(urls) - fill_rows}")
        
    except Exception as e:
        print(f"❌ 错误：{str(e)}")
        raise


if __name__ == "__main__":
    match_urls_to_excel()
