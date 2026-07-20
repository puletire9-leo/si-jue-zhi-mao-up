"""每月增量入库:对接表 xlsx → skill 数据分片。

用法:
  python ingest_month.py <xlsx或目录> --month YYYY-MM

流程:
  1. 遍历目录下所有 xlsx(或单文件),按日期递增排序
  2. 每张表按表头自适应列位(1-5 月无"采购数量",6 月新增)
  3. 兼容 sheet 命名 (Sheet1 / 英国 / 德国)
  4. 同 SKU 多次出现时以最晚日期版本为准
  5. SKU → lingxing_local_product 拼图 + 校验采购价
  6. 按开品理由关键词打手法标签,标题打主题标签
  7. 落四处:
     - data/monthly-snapshot/<month>.csv          (全量原始快照)
     - data/by-method/<name>/<month>.csv          (手法分片)
     - data/by-supplier/<supplier>/<month>.csv    (供应商分片)
     - data/skill-summary.csv + supplier-summary.csv (全局统计)
"""
from __future__ import annotations

import argparse
import csv
import io
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median

import openpyxl

SKILL_ROOT = Path(__file__).resolve().parent.parent
DATA = SKILL_ROOT / "data"
SNAPSHOT_DIR = DATA / "monthly-snapshot"
BY_METHOD_DIR = DATA / "by-method"
BY_SUPPLIER_DIR = DATA / "by-supplier"
SKILL_SUMMARY = DATA / "skill-summary.csv"
SUPPLIER_SUMMARY = DATA / "supplier-summary.csv"

# ===== 列位自适应 =====
# 老版(1-5 月)无 "采购数量":开品理由=9  采购数量=None
# 新版(6 月起)含 "采购数量":开品理由=10 采购数量=7  其余每列后移 1
# 检测:表头第 2 行索引 7 是否为 "采购数量"
def resolve_cols(header_row: tuple) -> dict[str, int | None]:
    """返回字段 → 列号映射;不存在的字段为 None。"""
    has_qty = len(header_row) > 7 and header_row[7] == "采购数量"
    if has_qty:
        return {
            "序号": 0, "时间": 1, "产品名称": 3, "SKU": 4, "开发": 5, "运营": 6,
            "采购数量": 7, "CPC段": 8, "市场价格段": 9, "开品理由": 10, "开发备注": 11,
            "竞品链接1": 12, "竞品链接2": 13, "竞品链接3": 14,
            "售价": 15, "空运利润$": 16, "空运利润率": 18,
            "海运利润$": 19, "海运利润率": 21,
            "站点": 22, "汇率": 23,
            "长cm": 26, "宽cm": 27, "高cm": 28, "体积重kg": 29, "毛重kg": 30,
            "采购价¥": 33,
            "供应商链接": 46, "供应商": 47,
            "材质": 49, "中文报关名": 50, "英文报关名": 51, "产品尺寸": 52,
        }
    return {
        "序号": 0, "时间": 1, "产品名称": 3, "SKU": 4, "开发": 5, "运营": 6,
        "采购数量": None, "CPC段": 7, "市场价格段": 8, "开品理由": 9, "开发备注": 10,
        "竞品链接1": 11, "竞品链接2": 12, "竞品链接3": 13,
        "售价": 14, "空运利润$": 15, "空运利润率": 17,
        "海运利润$": 18, "海运利润率": 20,
        "站点": 21, "汇率": 22,
        "长cm": 25, "宽cm": 26, "高cm": 27, "体积重kg": 28, "毛重kg": 29,
        "采购价¥": 32,
        "供应商链接": 45, "供应商": 46,
        "材质": 48, "中文报关名": 49, "英文报关名": 50,
        "产品尺寸": None,
    }

OUTPUT_FIELDS = [
    "来源sheet", "来源文件", "序号", "时间", "产品名称", "SKU", "开发", "运营",
    "采购数量", "CPC段", "市场价格段", "开品理由", "开发备注",
    "竞品链接1", "竞品链接2", "竞品链接3",
    "售价", "空运利润$", "空运利润率", "海运利润$", "海运利润率",
    "站点", "汇率",
    "长cm", "宽cm", "高cm", "体积重kg", "毛重kg",
    "采购价¥", "供应商链接", "供应商",
    "材质", "中文报关名", "英文报关名", "产品尺寸",
]

# ===== 手法关键词映射 =====
METHOD_PATTERNS: dict[str, list[str]] = {
    "01-供应商新款首发": [
        r"供应商.{0,4}(新款|新品|上新|新品发)", r"新款式", r"新款",
        r"^新品", r"新品发", r"上新", r"新设计", r"原创图案",
        r"热销", r"热卖", r"速上", r"销量.{0,4}(好|不错|很好|非常好|很大)",
        r"外网热卖", r"市场.{0,4}(大量上新|表现)", r"非常热销",
        r"esty.{0,4}热图", r"esty上面", r"这个.{0,4}(图案|元素).{0,4}(好卖|竞品少|市场)",
        r"图案.{0,4}好卖", r"款式.{0,4}销量", r"表现不错",
        r"价格优势",  # 价格优势是团队常用泛化信号,归入 S01 兜底
    ],
    "02-热销跨站复制": [
        r"美国站", r"美国上新", r"美国爆款", r"美国热销",
        r"英国站.{0,10}(好卖|销量|表现|无同款)",
        r"英国.{0,10}(热销|无竞品|好卖)",
        r"德国.{0,10}(尝试|销量|市场|FBA|FBM|有量|好卖)",
        r"英销", r"英德.{0,4}(好卖|销量|市场|好)",
        r"英国销量", r"美国销量",
        r"在英国出的还行", r"汉密尔顿", r"世界杯",
        r"美爆款",
    ],
    "03-自有链接拓款": [
        r"都是自己的链接", r"自己的链接", r"拓新地标", r"拓的新图案", r"拓的套装",
        r"拓品", r"拓款", r"新地标",
    ],
    "04-老品季节化改款": [
        r"秋季", r"秋天", r"秋色", r"秋叶", r"秋日", r"秋狐",
        r"符合秋", r"秋天质感", r"秋天元素",
        r"开学季", r"9月.{0,4}(高峰|流量)",
        r"节日", r"圣诞", r"万圣", r"春季", r"夏季", r"冬季",
    ],
    "05-元素载体重组": [
        r"换载体", r"新载体", r"元素.{0,6}(好卖|受欢迎|销量|表现|竞品少)",
        r"图案.{0,4}(不一样|已改|不同)", r"类似风格", r"元素改款", r"组合了", r"新组合",
        r"受欢迎的", r"越野车主题", r"露营车主题",
        r"系列.{0,10}(不错|好卖|表现)", r"豹纹.{0,4}系列",
    ],
    "06-套装数量密度": [
        r"性价比", r"款不重复", r"多.{0,3}(pcs|件|款)", r"比同数量",
        r"最有性价比", r"超绝性价比", r"最低价", r"套装",
        r"便宜.{0,10}(多|pcs|件)",
        r"数量.{0,4}(多|更多)", r"两件装",
    ],
    "07-异形审美精修": [
        r"异形好卖", r"更精致", r"更加精致", r"花纹.{0,4}精细", r"细节.{0,4}(精美|更精)",
        r"新配色", r"新颜色", r"新款颜色", r"新图案", r"自主改款", r"自组",
        r"图案小改", r"精致", r"精美", r"更新颖", r"很新颖",
    ],
    "08-FBM验证FBA补发": [
        r"FBM.{0,4}(上新|新款|新品|销量|表现|链接)",
        r"自发货.{0,4}(表现|销量|出单|有量|新品|能出|出得)",
        r"英国.{0,4}没有FBA", r"没有FBA.{0,4}速上", r"没有FBA",
        r"德国没有FBA",
    ],
}

THEME_PATTERNS: dict[str, list[str]] = {
    "T1-秋季": [r"秋", r"枫叶", r"落叶", r"罗宾", r"知更鸟", r"猫头鹰"],
    "T2-恐龙": [r"恐龙"],
    # T3 猫狗:严格化猫,避免"猫头鹰"命中(移到 T1)
    "T3-猫狗": [r"猫咪", r"小猫", r"喵", r"^猫", r"狗", r"腊肠", r"柴犬", r"黑猫", r"水豚", r"卡皮巴拉"],
    "T4-蝴蝶结粉色": [r"蝴蝶结", r"粉色", r"BOW"],
    "T5-高地牛": [r"高地牛", r"highland"],
    "T6-花朵": [r"花朵", r"向日葵", r"雏菊", r"玫瑰"],
    # 新增汽车电子标品主题
    "T7-汽车电子": [r"车载", r"汽车.{0,4}(配件|香薰|支架|开关|电压)", r"电瓶", r"传感器", r"点烟", r"套筒", r"扳手", r"氧传感"],
}


def compile_patterns(mapping: dict[str, list[str]]) -> dict[str, re.Pattern]:
    return {k: re.compile("|".join(v), re.IGNORECASE) for k, v in mapping.items()}


METHOD_RE = compile_patterns(METHOD_PATTERNS)
THEME_RE = compile_patterns(THEME_PATTERNS)


def extract_one_xlsx(xlsx_path: Path) -> list[dict]:
    """从一张 xlsx 抽记录,自动处理 Sheet1 / 英国 / 德国。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("WpsReserved"):
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            _title = next(rows_iter)      # row 0
            h1 = next(rows_iter)          # row 1
            _h2 = next(rows_iter)         # row 2 (subheader)
        except StopIteration:
            continue
        cols = resolve_cols(h1)
        for row in rows_iter:
            if not row:
                continue
            sku_v = row[cols["SKU"]] if cols["SKU"] < len(row) else None
            if sku_v in (None, ""):
                continue
            dev_v = row[cols["开发"]] if cols["开发"] < len(row) else None
            if not dev_v:
                continue
            rec = {"来源sheet": sheet_name, "来源文件": xlsx_path.name}
            for k, idx in cols.items():
                if idx is None or idx >= len(row):
                    rec[k] = None
                    continue
                v = row[idx]
                if isinstance(v, str) and v.startswith("=DISPIMG"):
                    v = ""
                rec[k] = v
            rec["SKU"] = str(rec["SKU"]).strip()
            out.append(rec)
    return out


def merge_snapshots(files: list[Path]) -> list[dict]:
    """多个 xlsx 按 SKU 去重,同 SKU 保留最新文件的版本(采购数量取最大)。"""
    # 用 dict 保留最后一次(files 已按日期升序)
    by_sku: dict[str, dict] = {}
    qty_max: dict[str, int] = {}
    total_rows = 0
    for f in files:
        rows = extract_one_xlsx(f)
        total_rows += len(rows)
        print(f"  [scan] {f.name}: {len(rows)} rows")
        for r in rows:
            sku = r["SKU"]
            # 记录最大采购数量
            q = r.get("采购数量")
            try:
                q_int = int(q) if q not in (None, "") else 0
            except (TypeError, ValueError):
                q_int = 0
            qty_max[sku] = max(qty_max.get(sku, 0), q_int)
            by_sku[sku] = r  # 后来的覆盖前面的
    print(f"  [merge] {total_rows} raw rows → {len(by_sku)} unique SKU")
    # 把累计最大采购数量回填
    for sku, r in by_sku.items():
        if qty_max[sku] > 0:
            r["采购数量"] = qty_max[sku]
    return list(by_sku.values())


def query_lingxing(skus: list[str]) -> dict[str, dict]:
    if not skus:
        return {}
    # 分批查询避免命令过长
    BATCH = 500
    all_out: dict[str, dict] = {}
    for i in range(0, len(skus), BATCH):
        batch = skus[i:i + BATCH]
        quoted = ",".join(f"'{s}'" for s in batch if s)
        sql = (
            "USE sijuelishi_dev; "
            "SELECT sku, product_developer, pic_url, cg_price, status_text, open_status "
            f"FROM lingxing_local_product WHERE sku IN ({quoted});"
        )
        cmd = [
            "docker", "exec", "dev-mysql", "mysql", "-uroot", "-proot",
            "--default-character-set=utf8mb4", "-B", "-e", sql,
        ]
        r = subprocess.run(cmd, capture_output=True, timeout=120)
        if r.returncode != 0:
            print(f"[ingest] mysql err: {r.stderr.decode('utf-8', 'replace')}", file=sys.stderr)
            continue
        reader = csv.DictReader(io.StringIO(r.stdout.decode("utf-8", "replace")), delimiter="\t")
        for row in reader:
            all_out[row["sku"]] = row
    return all_out


def match_methods(reason: str, title: str) -> list[str]:
    """理由为主,标题只允许 S04/S07 追加。"""
    reason = reason or ""
    title = title or ""
    reason_hits = {name for name, pat in METHOD_RE.items() if reason and pat.search(reason)}
    text_hits = {name for name, pat in METHOD_RE.items() if pat.search(f"{reason} {title}")}
    hits = set(reason_hits)
    for allow in ("04-老品季节化改款", "07-异形审美精修"):
        if allow in text_hits:
            hits.add(allow)
    return sorted(hits)


def match_themes(title: str) -> list[str]:
    text = title or ""
    return [name for name, pat in THEME_RE.items() if pat.search(text)]


def to_float(v):
    if v in (None, "", "NULL"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def enrich(rows: list[dict], lx: dict[str, dict]) -> list[dict]:
    for r in rows:
        sku = r["SKU"]
        lxr = lx.get(sku, {})
        r["lx_pic_url"] = lxr.get("pic_url", "")
        r["lx_cg_price"] = lxr.get("cg_price", "")
        r["lx_status_text"] = lxr.get("status_text", "")
        r["lx_open_status"] = lxr.get("open_status", "")
        methods = match_methods(r.get("开品理由", ""), r.get("产品名称", ""))
        themes = match_themes(r.get("产品名称", ""))
        r["匹配手法"] = "|".join(methods) if methods else "00-unmatched"
        r["匹配主题"] = "|".join(themes)
    return rows


def sanitize_dir(name: str) -> str:
    """供应商名清洗,避免文件系统非法字符。"""
    if not name:
        return "_missing"
    return re.sub(r'[<>:"/\\|?*\s]+', "_", name).strip("_")[:80]


def write_snapshot(rows: list[dict], month: str) -> Path:
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    dst = SNAPSHOT_DIR / f"{month}.csv"
    if not rows:
        return dst
    fields = OUTPUT_FIELDS + ["lx_pic_url", "lx_cg_price", "lx_status_text", "lx_open_status", "匹配手法", "匹配主题"]
    with dst.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    return dst


def write_by_method(rows: list[dict], month: str) -> dict[str, int]:
    BY_METHOD_DIR.mkdir(parents=True, exist_ok=True)
    method_rows: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        methods = (r["匹配手法"] or "00-unmatched").split("|")
        for m in methods:
            method_rows[m].append(r)
    stats: dict[str, int] = {}
    fields = OUTPUT_FIELDS + ["lx_pic_url", "lx_cg_price", "lx_status_text", "lx_open_status", "匹配手法", "匹配主题"]
    for method_name, subset in method_rows.items():
        mdir = BY_METHOD_DIR / method_name
        mdir.mkdir(parents=True, exist_ok=True)
        dst = mdir / f"{month}.csv"
        with dst.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            w.writeheader()
            w.writerows(subset)
        stats[method_name] = len(subset)
    return stats


# 供应商分片下界:当月 SKU 数低于此值的合并进 _tail
SUPPLIER_TAIL_THRESHOLD = 3


def write_by_supplier(rows: list[dict], month: str) -> dict[str, int]:
    """按供应商分片。≥ SUPPLIER_TAIL_THRESHOLD 的独立目录,否则合并到 _tail/。"""
    BY_SUPPLIER_DIR.mkdir(parents=True, exist_ok=True)
    sup_rows: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        sup = (r.get("供应商") or "_missing").strip() or "_missing"
        sup_rows[sup].append(r)
    stats: dict[str, int] = {}
    fields = OUTPUT_FIELDS + ["lx_pic_url", "lx_cg_price", "lx_status_text", "lx_open_status", "匹配手法", "匹配主题"]
    tail_rows: list[dict] = []
    for sup, subset in sup_rows.items():
        if len(subset) < SUPPLIER_TAIL_THRESHOLD:
            tail_rows.extend(subset)
            continue
        sdir = BY_SUPPLIER_DIR / sanitize_dir(sup)
        sdir.mkdir(parents=True, exist_ok=True)
        dst = sdir / f"{month}.csv"
        with dst.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            w.writeheader()
            w.writerows(subset)
        stats[sup] = len(subset)
    if tail_rows:
        tail_dir = BY_SUPPLIER_DIR / "_tail"
        tail_dir.mkdir(parents=True, exist_ok=True)
        dst = tail_dir / f"{month}.csv"
        with dst.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            w.writeheader()
            w.writerows(tail_rows)
        stats["_tail"] = len(tail_rows)
    return stats


def rebuild_skill_summary() -> Path:
    rows: list[dict] = []
    for method_dir in sorted(BY_METHOD_DIR.iterdir()):
        if not method_dir.is_dir():
            continue
        for csv_file in sorted(method_dir.glob("*.csv")):
            month = csv_file.stem
            with csv_file.open("r", encoding="utf-8-sig") as f:
                subset = list(csv.DictReader(f))
            if not subset:
                continue
            prices = [to_float(r.get("售价")) for r in subset]
            prices = [x for x in prices if x is not None]
            rates = [to_float(r.get("空运利润率")) for r in subset]
            rates = [x for x in rates if x is not None]
            devs = Counter(r.get("开发", "") for r in subset)
            rows.append({
                "方法": method_dir.name,
                "月份": month,
                "SKU数": len(subset),
                "售价中位": f"{median(prices):.2f}" if prices else "",
                "利润率中位": f"{median(rates):.4f}" if rates else "",
                "主要开发": ",".join(f"{k}({v})" for k, v in devs.most_common(3)),
            })
    if rows:
        with SKILL_SUMMARY.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
    return SKILL_SUMMARY


def rebuild_supplier_summary() -> Path:
    rows: list[dict] = []
    for sup_dir in sorted(BY_SUPPLIER_DIR.iterdir()):
        if not sup_dir.is_dir():
            continue
        # 汇总该供应商全部月份
        all_recs: list[dict] = []
        months: set[str] = set()
        for csv_file in sorted(sup_dir.glob("*.csv")):
            months.add(csv_file.stem)
            with csv_file.open("r", encoding="utf-8-sig") as f:
                all_recs.extend(csv.DictReader(f))
        if not all_recs:
            continue
        prices = [to_float(r.get("采购价¥")) for r in all_recs]
        prices = [x for x in prices if x is not None]
        rates = [to_float(r.get("空运利润率")) for r in all_recs]
        rates = [x for x in rates if x is not None]
        devs = Counter(r.get("开发", "") for r in all_recs)
        methods = Counter()
        for r in all_recs:
            for m in (r.get("匹配手法") or "").split("|"):
                if m:
                    methods[m] += 1
        # 供应商原始名(从任一记录里拿)
        raw_name = all_recs[0].get("供应商") or "_missing"
        rows.append({
            "供应商": raw_name,
            "SKU数": len(all_recs),
            "月份数": len(months),
            "月份列表": ",".join(sorted(months)),
            "采购价¥中位": f"{median(prices):.2f}" if prices else "",
            "利润率中位": f"{median(rates):.4f}" if rates else "",
            "主要开发": ",".join(f"{k}({v})" for k, v in devs.most_common(3)),
            "主要手法": ",".join(f"{k}({v})" for k, v in methods.most_common(3)),
        })
    rows.sort(key=lambda r: -r["SKU数"])
    if rows:
        with SUPPLIER_SUMMARY.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
    return SUPPLIER_SUMMARY


def gather_xlsx(path: Path, month: str | None = None) -> list[Path]:
    """收集目标 xlsx。若给了 month(YYYY-MM),只保留文件名以 'MM.' 开头的表。"""
    def _keep(p: Path) -> bool:
        if p.name.startswith("~"):
            return False
        if not month:
            return True
        mm = month.split("-")[1] if "-" in month else month
        # 匹配 "理实产品对接表<MM>." 或 "对接表<MM>." 前缀
        m = re.search(r"(\d{2})[.\-]", p.stem)
        return bool(m and m.group(1) == mm)

    if path.is_file():
        return [path]
    if path.is_dir():
        return sorted(p for p in path.glob("*.xlsx") if _keep(p))
    return []


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("target", help="对接表 .xlsx 或 目录路径")
    ap.add_argument("--month", required=True, help="月份标签 YYYY-MM")
    args = ap.parse_args()

    target = Path(args.target)
    files = gather_xlsx(target, args.month)
    if not files:
        print(f"[ingest] no xlsx found under {target}", file=sys.stderr)
        return 1

    print(f"[ingest] month={args.month} files={len(files)}")
    for f in files:
        print(f"  · {f.name}")

    rows = merge_snapshots(files)
    print(f"[ingest] merged unique SKU: {len(rows)}")

    skus = [r["SKU"] for r in rows]
    lx = query_lingxing(skus)
    hit = sum(1 for s in skus if s in lx)
    print(f"[ingest] lingxing hit: {hit}/{len(skus)} ({hit/len(skus):.1%})")

    enrich(rows, lx)

    reason_hit = sum(1 for r in rows if r.get("开品理由"))
    method_hit = sum(1 for r in rows if r["匹配手法"] != "00-unmatched")
    print(f"[ingest] 有开品理由 {reason_hit}/{len(rows)}  匹配到手法 {method_hit}/{len(rows)}")

    snap = write_snapshot(rows, args.month)
    print(f"[ingest] snapshot → {snap.name}")

    mstats = write_by_method(rows, args.month)
    print(f"[ingest] by-method 分片({len(mstats)} 个):")
    for k in sorted(mstats):
        print(f"    {k}: {mstats[k]}")

    sstats = write_by_supplier(rows, args.month)
    print(f"[ingest] by-supplier 分片({len(sstats)} 家供应商)")
    top = sorted(sstats.items(), key=lambda x: -x[1])[:8]
    for k, v in top:
        print(f"    {v:3d}  {k}")

    rebuild_skill_summary()
    rebuild_supplier_summary()
    print(f"[ingest] summary refreshed → {SKILL_SUMMARY.name}, {SUPPLIER_SUMMARY.name}")
    print("[ingest] done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
