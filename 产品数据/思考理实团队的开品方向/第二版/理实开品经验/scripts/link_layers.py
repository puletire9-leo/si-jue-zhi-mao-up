"""三层证据链粘合器:L1 开品 → L2 上架 → L3 经营。

只做索引,不复制原始数据。产出一份完整生命周期 csv,把 5,410 个 L1 SKU 挂上:
- L2 的 ASIN + 上架月 + FBA 首现月 + Listing 标签
- L3 的累计销量 + 结算销售额 + 结算利润 + 当前标签状态 + 币种

同一 SKU 可能挂多个 ASIN(不同站点/币种),每 (SKU, ASIN) 一行。

用法:
  python link_layers.py

产出:
  data/sku_full_lifecycle.csv    ← 一行 = 一对 (SKU, ASIN),含三层完整字段
  data/link_diagnostics.md       ← 覆盖率/断链分析报告
"""
from __future__ import annotations

import csv
import io
import sys
from collections import defaultdict
from pathlib import Path
from statistics import median

import openpyxl

SKILL_ROOT = Path(__file__).resolve().parent.parent
DATA = SKILL_ROOT / "data"
SNAPSHOT_DIR = DATA / "monthly-snapshot"
DST_CSV = DATA / "sku_full_lifecycle.csv"
DST_DIAG = DATA / "link_diagnostics.md"

# 领星模型路径
LX_MODEL = Path("F:/项目/si-jue-zhi-mao-up/产品数据/领星模型")
L2_BASE_CSV = LX_MODEL / "基础统一表/上架批次和批次ASIN集合永远锁定基础批次表/ASIN_FBA可售优先_商品信息创建时间兜底_模型分析起算月基准_2025-04至2026-06.csv"
L3_GBP_XLSX = LX_MODEL / "ASIN月度经营模型_FBA可售优先_2025-04至2026-06/03_审查报表/GBP/ASIN_FBA可售优先_商品信息创建时间兜底汇总_2025-04至2026-06.xlsx"
L3_EUR_XLSX = LX_MODEL / "ASIN月度经营模型_FBA可售优先_2025-04至2026-06/03_审查报表/EUR/ASIN_FBA可售优先_商品信息创建时间兜底汇总_2025-04至2026-06.xlsx"


def load_l1() -> dict[str, dict]:
    """L1:合并所有月份 snapshot,同 SKU 保留最新月份记录。"""
    l1: dict[str, dict] = {}
    if not SNAPSHOT_DIR.exists():
        print(f"[link] no L1 snapshot dir: {SNAPSHOT_DIR}", file=sys.stderr)
        return {}
    for csv_file in sorted(SNAPSHOT_DIR.glob("*.csv")):
        month = csv_file.stem
        with csv_file.open("r", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                sku = row.get("SKU", "").strip()
                if not sku:
                    continue
                row["L1月份"] = month
                l1[sku] = row  # 后写覆盖 → 保留最晚月份
    return l1


def load_l2() -> dict[str, list[dict]]:
    """L2:基础表,一个 SKU 可能挂多个 ASIN(不同 SKU 变体或站点)。"""
    # 现在的基础表 CSV 里每行是一个 ASIN,通过 基准SKU 反查
    by_sku: dict[str, list[dict]] = defaultdict(list)
    if not L2_BASE_CSV.exists():
        print(f"[link] no L2 base csv: {L2_BASE_CSV}", file=sys.stderr)
        return {}
    with L2_BASE_CSV.open("r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            sku = row.get("基准SKU", "").strip()
            if not sku:
                continue
            by_sku[sku].append(row)
    return by_sku


def load_l3_xlsx(path: Path, currency: str) -> dict[str, dict]:
    """L3:某币种 xlsx 所有月度 sheet 合并成 ASIN → 经营记录。"""
    out: dict[str, dict] = {}
    if not path.exists():
        return out
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        if "年" not in sn or "月" not in sn:
            continue
        ws = wb[sn]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            _title = next(rows_iter)  # row 0
            header = list(next(rows_iter))
        except StopIteration:
            continue
        for row in rows_iter:
            if not row or not row[0]:
                continue
            rec: dict = {}
            for i, h in enumerate(header):
                if h and i < len(row):
                    rec[h] = row[i]
            rec["币种"] = currency
            asin = rec.get("ASIN", "").strip() if rec.get("ASIN") else ""
            if asin:
                out[asin] = rec
    return out


def to_float(v):
    if v in (None, "", "NULL"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


OUT_FIELDS = [
    # ===== L1 开品层 =====
    "SKU", "L1月份", "开发", "产品名称", "供应商",
    "开品理由", "售价", "采购价¥", "空运利润率", "站点",
    "匹配手法", "匹配主题",
    "lx_pic_url",
    # ===== L2 上架层 =====
    "ASIN", "基准店铺", "创建时间", "最新Listing标签",
    "首次观察到FBA库存月", "首次观察到FBA可售月",
    "模型分析起算月", "月级FBA库存观察状态",
    # ===== L3 经营层 =====
    "币种", "累计销售量", "总销售额", "总结算销售额", "总结算利润", "总可用库存",
    "当前标签状态",
    # ===== 交叉指标 =====
    "L1→L2 状态", "L2→L3 状态", "生命周期状态",
]


def compute_lifecycle_state(l2_recs: list[dict], l3_rec: dict | None) -> tuple[str, str, str]:
    """三个层级的贯通状态。"""
    if not l2_recs:
        return ("L1 有 · L2 无", "N/A", "开品未进上架基础表")
    # L2 有
    has_fba = any(r.get("首次观察到FBA可售月") for r in l2_recs)
    l1_l2 = "L1 有 · L2 有 · " + ("有FBA证据" if has_fba else "仅创建兜底")
    if not l3_rec:
        return (l1_l2, "L2 有 · L3 无", "上架但无经营数据")
    # L3 有
    tag = l3_rec.get("当前标签状态", "")
    l2_l3 = f"L2 有 · L3 有 · {tag or '未知'}"
    profit = to_float(l3_rec.get("总结算利润"))
    if tag == "淘汰":
        lc = "已淘汰"
    elif profit is not None and profit > 0:
        lc = "留存且盈利"
    elif profit is not None and profit < 0:
        lc = "留存但亏损"
    else:
        lc = "留存 · 数据不足"
    return (l1_l2, l2_l3, lc)


def main() -> int:
    print("[link] loading L1 snapshots ...")
    l1 = load_l1()
    print(f"[link]   L1 SKU: {len(l1)}")

    print("[link] loading L2 base table ...")
    l2 = load_l2()
    print(f"[link]   L2 unique 基准SKU: {len(l2)}  ({sum(len(v) for v in l2.values())} 条 SKU-ASIN 映射)")

    print("[link] loading L3 GBP monthly ...")
    l3_gbp = load_l3_xlsx(L3_GBP_XLSX, "GBP")
    print(f"[link]   L3 GBP ASIN: {len(l3_gbp)}")

    print("[link] loading L3 EUR monthly ...")
    l3_eur = load_l3_xlsx(L3_EUR_XLSX, "EUR")
    print(f"[link]   L3 EUR ASIN: {len(l3_eur)}")

    l3_all = {**l3_gbp, **l3_eur}
    print(f"[link]   L3 合并 ASIN: {len(l3_all)}")

    # ===== 组装 =====
    output_rows: list[dict] = []
    covered_l1 = 0
    l1_l2_covered = 0
    l1_l2_l3_covered = 0

    for sku, l1_rec in l1.items():
        covered_l1 += 1
        l2_recs = l2.get(sku, [])
        if l2_recs:
            l1_l2_covered += 1
            for l2_rec in l2_recs:
                asin = l2_rec.get("ASIN", "")
                l3_rec = l3_all.get(asin)
                if l3_rec:
                    l1_l2_l3_covered += 1
                s1, s2, lc = compute_lifecycle_state([l2_rec], l3_rec)
                merged = {
                    # L1
                    "SKU": sku,
                    "L1月份": l1_rec.get("L1月份", ""),
                    "开发": l1_rec.get("开发", ""),
                    "产品名称": l1_rec.get("产品名称", ""),
                    "供应商": l1_rec.get("供应商", ""),
                    "开品理由": l1_rec.get("开品理由", ""),
                    "售价": l1_rec.get("售价", ""),
                    "采购价¥": l1_rec.get("采购价¥", ""),
                    "空运利润率": l1_rec.get("空运利润率", ""),
                    "站点": l1_rec.get("站点", ""),
                    "匹配手法": l1_rec.get("匹配手法", ""),
                    "匹配主题": l1_rec.get("匹配主题", ""),
                    "lx_pic_url": l1_rec.get("lx_pic_url", ""),
                    # L2
                    "ASIN": asin,
                    "基准店铺": l2_rec.get("基准店铺", ""),
                    "创建时间": l2_rec.get("创建时间", ""),
                    "最新Listing标签": l2_rec.get("最新Listing标签", ""),
                    "首次观察到FBA库存月": l2_rec.get("首次观察到FBA库存月", ""),
                    "首次观察到FBA可售月": l2_rec.get("首次观察到FBA可售月", ""),
                    "模型分析起算月": (l3_rec or {}).get("模型分析起算月", ""),
                    "月级FBA库存观察状态": l2_rec.get("月级FBA库存观察状态", ""),
                    # L3
                    "币种": (l3_rec or {}).get("币种", ""),
                    "累计销售量": (l3_rec or {}).get("累计销售量", ""),
                    "总销售额": (l3_rec or {}).get("总销售额", ""),
                    "总结算销售额": (l3_rec or {}).get("总结算销售额", ""),
                    "总结算利润": (l3_rec or {}).get("总结算利润", ""),
                    "总可用库存": (l3_rec or {}).get("总可用库存", ""),
                    "当前标签状态": (l3_rec or {}).get("当前标签状态", ""),
                    "L1→L2 状态": s1,
                    "L2→L3 状态": s2,
                    "生命周期状态": lc,
                }
                output_rows.append(merged)
        else:
            # 未进 L2 的 SKU 也保留一条,便于统计
            output_rows.append({
                "SKU": sku,
                "L1月份": l1_rec.get("L1月份", ""),
                "开发": l1_rec.get("开发", ""),
                "产品名称": l1_rec.get("产品名称", ""),
                "供应商": l1_rec.get("供应商", ""),
                "开品理由": l1_rec.get("开品理由", ""),
                "售价": l1_rec.get("售价", ""),
                "采购价¥": l1_rec.get("采购价¥", ""),
                "空运利润率": l1_rec.get("空运利润率", ""),
                "站点": l1_rec.get("站点", ""),
                "匹配手法": l1_rec.get("匹配手法", ""),
                "匹配主题": l1_rec.get("匹配主题", ""),
                "lx_pic_url": l1_rec.get("lx_pic_url", ""),
                "L1→L2 状态": "L1 有 · L2 无",
                "L2→L3 状态": "N/A",
                "生命周期状态": "开品未进上架基础表",
            })

    # 写全生命周期 CSV
    DST_CSV.parent.mkdir(parents=True, exist_ok=True)
    with DST_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUT_FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(output_rows)
    print(f"[link] wrote {len(output_rows)} rows → {DST_CSV.name}")

    # ===== 诊断报告 =====
    lines: list[str] = []
    lines.append("# 三层证据链贯通诊断\n")
    lines.append(f"- L1 开品 SKU 总数: **{covered_l1}**")
    lines.append(f"- L1 → L2 命中: **{l1_l2_covered} ({l1_l2_covered/covered_l1:.1%})**")
    lines.append(f"- L1 → L2 → L3 完整链路: **{l1_l2_l3_covered} 个 SKU-ASIN 组合**")
    lines.append(f"- 产出全生命周期 SKU-ASIN 记录: **{len(output_rows)}**")
    lines.append("")

    # 按 生命周期状态 分布
    from collections import Counter
    lc_c = Counter(r.get("生命周期状态", "") for r in output_rows)
    lines.append("## 生命周期状态分布")
    lines.append("")
    lines.append("| 状态 | 数量 | 占比 |")
    lines.append("|---|---:|---:|")
    for k, v in lc_c.most_common():
        lines.append(f"| {k} | {v} | {v/len(output_rows):.1%} |")
    lines.append("")

    # 按开发人分布(留存且盈利 vs 已淘汰)
    lines.append("## 开发人生命周期分布 (SKU-ASIN 组合数)")
    lines.append("")
    dev_lc: dict[str, Counter] = defaultdict(Counter)
    for r in output_rows:
        dev_lc[r.get("开发", "")][r.get("生命周期状态", "")] += 1
    lines.append("| 开发 | 总记录 | 留存且盈利 | 留存但亏损 | 已淘汰 | 未进 L2 | 无 L3 数据 |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|")
    for dev, c in sorted(dev_lc.items(), key=lambda x: -sum(x[1].values())):
        if not dev:
            continue
        total = sum(c.values())
        lines.append(
            f"| {dev} | {total} | {c.get('留存且盈利', 0)} | {c.get('留存但亏损', 0)} | {c.get('已淘汰', 0)} | {c.get('开品未进上架基础表', 0)} | {c.get('上架但无经营数据', 0)} |"
        )
    lines.append("")

    # 按手法看 L3 转化
    lines.append("## 手法 × 生命周期 (仅含进入 L3 的记录)")
    lines.append("")
    method_lc: dict[str, Counter] = defaultdict(Counter)
    for r in output_rows:
        if r.get("生命周期状态") in ("开品未进上架基础表", "上架但无经营数据"):
            continue
        for m in (r.get("匹配手法") or "").split("|"):
            if m:
                method_lc[m][r.get("生命周期状态", "")] += 1
    lines.append("| 手法 | 总 | 留存且盈利 | 留存但亏损 | 已淘汰 |")
    lines.append("|---|---:|---:|---:|---:|")
    for m in sorted(method_lc):
        c = method_lc[m]
        total = sum(c.values())
        lines.append(
            f"| {m} | {total} | {c.get('留存且盈利', 0)} ({c.get('留存且盈利', 0)/total:.0%}) | {c.get('留存但亏损', 0)} ({c.get('留存但亏损', 0)/total:.0%}) | {c.get('已淘汰', 0)} ({c.get('已淘汰', 0)/total:.0%}) |"
        )

    DST_DIAG.write_text("\n".join(lines), encoding="utf-8")
    print(f"[link] diag → {DST_DIAG.name}")
    print("[link] done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
