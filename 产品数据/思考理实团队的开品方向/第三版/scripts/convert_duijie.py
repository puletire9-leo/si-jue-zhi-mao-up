# -*- coding: utf-8 -*-
"""把理实产品对接表 xlsx 转成 AI 可快速读懂的 JSONL + 干净 CSV + 按 SKU 命名的图片。

输入: 一个对接表 xlsx (含 WPS DISPIMG 图片)
输出 (在 --out 目录下):
  products.jsonl     每行一个产品, 核心字段, AI 首选
  products.csv       同数据的表格版, 人看
  images/<SKU>.<ext> 每个产品的主图, 按 SKU 命名
  _convert_report.md 转换诊断

DISPIMG 图片链路:
  单元格公式 =DISPIMG("ID_xxx") 或 =_xlfn.DISPIMG("ID_xxx")
   -> xl/cellimages.xml   <xdr:cNvPr name="ID_xxx"/> ... <a:blip r:embed="rIdN"/>
   -> xl/_rels/cellimages.xml.rels  rIdN -> media/imageN.ext

用法:
  python convert_duijie.py <xlsx> --out <目录>
"""
from __future__ import annotations
import argparse, csv, io, json, re, sys, zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# 我们要保留的核心字段: 表头原名 -> 干净键名
# 表头是 row2(主)+row3(子) 两行, 这里用拍平后的定位(按列索引取)
# 列索引见 inspect: 0序号 1时间 2图片 3产品名称 4SKU 5开发 6运营 7采购数量
#   8CPC段 9市场价格段 10开品理由 11开发备注 12-14竞品链接 15售价
#   16空运利润$ 17空运利润¥ 18空运利润率 19海运利润$ 20海运利润¥ 21海运利润率
#   22站点 23汇率 24VAT税率 25佣金比例 26长 27宽 28高 29体积重 30毛重
#   ... 47供应商 48海关编码 49材质 50中文报关名 51英文报关名 52产品尺寸
CORE_COLS = {
    0:  "序号",
    1:  "时间_excel",
    3:  "产品名称",
    4:  "SKU",
    5:  "开发",
    6:  "运营",
    7:  "采购数量",
    8:  "CPC段",
    9:  "市场价格段",
    10: "开品理由",
    11: "开发备注",
    12: "竞品链接1",
    13: "竞品链接2",
    14: "竞品链接3",
    15: "售价",
    16: "空运利润_$",
    18: "空运利润率",
    19: "海运利润_$",
    21: "海运利润率",
    22: "站点",
    26: "长cm",
    27: "宽cm",
    28: "高cm",
    30: "毛重kg",
    46: "供应商链接",
    47: "供应商",
    48: "海关编码",
    49: "材质",
    50: "中文报关名",
    51: "英文报关名",
    52: "产品尺寸",
}
IMG_COL = 2          # 图片列
DATA_START_ROW = 4   # 第4行开始是数据

NS = {
    "a":   "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r":   "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

DISPIMG_RE = re.compile(r'DISPIMG\(\s*"([^"]+)"', re.IGNORECASE)


def build_image_map(xlsx: Path) -> dict[str, bytes]:
    """返回 DISPIMG ID -> 图片字节。"""
    z = zipfile.ZipFile(xlsx)
    names = set(z.namelist())
    if "xl/cellimages.xml" not in names:
        z.close()
        return {}

    # rId -> media path
    rid_to_media: dict[str, str] = {}
    rels = ET.fromstring(z.read("xl/_rels/cellimages.xml.rels"))
    for rel in rels.findall("rel:Relationship", NS):
        target = rel.get("Target")
        if not target.startswith("xl/"):
            target = "xl/" + target.lstrip("/")
        rid_to_media[rel.get("Id")] = target

    # DISPIMG ID(name) -> rId
    id_to_bytes: dict[str, bytes] = {}
    cimgs = ET.fromstring(z.read("xl/cellimages.xml"))
    for ci in cimgs:
        pic = ci.find("xdr:pic", NS)
        if pic is None:
            continue
        cnvpr = pic.find("xdr:nvPicPr/xdr:cNvPr", NS)
        blip = pic.find("xdr:blipFill/a:blip", NS)
        if cnvpr is None or blip is None:
            continue
        img_id = cnvpr.get("name")
        rid = blip.get(f"{{{NS['r']}}}embed")
        media = rid_to_media.get(rid)
        if img_id and media and media in names:
            id_to_bytes[img_id] = z.read(media)
    z.close()
    return id_to_bytes


def raw_dispimg_ids(xlsx: Path, sheet_name: str) -> dict[int, str]:
    """用只读公式模式拿到 图片列 每行的 DISPIMG ID (行号1基 -> ID)。

    openpyxl data_only 读公式会得到 None, 需要 data_only=False 读公式文本。
    """
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=False)
    ws = wb[sheet_name]
    result: dict[int, str] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, min_col=IMG_COL+1, max_col=IMG_COL+1):
        for cell in row:
            v = cell.value
            if isinstance(v, str):
                m = DISPIMG_RE.search(v)
                if m:
                    result[cell.row] = m.group(1)
    wb.close()
    return result


def excel_serial_to_date(v):
    """Excel 日期序列 -> YYYY-MM-DD。46219 这种。"""
    try:
        n = int(float(v))
    except (TypeError, ValueError):
        return v
    if n < 20000 or n > 60000:
        return v
    from datetime import date, timedelta
    return (date(1899, 12, 30) + timedelta(days=n)).isoformat()


def clean_num(v):
    if v in (None, "", "NULL"):
        return None
    if isinstance(v, (int, float)):
        return round(v, 4) if isinstance(v, float) else v
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return v


def convert(xlsx: Path, out: Path):
    out.mkdir(parents=True, exist_ok=True)
    img_dir = out / "images"
    img_dir.mkdir(exist_ok=True)

    id_to_bytes = build_image_map(xlsx)
    print(f"[img] cellimages 解出 {len(id_to_bytes)} 张")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    all_products = []
    report_lines = [f"# 转换诊断 · {xlsx.name}\n"]

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("WpsReserved"):
            continue
        ws = wb[sheet_name]
        # 站点标识: sheet 名(英国/德国)
        row_to_id = raw_dispimg_ids(xlsx, sheet_name)
        n_rows = 0
        n_img = 0
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=False):
            cells = row
            get = lambda i: cells[i].value if i < len(cells) else None
            sku = get(4)
            name = get(3)
            if not sku and not name:
                continue
            n_rows += 1
            rec = {"sheet": sheet_name}
            for idx, key in CORE_COLS.items():
                val = get(idx)
                if key == "时间_excel":
                    rec["时间"] = excel_serial_to_date(val)
                    continue
                if key in ("售价", "空运利润_$", "空运利润率", "海运利润_$",
                           "海运利润率", "长cm", "宽cm", "高cm", "毛重kg", "采购数量"):
                    rec[key] = clean_num(val)
                else:
                    rec[key] = val.strip() if isinstance(val, str) else val
            # 利润率转百分比显示
            for rk in ("空运利润率", "海运利润率"):
                if isinstance(rec.get(rk), (int, float)):
                    rec[rk + "_pct"] = round(rec[rk] * 100, 2)
            # 图片
            row_no = cells[0].row
            img_id = row_to_id.get(row_no)
            img_file = ""
            if img_id and img_id in id_to_bytes:
                data = id_to_bytes[img_id]
                ext = "png" if data[:8] == b"\x89PNG\r\n\x1a\n" else "jpg"
                safe_sku = re.sub(r"[^\w\-]", "_", str(sku or f"row{row_no}"))
                img_file = f"{safe_sku}.{ext}"
                (img_dir / img_file).write_bytes(data)
                n_img += 1
            rec["图片文件"] = img_file
            all_products.append(rec)
        report_lines.append(f"- {sheet_name}: {n_rows} 个产品, {n_img} 张图")
        print(f"[{sheet_name}] {n_rows} 产品, {n_img} 图")
    wb.close()

    # 写 JSONL
    jsonl_path = out / "products.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as f:
        for rec in all_products:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # 写 CSV
    csv_path = out / "products.csv"
    if all_products:
        keys = list(all_products[0].keys())
        # 保证所有键都在
        for rec in all_products:
            for k in rec:
                if k not in keys:
                    keys.append(k)
        with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
            w.writeheader()
            w.writerows(all_products)

    report_lines.append(f"\n合计: {len(all_products)} 个产品")
    report_lines.append(f"产出: {jsonl_path.name} / {csv_path.name} / images/")
    (out / "_convert_report.md").write_text("\n".join(report_lines), encoding="utf-8")
    print(f"[done] {len(all_products)} 产品 -> {out}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    convert(Path(args.xlsx), Path(args.out))


if __name__ == "__main__":
    main()
