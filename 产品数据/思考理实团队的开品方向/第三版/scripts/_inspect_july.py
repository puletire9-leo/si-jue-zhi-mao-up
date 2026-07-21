# -*- coding: utf-8 -*-
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

f = r"E:\项目\si-jue-zhi-mao-up\产品数据\产品表\理实产品开发表\7月份\理实产品对接表07.16.xlsx"
wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"\n=== {ws.title} dims={ws.max_row}x{ws.max_column} ===")
    rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    for i, r in enumerate(rows):
        vals = [str(c)[:22] if c is not None else "" for c in r]
        print(f"  row{i+1}:", vals)
wb.close()
