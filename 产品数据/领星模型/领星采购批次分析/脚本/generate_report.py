# -*- coding: utf-8 -*-
"""
脚本2: 生成备货率分析报表（可复用主脚本）

用途: 读取SKU清单，查询领星数据库，生成备货率+利润分析报表
      1个汇总Sheet（开发人分组，无毛利）+ N个明细Sheet（每月×开发人×国家）

配置: 修改下方 MONTHS_CONFIG 和 OUT_PATH 即可

核心口径:
    - 批次单位 = 采购单(PO order_sn) —— 真正下单采购才算备货
    - 批次边界 = 采购单下单时间(order_time)，同SKU同国同日多PO合并为1批
    - 有效单过滤 = 排除作废单 status NOT IN (-1,124) + item.is_delete=1
    - 国家归属 = 通过PO明细.plan_sn关联采购计划.creator_real_name (余江燕=DE, 其他=UK)
    - 利润归属 = 周表按marketplace归UK/DE, 各国独立按采购单下单日分Q1/Q2/Q3
    - 组合SKU = is_combo=1(算入分母,蓝底); 辅料 = is_aux=1(排除,灰底)

口径修正(2026-08-14): 之前用采购计划(PPG)算备货, 把"有计划无订单"的SKU也算成备货, 高估备货率。
    改为采购单(PO)口径后, 只有实际下过采购单的SKU才算有备货。

依赖: pip install pymysql openpyxl
数据库: 读 config/secrets 的 MYSQL_PASSWORD，不要写在脚本里
"""
import os, glob, datetime, openpyxl
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
import concurrent.futures, urllib.request
from query_database import get_conn

# 图片本地缓存目录（按ASIN命名，如 B0XXX.jpg）
IMG_DIR = r"E:\项目\si-jue-zhi-mao-up\tmp\lingxing_images"
EMBED_IMAGE = False   # 是否在明细第一列嵌入图片
IMG_SIZE = 60        # 嵌入图片边长(px)

# ============ 配置区 ============
SKU_BASE = r"E:\项目\si-jue-zhi-mao-up\产品数据\领星模型\领星采购批次分析\SKU清单"

# 月份配置: (月份名, SKU清单目录) —— 会自动扫描该目录下所有 {月份}新建sku-*.xlsx
MONTHS_CONFIG = [
    ("1-2月", os.path.join(SKU_BASE, "1-2月SKU明细")),
    ("3月", os.path.join(SKU_BASE, "3月SKU明细")),
    ("4月", os.path.join(SKU_BASE, "4月SKU明细")),
    ("5月", os.path.join(SKU_BASE, "5月SKU明细")),
    ("6月", os.path.join(SKU_BASE, "6月SKU明细")),
    ("7月", os.path.join(SKU_BASE, "7月SKU明细")),
]

OUT_PATH = r"E:\项目\si-jue-zhi-mao-up\产品数据\领星模型\领星采购批次分析\报表\1-7月备货分析_全部开发人.xlsx"
# ============ 配置区结束 ============

HEADERS_DETAIL = (["图片"] if EMBED_IMAGE else []) + [
           "SKU","中文名称","类型","批次数","各批采购单日期",
           "Q1量","Q2量","Q3量","备注",
           "Q1销量","Q1销售额","Q1毛利","Q1毛利率(%)",
           "Q2销量","Q2销售额","Q2毛利","Q2毛利率(%)",
           "Q3销量","Q3销售额","Q3毛利","Q3毛利率(%)",
           "汇总销量","汇总销售额","汇总毛利","汇总毛利率(%)"]

def get_dev_from_filename(fn):
    """从文件名提取开发人简称"""
    for dev in ["龙梦临","周沁仪","宋","陈杨","蒋舒","黄雨珊","张子轩","刘淼","韩金路","夏浩宇"]:
        if dev in fn:
            return dev
    return None

def style_header(ws, headers):
    is_detail = len(headers) >= 15
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor="C00000" if not is_detail else "4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    if is_detail:
        if EMBED_IMAGE:
            ws.column_dimensions["A"].width = 10   # 图片列
            ws.freeze_panes = "C2"
        else:
            ws.freeze_panes = "B2"
    else:
        ws.freeze_panes = "B2"

def img_path_of(asin):
    """按ASIN找本地图片路径，找不到返回None"""
    if not asin: return None
    for ext in (".jpg", ".jpeg", ".png", ".webp"):
        p = os.path.join(IMG_DIR, f"{asin}{ext}")
        if os.path.exists(p) and os.path.getsize(p) > 100:
            return p
    return None

def download_image(args):
    """下载单张图片到 IMG_DIR/{asin}.jpg"""
    asin, url = args
    if not url: return (asin, None)
    dst = os.path.join(IMG_DIR, f"{asin}.jpg")
    if os.path.exists(dst) and os.path.getsize(dst) > 100:
        return (asin, dst)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            data = r.read()
        if len(data) > 100:
            with open(dst, "wb") as f:
                f.write(data)
            return (asin, dst)
    except Exception:
        pass
    return (asin, None)

def pct(gp, amt): return round(gp/amt*100, 2) if amt and amt != 0 else None
def rnd(v): return round(v, 2) if v else None

def write_detail_row(ws, er, sku, ctry, prod, sku_country_batches, agg, sku_img=None):
    p = prod.get(sku, {})
    kind = p.get("kind", "")
    bl = sku_country_batches[sku][ctry]
    qrs = [b[2] for b in bl]
    dates = [str(b[0]) for b in bl]
    a = agg[sku][ctry]
    def grp(idx):
        v = a.get(idx, [0,0,0.0])
        return v[0] or None, rnd(v[1]), rnd(v[2]), pct(v[2], v[1])
    tot = [0,0,0.0]
    for k, v in a.items():
        if k == 99: continue
        tot[0]+=v[0]; tot[1]+=v[1]; tot[2]+=v[2]
    v99 = a.get(99, [0,0,0.0])
    tot[0]+=v99[0]; tot[1]+=v99[1]; tot[2]+=v99[2]
    # 第1列图片（嵌入），数据从第2列开始
    img_p = (sku_img or {}).get(sku)
    if EMBED_IMAGE and img_p:
        try:
            im = XLImage(img_p)
            im.width, im.height = IMG_SIZE, IMG_SIZE
            ws.add_image(im, f"A{er}")
            ws.row_dimensions[er].height = IMG_SIZE * 0.78
        except Exception:
            pass
    col = 2 if EMBED_IMAGE else 1
    ws.cell(row=er, column=col, value=sku); col+=1
    ws.cell(row=er, column=col, value=p.get("name")); col+=1
    ws.cell(row=er, column=col, value=kind); col+=1
    ws.cell(row=er, column=col, value=len(bl)); col+=1
    ws.cell(row=er, column=col, value=" / ".join(dates) if dates else ""); col+=1
    ws.cell(row=er, column=col, value=qrs[0] if len(qrs)>=1 else None); col+=1
    ws.cell(row=er, column=col, value=qrs[1] if len(qrs)>=2 else None); col+=1
    ws.cell(row=er, column=col, value=qrs[2] if len(qrs)>=3 else None); col+=1
    ws.cell(row=er, column=col, value="采购数量"); col+=1
    for idx in (1,2,3):
        for v in grp(idx):
            ws.cell(row=er, column=col, value=v); col+=1
    for v in (tot[0] or None, rnd(tot[1]), rnd(tot[2]), pct(tot[2], tot[1])):
        ws.cell(row=er, column=col, value=v); col+=1
    fill = None
    if kind == "组合": fill = "BDD7EE"
    elif kind == "辅料": fill = "D9D9D9"
    if fill:
        for c in range(1, len(HEADERS_DETAIL)+1):
            ws.cell(row=er, column=c).fill = PatternFill("solid", fgColor=fill)


def main():
    print("=== 备货率分析报表生成 ===")

    conn = get_conn()
    def q(sql, args=None):
        with conn.cursor() as c:
            c.execute(sql, args); return c.fetchall()

    # 1. 收集所有月份SKU: dev_month_skus[dev][month] = [skus]
    dev_month_skus = defaultdict(lambda: defaultdict(list))
    all_skus_set = set()
    month_names = []

    for month_name, dir_path in MONTHS_CONFIG:
        month_names.append(month_name)
        for f in glob.glob(os.path.join(dir_path, "*.xlsx")):
            dev = get_dev_from_filename(os.path.basename(f))
            if not dev: continue
            wb_r = openpyxl.load_workbook(f, read_only=True); ws_r = wb_r.active
            skus = []
            for row in ws_r.iter_rows(min_row=1, values_only=True):
                v = row[0]
                if v is None: continue
                sv = str(v).strip()
                if sv and any(c.isdigit() for c in sv):
                    skus.append(sv); all_skus_set.add(sv)
            wb_r.close()
            uniq, seen = [], set()
            for s in skus:
                if s not in seen: seen.add(s); uniq.append(s)
            dev_month_skus[dev][month_name] = uniq

    all_skus = list(all_skus_set)
    print(f"合计去重SKU: {len(all_skus)}")
    if not all_skus:
        conn.close(); print("无SKU，退出"); return

    ph = ",".join(["%s"]*len(all_skus))

    # 2. 主数据
    prod = {}
    for r in q(f"SELECT sku, product_name FROM lingxing_local_product WHERE sku IN ({ph})", all_skus):
        prod[r[0]] = {"name": r[1]}

    combo_skus, aux_skus = set(), set()
    for r in q(f"SELECT DISTINCT sku FROM lingxing_purchase_plan WHERE is_combo=1 AND sku IN ({ph})", all_skus):
        combo_skus.add(r[0])
    for r in q(f"SELECT DISTINCT sku FROM lingxing_purchase_plan WHERE is_aux=1 AND sku IN ({ph})", all_skus):
        aux_skus.add(r[0])
    for s in all_skus:
        if s in combo_skus: prod.setdefault(s,{})["kind"]="组合"
        elif s in aux_skus: prod.setdefault(s,{})["kind"]="辅料"
        else: prod.setdefault(s,{})["kind"]=""

    # 3. 备货批次 = 采购单(PO), 按 order_sn 去重
    #    口径修正(2026-08): 之前用采购计划(PPG)算备货, 把"有计划无订单"的SKU也算成备货, 高估了备货率
    #    只有真正下了采购单(PO)才算备货; 一个有效采购单 = 一个备货批次
    #    国家归属: 通过 PO明细.plan_sn 关联采购计划.creator_real_name (余江燕=DE, 其他=UK)
    #    订单数量: quantity_real(实际采购量)
    po_sql = f"""
    SELECT i.sku, o.order_sn, DATE(o.order_time) AS po_date,
           p.creator_real_name, SUM(i.quantity_real) AS qty_real
    FROM lingxing_purchase_order_item i
    JOIN lingxing_purchase_order o ON o.order_sn = i.order_sn
    LEFT JOIN lingxing_purchase_plan p ON p.plan_sn = i.plan_sn AND p.status = -2
    WHERE i.sku IN ({ph})
      AND o.status NOT IN (-1, 124)        -- 排除作废采购单
      AND (i.is_delete IS NULL OR i.is_delete = 0)
      AND o.order_time IS NOT NULL
    GROUP BY i.sku, o.order_sn, DATE(o.order_time), p.creator_real_name
    """
    po_rows = q(po_sql, all_skus)

    sku_country_date_po = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0, []])))
    for sku, osn, po_date, creator, qty in po_rows:
        if not po_date: continue
        # 国家归属: 优先用关联采购计划的创建人, 关联不上默认UK
        ctry = "DE" if creator == "余江燕" else "UK"
        sku_country_date_po[sku][ctry][po_date][0] += int(qty or 0)
        sku_country_date_po[sku][ctry][po_date][1].append(osn)

    # 转批次列表: 同SKU同国同日多PO合并为1批, 按日期排序
    sku_country_batches = defaultdict(lambda: defaultdict(list))
    for sku in sku_country_date_po:
        for ctry in sku_country_date_po[sku]:
            for dt in sorted(sku_country_date_po[sku][ctry].keys()):
                qty, osns = sku_country_date_po[sku][ctry][dt]
                sku_country_batches[sku][ctry].append((dt, ", ".join(set(osns)), qty, 0))

    # 4. 周表利润
    conn2 = get_conn()
    with conn2.cursor() as cur:
        cur.execute(f"""SELECT l.local_sku, l.marketplace, w.week_start, w.volume, w.amount, w.gross_profit
                        FROM lingxing_listing l JOIN lingxing_sku_weekly_performance w ON w.asin=l.asin
                        WHERE l.local_sku IN ({ph})""", all_skus)
        wk_rows = cur.fetchall()
    conn2.close()

    def ctry_of_mkt(m):
        return "UK" if m=="英国" else ("DE" if m=="德国" else None)
    def parse_d(s):
        if isinstance(s, datetime.date): return s
        if not s: return None
        try: return datetime.date.fromisoformat(str(s)[:10])
        except: return None

    agg = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0,0,0.0])))
    for sku, mkt, ws, vol, amt, gp in wk_rows:
        ctry = ctry_of_mkt(mkt)
        if not ctry: continue
        bl = sku_country_batches[sku][ctry]
        ws_d = parse_d(ws)
        idx = 0
        if ws_d is not None and bl:
            for i, b in enumerate(bl, 1):
                if b[0] is not None and ws_d >= b[0]: idx = i
                elif b[0] is not None and ws_d < b[0]: break
            if idx == 0: idx = 1
        elif not bl: idx = 99
        else: idx = 1
        agg[sku][ctry][idx][0] += vol or 0
        agg[sku][ctry][idx][1] += float(amt or 0)
        agg[sku][ctry][idx][2] += float(gp or 0)
    conn.close()

    def ctry_stat(skus, ctry):
        ge1=ge2=ge3=0
        for s in skus:
            if prod.get(s,{}).get("kind","")=="辅料": continue
            n = len(sku_country_batches[s][ctry])
            if n>=1: ge1+=1
            if n>=2: ge2+=1
            if n>=3: ge3+=1
        n_main = sum(1 for s in skus if prod.get(s,{}).get("kind","")!="辅料")
        return dict(n_main=n_main, ge1=ge1, ge2=ge2, ge3=ge3,
                    r2=(ge2/ge1 if ge1 else 0), r3=(ge3/ge1 if ge1 else 0))

    # 4.5 图片映射: sku -> 本地图片路径
    #   优先 listing.asin (lingxing_images缓存/small_image_url下载)
    #   兜底 local_product.pic_url (无listing的SKU, 按sku命名下载)
    sku_img = {}
    if EMBED_IMAGE:
        os.makedirs(IMG_DIR, exist_ok=True)
        conn3 = get_conn()
        with conn3.cursor() as cur:
            cur.execute(f"""SELECT local_sku, asin, small_image_url
                            FROM lingxing_listing WHERE local_sku IN ({ph})""", all_skus)
            listing_img_rows = cur.fetchall()
            cur.execute(f"""SELECT sku, pic_url FROM lingxing_local_product WHERE sku IN ({ph})""", all_skus)
            prod_pic = {r[0]: r[1] for r in cur.fetchall()}
        conn3.close()
        # 每个sku选一个asin(优先已有本地图的)
        sku_asin_url = {}
        for lsku, asin, url in listing_img_rows:
            if not asin: continue
            if lsku not in sku_asin_url or img_path_of(asin):
                sku_asin_url[lsku] = (asin, url)
        # 用asin本地缓存 / 攒下载任务(按asin命名)
        need_download = []   # (key, url) key=asin或sku
        for lsku, (asin, url) in sku_asin_url.items():
            p = img_path_of(asin)
            if p:
                sku_img[lsku] = p
            elif url:
                need_download.append((asin, url))
        # 无asin本地图 + 无listing 的SKU, 用 pic_url 按sku命名下载
        pic_task_sku = {}   # download key(sku) -> lsku
        for lsku in all_skus:
            if lsku in sku_img: continue
            if lsku in sku_asin_url and sku_asin_url[lsku][1]: continue  # 已在asin下载队列
            pu = prod_pic.get(lsku)
            if pu:
                # 先看本地有没有按sku命名的
                p = img_path_of(lsku)
                if p:
                    sku_img[lsku] = p
                else:
                    need_download.append((lsku, pu))
                    pic_task_sku[lsku] = lsku
        print(f"  图片: 本地命中 {len(sku_img)}, 待下载 {len(need_download)}")
        # 并发下载
        if need_download:
            asin2sku = defaultdict(list)
            for lsku, (asin, url) in sku_asin_url.items():
                asin2sku[asin].append(lsku)
            with concurrent.futures.ThreadPoolExecutor(max_workers=16) as ex:
                for key, dst in ex.map(download_image, need_download):
                    if not dst: continue
                    # key可能是asin(映射多个sku)或sku
                    if key in asin2sku:
                        for lsku in asin2sku[key]:
                            if lsku not in sku_img:
                                sku_img[lsku] = dst
                    if key in pic_task_sku:
                        sku_img[pic_task_sku[key]] = dst
        print(f"  图片: 最终有图SKU {len(sku_img)} / {len(all_skus)}")

    # 5. 生成Excel
    wb = Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet(title="汇总")
    style_header(ws_sum, ["开发人","月份","国家","主SKU","有采购","≥2批","≥3批","二次备货率","三次备货率"])

    def write_sum_row(er, dev, month, ctry, st, fill=None, bold=False, italic=False):
        ws_sum.cell(row=er, column=1, value=dev)
        ws_sum.cell(row=er, column=2, value=month)
        ws_sum.cell(row=er, column=3, value=ctry)
        ws_sum.cell(row=er, column=4, value=st["n_main"])
        ws_sum.cell(row=er, column=5, value=st["ge1"])
        ws_sum.cell(row=er, column=6, value=st["ge2"])
        ws_sum.cell(row=er, column=7, value=st["ge3"])
        ws_sum.cell(row=er, column=8, value=round(st["r2"]*100,1)).number_format='0.0"%"'
        ws_sum.cell(row=er, column=9, value=round(st["r3"]*100,1)).number_format='0.0"%"'
        if bold: ws_sum.cell(row=er, column=1).font = Font(bold=True)
        if italic: ws_sum.cell(row=er, column=1).font = Font(italic=True)
        if fill:
            for c in range(1,10):
                ws_sum.cell(row=er, column=c).fill = PatternFill("solid", fgColor=fill)

    er = 2
    # 按开发人分组
    for dev in sorted(dev_month_skus.keys()):
        for month in month_names:
            skus = dev_month_skus[dev].get(month, [])
            if not skus: continue
            for ctry in ("UK","DE"):
                st = ctry_stat(skus, ctry)
                if st["ge1"] == 0: continue
                write_sum_row(er, dev, month, ctry, st); er += 1
        # 该开发人跨月合计（仅当有多月数据时）
        all_dev = []
        for m in month_names: all_dev.extend(dev_month_skus[dev].get(m, []))
        all_dev = list(set(all_dev))
        month_span = f"{month_names[0]}-{month_names[-1]}合计" if len(month_names)>1 else f"{month_names[0]}"
        if len(month_names) > 1:
            for ctry in ("UK","DE"):
                st = ctry_stat(all_dev, ctry)
                if st["ge1"] == 0: continue
                write_sum_row(er, dev, month_span, ctry, st, fill="E7E6E6", italic=True); er += 1

    # 全局各月总计
    if len(dev_month_skus) > 1:
        for month in month_names:
            m_all = []
            for dev in dev_month_skus: m_all.extend(dev_month_skus[dev].get(month, []))
            m_all = list(set(m_all))
            for ctry in ("UK","DE"):
                st = ctry_stat(m_all, ctry)
                if st["ge1"] == 0: continue
                write_sum_row(er, "总计", month, ctry, st, fill="FFF2CC", bold=True); er += 1
        # 最终总计
        if len(month_names) > 1:
            for ctry in ("UK","DE"):
                st = ctry_stat(all_skus, ctry)
                if st["ge1"] == 0: continue
                month_span = f"{month_names[0]}-{month_names[-1]}合计"
                write_sum_row(er, "总计", month_span, ctry, st, fill="FFE699", bold=True); er += 1

    # 明细Sheet
    for month in month_names:
        for dev in sorted(dev_month_skus.keys()):
            skus = dev_month_skus[dev].get(month, [])
            if not skus: continue
            for ctry in ("UK","DE"):
                rows_skus = [s for s in skus if len(sku_country_batches[s][ctry]) > 0]
                if not rows_skus: continue
                ws_d = wb.create_sheet(title=f"{month}-{dev[:4]}-{ctry}")
                style_header(ws_d, HEADERS_DETAIL)
                rows_skus.sort(key=lambda s: (-len(sku_country_batches[s][ctry]), s))
                for i, sku in enumerate(rows_skus):
                    write_detail_row(ws_d, i+2, sku, ctry, prod, sku_country_batches, agg, sku_img)

    wb.save(OUT_PATH)
    print(f"已保存: {OUT_PATH}")
    print(f"Sheet总数: {len(wb.worksheets)}")


if __name__ == "__main__":
    main()
