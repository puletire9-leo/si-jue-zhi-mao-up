import openpyxl, subprocess

wb = openpyxl.load_workbook(r'e:\项目\si-jue-zhi-mao-up\scripts\八爪鱼新品榜asin\榜单抓数据处理\英国\榜单链接采榜单数据-英国.xlsx')
ws = wb.active
file_asins = set()
for row_idx in range(2, ws.max_row + 1):
    asin = ws.cell(row=row_idx, column=2).value
    if asin and str(asin).strip():
        file_asins.add(str(asin).strip())

result = subprocess.run([
    'docker', 'exec', 'dev-mysql', 'mysql', '-usijue', '-psijue123456', 'sijuelishi_dev',
    '-e', 'SELECT DISTINCT asin FROM competitor_products UNION SELECT DISTINCT parent_asin FROM competitor_products WHERE parent_asin IS NOT NULL AND parent_asin != ""'
], capture_output=True, text=True)

db_asins = set(line.strip() for line in result.stdout.strip().split('\n')[1:] if line.strip())
overlap = file_asins & db_asins
print(f'File ASINs: {len(file_asins)}')
print(f'DB (asin + parent_asin): {len(db_asins)}')
print(f'Overlap: {len(overlap)}')
print(f'File ASINs NOT in DB: {len(file_asins - db_asins)}')
