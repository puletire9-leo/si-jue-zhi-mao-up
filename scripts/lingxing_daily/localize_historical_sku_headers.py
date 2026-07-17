"""Localize existing user-facing historical SKU files without changing data rows."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
BASE_DIR = ROOT / "\u4ea7\u54c1\u6570\u636e" / "\u9886\u661f\u6570\u636eapi" / "\u9886\u661f25\u5e74\u523026\u5e746\u6708\u6240\u6709\u6570\u636e\uff0c\u4ee5\u6bcf\u6708\u6570\u636e" / "\u5386\u53f2SKU\u4e0a\u67b6\u57fa\u7840\u6570\u636e_2025-04\u81f32026-06"

LABELS = {
    "marketplace": "站点", "country": "国家", "store_name": "店铺", "sku": "SKU",
    "first_fba_month": "首次FBA可售观察月", "first_fba_window_start": "首次观察窗口开始日",
    "first_fba_window_end": "首次观察窗口结束日", "first_fba_max": "首次观察月最大FBA可售库存",
    "observed_active_months": "观察期内FBA可售月份数", "first_fba_status": "首次FBA可售判定",
    "first_observed_month": "首次观察月份", "first_observed_window_start": "首次观察窗口开始日",
    "first_observed_window_end": "首次观察窗口结束日", "ever_fba_available": "观察期内是否曾FBA可售",
    "presence_status": "月度观察状态", "mskus": "MSKU列表", "asins": "ASIN列表",
    "parent_asins": "父ASIN列表", "listing_tags": "Listing标签", "product_names": "品名", "titles": "商品标题",
    "first_fba_month": "首次FBA可售观察月", "new_store_sku_count": "新增店铺SKU数",
}
VALUE_LABELS = {
    "LEFT_CENSORED_AT_DATA_START": "数据起点已可售，真实首次上架月未知",
    "MONTHLY_FIRST_FBA_OBSERVED": "月度数据中首次观察到FBA可售",
    "NEVER_FBA_OBSERVED": "观察期内未出现FBA可售",
}


def localize_value(value):
    return VALUE_LABELS.get(value, value) if isinstance(value, str) else value


def localize_csv(path: Path) -> None:
    with path.open(encoding="utf-8-sig", newline="") as source:
        rows = list(csv.reader(source))
    if not rows:
        return
    rows[0] = [LABELS.get(value, value) for value in rows[0]]
    for row in rows[1:]:
        for index, value in enumerate(row):
            row[index] = localize_value(value)
    with path.open("w", encoding="utf-8-sig", newline="") as destination:
        csv.writer(destination).writerows(rows)


def localize_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.value = LABELS.get(cell.value, cell.value)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = localize_value(cell.value)
    workbook.save(path)


def main() -> None:
    overview = BASE_DIR / "01_\u5feb\u901f\u603b\u89c8" / "\u5386\u53f2SKU_\u6708\u5ea6\u65b0\u589e\u6c47\u603b_2025-04\u81f32026-06.csv"
    baseline = BASE_DIR / "02_\u5386\u53f2SKU\u4e0a\u67b6\u57fa\u7ebf"
    files = [
        overview,
        baseline / "\u5386\u53f2SKU_\u9996\u6b21FBA\u53ef\u552e\u6392\u5e03\u8868_2025-04\u81f32026-06.csv",
        baseline / "\u5386\u53f2SKU_\u5168\u91cf\u6708\u5ea6\u89c2\u5bdf\u8868_2025-04\u81f32026-06.csv",
    ]
    for path in files:
        localize_csv(path)
    for path in baseline.glob("*.xlsx"):
        localize_workbook(path)
    print("localized historical SKU headers")


if __name__ == "__main__":
    main()
