#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create the SKU-led Q1 -> ASIN -> first-month sales / Q2 / payback audit.

The purchase order is the source of truth for the first and second purchase
batches.  ASIN is a mapped business-result field; it must not filter the
purchase sample before SKU-level data quality has been audited.
"""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from python_calamine import load_workbook

from lingxing_base_access import dec, load_asin_baseline, mysql_env
from lingxing_model_paths import LEGACY_MODEL_ARCHIVE_ROOT, LINGXING_DATA_ROOT, MONTHLY_MODEL_DIR


ROOT = Path(__file__).resolve().parents[2]
LINGXING_ROOT = LINGXING_DATA_ROOT
ASIN_MODEL_ROOT = MONTHLY_MODEL_DIR
OUTPUT_ROOT = LEGACY_MODEL_ARCHIVE_ROOT / "SKU首批单量_ASIN映射_首月出单率_二批回正模型_第一版测试_2026-07-13"
MONTHLY_SOURCE = LINGXING_ROOT / "领星25年到26年6月所有数据，以每月数据"
DATA_CUTOFF = "2026-06"

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")


def add_month(month: str, offset: int = 1) -> str:
    year, value = map(int, month.split("-"))
    point = year * 12 + value - 1 + offset
    return f"{point // 12:04d}-{point % 12 + 1:02d}"


def load_monthly_sales(asins: dict[str, dict[str, Any]]) -> None:
    for path in sorted(MONTHLY_SOURCE.glob("*.xlsx")):
        match = re.search(r"(20\d{2}-\d{2})-\d{2}~", path.name)
        if not match:
            continue
        month = match.group(1)
        if month > DATA_CUTOFF:
            continue
        sheet = load_workbook(path).get_sheet_by_index(0)
        rows = sheet.iter_rows()
        headers = [str(value or "").strip() for value in next(rows)]
        required = ("ASIN", "国家", "销量")
        if not all(field in headers for field in required):
            raise ValueError(f"月表字段缺失：{path.name}")
        positions = {field: headers.index(field) for field in required}
        for row in rows:
            asin = str(row[positions["ASIN"]] or "").strip()
            record = asins.get(asin)
            if record is None:
                continue
            record["monthly_sales"][month] += dec(row[positions["销量"]])
            record["markets"].add(str(row[positions["国家"]] or "").strip())


def load_purchase_facts() -> tuple[dict[str, list[dict[str, Any]]], list[list[Any]]]:
    """Group every effective completed purchase by SKU + order number.

    This intentionally does *not* join to ASIN.  That join happens after the
    SKU first batch has been determined, preserving all team purchasing facts.
    """
    sql = """
        SELECT o.order_sn, COALESCE(o.order_time, o.create_time) AS purchase_time,
               i.sku, i.quantity_real, i.quantity_entry, i.item_id
        FROM lingxing_purchase_order_item i
        JOIN lingxing_purchase_order o ON o.order_sn = i.order_sn
        WHERE o.status = 9
          AND o.status_shipped = 3
          AND COALESCE(i.is_delete, 0) = 0
          AND COALESCE(i.quantity_real, 0) > 0
        ORDER BY COALESCE(o.order_time, o.create_time), o.order_sn, i.item_id
    """
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    monthly: Counter[str] = Counter()
    with pymysql.connect(**mysql_env()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(sql)
            for order_sn, purchase_time, sku, quantity_real, quantity_entry, item_id in cursor.fetchall():
                sku = str(sku or "").strip()
                if not sku:
                    continue
                key = (sku, str(order_sn))
                fact = grouped.setdefault(key, {
                    "采购单号": str(order_sn), "采购时间": purchase_time, "SKU": sku,
                    "实际下单量": Decimal(0), "实际入库量": Decimal(0), "子项数": 0,
                })
                fact["实际下单量"] += dec(quantity_real)
                fact["实际入库量"] += dec(quantity_entry)
                fact["子项数"] += 1
                if purchase_time:
                    monthly[purchase_time.strftime("%Y-%m")] += 1
    by_sku: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for fact in grouped.values():
        by_sku[str(fact["SKU"])].append(fact)
    for facts in by_sku.values():
        facts.sort(key=lambda item: (item["采购时间"] or datetime.max, item["采购单号"]))
    monthly_rows = [[month, count] for month, count in sorted(monthly.items())]
    return by_sku, monthly_rows


def load_payback(asins: dict[str, dict[str, Any]]) -> dict[str, dict[str, str]]:
    sql = """
        SELECT asin, currency_code, data_date, gross_profit
        FROM lingxing_profit_asin
        WHERE data_date >= %s AND data_date < %s
        ORDER BY asin, currency_code, data_date
    """
    daily: dict[tuple[str, str], list[tuple[date, Decimal]]] = defaultdict(list)
    with pymysql.connect(**mysql_env()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(sql, (date(2025, 4, 1), date(2026, 7, 1)))
            for asin, currency, data_date, gross_profit in cursor.fetchall():
                asin = str(asin or "").strip()
                if asin in asins and currency in {"GBP", "EUR"} and data_date:
                    daily[(asin, str(currency))].append((data_date, dec(gross_profit)))
    result: dict[str, dict[str, str]] = defaultdict(dict)
    for (asin, currency), rows in daily.items():
        start = str(asins[asin]["first_fba_month"] or "")
        if not start:
            continue
        cumulative = Decimal(0)
        for data_date, profit in rows:
            if data_date.strftime("%Y-%m") < start:
                continue
            cumulative += profit
            if cumulative > 0:
                result[asin][currency] = data_date.strftime("%Y-%m")
                break
    return result


def q1_bucket(value: Decimal | None) -> str:
    if value is None:
        return "无 Q1"
    if value <= 10:
        return "1-10"
    if value <= 15:
        return "11-15"
    if value <= 20:
        return "16-20"
    if value <= 30:
        return "21-30"
    return ">30"


def build_rows(
    asins: dict[str, dict[str, Any]], sku_to_asins: dict[str, set[str]],
    purchases: dict[str, list[dict[str, Any]]], payback: dict[str, dict[str, str]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    # Scope comes exclusively from the formal ASIN baseline, which has already
    # been Listing-tag filtered.  Never use the all-developer 15,200 SKU audit
    # table as this model's population.
    for sku in sorted(set(purchases) & set(sku_to_asins)):
        orders = purchases[sku]
        q1 = orders[0]
        q2 = orders[1] if len(orders) > 1 else None
        candidates = sorted(sku_to_asins.get(sku, set()))
        asin = candidates[0] if len(candidates) == 1 else ""
        record = asins.get(asin) if asin else None
        first_fba = str(record["first_fba_month"]) if record else ""
        model_start = str(record["model_start_month"]) if record else ""
        start_basis = str(record["start_basis"]) if record else ""
        first_full = add_month(model_start) if model_start else ""
        first_sales = record["monthly_sales"].get(first_full, Decimal(0)) if record and first_full <= DATA_CUTOFF else None
        q1_before_start = bool(q1["采购时间"] and model_start and q1["采购时间"].strftime("%Y-%m") <= model_start)
        if not candidates:
            sample_status = "SKU 未映射到标签范围 ASIN"
        elif len(candidates) > 1:
            sample_status = "SKU 映射多个标签范围 ASIN，待确认"
        elif not model_start:
            sample_status = "ASIN 缺少模型起算月"
        elif not q1_before_start:
            sample_status = "Q1 采购晚于 ASIN 模型起算月"
        elif first_full > DATA_CUTOFF:
            sample_status = "首个完整月观察不足"
        else:
            sample_status = "有效样本"
        candidate_developers = " | ".join(sorted({str(asins[item]["developer"]) for item in candidates}))
        rows.append({
            "SKU": sku, "开发人": str(record["developer"]) if record else candidate_developers,
            "映射 ASIN": asin,
            "ASIN 候选数": len(candidates), "ASIN 候选列表": " | ".join(candidates),
            "Q1 采购单号": q1["采购单号"], "Q1 采购时间": q1["采购时间"].strftime("%Y-%m-%d") if q1["采购时间"] else "",
            "Q1 实际下单量": q1["实际下单量"], "Q1 实际入库量": q1["实际入库量"], "Q1 分桶": q1_bucket(q1["实际下单量"]),
            "Q2 采购单号": q2["采购单号"] if q2 else "", "Q2 采购时间": q2["采购时间"].strftime("%Y-%m-%d") if q2 and q2["采购时间"] else "",
            "Q2 实际下单量": q2["实际下单量"] if q2 else None, "是否进入二批": "是" if q2 else "否",
            "首次 FBA 可售月": first_fba, "ASIN 模型起算月": model_start,
            "ASIN 模型起算依据": start_basis, "首个完整观察月": first_full,
            "首月销量": first_sales, "首月平均日销量": first_sales / Decimal(30) if first_sales is not None else None,
            "ASIN 国家集合": " | ".join(sorted(record["markets"])) if record else "",
            "GBP 财务回正月": payback.get(asin, {}).get("GBP", "") if asin else "",
            "EUR 财务回正月": payback.get(asin, {}).get("EUR", "") if asin else "",
            "最新 Listing 标签": str(record["label"]) if record else "", "样本状态": sample_status,
        })
    return rows


def summary_by_bucket(rows: list[dict[str, Any]]) -> list[list[Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row["样本状态"] == "有效样本":
            groups[str(row["Q1 分桶"])].append(row)
    output: list[list[Any]] = []
    for bucket in ("1-10", "11-15", "16-20", "21-30", ">30"):
        group = groups.get(bucket, [])
        q1_total = sum((dec(row["Q1 实际下单量"]) for row in group), Decimal(0))
        sales_total = sum((dec(row["首月销量"]) for row in group), Decimal(0))
        q2_count = sum(row["是否进入二批"] == "是" for row in group)
        gbp_positive = sum(bool(row["GBP 财务回正月"]) for row in group)
        eur_positive = sum(bool(row["EUR 财务回正月"]) for row in group)
        output.append([
            bucket, len(group), q1_total, q1_total / len(group) if group else None, sales_total,
            sales_total / (len(group) * 30) if group else None, q2_count, q2_count / len(group) if group else None,
            gbp_positive, gbp_positive / len(group) if group else None, eur_positive, eur_positive / len(group) if group else None,
        ])
    return output


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    headers = list(rows[0]) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def style_table(sheet: Any, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"].font = Font(bold=True, size=14, color="0B3D47")
    sheet.append(headers)
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    for row in rows:
        sheet.append([float(value) if isinstance(value, Decimal) else value for value in row])
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(2, sheet.max_row)}"
    sheet.sheet_view.showGridLines = False
    for number, header in enumerate(headers, 1):
        letter = get_column_letter(number)
        sheet.column_dimensions[letter].width = min(max(len(header) + 4, 13), 36)
        if "率" in header:
            for cell in sheet[letter][2:]:
                cell.number_format = "0.00%;[Red]-0.00%"
        elif any(word in header for word in ("销量", "量", "数")):
            for cell in sheet[letter][2:]:
                cell.number_format = "#,##0.00;[Red]-#,##0.00"


def write_outputs(rows: list[dict[str, Any]], monthly_purchase: list[list[Any]]) -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    write_csv(OUTPUT_ROOT / "02_SKU_Q1映射ASIN与首月结果明细.csv", rows)
    exception_rows = [row for row in rows if row["样本状态"] in {"SKU 未映射到标签范围 ASIN", "SKU 映射多个标签范围 ASIN，待确认"}]
    write_csv(OUTPUT_ROOT / "03_SKU_ASIN映射异常.csv", exception_rows)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "总览"
    statuses = Counter(str(row["样本状态"]) for row in rows)
    overview_rows = [
        ["标签范围 ASIN 已关联的有效采购 SKU", len(rows)],
        ["SKU 唯一映射标签范围 ASIN", sum(row["ASIN 候选数"] == 1 for row in rows)],
        ["SKU 映射多个标签范围 ASIN", statuses["SKU 映射多个标签范围 ASIN，待确认"]],
        ["使用 FBA 可售起算", sum("FBA可售" in str(row["ASIN 模型起算依据"]) for row in rows)],
        ["使用创建时间兜底", sum("创建时间" in str(row["ASIN 模型起算依据"]) for row in rows)],
        ["Q1 不晚于 ASIN 模型起算月", sum(row["样本状态"] in {"有效样本", "首个完整月观察不足"} for row in rows)],
        ["可用于 Q1 首月模型的有效 SKU", statuses["有效样本"]],
        ["首个完整月观察不足", statuses["首个完整月观察不足"]],
    ]
    style_table(overview, "SKU 首批单量—ASIN 映射—首月销量—二批/回正：第一版测试", ["指标", "数量"], overview_rows)
    overview.append([])
    overview.append(["先读这一页"])
    overview["A12"].fill = NOTE_FILL
    overview["A12"].font = Font(bold=True)
    overview.append(["模型范围", "仅使用正式《ASIN模型分析起算月基准》中的 Listing 标签范围 ASIN；不使用“8 位开发人的全部本地 SKU”作为范围。"])
    overview.append(["模型入口", "在标签范围 ASIN 的基准 SKU 中寻找有效采购记录；每个 SKU 的首个有效完成采购单为 Q1。"])
    overview.append(["ASIN 起算", "首次 FBA 可售月优先；全期未观察到 FBA 可售时，商品信息创建时间所在月兜底。首月销量按该起算月之后的首个完整自然月取数。"])
    overview.append(["采购覆盖", "数据库当前仅有 2026-01 至 2026-07 有效完成采购；2025 无采购历史，不能用于 Q1 训练。"])
    overview.append(["首月", "首次 FBA 可售月之后的首个完整自然月；首月平均日销量 = 首月销量 ÷ 30。销量口径，不使用订单量。"])
    overview.append(["二批", "同一 SKU 的第二个有效完成采购单；存在即记为进入二批。"])
    overview.append(["回正", "累计领星结算毛利润从首次 FBA 可售月起首次大于 0；GBP、EUR 分开，不能视为资金回本。"])
    overview.column_dimensions["A"].width = 34
    overview.column_dimensions["B"].width = 110

    bucket_sheet = workbook.create_sheet("Q1分桶汇总")
    bucket_headers = ["Q1 分桶", "有效 SKU 数", "Q1 下单量合计", "平均 Q1 下单量", "首月销量合计", "首月平均日销量", "进入二批 SKU 数", "二批率", "GBP 财务回正 SKU 数", "GBP 财务回正率", "EUR 财务回正 SKU 数", "EUR 财务回正率"]
    style_table(bucket_sheet, "仅有效样本：Q1 首批量分桶表现", bucket_headers, summary_by_bucket(rows))

    detail_sheet = workbook.create_sheet("SKU明细")
    detail_headers = list(rows[0]) if rows else []
    style_table(detail_sheet, "每个团队采购 SKU 的 Q1、ASIN 映射、首月销量、Q2 与回正明细", detail_headers, [[row[header] for header in detail_headers] for row in rows])

    quality_sheet = workbook.create_sheet("数据质量")
    quality_rows = [[name, count] for name, count in sorted(statuses.items())]
    quality_rows += [[f"采购子项月份：{month}", count] for month, count in monthly_purchase]
    style_table(quality_sheet, "样本状态与有效完成采购子项月份覆盖", ["项目", "数量"], quality_rows)

    notes = workbook.create_sheet("口径说明")
    notes_rows = [
        ["模型范围", "正式 ASIN 模型分析起算月基准中已完成 Listing 标签筛选的 ASIN；不扩展到全部开发人本地 SKU。"],
        ["模型主键", "采购层为 SKU。每个标签范围 ASIN 的基准 SKU 中，第一个有效完成采购单为 Q1，第二个为 Q2。"],
        ["有效完成采购", "采购单 status=9、status_shipped=3，子项未删除且 quantity_real>0。Q1 使用实际下单量 quantity_real，实际入库量 quantity_entry 仅作到货核查。"],
        ["样本顺序", "标签范围 ASIN 月度基准 → 基准 SKU 的 Q1/Q2 → ASIN 的 FBA/创建时间起算月 → 首月销量/利润结果。SKU 一对多 ASIN 不强行合并。"],
        ["有效样本", "SKU 唯一映射标签范围 ASIN；Q1 所在月不晚于 ASIN 模型起算月；且首个完整自然月不晚于 2026-06。模型起算月按 FBA 可售优先、创建时间兜底。"],
        ["限制", "采购事实从 2026-01 才开始；3–6 月为主要样本窗口。2026-06 首次可售的 ASIN 没有完整首月，暂不进入首月模型。"],
    ]
    style_table(notes, "口径与限制", ["主题", "说明"], notes_rows)
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 120
    workbook.save(OUTPUT_ROOT / "01_SKU_Q1首批单量模型_第一版测试.xlsx")

    (OUTPUT_ROOT / "00_先看这里.md").write_text(
        "# SKU 首批单量—ASIN 映射—首月销量—二批/回正模型（第一版测试）\n\n"
        "本版本直接复用已完成 Listing 标签筛选的 **ASIN 模型分析起算月基准**："
        "FBA 可售首现优先、商品信息创建时间兜底。以其基准 SKU 的首个有效完成采购单为 Q1，"
        "再读取同一 ASIN 的首月销量和财务结果。\n\n"
        "## 阅读顺序\n\n"
        "1. `01_SKU_Q1首批单量模型_第一版测试.xlsx` 的“总览”和“Q1分桶汇总”。\n"
        "2. 需要核查单品时，筛选“SKU明细”。\n"
        "3. 映射异常只在 `03_SKU_ASIN映射异常.csv` 审查，不强行归并。\n\n"
        "## 关键限制\n\n"
        "- 采购事实当前从 2026-01 才开始，2025 没有采购历史；3–6 月才是主要训练窗口。\n"
        "- 首月是首次 FBA 可售后的首个完整自然月；首月平均日销量 = 首月销量 ÷ 30。\n"
        "- 财务回正是累计领星结算毛利润首次大于 0，GBP/EUR 独立，不等同于资金回本。\n",
        encoding="utf-8",
    )


def main() -> None:
    asins, sku_to_asins, _sku_to_developers = load_asin_baseline()
    load_monthly_sales(asins)
    purchases, monthly_purchase = load_purchase_facts()
    payback = load_payback(asins)
    rows = build_rows(asins, sku_to_asins, purchases, payback)
    write_outputs(rows, monthly_purchase)
    print(f"团队有效采购 SKU：{len(rows)}")
    print(f"有效 Q1 首月样本：{sum(row['样本状态'] == '有效样本' for row in rows)}")
    print(f"输出目录：{OUTPUT_ROOT}")


if __name__ == "__main__":
    main()
