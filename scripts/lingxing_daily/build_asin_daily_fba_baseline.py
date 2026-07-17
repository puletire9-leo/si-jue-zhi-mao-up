#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build a compact ASIN FBA-available observation baseline from daily fast-test CSVs."""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DAILY_ROOT = ROOT / "产品数据" / "领星数据api" / "领星26年1到6月所有数据所有"
FAST_MANIFEST = DAILY_ROOT / "派生日数据_v1" / "manifest.json"
OUTPUT_DIR = DAILY_ROOT / "ASIN_FBA可售时间日基准_2026-01至2026-06"
OUTPUT_FILE = OUTPUT_DIR / "ASIN_FBA可售时间日基准.csv"
ANNUAL_SOURCE_IDS = ("934061734342180864", "934062080035405824")
TEAM_DEVELOPERS = {"刘淼", "周沁仪", "宋凤莉", "张子轩", "蒋舒", "陈杨", "黄雨珊", "龙梦临"}


def number(value: str) -> Decimal:
    try:
        return Decimal(value or "0")
    except InvalidOperation:
        return Decimal(0)


def joined(values: set[str]) -> str:
    return " | ".join(sorted(value for value in values if value))


def load_annual_asin_base() -> dict[str, dict[str, set[str]]]:
    files = []
    for source_id in ANNUAL_SOURCE_IDS:
        files.extend(ROOT.rglob(f"*{source_id}.xlsx"))
    base: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for path in files:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        positions = {field: headers.index(field) for field in ("ASIN", "SKU", "店铺", "开发人", "创建时间", "listing标签")}
        for row in rows:
            asin = str(row[positions["ASIN"]] or "").strip()
            created_at = str(row[positions["创建时间"]] or "").strip()
            # Match the verified legacy baseline: only ASINs created from 2025-04 onward.
            if not asin or created_at[:7] < "2025-04":
                continue
            for field, position in positions.items():
                if field != "ASIN":
                    base[asin][field].add(str(row[position] or "").strip())
        workbook.close()
    return base


def build_baseline() -> tuple[int, int, str, str]:
    base = load_annual_asin_base()
    manifest = json.loads(FAST_MANIFEST.read_text(encoding="utf-8"))
    daily = defaultdict(lambda: {
        "first_observed": "", "last_observed": "", "rows": 0, "fba_rows": 0,
        "first_fba": "", "first_stores": set(), "first_countries": set(), "first_skus": set(), "first_fba_total": Decimal(0),
    })
    coverage_start = ""
    coverage_end = ""
    for relative in manifest["modelFiles"].values():
        path = FAST_MANIFEST.parent / relative
        with path.open(encoding="utf-8-sig", newline="") as source:
            for row in csv.DictReader(source):
                asin = row["asin"].strip()
                if asin not in base:
                    continue
                data_date = row["data_date"]
                point = daily[asin]
                point["rows"] += 1
                point["first_observed"] = min(point["first_observed"] or data_date, data_date)
                point["last_observed"] = max(point["last_observed"], data_date)
                coverage_start = min(coverage_start or data_date, data_date)
                coverage_end = max(coverage_end, data_date)
                available = number(row["fba_available"])
                if available <= 0:
                    continue
                point["fba_rows"] += 1
                if not point["first_fba"] or data_date < point["first_fba"]:
                    point["first_fba"] = data_date
                    point["first_stores"] = {row["store_name"]}
                    point["first_countries"] = {row["country"]}
                    point["first_skus"] = {row["sku"]}
                    point["first_fba_total"] = available
                elif data_date == point["first_fba"]:
                    point["first_stores"].add(row["store_name"])
                    point["first_countries"].add(row["country"])
                    point["first_skus"].add(row["sku"])
                    point["first_fba_total"] += available

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    columns = [
        "ASIN", "基准SKU", "基准店铺", "开发人", "创建时间", "最新Listing标签", "是否目标8开发人",
        "日级观察开始日", "日级观察结束日", "日级记录行数", "首次观察到FBA可售日", "首日FBA可售店铺", "首日FBA可售国家",
        "首日FBA可售SKU", "首日FBA可售库存合计", "FBA可售日级记录行数", "日级FBA观察状态", "首日观察说明", "日级数据覆盖截止日",
    ]
    observed_fba = 0
    with OUTPUT_FILE.open("w", encoding="utf-8-sig", newline="") as destination:
        writer = csv.DictWriter(destination, fieldnames=columns)
        writer.writeheader()
        for asin, identity in sorted(base.items()):
            point = daily[asin]
            first_fba = point["first_fba"]
            if first_fba:
                observed_fba += 1
                state = "已观察到FBA可售"
                note = "观察窗口左截断，真实首次可售日可能更早" if first_fba == coverage_start else "日级首次观察到FBA可售，不等同于真实首次上架日"
            elif point["rows"]:
                state = "有日级记录但未观察到FBA可售"
                note = "观察窗口内 FBA-可售始终为 0"
            else:
                state = "日级数据无此ASIN"
                note = "该 ASIN 未出现在 2026-01 至 2026-06 日级数据"
            developers = identity["开发人"]
            writer.writerow({
                "ASIN": asin, "基准SKU": joined(identity["SKU"]), "基准店铺": joined(identity["店铺"]), "开发人": joined(developers),
                "创建时间": joined(identity["创建时间"]), "最新Listing标签": joined(identity["listing标签"]),
                "是否目标8开发人": "是" if developers & TEAM_DEVELOPERS else "否", "日级观察开始日": point["first_observed"],
                "日级观察结束日": point["last_observed"], "日级记录行数": point["rows"], "首次观察到FBA可售日": first_fba,
                "首日FBA可售店铺": joined(point["first_stores"]), "首日FBA可售国家": joined(point["first_countries"]),
                "首日FBA可售SKU": joined(point["first_skus"]), "首日FBA可售库存合计": point["first_fba_total"],
                "FBA可售日级记录行数": point["fba_rows"], "日级FBA观察状态": state, "首日观察说明": note,
                "日级数据覆盖截止日": coverage_end,
            })
    return len(base), observed_fba, coverage_start, coverage_end


if __name__ == "__main__":
    base_count, fba_count, start, end = build_baseline()
    print(f"annual_asins={base_count} observed_fba_asins={fba_count} coverage={start}~{end} output={OUTPUT_FILE}")
