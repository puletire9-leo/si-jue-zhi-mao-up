#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the ASIN monthly FBA-inventory first-observation baseline."""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from python_calamine import load_workbook as load_calamine_workbook

from lingxing_model_paths import (
    COHORT_BASE_DIR,
    FBA_INVENTORY_BASELINE,
    LINGXING_DATA_ROOT,
    NO_FBA_INVENTORY_ASINS,
)


ROOT = Path(__file__).resolve().parents[2]
ANNUAL_SOURCE_IDS = ("934061734342180864", "934062080035405824")
WINDOW = re.compile(r"(20\d{2}-\d{2})-\d{2}~")

# 原始 Excel 保持在数据源目录；公共基础事实统一写入“领星模型/基础统一表”。
MONTHLY_SOURCE_DIR = LINGXING_DATA_ROOT / "领星25年到26年6月所有数据，以每月数据"
OUTPUT_DIR = COHORT_BASE_DIR
OUTPUT_FILE = FBA_INVENTORY_BASELINE
NO_INVENTORY_ASIN_FILE = NO_FBA_INVENTORY_ASINS


def number(value: object) -> Decimal:
    try:
        return Decimal(str(value or "0"))
    except InvalidOperation:
        return Decimal(0)


def joined(values: set[str]) -> str:
    return " | ".join(sorted(value for value in values if value))


def annual_asin_base() -> dict[str, dict[str, set[str]]]:
    files = []
    for source_id in ANNUAL_SOURCE_IDS:
        files.extend(ROOT.rglob(f"*{source_id}.xlsx"))
    base: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for path in files:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        positions = {field: headers.index(field) for field in ("ASIN", "SKU", "店铺", "开发人", "创建时间", "listing标签")}
        for row in rows:
            asin = str(row[positions["ASIN"]] or "").strip()
            created_at = str(row[positions["创建时间"]] or "").strip()
            if not asin or created_at[:7] < "2025-04":
                continue
            for field, position in positions.items():
                if field != "ASIN":
                    base[asin][field].add(str(row[position] or "").strip())
        workbook.close()
    return base


def update_first(point, month: str, row, positions, field: str, prefix: str) -> None:
    if point[prefix] and month > point[prefix]:
        return
    if not point[prefix] or month < point[prefix]:
        point[prefix] = month
        point[f"{prefix}_stores"] = {str(row[positions["店铺"]] or "").strip()}
        point[f"{prefix}_countries"] = {str(row[positions["国家"]] or "").strip()}
        point[f"{prefix}_skus"] = {str(row[positions["SKU"]] or "").strip()}
        point[f"{prefix}_total"] = number(row[positions[field]])
    else:
        point[f"{prefix}_stores"].add(str(row[positions["店铺"]] or "").strip())
        point[f"{prefix}_countries"].add(str(row[positions["国家"]] or "").strip())
        point[f"{prefix}_skus"].add(str(row[positions["SKU"]] or "").strip())
        point[f"{prefix}_total"] += number(row[positions[field]])


def build() -> tuple[int, int, int]:
    base = annual_asin_base()
    points = defaultdict(lambda: {
        "first_inventory": "", "first_inventory_stores": set(), "first_inventory_countries": set(), "first_inventory_skus": set(), "first_inventory_total": Decimal(0),
        "first_available": "", "first_available_stores": set(), "first_available_countries": set(), "first_available_skus": set(), "first_available_total": Decimal(0),
    })
    files = []
    for path in MONTHLY_SOURCE_DIR.glob("*.xlsx"):
        match = WINDOW.search(path.name)
        if match:
            files.append((match.group(1), path))
    for month, path in sorted(files):
        rows = load_calamine_workbook(path).get_sheet_by_index(0).iter_rows()
        headers = [str(value or "") for value in next(rows)]
        fields = ("ASIN", "SKU", "店铺", "国家", "FBA库存", "FBA-可售")
        positions = {field: headers.index(field) for field in fields}
        for row in rows:
            asin = str(row[positions["ASIN"]] or "").strip()
            if asin not in base:
                continue
            point = points[asin]
            if number(row[positions["FBA库存"]]) > 0:
                update_first(point, month, row, positions, "FBA库存", "first_inventory")
            if number(row[positions["FBA-可售"]]) > 0:
                update_first(point, month, row, positions, "FBA-可售", "first_available")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    headers = [
        "ASIN", "基准SKU", "基准店铺", "开发人", "创建时间", "最新Listing标签", "首次观察到FBA库存月", "库存首现店铺", "库存首现国家", "库存首现SKU", "库存首现月FBA库存合计",
        "首次观察到FBA可售月", "可售首现店铺", "可售首现国家", "可售首现SKU", "可售首现月FBA可售合计", "月级FBA库存观察状态", "数据覆盖截止月",
    ]
    inventory_count = 0
    available_count = 0
    no_inventory_asins: list[str] = []
    with OUTPUT_FILE.open("w", encoding="utf-8-sig", newline="") as destination:
        writer = csv.DictWriter(destination, fieldnames=headers)
        writer.writeheader()
        for asin, identity in sorted(base.items()):
            point = points[asin]
            if point["first_inventory"]:
                inventory_count += 1
                status = "已观察到FBA库存"
            else:
                status = "月表未观察到FBA库存"
                no_inventory_asins.append(asin)
            if point["first_available"]:
                available_count += 1
            writer.writerow({
                "ASIN": asin, "基准SKU": joined(identity["SKU"]), "基准店铺": joined(identity["店铺"]), "开发人": joined(identity["开发人"]),
                "创建时间": joined(identity["创建时间"]), "最新Listing标签": joined(identity["listing标签"]),
                "首次观察到FBA库存月": point["first_inventory"], "库存首现店铺": joined(point["first_inventory_stores"]),
                "库存首现国家": joined(point["first_inventory_countries"]), "库存首现SKU": joined(point["first_inventory_skus"]), "库存首现月FBA库存合计": point["first_inventory_total"],
                "首次观察到FBA可售月": point["first_available"], "可售首现店铺": joined(point["first_available_stores"]),
                "可售首现国家": joined(point["first_available_countries"]), "可售首现SKU": joined(point["first_available_skus"]), "可售首现月FBA可售合计": point["first_available_total"],
                "月级FBA库存观察状态": status, "数据覆盖截止月": "2026-06",
            })
    NO_INVENTORY_ASIN_FILE.write_text("\n".join(no_inventory_asins) + "\n", encoding="utf-8")
    written_count = sum(1 for line in NO_INVENTORY_ASIN_FILE.read_text(encoding="utf-8").splitlines() if line)
    if written_count != len(no_inventory_asins):
        raise RuntimeError(f"未库存 ASIN txt 行数异常：期望 {len(no_inventory_asins)}，实际 {written_count}")
    return len(base), inventory_count, available_count, written_count


if __name__ == "__main__":
    total, inventory, available, no_inventory = build()
    print(f"annual_asins={total} fba_inventory_asins={inventory} fba_available_asins={available} no_inventory_asins={no_inventory} output={OUTPUT_FILE}")
