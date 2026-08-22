#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the first ASIN Q1 -> first-month sales -> Q2/payback audit dataset.

The model entity is ASIN. Purchase SKU is only a source mapping key: a purchase
item is included only when its SKU maps to exactly one team ASIN in the ASIN
baseline. Ambiguous mappings remain in a separate quality file.
"""

from __future__ import annotations

import csv
import os
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pymysql
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from python_calamine import load_workbook


ROOT = Path(__file__).resolve().parents[2]
LINGXING_ROOT = ROOT / "产品数据" / "领星数据api"
MODEL_ROOT = LINGXING_ROOT / "ASIN首批单量_首月出单率_二批回正模型_第一版测试_2026-07-13"
MONTHLY_SOURCE = LINGXING_ROOT / "领星25年到26年6月所有数据，以每月数据"
BASELINE = next(
    ROOT.rglob("ASIN_FBA可售优先_商品信息创建时间兜底_模型分析起算月基准_2025-04至2026-06.csv"),
    None,
)
TEAM_DEVELOPERS = frozenset({"蒋舒", "陈杨", "宋凤莉", "刘淼", "龙梦临", "周沁仪", "张子轩", "黄雨珊"})
DATA_START = "2025-04"
DATA_CUTOFF = "2026-06"
MONTHS = [f"2025-{month:02d}" for month in range(4, 13)] + [f"2026-{month:02d}" for month in range(1, 7)]

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
NEGATIVE_FILL = PatternFill("solid", fgColor="F4CCCC")
POSITIVE_FILL = PatternFill("solid", fgColor="D9EAD3")


def decimal(value: object) -> Decimal:
    try:
        return Decimal(str(value or "").strip() or "0")
    except InvalidOperation:
        return Decimal(0)


def add_month(month: str, offset: int = 1) -> str:
    year, value = map(int, month.split("-"))
    point = year * 12 + value - 1 + offset
    return f"{point // 12:04d}-{point % 12 + 1:02d}"


def split_values(value: str) -> set[str]:
    return {item.strip() for item in (value or "").split(" | ") if item.strip()}


def team_developer(value: str) -> str | None:
    names = {item.strip() for item in (value or "").replace("，", ",").split(",") if item.strip()}
    matched = names & TEAM_DEVELOPERS
    if len(matched) == 1 and len(names) == 1:
        return next(iter(matched))
    if matched:
        return "多人归属待确认"
    return None


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        return list(csv.DictReader(source))


def mysql_env() -> dict[str, object]:
    values: dict[str, str] = {}
    for path in (
        ROOT / ".env",
        ROOT / "config/public/prod.env",
        ROOT / "config/secrets/prod.env",
    ):
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            text = line.strip()
            if text and not text.startswith("#") and "=" in text:
                key, value = text.split("=", 1)
                values[key.strip()] = value.strip().strip('"').strip("'")
    values.update({key: value for key, value in os.environ.items() if value})
    host = values.get("MYSQL_HOST", "127.0.0.1")
    if host == "mysql":
        host = values.get("MYSQL_HOST_EXTERNAL", "127.0.0.1")
    return {
        "host": host,
        "port": int(values.get("MYSQL_PORT_EXTERNAL", values.get("MYSQL_PORT", "3310"))),
        "user": values.get("MYSQL_USERNAME", values.get("MYSQL_USER", "sijue")),
        "password": values["MYSQL_PASSWORD"],
        "database": values.get("MYSQL_DATABASE", "sijuelishi"),
        "charset": "utf8mb4",
    }


def load_asin_baseline() -> tuple[dict[str, dict[str, object]], dict[str, set[str]]]:
    if BASELINE is None:
        raise FileNotFoundError("未找到 ASIN FBA可售优先模型起算月基准")
    asins: dict[str, dict[str, object]] = {}
    sku_to_asins: dict[str, set[str]] = defaultdict(set)
    for row in read_csv(BASELINE):
        developer = team_developer(row.get("开发人", ""))
        asin = row.get("ASIN", "").strip()
        if not asin or developer is None:
            continue
        first_fba_month = row.get("首次观察到FBA可售月", "").strip()
        asins[asin] = {
            "asin": asin,
            "developer": developer,
            "skus": split_values(row.get("基准SKU", "")),
            "first_fba_month": first_fba_month,
            "model_start_month": row.get("模型分析起算月", "").strip(),
            "start_basis": row.get("模型分析起算依据", "").strip(),
            "label": row.get("最新Listing标签", "").strip(),
            "status": "淘汰" if "欧洲精铺2025淘汰" in row.get("最新Listing标签", "") else "留存",
            "monthly_sales": defaultdict(Decimal),
            "monthly_available": defaultdict(Decimal),
            "markets": set(),
        }
        for sku in asins[asin]["skus"]:
            sku_to_asins[sku].add(asin)
    return asins, sku_to_asins


def load_monthly_sales(asins: dict[str, dict[str, object]]) -> None:
    for path in sorted(MONTHLY_SOURCE.glob("*.xlsx")):
        name = path.name
        month = name[name.find("20"):name.find("20") + 7]
        if month not in MONTHS:
            continue
        rows = load_workbook(path).get_sheet_by_index(0).iter_rows()
        headers = [str(value or "") for value in next(rows)]
        fields = ("ASIN", "国家", "销量", "FBA-可售")
        positions = {field: headers.index(field) for field in fields}
        for row in rows:
            asin = str(row[positions["ASIN"]] or "").strip()
            record = asins.get(asin)
            if record is None:
                continue
            record["monthly_sales"][month] += decimal(row[positions["销量"]])
            record["monthly_available"][month] += decimal(row[positions["FBA-可售"]])
            record["markets"].add(str(row[positions["国家"]] or "").strip())


def load_purchase_facts(
    sku_to_asins: dict[str, set[str]],
) -> tuple[dict[str, list[dict[str, object]]], list[dict[str, object]]]:
    sql = """
        SELECT o.order_sn, COALESCE(o.order_time, o.create_time),
               i.sku, i.quantity_real, i.quantity_entry
        FROM lingxing_purchase_order_item i
        JOIN lingxing_purchase_order o ON o.order_sn = i.order_sn
        WHERE o.status = 9
          AND o.status_shipped = 3
          AND COALESCE(i.is_delete, 0) = 0
          AND COALESCE(i.quantity_real, 0) > 0
        ORDER BY COALESCE(o.order_time, o.create_time), o.order_sn, i.item_id
    """
    by_asin_order: dict[tuple[str, str], dict[str, object]] = {}
    quality_rows: list[dict[str, object]] = []
    with pymysql.connect(**mysql_env()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(sql)
            for order_sn, order_time, sku, quantity_real, quantity_entry in cursor.fetchall():
                value = str(sku or "").strip()
                candidates = sorted(sku_to_asins.get(value, set()))
                if len(candidates) != 1:
                    quality_rows.append({
                        "采购单号": order_sn,
                        "采购时间": order_time.isoformat(sep=" ") if order_time else "",
                        "SKU": value,
                        "quantity_real": quantity_real or 0,
                        "quantity_entry": quantity_entry or 0,
                        "映射状态": "SKU未映射到团队ASIN" if not candidates else "SKU映射多个团队ASIN",
                        "候选ASIN": " | ".join(candidates),
                    })
                    continue
                asin = candidates[0]
                key = (asin, str(order_sn))
                group = by_asin_order.setdefault(key, {
                    "order_sn": str(order_sn),
                    "order_time": order_time,
                    "quantity_real": Decimal(0),
                    "quantity_entry": Decimal(0),
                    "skus": set(),
                })
                group["quantity_real"] += decimal(quantity_real)
                group["quantity_entry"] += decimal(quantity_entry)
                group["skus"].add(value)
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for (asin, _order_sn), item in by_asin_order.items():
        grouped[asin].append(item)
    for asin in grouped:
        grouped[asin].sort(key=lambda item: (item["order_time"] or datetime.max, item["order_sn"]))
    return grouped, quality_rows


def load_payback(asins: dict[str, dict[str, object]]) -> dict[str, dict[str, str]]:
    sql = """
        SELECT asin, currency_code, data_date, gross_profit
        FROM lingxing_profit_asin
        WHERE data_date >= %s AND data_date < %s
        ORDER BY asin, currency_code, data_date
    """
    daily: dict[tuple[str, str], list[tuple[date, Decimal]]] = defaultdict(list)
    with pymysql.connect(**mysql_env()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(sql, (date(2025, 4, 1), date(2026, 7, 1)))
            for asin, currency, data_date, gross_profit in cursor.fetchall():
                key = str(asin or "").strip()
                if key in asins and currency in {"GBP", "EUR"} and data_date:
                    daily[(key, str(currency))].append((data_date, decimal(gross_profit)))
    result: dict[str, dict[str, str]] = defaultdict(dict)
    for (asin, currency), rows in daily.items():
        start_month = str(asins[asin]["first_fba_month"] or "")
        if not start_month:
            continue
        cumulative = Decimal(0)
        for data_date, profit in rows:
            if data_date.strftime("%Y-%m") < start_month:
                continue
            cumulative += profit
            if cumulative > 0:
                result[asin][currency] = data_date.strftime("%Y-%m")
                break
    return result


def bucket(quantity: Decimal | None) -> str:
    if quantity is None:
        return "无Q1"
    if quantity <= 10:
        return "1-10"
    if quantity <= 15:
        return "11-15"
    if quantity <= 20:
        return "16-20"
    if quantity <= 30:
        return "21-30"
    return ">30"


def build_detail_rows(
    asins: dict[str, dict[str, object]], purchases: dict[str, list[dict[str, object]]], payback: dict[str, dict[str, str]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for asin, record in sorted(asins.items()):
        orders = purchases.get(asin, [])
        q1 = orders[0] if orders else None
        q2 = orders[1] if len(orders) > 1 else None
        first_fba_month = str(record["first_fba_month"])
        first_full_month = add_month(first_fba_month) if first_fba_month else ""
        first_full_sales = record["monthly_sales"].get(first_full_month, Decimal(0)) if first_full_month <= DATA_CUTOFF else Decimal(0)
        q1_before_fba = bool(q1 and first_fba_month and q1["order_time"] and q1["order_time"].strftime("%Y-%m") <= first_fba_month)
        if not q1:
            sample_status = "未找到唯一映射的有效采购Q1"
        elif not first_fba_month:
            sample_status = "未观察到FBA可售，不进入Q1首月模型"
        elif not q1_before_fba:
            sample_status = "Q1采购时间晚于首次FBA可售"
        elif first_full_month > DATA_CUTOFF:
            sample_status = "首个完整月观察不足"
        else:
            sample_status = "有效样本"
        rows.append({
            "ASIN": asin,
            "开发人": record["developer"],
            "基准SKU": " | ".join(sorted(record["skus"])),
            "市场集合": " | ".join(sorted(record["markets"])),
            "首次FBA可售月": first_fba_month,
            "首个完整月": first_full_month,
            "首月销量": first_full_sales,
            "首月出单率": first_full_sales / Decimal(30) if first_full_month <= DATA_CUTOFF else None,
            "Q1采购单号": q1["order_sn"] if q1 else "",
            "Q1采购时间": q1["order_time"].strftime("%Y-%m-%d") if q1 and q1["order_time"] else "",
            "Q1采购SKU": " | ".join(sorted(q1["skus"])) if q1 else "",
            "Q1实际下单量": q1["quantity_real"] if q1 else None,
            "Q1实际入库量": q1["quantity_entry"] if q1 else None,
            "Q1分桶": bucket(q1["quantity_real"]) if q1 else "无Q1",
            "Q2采购单号": q2["order_sn"] if q2 else "",
            "Q2实际下单量": q2["quantity_real"] if q2 else None,
            "是否进入二批": "是" if q2 else "否",
            "GBP财务回正月": payback.get(asin, {}).get("GBP", ""),
            "EUR财务回正月": payback.get(asin, {}).get("EUR", ""),
            "最新标签状态": record["status"],
            "样本状态": sample_status,
        })
    return rows


def q1_summary(rows: list[dict[str, object]]) -> list[list[object]]:
    buckets: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        if row["样本状态"] == "有效样本":
            buckets[str(row["Q1分桶"])].append(row)
    result = []
    for name in ("1-10", "11-15", "16-20", "21-30", ">30"):
        group = buckets.get(name, [])
        q1_total = sum((decimal(row["Q1实际下单量"]) for row in group), Decimal(0))
        first_sales = sum((decimal(row["首月销量"]) for row in group), Decimal(0))
        q2_count = sum(row["是否进入二批"] == "是" for row in group)
        gbp_positive = sum(bool(row["GBP财务回正月"]) for row in group)
        eur_positive = sum(bool(row["EUR财务回正月"]) for row in group)
        result.append([
            name, len(group), q1_total, q1_total / len(group) if group else None, first_sales,
            first_sales / (len(group) * 30) if group else None,
            q2_count, Decimal(q2_count) / len(group) if group else None,
            gbp_positive, Decimal(gbp_positive) / len(group) if group else None,
            eur_positive, Decimal(eur_positive) / len(group) if group else None,
        ])
    return result


def style_table(sheet, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"].font = Font(bold=True, size=14, color="0B3D47")
    sheet.append(headers)
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    for row in rows:
        sheet.append([float(value) if isinstance(value, Decimal) else value for value in row])
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(2, sheet.max_row)}"
    sheet.sheet_view.showGridLines = False
    for index, header in enumerate(headers, start=1):
        column = get_column_letter(index)
        sheet.column_dimensions[column].width = min(max(len(header) + 4, 13), 30)
        if "率" in header:
            for cell in sheet[column][2:]:
                cell.number_format = '0.00%;[Red]-0.00%'
        elif any(word in header for word in ("销量", "量")):
            for cell in sheet[column][2:]:
                cell.number_format = '#,##0.00;[Red]-#,##0.00'


def write_outputs(rows: list[dict[str, object]], quality_rows: list[dict[str, object]]) -> None:
    MODEL_ROOT.mkdir(parents=True, exist_ok=True)
    detail_file = MODEL_ROOT / "02_ASIN_Q1映射与首月结果明细.csv"
    quality_file = MODEL_ROOT / "03_采购SKU_ASIN映射异常.csv"
    headers = list(rows[0]) if rows else []
    with detail_file.open("w", encoding="utf-8-sig", newline="") as destination:
        writer = csv.DictWriter(destination, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    with quality_file.open("w", encoding="utf-8-sig", newline="") as destination:
        writer = csv.DictWriter(destination, fieldnames=["采购单号", "采购时间", "SKU", "quantity_real", "quantity_entry", "映射状态", "候选ASIN"])
        writer.writeheader()
        writer.writerows(quality_rows)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "总览"
    status = Counter(str(row["样本状态"]) for row in rows)
    overview_rows = [
        ["团队 ASIN 总数", len(rows)],
        ["已观察到 FBA 可售", sum(bool(row["首次FBA可售月"]) for row in rows)],
        ["唯一映射采购 Q1", sum(bool(row["Q1采购单号"]) for row in rows)],
        ["有效 Q1 首月样本", status["有效样本"]],
        ["SKU→ASIN 映射异常采购行", len(quality_rows)],
        ["Q1 晚于首次 FBA 可售", status["Q1采购时间晚于首次FBA可售"]],
        ["首月观察不足", status["首个完整月观察不足"]],
    ]
    style_table(overview, "ASIN 首批单量—首月出单率—二批/回正模型：第一版测试", ["指标", "数量"], overview_rows)

    summary = workbook.create_sheet("Q1分桶汇总")
    summary_headers = ["Q1分桶", "有效ASIN数", "Q1下单量合计", "平均Q1下单量", "首月销量合计", "首月出单率", "进入二批ASIN数", "二批率", "GBP财务回正ASIN数", "GBP回正率", "EUR财务回正ASIN数", "EUR回正率"]
    style_table(summary, "Q1 首批量分桶结果（仅有效样本）", summary_headers, q1_summary(rows))

    detail_sheet = workbook.create_sheet("ASIN明细")
    style_table(detail_sheet, "ASIN Q1 映射、首月出单、二批与回正明细", headers, [[row[key] for key in headers] for row in rows])

    quality_sheet = workbook.create_sheet("数据质量")
    quality_headers = ["样本状态", "ASIN数"]
    style_table(quality_sheet, "ASIN 样本状态统计", quality_headers, [[name, count] for name, count in sorted(status.items())])

    notes = workbook.create_sheet("口径说明")
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 108
    notes_rows = [
        ("模型主体", "ASIN 是唯一模型对象；采购 SKU 仅作来源映射。"),
        ("Q1", "唯一映射到 ASIN 的首个有效完成采购单；同一采购单内映射到同一 ASIN 的多个 SKU 数量合并。Q1实际下单量取 quantity_real，实际入库量取 quantity_entry。"),
        ("首月出单率", "首次 FBA 可售后的首个完整自然月销量 ÷ 30。分桶汇总再除以该桶有效 ASIN 数，按业务约定显示为百分比。"),
        ("有效样本", "有唯一映射 Q1、已观察到 FBA 可售、Q1采购时间不晚于首次 FBA 可售月、且首个完整月已落在 2026-06 截止月内。"),
        ("二批", "同一 ASIN 的第二个唯一映射有效采购单；仅表示进入二批，不代表二批已到仓。"),
        ("回正", "按 ASIN 分币种累计领星 gross_profit 首次大于0的月份；GBP 与 EUR 不相加。"),
        ("限制", "采购子项 sid 大量缺失；本版通过 SKU→ASIN 唯一映射归集，映射冲突和时间异常不进入有效样本。首月为月度代理，日级精确30天窗口后续再补。"),
    ]
    notes.append(["口径", "说明"])
    for row in notes_rows:
        notes.append(row)
    for cell in notes[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
    for row in notes.iter_rows(min_row=2):
        row[0].fill = NOTE_FILL
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(MODEL_ROOT / "01_ASIN_Q1首批单量模型_第一版测试.xlsx")

    (MODEL_ROOT / "00_先看这里.md").write_text(
        "# ASIN 首批单量—首月出单率—二批/回正模型：第一版测试\n\n"
        "本目录是 Q1 模型的样本审查，不是首批量推荐结论。\n\n"
        "阅读顺序：先看 `01_ASIN_Q1首批单量模型_第一版测试.xlsx` 的总览和 Q1分桶汇总，再按 ASIN明细核对，最后查看采购 SKU 映射异常。\n\n"
        "本版仅将唯一映射的采购 SKU 归集到 ASIN；采购时间晚于首次 FBA 可售、首月观察不足或无 FBA 可售证据的对象不进入有效样本。\n",
        encoding="utf-8",
    )


def main() -> None:
    asins, sku_to_asins = load_asin_baseline()
    load_monthly_sales(asins)
    purchases, quality_rows = load_purchase_facts(sku_to_asins)
    payback = load_payback(asins)
    rows = build_detail_rows(asins, purchases, payback)
    write_outputs(rows, quality_rows)
    status = Counter(str(row["样本状态"]) for row in rows)
    print(f"team_asins={len(rows)} valid_samples={status['有效样本']} mapping_exceptions={len(quality_rows)} output={MODEL_ROOT}")


if __name__ == "__main__":
    main()
