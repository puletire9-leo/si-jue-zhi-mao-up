#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the full ASIN lifecycle / cohort model from Lingxing fact tables.

Business key: ASIN.  SKU is retained only to attach completed purchase-plan
quantities (Q1/Q2).  FBA cycles are identified solely from weekly FBA
sellable availability; finance uses daily gross_profit and is never converted
between GBP and EUR.
"""

from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from statistics import mean
from typing import Any, Iterable

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lingxing_base_access import dec, load_asin_baseline, load_completed_plans, mysql_env
from lingxing_model_paths import ASIN_START_BASELINE, LIFECYCLE_MODEL_DIR


MODEL_START = date(2025, 4, 1)
MODEL_END = date(2026, 6, 30)
CURRENCIES = ("GBP", "EUR")
OUTPUT_ROOT = LIFECYCLE_MODEL_DIR

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
WHITE_BOLD = Font(color="FFFFFF", bold=True)


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def load_fixed_currency_cohorts() -> dict[str, dict[str, str]]:
    """从数据库 lingxing_product_unified 统一表读取固定上架批次及站点归属。"""
    # 数据源迁移到统一表；币种直接用 country(UK/DE)，比原 base_store/country 拼串更准
    sql = """
        SELECT asin, model_start_month, country
        FROM lingxing_product_unified
    """
    result: dict[str, dict[str, str]] = {currency: {} for currency in CURRENCIES}
    with pymysql.connect(**mysql_env()) as conn:
        with conn.cursor() as cur:
            cur.execute(sql)
            for asin, month, country in cur.fetchall():
                asin = str(asin or "").strip()
                month = str(month or "").strip()
                if not asin or not MODEL_START.strftime("%Y-%m") <= month <= MODEL_END.strftime("%Y-%m"):
                    continue
                site = str(country or "").upper()
                if "UK" in site or "英国" in str(country or ""):
                    result["GBP"][asin] = month
                elif "DE" in site or "德国" in str(country or ""):
                    result["EUR"][asin] = month
    return result


def align_rows_to_fixed_cohorts(
    life_by_currency: dict[str, list[dict[str, Any]]],
    fixed_cohorts: dict[str, dict[str, str]],
) -> dict[str, list[dict[str, Any]]]:
    aligned: dict[str, list[dict[str, Any]]] = {currency: [] for currency in CURRENCIES}
    for currency in CURRENCIES:
        membership = fixed_cohorts[currency]
        for row in life_by_currency[currency]:
            cohort = membership.get(str(row["ASIN"]))
            if cohort is None:
                continue
            copied = dict(row)
            copied["上架批次"] = cohort
            aligned[currency].append(copied)
    return aligned


def money(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01")))


def listing_state(tags: str) -> str:
    value = str(tags or "")
    if "待淘汰" in value:
        return "待淘汰"
    if "淘汰" in value:
        return "已淘汰"
    return "在售/未标注"


def relative_week(start: date, point: date) -> int:
    return max(1, (point - start).days // 7 + 1)


def first_financial_positive_date(values: Iterable[tuple[date, Decimal]]) -> date | None:
    """返回累计利润回正后不再跌回负数的首次日期。"""
    cumulative = Decimal(0)
    cumulative_rows: list[tuple[date, Decimal]] = []
    for day, profit in sorted(values):
        cumulative += profit
        cumulative_rows.append((day, cumulative))
    if not cumulative_rows or cumulative_rows[-1][1] < 0:
        return None
    suffix_min = cumulative_rows[-1][1]
    stable_date = cumulative_rows[-1][0]
    for day, value in reversed(cumulative_rows):
        suffix_min = min(suffix_min, value)
        if suffix_min >= 0:
            stable_date = day
        else:
            break
    return stable_date


def cycle_financial_payback_days(
    financial: Iterable[dict[str, Any]], start: date, end: date,
) -> int | None:
    """Return calendar days to financial break-even within one FBA cycle.

    The cumulative amount always restarts at this cycle's FBA start.  The
    cycle must still be non-negative at its end, and the returned date is the
    first date after which cumulative profit never falls below zero again.
    """
    cumulative = Decimal(0)
    cumulative_rows: list[tuple[date, Decimal]] = []
    for item in sorted(financial, key=lambda row: row["date"]):
        if not start <= item["date"] <= end:
            continue
        cumulative += item["profit"]
        cumulative_rows.append((item["date"], cumulative))
    if not cumulative_rows or cumulative_rows[-1][1] < 0:
        return None
    suffix_min = cumulative_rows[-1][1]
    stable_date = cumulative_rows[-1][0]
    for day, value in reversed(cumulative_rows):
        suffix_min = min(suffix_min, value)
        if suffix_min >= 0:
            stable_date = day
        else:
            break
    return (stable_date - start).days + 1


def fba_cycle_name(cycle_no: int) -> str:
    names = {
        1: "首批FBA周期",
        2: "二批FBA周期",
        3: "三批FBA周期",
        4: "四批FBA周期",
        5: "五批FBA周期",
        6: "六批FBA周期",
        7: "七批FBA周期",
        8: "八批FBA周期",
        9: "九批FBA周期",
        10: "十批FBA周期",
    }
    return names.get(cycle_no, f"第{cycle_no}批FBA周期")


def fba_cycle_sheet_name(cycle_no: int) -> str:
    return fba_cycle_name(cycle_no).replace("周期", "")


def derive_fba_cycles(rows: list[dict[str, Any]], cutoff: date, fallback_start: date) -> list[dict[str, Any]]:
    """Derive FBA availability cycles without treating continuous stock as a refill.

    A later cycle starts only after an explicit observed zero followed by a
    subsequent positive FBA observation.  Missing records are never zeros.
    """
    rows = sorted(rows, key=lambda row: (row["start"], row["end"]))
    positive = [row for row in rows if row["available"] > 0]
    if not positive:
        return [{
            "cycle_no": 1,
            "cycle_type": "首批",
            "start": fallback_start,
            "end": cutoff,
            "start_source": "创建时间兜底",
            "stockout_start": None,
        }]

    cycles: list[dict[str, Any]] = []
    cycle_start = positive[0]["start"]
    cycle_type = "首批"
    zero_start: date | None = None
    last_positive_end = positive[0]["end"]

    for row in rows:
        if row["start"] < cycle_start:
            continue
        if row["available"] > 0:
            if zero_start is not None:
                cycles.append({
                    "cycle_no": len(cycles) + 1,
                    "cycle_type": cycle_type,
                    "start": cycle_start,
                    "end": zero_start - timedelta(days=1),
                    "start_source": "周级FBA首次可售" if not cycles else "断货后再次FBA可售",
                    "stockout_start": zero_start,
                })
                cycle_start = row["start"]
                cycle_type = "二批候选" if len(cycles) == 1 else f"第{len(cycles) + 1}批候选"
                zero_start = None
            last_positive_end = row["end"]
        elif zero_start is None and last_positive_end >= cycle_start:
            zero_start = row["start"]

    cycles.append({
        "cycle_no": len(cycles) + 1,
        "cycle_type": cycle_type,
        "start": cycle_start,
        "end": (zero_start - timedelta(days=1)) if zero_start else cutoff,
        "start_source": "周级FBA首次可售" if not cycles else "断货后再次FBA可售",
        "stockout_start": zero_start,
    })
    return cycles


def sql_chunks(values: set[str], size: int = 500) -> Iterable[list[str]]:
    ordered = sorted(values)
    for index in range(0, len(ordered), size):
        yield ordered[index:index + size]


def canonical_week_start(start: date) -> bool:
    """Exclude an older ad-hoc overlapping validation week in 2026-03."""
    if start > date(2026, 3, 31):
        return True
    return (start - MODEL_START).days % 7 == 0


def load_weekly(asins: set[str]) -> dict[str, list[dict[str, Any]]]:
    by_asin_window: dict[tuple[str, date, date], dict[str, Any]] = {}
    with pymysql.connect(**mysql_env()) as connection:
        with connection.cursor() as cursor:
            for chunk in sql_chunks(asins):
                placeholders = ", ".join(["%s"] * len(chunk))
                cursor.execute(
                    f"""
                    SELECT asin, currency_code, week_start, week_end,
                           MAX(COALESCE(afn_fulfillable_quantity, 0)),
                           SUM(COALESCE(volume, 0)), SUM(COALESCE(amount, 0))
                    FROM lingxing_sku_weekly_performance
                    WHERE asin IN ({placeholders})
                      AND week_start >= %s AND week_end <= %s
                      AND currency_code IN ('GBP', 'EUR')
                    GROUP BY asin, currency_code, week_start, week_end
                    """,
                    [*chunk, MODEL_START, MODEL_END],
                )
                for asin, currency, start, end, available, volume, amount in cursor.fetchall():
                    if not asin or not start or not end or not canonical_week_start(start):
                        continue
                    value = by_asin_window.setdefault((str(asin), start, end), {
                        "start": start, "end": end, "available": Decimal(0), "currencies": set(), "currency_metrics": {},
                    })
                    value["available"] = max(value["available"], dec(available))
                    value["currencies"].add(str(currency))
                    value["currency_metrics"][str(currency)] = {"volume": dec(volume), "amount": dec(amount)}
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for (asin, _start, _end), row in by_asin_window.items():
        result[asin].append(row)
    for rows in result.values():
        rows.sort(key=lambda row: (row["start"], row["end"]))
    return result


def load_finance(asins: set[str]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    result: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    with pymysql.connect(**mysql_env()) as connection:
        with connection.cursor() as cursor:
            for chunk in sql_chunks(asins):
                placeholders = ", ".join(["%s"] * len(chunk))
                cursor.execute(
                    f"""
                    SELECT asin, currency_code, data_date,
                           SUM(COALESCE(total_sales_quantity, 0)),
                           SUM(COALESCE(total_sales_amount, 0)),
                           SUM(COALESCE(total_ads_cost, 0)),
                           SUM(COALESCE(total_cost, 0)),
                           SUM(COALESCE(cg_price, 0)),
                           SUM(COALESCE(cg_transport_costs, 0)),
                           SUM(COALESCE(gross_profit, 0))
                    FROM lingxing_profit_asin
                    WHERE asin IN ({placeholders})
                      AND data_date >= %s AND data_date <= %s
                      AND currency_code IN ('GBP', 'EUR')
                    GROUP BY asin, currency_code, data_date
                    ORDER BY asin, currency_code, data_date
                    """,
                    [*chunk, MODEL_START, MODEL_END],
                )
                for asin, currency, day, quantity, sales, ads, cost, cg_price, cg_transport, profit in cursor.fetchall():
                    result[(str(asin), str(currency))].append({
                        "date": day, "quantity": dec(quantity), "sales": dec(sales),
                        "ads": dec(ads), "cost": dec(cost),
                        "cg_price": dec(cg_price), "cg_transport": dec(cg_transport),
                        "profit": dec(profit),
                    })
    return result


def allocate_plans(asin: str, record: dict[str, Any], sku_to_asins: dict[str, set[str]], plans: dict[str, list[dict[str, Any]]]) -> tuple[Decimal | None, Decimal | None, str, set[str]]:
    q1 = Decimal(0)
    q2 = Decimal(0)
    mapped = False
    split_candidates: set[str] = set()
    for sku in sorted(record["skus"]):
        candidates = sku_to_asins.get(sku, set())
        facts = plans.get(sku, [])
        if asin not in candidates or not facts:
            continue
        mapped = True
        split_candidates.update(candidates)
        divisor = Decimal(len(candidates))
        q1 += facts[0]["quantity"] / divisor
        if len(facts) > 1:
            q2 += facts[1]["quantity"] / divisor
    if not mapped:
        return None, None, "无已完成采购计划", set()
    return q1, q2 if q2 > 0 else None, ("SKU多ASIN等额分摊" if len(split_candidates) > 1 else "唯一SKU映射"), split_candidates


def metric_total(rows: Iterable[dict[str, Any]]) -> dict[str, Decimal]:
    values = list(rows)
    return {
        "quantity": sum((row["quantity"] for row in values), Decimal(0)),
        "sales": sum((row["sales"] for row in values), Decimal(0)),
        "ads": sum((row["ads"] for row in values), Decimal(0)),
        "cost": sum((row["cost"] for row in values), Decimal(0)),
        "profit": sum((row["profit"] for row in values), Decimal(0)),
    }


def finance_metrics(rows: list[dict[str, Any]], start: date, end: date) -> dict[str, Decimal]:
    total = metric_total(row for row in rows if start <= row["date"] <= end)
    total["other"] = total["profit"] - total["sales"] - total["ads"] - total["cost"]
    return total


def lifecycle_rows(
    asins: dict[str, dict[str, Any]], sku_to_asins: dict[str, set[str]], plans: dict[str, list[dict[str, Any]]],
    weekly: dict[str, list[dict[str, Any]]], finance: dict[tuple[str, str], list[dict[str, Any]]],
    fixed_cohorts: dict[str, dict[str, str]],
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, dict[str, Any]], list[dict[str, Any]]]:
    by_currency: dict[str, list[dict[str, Any]]] = {currency: [] for currency in CURRENCIES}
    asin_summary: dict[str, dict[str, Any]] = {}
    quality: list[dict[str, Any]] = []
    for asin, record in sorted(asins.items()):
        fallback_month = str(record.get("model_start_month") or "")
        if not fallback_month or fallback_month > MODEL_END.strftime("%Y-%m"):
            continue
        fallback = date.fromisoformat(f"{fallback_month}-01")
        rows = weekly.get(asin, [])
        # 未补齐的周级 FBA 记录不是“持续有货”。每个 ASIN 的周期最多只能
        # 计算到该 ASIN 最后一条周级观察；事实表截止日由数据库实际数据决定。
        observed_end = min(MODEL_END, max((row["end"] for row in rows), default=MODEL_END))
        cycles = derive_fba_cycles(rows, observed_end, fallback)
        q1, q2, allocation, candidates = allocate_plans(asin, record, sku_to_asins, plans)
        first = cycles[0]
        first_start = first["start"]
        latest = rows[-1] if rows else None
        currencies = {currency for row in rows for currency in row["currencies"]}
        currencies.update(currency for key_asin, currency in finance if key_asin == asin)
        currencies.update(currency for currency in CURRENCIES if asin in fixed_cohorts[currency])
        label_status = listing_state(str(record.get("label") or ""))
        first_30_sales = sum((sum((metric["volume"] for metric in row["currency_metrics"].values()), Decimal(0)) for row in rows if first_start <= row["start"] <= first_start + timedelta(days=29)), Decimal(0))
        summary = {
            "asin": asin,
            "developer": record["developer"],
            "sku": " | ".join(sorted(record["skus"])),
            "launch_date": first_start,
            # 上架批次使用已经统一的基础批次月份；周级 FBA 只负责周期起止。
            "launch_month": fallback_month,
            "start_source": first["start_source"],
            "label": str(record.get("label") or ""),
            "label_status": label_status,
            "q1": q1,
            "q2": q2,
            "allocation": allocation,
            "allocation_candidates": candidates,
            "latest_fba": latest["available"] if latest else Decimal(0),
            "latest_fba_week": latest["start"] if latest else None,
            "first_30_sales": first_30_sales,
            "cycles": cycles,
        }
        asin_summary[asin] = summary
        if first["start_source"] == "创建时间兜底" or allocation == "SKU多ASIN等额分摊":
            quality.append({
                "ASIN": asin, "开发人": record["developer"], "基准SKU": summary["sku"],
                "问题": "FBA起点创建时间兜底" if first["start_source"] == "创建时间兜底" else "Q1/Q2多ASIN等额分摊",
                "说明": "全期间未观察到周级FBA可售" if first["start_source"] == "创建时间兜底" else "采购计划量按候选ASIN等额分摊，需人工核验",
            })
        for currency in sorted(currencies & set(CURRENCIES)):
            financial = finance.get((asin, currency), [])
            payback = first_financial_positive_date([(row["date"], row["profit"]) for row in financial if row["date"] >= first_start])
            cumulative = finance_metrics(financial, first_start, MODEL_END)
            summary[f"finance_{currency}"] = cumulative
            summary[f"payback_{currency}"] = payback
            summary[f"performance_{currency}"] = {
                "volume": sum((row["currency_metrics"].get(currency, {}).get("volume", Decimal(0)) for row in rows if row["start"] >= first_start), Decimal(0)),
                "amount": sum((row["currency_metrics"].get(currency, {}).get("amount", Decimal(0)) for row in rows if row["start"] >= first_start), Decimal(0)),
            }
            for cycle in cycles:
                cycle_weekly = [row for row in rows if cycle["start"] <= row["start"] <= cycle["end"]]
                cycle_finance = [row for row in financial if cycle["start"] <= row["date"] <= cycle["end"]]
                metrics = finance_metrics(financial, cycle["start"], cycle["end"])
                payback_days = cycle_financial_payback_days(financial, cycle["start"], cycle["end"])
                by_currency[currency].append({
                    "ASIN": asin,
                    "开发人": record["developer"],
                    "基准SKU": summary["sku"],
                    "上架批次": summary["launch_month"],
                    "首次FBA可售日期": first_start,
                    "上架起点依据": first["start_source"],
                    "当前标签状态": label_status,
                    "FBA周期序号": cycle["cycle_no"],
                    "FBA周期类型": cycle["cycle_type"],
                    "周期开始": cycle["start"],
                    "周期结束": cycle["end"],
                    "前次断货观察周": cycle["stockout_start"] or "",
                    "Q1计划量": q1 if q1 is not None else "",
                    "Q2计划量": q2 if q2 is not None else "",
                    "采购计划关联": allocation,
                    "周期FBA可售周数": sum(row["available"] > 0 for row in cycle_weekly),
                    "售卖周期天数": (cycle["end"] - cycle["start"]).days + 1,
                    "周期财务回本天数": payback_days if payback_days is not None else "",
                    "周期销量": sum((row["currency_metrics"].get(currency, {}).get("volume", Decimal(0)) for row in cycle_weekly), Decimal(0)),
                    "周期产品表现销售额": sum((row["currency_metrics"].get(currency, {}).get("amount", Decimal(0)) for row in cycle_weekly), Decimal(0)),
                    "有财务记录": bool(cycle_finance),
                    "财务销量": metrics["quantity"],
                    "结算销售额": metrics["sales"],
                    "广告费支出": abs(metrics["ads"]),
                    "财务总成本支出": abs(metrics["cost"]),
                    "其他结算支出/调整": abs(metrics["other"]),
                    "结算毛利润": metrics["profit"],
                    "累计结算毛利润": cumulative["profit"],
                    "财务回正日期": payback or "",
                    "最新观察FBA可售库存": summary["latest_fba"],
                    "最新FBA观察周": summary["latest_fba_week"] or "",
                })
    return by_currency, asin_summary, quality


def summarize_cycle_finance(
    rows: list[dict[str, Any]], *, by_cohort: bool = True,
) -> list[dict[str, Any]]:
    """按“上架批次 × FBA 周期”汇总财务与经营时间。

    每个 ASIN 的周期起止日期不同，因此模型表展示该上架批次内的开始/结束
    日期范围；精确到单个 ASIN 的日期仍保留在基础工作簿中。
    """
    bucket: dict[tuple[str, str], dict[str, Any]] = {}

    def date_range(values: list[date]) -> str:
        if not values:
            return ""
        first = min(values).isoformat()
        last = max(values).isoformat()
        return first if first == last else f"{first} 至 {last}"

    for row in rows:
        cycle_type = str(row["FBA周期类型"])
        # 该 ASIN 没有任何周级 FBA 可售观察，只是以商品创建时间补齐了
        # 基础主档；它不是 FBA 周期，不能混入首批/二批的经营结论。
        if cycle_type == "未观察到FBA":
            continue
        cycle_no = int(row["FBA周期序号"])
        display_type = fba_cycle_name(cycle_no)
        cohort = str(row["上架批次"]) if by_cohort else "全部上架批次"
        key = (cohort, display_type)
        value = bucket.setdefault(key, {
            "上架批次": key[0], "FBA周期": key[1], "FBA批次序号": cycle_no,
            "finance_asins": set(),
            "cycle_asins": set(), "cycle_starts": [], "cycle_ends": [],
            "sale_days": [], "payback_days": [],
            "财务销量": Decimal(0), "销售额": Decimal(0), "广告费支出": Decimal(0),
            "财务总成本支出": Decimal(0), "其他结算支出/调整": Decimal(0), "profits": [],
        })
        value["cycle_asins"].add(row["ASIN"])
        if isinstance(row["周期开始"], date):
            value["cycle_starts"].append(row["周期开始"])
        if isinstance(row["周期结束"], date):
            value["cycle_ends"].append(row["周期结束"])
        value["sale_days"].append(dec(row["售卖周期天数"]))
        if row["周期财务回本天数"] != "":
            value["payback_days"].append(dec(row["周期财务回本天数"]))
        if not row["有财务记录"]:
            continue
        value["finance_asins"].add(row["ASIN"])
        value["财务销量"] += dec(row["财务销量"])
        value["销售额"] += dec(row["结算销售额"])
        value["广告费支出"] += dec(row["广告费支出"])
        value["财务总成本支出"] += dec(row["财务总成本支出"])
        value["其他结算支出/调整"] += dec(row["其他结算支出/调整"])
        value["profits"].append(dec(row["结算毛利润"]))

    output: list[dict[str, Any]] = []
    for value in sorted(
        bucket.values(),
        key=lambda item: (item["FBA批次序号"], item["上架批次"]),
    ):
        profits = value["profits"]
        losses = [profit for profit in profits if profit < 0]
        gains = [profit for profit in profits if profit > 0]
        output.append({
            "上架批次": value["上架批次"],
            "FBA周期": value["FBA周期"],
            "周期ASIN数": len(value["cycle_asins"]),
            "周期开始范围": date_range(value["cycle_starts"]),
            "周期结束/观察截止范围": date_range(value["cycle_ends"]),
            "平均售卖周期天数": sum(value["sale_days"], Decimal(0)) / len(value["sale_days"]) if value["sale_days"] else Decimal(0),
            "有财务记录ASIN数": len(value["finance_asins"]),
            "财务销量": value["财务销量"],
            "销售额": value["销售额"],
            "广告费支出": value["广告费支出"],
            "财务总成本支出": value["财务总成本支出"],
            "其他结算支出/调整": value["其他结算支出/调整"],
            "亏损ASIN数": len(losses),
            "亏损金额": abs(sum(losses, Decimal(0))),
            "盈利ASIN数": len(gains),
            "盈利金额": sum(gains, Decimal(0)),
            "持平ASIN数": sum(profit == 0 for profit in profits),
            "领星结算毛利润": sum(profits, Decimal(0)),
            "批次整体是否回本": "已回本" if sum(profits, Decimal(0)) >= 0 else "未回本",
            "财务回本ASIN数": len(value["payback_days"]),
            "平均财务回本天数": sum(value["payback_days"], Decimal(0)) / len(value["payback_days"]) if value["payback_days"] else "",
        })
    return output


def summarize_all_cycles_total(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """汇总“所有周期汇总”Sheet；数量按 ASIN-周期记录计数。"""
    if not rows:
        return {}

    def boundary(field: str, take_max: bool) -> str:
        values: list[str] = []
        for row in rows:
            text_value = str(row.get(field) or "")
            if not text_value:
                continue
            parts = [part.strip() for part in text_value.split("至")]
            values.extend(parts)
        if not values:
            return ""
        return max(values) if take_max else min(values)

    cycle_count = sum((int(row["周期ASIN数"]) for row in rows), 0)
    payback_count = sum((int(row["财务回本ASIN数"]) for row in rows), 0)
    total_profit = sum((dec(row["领星结算毛利润"]) for row in rows), Decimal(0))
    weighted_sale_days = sum(
        (dec(row["平均售卖周期天数"]) * int(row["周期ASIN数"]) for row in rows),
        Decimal(0),
    )
    weighted_payback_days = sum(
        (
            dec(row["平均财务回本天数"]) * int(row["财务回本ASIN数"])
            for row in rows
            if row["平均财务回本天数"] != ""
        ),
        Decimal(0),
    )
    return {
        "上架批次": "合计",
        "FBA周期": "全部FBA周期",
        "周期ASIN数": cycle_count,
        "周期开始范围": f"{boundary('周期开始范围', False)} 至 {boundary('周期开始范围', True)}",
        "周期结束/观察截止范围": f"{boundary('周期结束/观察截止范围', False)} 至 {boundary('周期结束/观察截止范围', True)}",
        "平均售卖周期天数": weighted_sale_days / cycle_count if cycle_count else Decimal(0),
        "有财务记录ASIN数": sum((int(row["有财务记录ASIN数"]) for row in rows), 0),
        "财务销量": sum((dec(row["财务销量"]) for row in rows), Decimal(0)),
        "销售额": sum((dec(row["销售额"]) for row in rows), Decimal(0)),
        "广告费支出": sum((dec(row["广告费支出"]) for row in rows), Decimal(0)),
        "财务总成本支出": sum((dec(row["财务总成本支出"]) for row in rows), Decimal(0)),
        "其他结算支出/调整": sum((dec(row["其他结算支出/调整"]) for row in rows), Decimal(0)),
        "亏损ASIN数": sum((int(row["亏损ASIN数"]) for row in rows), 0),
        "亏损金额": sum((dec(row["亏损金额"]) for row in rows), Decimal(0)),
        "盈利ASIN数": sum((int(row["盈利ASIN数"]) for row in rows), 0),
        "盈利金额": sum((dec(row["盈利金额"]) for row in rows), Decimal(0)),
        "持平ASIN数": sum((int(row["持平ASIN数"]) for row in rows), 0),
        "领星结算毛利润": total_profit,
        "批次整体是否回本": "已回本" if total_profit >= 0 else "未回本",
        "财务回本ASIN数": payback_count,
        "平均财务回本天数": weighted_payback_days / payback_count if payback_count else "",
    }


def field_explanation_rows(currency: str) -> list[dict[str, str]]:
    """面向业务人员的大白话字段说明，作为汇总工作簿第一个 Sheet。"""
    source_baseline = "基础统一表中的 ASIN 模型分析起算月基准"
    source_weekly = "数据库 lingxing_sku_weekly_performance 周级产品表现表"
    source_finance = "数据库 lingxing_profit_asin 领星财务 ASIN 日数据"
    return [
        {
            "字段": "上架批次",
            "大白话是什么意思": "这批 ASIN 被归到哪个上架月份，例如 2026-01。",
            "具体怎么算": "直接读取统一基础表的“模型分析起算月”：优先用月表首次 FBA 可售月份；全期没有 FBA 可售时才用商品信息创建月份兜底。周数据不能改这个月份。",
            "数据从哪里来": source_baseline,
            "容易误解的地方": "这是模型归批月份，不是亚马逊后台精确到分钟的真实上架时间。",
        },
        {
            "字段": "FBA周期",
            "大白话是什么意思": "同一个 ASIN 的第几批可售库存经营阶段，例如首批、二批、三批。",
            "具体怎么算": "首批从周表首次 FBA可售>0 开始；只有明确观察到 FBA可售=0，之后再次>0，才开始下一批。没有周记录不能当成断货。全期没有周级可售证据时，用统一起算时间兜底为首批。",
            "数据从哪里来": source_weekly + "；无周级可售时使用统一基础表兜底",
            "容易误解的地方": "这里的二批、三批是 FBA 库存周期，不是采购计划编号，也不是简单按自然月切分。",
        },
        {
            "字段": "周期ASIN数",
            "大白话是什么意思": "这一行实际包含多少个 ASIN 周期。",
            "具体怎么算": "在“上架批次+FBA周期”这一组内，对 ASIN 去重计数。首 Sheet 最后一行是所有周期的 ASIN-周期记录数相加。",
            "数据从哪里来": "统一 ASIN 集合 + 周级 FBA 周期识别结果",
            "容易误解的地方": "首 Sheet 合计不是唯一 ASIN 总数；同一 ASIN 有首批和二批时会分别计一次。",
        },
        {
            "字段": "周期开始范围",
            "大白话是什么意思": "这一组 ASIN 最早什么时候开始该周期、最晚什么时候开始该周期。",
            "具体怎么算": "取组内每个 ASIN 周期开始日期的最小值和最大值，显示为“最早日期 至 最晚日期”。",
            "数据从哪里来": source_weekly,
            "容易误解的地方": "这是开始日期的范围，不是一个 ASIN 连续售卖了这么久。单个 ASIN 日期请看 ASIN 明细表。",
        },
        {
            "字段": "周期结束/观察截止范围",
            "大白话是什么意思": "这一组 ASIN 的周期最早结束到最晚结束的日期范围。",
            "具体怎么算": "有明确断货时，结束日为断货观察周之前一天；没有观察到下一次断货时，结束到该 ASIN 最后可用观察日或模型截止日。再取组内最小和最大结束日。",
            "数据从哪里来": source_weekly,
            "容易误解的地方": "写到观察截止不代表货一定在当天才卖完，只代表现有数据只能观察到这里。",
        },
        {
            "字段": "平均售卖周期天数",
            "大白话是什么意思": "这一组 ASIN 平均每个库存周期持续多少天。",
            "具体怎么算": "先对每个 ASIN 算“周期结束日-周期开始日+1”，再把这些天数做普通平均。首 Sheet 合计按周期ASIN数加权。",
            "数据从哪里来": "周级 FBA 周期开始、结束日期计算",
            "容易误解的地方": "是自然日，不是有订单的天数，也不是库存实际在库天数。",
        },
        {
            "字段": "有财务记录ASIN数",
            "大白话是什么意思": "这一组里有多少个 ASIN 在对应周期内找到了领星财务记录。",
            "具体怎么算": "ASIN 在该周期日期范围内，只要至少有一条对应币种的财务日数据，就计 1 个。",
            "数据从哪里来": source_finance,
            "容易误解的地方": "没有财务记录不等于利润为 0，而是本次财务事实表没有匹配到记录。",
        },
        {
            "字段": "财务销量",
            "大白话是什么意思": "领星财务口径下，这些 ASIN 在对应周期卖了多少件。",
            "具体怎么算": "按 ASIN、{0} 币种和周期日期范围，汇总财务日数据的 total_sales_quantity。".format(currency),
            "数据从哪里来": source_finance + " 的 total_sales_quantity",
            "容易误解的地方": "这是财务表销量，不是周级产品表现表销量；退款、结算时间差可能造成两个口径不完全相同。",
        },
        {
            "字段": "销售额",
            "大白话是什么意思": "领星财务口径确认的结算销售金额。",
            "具体怎么算": "按 ASIN、{0} 币种和周期日期范围，汇总 total_sales_amount。".format(currency),
            "数据从哪里来": source_finance + " 的 total_sales_amount",
            "容易误解的地方": "金额单位只可能是当前文件的 {0}，不会与另一币种相加。".format(currency),
        },
        {
            "字段": "广告费支出",
            "大白话是什么意思": "对应周期内被领星财务归到广告费的金额。",
            "具体怎么算": "汇总财务日数据 total_ads_cost。原始支出通常是负数，报表取绝对值后用正数展示支出大小。",
            "数据从哪里来": source_finance + " 的 total_ads_cost",
            "容易误解的地方": "表里显示 100 表示支出了 100，不表示增加了 100 利润。",
        },
        {
            "字段": "财务总成本支出",
            "大白话是什么意思": "领星财务表中除广告费外归入 total_cost 的总成本金额。",
            "具体怎么算": "汇总财务日数据 total_cost。原始支出通常是负数，报表取绝对值后用正数展示。",
            "数据从哪里来": source_finance + " 的 total_cost",
            "容易误解的地方": "它是领星财务接口给出的成本口径，不等于采购计划 Q1/Q2 数量，也不等于单独的采购现金支出。",
        },
        {
            "字段": "其他结算支出/调整",
            "大白话是什么意思": "销售额、广告费、总成本之外，为了与领星结算毛利润对平所剩下的其他费用或调整金额。",
            "具体怎么算": "每个 ASIN 周期先算 abs(gross_profit-total_sales_amount-total_ads_cost-total_cost)，再汇总。这里只展示差额大小。",
            "数据从哪里来": source_finance + " 中销售额、广告费、总成本和 gross_profit 的对账差额",
            "容易误解的地方": "这不是领星直接返回的一个独立字段；取了绝对值，只能看调整规模，不能从本列单独判断是加项还是减项。",
        },
        {
            "字段": "亏损ASIN数",
            "大白话是什么意思": "这一组里有财务记录且周期结算毛利润小于 0 的 ASIN 数量。",
            "具体怎么算": "先逐 ASIN 汇总该周期 gross_profit；结果<0 计入亏损。",
            "数据从哪里来": source_finance + " 的 gross_profit",
            "容易误解的地方": "统计单位是 ASIN-周期。同一 ASIN 首批亏损、二批盈利，会分别进入不同周期统计。",
        },
        {
            "字段": "亏损金额",
            "大白话是什么意思": "所有亏损 ASIN 一共亏了多少钱。",
            "具体怎么算": "把周期结算毛利润<0 的 ASIN 利润相加，再取绝对值，用正数显示亏损规模。",
            "数据从哪里来": source_finance + " 的 gross_profit",
            "容易误解的地方": "显示 1,000 表示合计亏损 1,000；它不会带负号。",
        },
        {
            "字段": "盈利ASIN数",
            "大白话是什么意思": "这一组里有财务记录且周期结算毛利润大于 0 的 ASIN 数量。",
            "具体怎么算": "先逐 ASIN 汇总该周期 gross_profit；结果>0 计入盈利。",
            "数据从哪里来": source_finance + " 的 gross_profit",
            "容易误解的地方": "盈利只表示领星结算毛利润为正，不等于采购现金已经全部收回。",
        },
        {
            "字段": "盈利金额",
            "大白话是什么意思": "所有盈利 ASIN 一共赚了多少钱。",
            "具体怎么算": "把周期结算毛利润>0 的 ASIN 利润直接相加。",
            "数据从哪里来": source_finance + " 的 gross_profit",
            "容易误解的地方": "亏损 ASIN 不参与本列；整批净利润要看“领星结算毛利润”。",
        },
        {
            "字段": "持平ASIN数",
            "大白话是什么意思": "有财务记录，但周期结算毛利润正好等于 0 的 ASIN 数量。",
            "具体怎么算": "逐 ASIN 汇总周期 gross_profit，结果=0 计入持平。",
            "数据从哪里来": source_finance + " 的 gross_profit",
            "容易误解的地方": "没有财务记录的 ASIN 不算持平。",
        },
        {
            "字段": "领星结算毛利润",
            "大白话是什么意思": "领星财务最终给出的这组 ASIN 在对应周期到底赚了还是亏了。",
            "具体怎么算": "按 ASIN、{0} 币种和周期日期范围直接汇总 gross_profit；模型不自行用销售额减成本重新造利润。".format(currency),
            "数据从哪里来": source_finance + " 的 gross_profit",
            "容易误解的地方": "这是结算毛利润，不是采购现金回本，也不是把“盈利金额-亏损金额”以外的另一套利润。",
        },
        {
            "字段": "批次整体是否回本",
            "大白话是什么意思": "这一整行 ASIN 合起来，结算毛利润最终有没有回到 0 以上。",
            "具体怎么算": "本行领星结算毛利润>=0 显示“已回本”；<0 显示“未回本”。",
            "数据从哪里来": "本行领星结算毛利润计算结果",
            "容易误解的地方": "这是整批状态。整批未回本时，里面仍可能有部分单个 ASIN 已回本。",
        },
        {
            "字段": "财务回本ASIN数",
            "大白话是什么意思": "这一组里有多少个单独 ASIN 在该周期结束时累计结算毛利润仍不低于 0。",
            "具体怎么算": "每个 ASIN 从该 FBA 周期开始日重新累计每日 gross_profit；周期结束仍<0 算未回本，结束时>=0 才算已回本。",
            "数据从哪里来": source_finance + " 的每日 gross_profit + 单个 ASIN 周期日期",
            "容易误解的地方": "这是单个 ASIN 数量，不是整批是否回本。整批亏损时该列仍可能大于 0。",
        },
        {
            "字段": "平均财务回本天数",
            "大白话是什么意思": "已经回本的那些 ASIN，平均花了多少天实现稳定回正。",
            "具体怎么算": "对每个已回本 ASIN，取累计利润回正后再也没有跌回负数的首次日期；回本天数=该日期-周期开始日+1，再对已回本 ASIN 做平均。首 Sheet 合计按回本ASIN数加权。",
            "数据从哪里来": source_finance + " 的每日 gross_profit + 单个 ASIN 周期开始日",
            "容易误解的地方": "未回本 ASIN 不参与平均；不是从商品创建日计算，也不是采购资金回本天数。",
        },
    ]


def month_number(month: str) -> int:
    year, value = map(int, month.split("-"))
    return year * 12 + value - 1


def month_text(number: int) -> str:
    return f"{number // 12:04d}-{number % 12 + 1:02d}"


def months_between(start: str, end: str) -> list[str]:
    return [month_text(value) for value in range(month_number(start), month_number(end) + 1)]


def stable_recovery_index(cumulative_values: list[Decimal]) -> int | None:
    """返回稳定回正的相对月份下标；最终仍亏损时返回 None。"""
    if not cumulative_values or cumulative_values[-1] < 0:
        return None
    suffix_min = cumulative_values[-1]
    stable = len(cumulative_values) - 1
    for index in range(len(cumulative_values) - 1, -1, -1):
        suffix_min = min(suffix_min, cumulative_values[index])
        if suffix_min >= 0:
            stable = index
        else:
            break
    return stable


def allocate_q123(
    asin: str,
    record: dict[str, Any],
    sku_to_asins: dict[str, set[str]],
    plans: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    quantities = [Decimal(0), Decimal(0), Decimal(0)]
    mapped = False
    split = False
    completed_batches = 0
    plan_numbers: list[set[str]] = [set(), set(), set()]
    for sku in sorted(record["skus"]):
        candidates = sku_to_asins.get(sku, set())
        if asin not in candidates or not candidates:
            continue
        facts: list[dict[str, Any]] = []
        for fact in plans.get(sku, []):
            point = fact.get("create_time")
            if hasattr(point, "date"):
                point = point.date()
            if point and point <= MODEL_END:
                facts.append(fact)
        if not facts:
            continue
        mapped = True
        split = split or len(candidates) > 1
        completed_batches = max(completed_batches, len(facts))
        divisor = Decimal(len(candidates))
        for index, fact in enumerate(facts[:3]):
            quantities[index] += dec(fact["quantity"]) / divisor
            if fact.get("plan_sn"):
                plan_numbers[index].add(str(fact["plan_sn"]))
    return {
        "q1": quantities[0] if mapped and quantities[0] > 0 else None,
        "q2": quantities[1] if mapped and quantities[1] > 0 else None,
        "q3": quantities[2] if mapped and quantities[2] > 0 else None,
        "completed_batches": completed_batches,
        "allocation": "SKU多ASIN等额分摊" if split else ("唯一SKU映射" if mapped else "无已完成采购计划"),
        "q1_plans": " | ".join(sorted(plan_numbers[0])),
        "q2_plans": " | ".join(sorted(plan_numbers[1])),
        "q3_plans": " | ".join(sorted(plan_numbers[2])),
    }


def monthly_finance_by_asin(
    rows: list[dict[str, Any]], launch_month: str,
) -> dict[str, dict[str, Decimal]]:
    result: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {
            "quantity": Decimal(0), "sales": Decimal(0), "ads": Decimal(0),
            "cost": Decimal(0), "profit": Decimal(0), "records": Decimal(0),
        }
    )
    for row in rows:
        month = row["date"].strftime("%Y-%m")
        if month < launch_month or month > MODEL_END.strftime("%Y-%m"):
            continue
        value = result[month]
        value["quantity"] += dec(row["quantity"])
        value["sales"] += dec(row["sales"])
        value["ads"] += dec(row["ads"])
        value["cost"] += dec(row["cost"])
        value["profit"] += dec(row["profit"])
        value["records"] += 1
    return result


def monthly_cohort_explanations(currency: str) -> list[dict[str, str]]:
    return [
        {"字段": "上架批次", "大白话是什么意思": "同一个上架月份的 ASIN 集合。", "具体怎么算": "直接读取统一基础表的模型分析起算月，模型不重新分配。", "数据从哪里来": "基础统一表", "容易误解的地方": "不是采购计划创建月。"},
        {"字段": "Q1/Q2/Q3有计划ASIN数", "大白话是什么意思": "这批 ASIN 中有第1、第2、第3个已完成采购计划的数量。", "具体怎么算": "按SKU把已完成采购计划依创建时间排序，第1个为Q1、第2个为Q2、第3个为Q3，再映射回ASIN。", "数据从哪里来": "lingxing_purchase_plan", "容易误解的地方": "只有已完成计划才算；没有计划不等于没有实际销售。"},
        {"字段": "Q1/Q2/Q3计划量", "大白话是什么意思": "这批 ASIN 第1、第2、第3次已完成采购计划分别计划采购多少件。", "具体怎么算": "汇总 quantity_plan；SKU只对应一个ASIN时全量归属，SKU对应多个ASIN时暂按ASIN数等额分摊。", "数据从哪里来": "lingxing_purchase_plan.quantity_plan", "容易误解的地方": "这是计划采购量，不是FBA库存、入库量或销量。"},
        {"字段": "Q1-Q3完整ASIN数", "大白话是什么意思": "同时找到了Q1、Q2、Q3三个已完成采购计划的ASIN数。", "具体怎么算": "单个ASIN的Q1、Q2、Q3计划量都大于0才计入。", "数据从哪里来": "采购计划映射结果", "容易误解的地方": "新品尚未发生Q2/Q3时不完整是正常现象。"},
        {"字段": "经营第N个月", "大白话是什么意思": "从上架批次月份开始往后数的第几个月。", "具体怎么算": "上架月为第1个月，下一个自然月为第2个月，依次计算到2026-06。", "数据从哪里来": "统一上架批次+财务自然月", "容易误解的地方": "这里不再按断货或FBA补货切周期。"},
        {"字段": "当月结算毛利润", "大白话是什么意思": "这批ASIN在该自然月赚了或亏了多少钱。", "具体怎么算": "按ASIN、{0}币种和自然月汇总gross_profit。".format(currency), "数据从哪里来": "lingxing_profit_asin.gross_profit", "容易误解的地方": "这是领星结算毛利润，不是采购现金流。"},
        {"字段": "累计结算毛利润", "大白话是什么意思": "从上架月一直累计到该月，一共赚了或亏了多少钱。", "具体怎么算": "第1个月利润+第2个月利润+……+当前月份利润。", "数据从哪里来": "逐月gross_profit累计", "容易误解的地方": "后续月份亏损会把之前盈利重新亏回去。"},
        {"字段": "稳定回本月数", "大白话是什么意思": "这批ASIN上架几个月后，累计利润回到0以上并且以后再也没有跌回负数。", "具体怎么算": "截至2026-06累计利润必须>=0，再从后往前找最后一次由负转为非负后的月份。", "数据从哪里来": "批次逐月累计结算毛利润", "容易误解的地方": "最终仍亏损时留空；中途短暂盈利但后来又亏损不算稳定回本。"},
        {"字段": "截至当前累计利润", "大白话是什么意思": "这批ASIN从上架到2026-06最终合计赚了或亏了多少钱。", "具体怎么算": "汇总上架月到2026-06全部gross_profit。", "数据从哪里来": "lingxing_profit_asin.gross_profit", "容易误解的地方": "GBP和EUR分别统计，不能相加。"},
        {"字段": "单个ASIN已稳定回本数", "大白话是什么意思": "这批里面最终有多少个单独ASIN已经稳定回正。", "具体怎么算": "每个ASIN单独累计利润，最终>=0并找到稳定回正月份才计入。", "数据从哪里来": "单个ASIN逐月gross_profit", "容易误解的地方": "批次整体未回本时，仍可能有部分ASIN单独回本。"},
    ]


def build_monthly_cohort_model(
    asins: dict[str, dict[str, Any]],
    sku_to_asins: dict[str, set[str]],
    plans: dict[str, list[dict[str, Any]]],
    finance: dict[tuple[str, str], list[dict[str, Any]]],
    fixed_cohorts: dict[str, dict[str, str]],
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    cutoff_month = MODEL_END.strftime("%Y-%m")
    for currency in CURRENCIES:
        details: list[dict[str, Any]] = []
        detail_internal: dict[str, dict[str, Any]] = {}
        for asin, launch_month in sorted(fixed_cohorts[currency].items(), key=lambda item: (item[1], item[0])):
            record = asins.get(asin)
            if record is None or not MODEL_START.strftime("%Y-%m") <= launch_month <= cutoff_month:
                continue
            purchase = allocate_q123(asin, record, sku_to_asins, plans)
            monthly = monthly_finance_by_asin(finance.get((asin, currency), []), launch_month)
            months = months_between(launch_month, cutoff_month)
            cumulative_values: list[Decimal] = []
            cumulative = Decimal(0)
            for month in months:
                cumulative += monthly.get(month, {}).get("profit", Decimal(0))
                cumulative_values.append(cumulative)
            has_finance = any(value.get("records", 0) > 0 for value in monthly.values())
            recovery_index = stable_recovery_index(cumulative_values) if has_finance else None
            current_profit = cumulative_values[-1] if cumulative_values else Decimal(0)
            current_status = "无财务记录" if not has_finance else ("已稳定回本" if recovery_index is not None else "未回本")
            item = {
                "上架批次": launch_month,
                "ASIN": asin,
                "开发人": record["developer"],
                "基准SKU": " | ".join(sorted(record["skus"])),
                "当前标签状态": listing_state(str(record.get("label") or "")),
                "Q1计划量": purchase["q1"] if purchase["q1"] is not None else "",
                "Q2计划量": purchase["q2"] if purchase["q2"] is not None else "",
                "Q3计划量": purchase["q3"] if purchase["q3"] is not None else "",
                "Q1采购计划号": purchase["q1_plans"],
                "Q2采购计划号": purchase["q2_plans"],
                "Q3采购计划号": purchase["q3_plans"],
                "已完成采购计划批次数": purchase["completed_batches"],
                "Q1-Q3是否完整": "是" if all(purchase[key] is not None for key in ("q1", "q2", "q3")) else "否",
                "采购计划分摊方式": purchase["allocation"],
                "有财务记录月份数": sum(value.get("records", 0) > 0 for value in monthly.values()),
                "首月结算毛利润": monthly.get(launch_month, {}).get("profit", Decimal(0)),
                "最近月结算毛利润": monthly.get(cutoff_month, {}).get("profit", Decimal(0)),
                "截至当前累计利润": current_profit,
                "当前回本状态": current_status,
                "稳定回本月数": recovery_index + 1 if recovery_index is not None else "",
                "稳定回本自然月": months[recovery_index] if recovery_index is not None else "",
            }
            details.append(item)
            detail_internal[asin] = {"row": item, "monthly": monthly, "months": months, "cumulative": cumulative_values}

        by_cohort: dict[str, list[str]] = defaultdict(list)
        for item in details:
            by_cohort[str(item["上架批次"])].append(str(item["ASIN"]))

        overview: list[dict[str, Any]] = []
        timelines: dict[str, list[dict[str, Any]]] = {}
        for launch_month, cohort_asins in sorted(by_cohort.items()):
            cohort_details = [detail_internal[asin] for asin in cohort_asins]
            asin_count = len(cohort_details)
            q_counts = {
                q: sum(detail["row"][f"{q}计划量"] != "" for detail in cohort_details)
                for q in ("Q1", "Q2", "Q3")
            }
            q_totals = {
                q: sum((dec(detail["row"][f"{q}计划量"]) for detail in cohort_details), Decimal(0))
                for q in ("Q1", "Q2", "Q3")
            }
            complete_count = sum(detail["row"]["Q1-Q3是否完整"] == "是" for detail in cohort_details)
            months = months_between(launch_month, cutoff_month)
            cumulative = Decimal(0)
            cumulative_values: list[Decimal] = []
            timeline: list[dict[str, Any]] = []
            previous_profit: Decimal | None = None
            any_finance = False
            for index, month in enumerate(months, start=1):
                month_values = [detail["monthly"].get(month, {}) for detail in cohort_details]
                finance_values = [value for value in month_values if value.get("records", 0) > 0]
                any_finance = any_finance or bool(finance_values)
                month_profit = sum((value.get("profit", Decimal(0)) for value in month_values), Decimal(0))
                cumulative += month_profit
                cumulative_values.append(cumulative)
                asin_cumulative = [detail["cumulative"][index - 1] for detail in cohort_details]
                timeline.append({
                    "上架批次": launch_month,
                    "批次ASIN数": asin_count,
                    "Q1有计划ASIN数": q_counts["Q1"], "Q1计划量": q_totals["Q1"],
                    "Q2有计划ASIN数": q_counts["Q2"], "Q2计划量": q_totals["Q2"],
                    "Q3有计划ASIN数": q_counts["Q3"], "Q3计划量": q_totals["Q3"],
                    "Q1-Q3完整ASIN数": complete_count,
                    "经营第几个月": index,
                    "财务自然月": month,
                    "当月有财务记录ASIN数": len(finance_values),
                    "当月财务销量": sum((value.get("quantity", Decimal(0)) for value in month_values), Decimal(0)),
                    "当月销售额": sum((value.get("sales", Decimal(0)) for value in month_values), Decimal(0)),
                    "当月结算毛利润": month_profit,
                    "较上月利润变化": month_profit - previous_profit if previous_profit is not None else "",
                    "累计结算毛利润": cumulative,
                    "截至该月累计盈利ASIN数": sum(value > 0 for value in asin_cumulative),
                    "截至该月累计亏损ASIN数": sum(value < 0 for value in asin_cumulative),
                    "截至该月累计持平ASIN数": sum(value == 0 for value in asin_cumulative),
                    "批次截至该月状态": "无财务记录" if not any_finance else ("累计已回本" if cumulative >= 0 else "累计未回本"),
                })
                previous_profit = month_profit
            recovery_index = stable_recovery_index(cumulative_values) if any_finance else None
            first_three = [timeline[index]["当月结算毛利润"] if index < len(timeline) else "" for index in range(3)]
            overview.append({
                "上架批次": launch_month,
                "ASIN总数": asin_count,
                "Q1有计划ASIN数": q_counts["Q1"], "Q1覆盖率": Decimal(q_counts["Q1"]) / asin_count if asin_count else None, "Q1计划量": q_totals["Q1"],
                "Q2有计划ASIN数": q_counts["Q2"], "Q2覆盖率": Decimal(q_counts["Q2"]) / asin_count if asin_count else None, "Q2计划量": q_totals["Q2"],
                "Q3有计划ASIN数": q_counts["Q3"], "Q3覆盖率": Decimal(q_counts["Q3"]) / asin_count if asin_count else None, "Q3计划量": q_totals["Q3"],
                "Q1-Q3完整ASIN数": complete_count,
                "首月结算毛利润": first_three[0], "第2个月结算毛利润": first_three[1], "第3个月结算毛利润": first_three[2],
                "稳定回本月数": recovery_index + 1 if recovery_index is not None else "",
                "稳定回本自然月": months[recovery_index] if recovery_index is not None else "",
                "截至当前累计利润": cumulative_values[-1] if cumulative_values else Decimal(0),
                "批次当前状态": "无财务记录" if not any_finance else ("已稳定回本" if recovery_index is not None else "未回本"),
                "最近月结算毛利润": timeline[-1]["当月结算毛利润"] if timeline else Decimal(0),
                "单个ASIN已稳定回本数": sum(detail["row"]["当前回本状态"] == "已稳定回本" for detail in cohort_details),
            })
            timelines[launch_month] = timeline
        result[currency] = {"overview": overview, "timelines": timelines, "details": details}
    return result


def monthly_overview_total(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {}
    asin_count = sum(int(row["ASIN总数"]) for row in rows)
    total: dict[str, Any] = {
        "上架批次": "合计",
        "ASIN总数": asin_count,
        "Q1有计划ASIN数": sum(int(row["Q1有计划ASIN数"]) for row in rows),
        "Q1计划量": sum((dec(row["Q1计划量"]) for row in rows), Decimal(0)),
        "Q2有计划ASIN数": sum(int(row["Q2有计划ASIN数"]) for row in rows),
        "Q2计划量": sum((dec(row["Q2计划量"]) for row in rows), Decimal(0)),
        "Q3有计划ASIN数": sum(int(row["Q3有计划ASIN数"]) for row in rows),
        "Q3计划量": sum((dec(row["Q3计划量"]) for row in rows), Decimal(0)),
        "Q1-Q3完整ASIN数": sum(int(row["Q1-Q3完整ASIN数"]) for row in rows),
        "首月结算毛利润": sum((dec(row["首月结算毛利润"]) for row in rows), Decimal(0)),
        "第2个月结算毛利润": sum((dec(row["第2个月结算毛利润"]) for row in rows if row["第2个月结算毛利润"] != ""), Decimal(0)),
        "第3个月结算毛利润": sum((dec(row["第3个月结算毛利润"]) for row in rows if row["第3个月结算毛利润"] != ""), Decimal(0)),
        "稳定回本月数": "",
        "稳定回本自然月": "",
        "截至当前累计利润": sum((dec(row["截至当前累计利润"]) for row in rows), Decimal(0)),
        "最近月结算毛利润": sum((dec(row["最近月结算毛利润"]) for row in rows), Decimal(0)),
        "单个ASIN已稳定回本数": sum(int(row["单个ASIN已稳定回本数"]) for row in rows),
    }
    for q in ("Q1", "Q2", "Q3"):
        total[f"{q}覆盖率"] = Decimal(total[f"{q}有计划ASIN数"]) / asin_count if asin_count else None
    total["批次当前状态"] = "已盈利" if total["截至当前累计利润"] >= 0 else "累计亏损"
    return total


def allocate_finance_to_q_batches(
    financial: list[dict[str, Any]], purchase: dict[str, Any], launch_month: str,
) -> dict[str, Any]:
    """按累计财务销量依次落入 Q1/Q2/Q3，并计算每批真实投入后的回收余额。"""
    batch_names = ("Q1", "Q2", "Q3")
    capacities = {name: dec(purchase[name.lower()] or 0) for name in batch_names}
    monthly: dict[str, dict[str, dict[str, Decimal]]] = {
        name: defaultdict(lambda: {
            "quantity": Decimal(0), "sales": Decimal(0), "profit": Decimal(0),
            "recognized_cost": Decimal(0), "contribution": Decimal(0),
        })
        for name in (*batch_names, "超出Q3")
    }
    positive_quantity = sum((max(dec(row["quantity"]), Decimal(0)) for row in financial), Decimal(0))
    recognized_total = sum(
        (abs(dec(row.get("cg_price"))) + abs(dec(row.get("cg_transport"))) for row in financial),
        Decimal(0),
    )
    fallback_unit_cost = recognized_total / positive_quantity if positive_quantity > 0 else Decimal(0)
    position = Decimal(0)

    def bucket_for(current: Decimal) -> tuple[str, Decimal | None]:
        boundary = Decimal(0)
        for name in batch_names:
            capacity = capacities[name]
            if capacity <= 0:
                continue
            if current < boundary + capacity:
                return name, boundary + capacity - current
            boundary += capacity
        return "超出Q3", None

    for row in sorted(financial, key=lambda item: item["date"]):
        month = row["date"].strftime("%Y-%m")
        if month < launch_month or month > MODEL_END.strftime("%Y-%m"):
            continue
        quantity = dec(row["quantity"])
        metrics = {
            "quantity": quantity,
            "sales": dec(row["sales"]),
            "profit": dec(row["profit"]),
            "recognized_cost": abs(dec(row.get("cg_price"))) + abs(dec(row.get("cg_transport"))),
        }
        metrics["contribution"] = metrics["profit"] + metrics["recognized_cost"]
        if quantity > 0:
            remaining = quantity
            while remaining > 0:
                name, available = bucket_for(position)
                portion = remaining if available is None else min(remaining, available)
                ratio = portion / quantity
                value = monthly[name][month]
                for field in ("sales", "profit", "recognized_cost", "contribution"):
                    value[field] += metrics[field] * ratio
                value["quantity"] += portion
                position += portion
                remaining -= portion
        else:
            # 无销量费用或退款归到当时正在销售的批次，不凭空创造新批次。
            name, _available = bucket_for(max(position - Decimal("0.000001"), Decimal(0)))
            value = monthly[name][month]
            for field, amount in metrics.items():
                value[field] += amount
            position = max(Decimal(0), position + quantity)

    batches: dict[str, dict[str, Any]] = {}
    cutoff_month = MODEL_END.strftime("%Y-%m")
    for name in batch_names:
        planned = capacities[name]
        values = monthly[name]
        sold = sum((value["quantity"] for value in values.values()), Decimal(0))
        recognized = sum((value["recognized_cost"] for value in values.values()), Decimal(0))
        unit_cost = recognized / sold if sold > 0 and recognized > 0 else fallback_unit_cost
        investment = planned * unit_cost if planned > 0 and unit_cost > 0 else None
        first_sale_month = min((month for month, value in values.items() if value["quantity"] > 0), default=None)
        start_month = launch_month if name == "Q1" and planned > 0 else first_sale_month
        balance_rows: list[tuple[str, Decimal]] = []
        balance = -investment if investment is not None and start_month is not None else Decimal(0)
        if start_month is not None:
            for month in months_between(start_month, cutoff_month):
                balance += values.get(month, {}).get("contribution", Decimal(0))
                balance_rows.append((month, balance))
        recovery_index = stable_recovery_index([value for _month, value in balance_rows]) if investment is not None else None
        if planned <= 0:
            status = "无已完成采购计划"
        elif investment is None:
            status = "财务成本不足，无法判断"
        elif start_month is None:
            status = "未开始售卖"
        elif recovery_index is None:
            status = "未回本"
        else:
            status = "已回本"
        batches[name] = {
            "planned": planned,
            "sold": sold,
            "unit_cost": unit_cost if unit_cost > 0 else None,
            "investment": investment,
            "gross_profit": sum((value["profit"] for value in values.values()), Decimal(0)),
            "contribution": sum((value["contribution"] for value in values.values()), Decimal(0)),
            "current_balance": balance if investment is not None and start_month is not None else None,
            "start_month": start_month,
            "recovery_months": recovery_index + 1 if recovery_index is not None else None,
            "recovery_natural_month": balance_rows[recovery_index][0] if recovery_index is not None else None,
            "status": status,
            "monthly": values,
            "balance_rows": dict(balance_rows),
        }
    return {
        "batches": batches,
        "overflow": monthly["超出Q3"],
        "total_gross_profit": sum((dec(row["profit"]) for row in financial if launch_month <= row["date"].strftime("%Y-%m") <= cutoff_month), Decimal(0)),
    }


def build_combined_q_profit_model(
    asins: dict[str, dict[str, Any]],
    sku_to_asins: dict[str, set[str]],
    plans: dict[str, list[dict[str, Any]]],
    finance: dict[tuple[str, str], list[dict[str, Any]]],
    fixed_cohorts: dict[str, dict[str, str]],
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    cutoff_month = MODEL_END.strftime("%Y-%m")
    for currency in CURRENCIES:
        internal: list[dict[str, Any]] = []
        for asin, launch_month in sorted(fixed_cohorts[currency].items(), key=lambda item: (item[1], item[0])):
            record = asins.get(asin)
            if record is None or not MODEL_START.strftime("%Y-%m") <= launch_month <= cutoff_month:
                continue
            purchase = allocate_q123(asin, record, sku_to_asins, plans)
            allocated = allocate_finance_to_q_batches(finance.get((asin, currency), []), purchase, launch_month)
            detail: dict[str, Any] = {
                "上架批次": launch_month, "ASIN": asin, "开发人": record["developer"],
                "基准SKU": " | ".join(sorted(record["skus"])),
                "当前标签状态": listing_state(str(record.get("label") or "")),
                "采购计划分摊方式": purchase["allocation"],
                "截至当前全部结算毛利润": allocated["total_gross_profit"],
            }
            for name in ("Q1", "Q2", "Q3"):
                batch = allocated["batches"][name]
                detail.update({
                    f"{name}计划量": batch["planned"] if batch["planned"] > 0 else "",
                    f"{name}已售量": batch["sold"],
                    f"{name}售罄率": batch["sold"] / batch["planned"] if batch["planned"] > 0 else None,
                    f"{name}财务单位采购及运输成本": batch["unit_cost"] if batch["unit_cost"] is not None else "",
                    f"{name}估算投入成本": batch["investment"] if batch["investment"] is not None else "",
                    f"{name}结算毛利润": batch["gross_profit"],
                    f"{name}当前回收余额": batch["current_balance"] if batch["current_balance"] is not None else "",
                    f"{name}回本状态": batch["status"],
                    f"{name}开始售卖月": batch["start_month"] or "",
                    f"{name}稳定回本月数": batch["recovery_months"] or "",
                    f"{name}稳定回本自然月": batch["recovery_natural_month"] or "",
                })
            internal.append({"detail": detail, "allocated": allocated, "launch_month": launch_month})

        by_cohort: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for item in internal:
            by_cohort[item["launch_month"]].append(item)
        overview: list[dict[str, Any]] = []
        q_summary: list[dict[str, Any]] = []
        timelines: dict[str, list[dict[str, Any]]] = {}
        for launch_month, cohort in sorted(by_cohort.items()):
            months = months_between(launch_month, cutoff_month)
            timeline: list[dict[str, Any]] = []
            cohort_batches: dict[str, dict[str, Any]] = {}
            for name in ("Q1", "Q2", "Q3"):
                batch_items = [item["allocated"]["batches"][name] for item in cohort]
                planned_items = [item for item in batch_items if item["planned"] > 0]
                started_items = [item for item in planned_items if item["start_month"] is not None]
                cost_items = [item for item in started_items if item["investment"] is not None]
                monthly_net: dict[str, Decimal] = defaultdict(Decimal)
                monthly_profit: dict[str, Decimal] = defaultdict(Decimal)
                monthly_qty: dict[str, Decimal] = defaultdict(Decimal)
                for batch in started_items:
                    for month, value in batch["monthly"].items():
                        monthly_profit[month] += value["profit"]
                        monthly_qty[month] += value["quantity"]
                for batch in cost_items:
                    monthly_net[str(batch["start_month"])] -= dec(batch["investment"])
                    for month, value in batch["monthly"].items():
                        monthly_net[month] += value["contribution"]
                balance = Decimal(0)
                balances: list[Decimal] = []
                for month in months:
                    balance += monthly_net[month]
                    balances.append(balance)
                first_start_month = min((str(item["start_month"]) for item in cost_items), default=None)
                start_index = months.index(first_start_month) if first_start_month in months else 0
                local_recovery = (
                    stable_recovery_index(balances[start_index:])
                    if cost_items
                    else None
                )
                recovery_index = start_index + local_recovery if local_recovery is not None else None
                if not planned_items:
                    status = "无已完成采购计划"
                elif not started_items:
                    status = "未开始售卖"
                elif len(cost_items) < len(started_items):
                    status = "成本已覆盖部分已回本" if recovery_index is not None else "成本已覆盖部分未回本"
                elif recovery_index is None:
                    status = "已开始部分未回本" if len(started_items) < len(planned_items) else "未回本"
                else:
                    status = "已开始部分已回本" if len(started_items) < len(planned_items) else "已回本"
                value = {
                    "上架批次": launch_month, "Q批次": name, "批次ASIN总数": len(cohort),
                    "有该Q计划ASIN数": len(planned_items), "已开始售卖ASIN数": len(started_items),
                    "尚未开始售卖ASIN数": len(planned_items) - len(started_items),
                    "成本可计算ASIN数": len(cost_items),
                    "计划量": sum((item["planned"] for item in planned_items), Decimal(0)),
                    "已售量": sum((item["sold"] for item in planned_items), Decimal(0)),
                    "售罄率": sum((item["sold"] for item in planned_items), Decimal(0)) / sum((item["planned"] for item in planned_items), Decimal(0)) if planned_items and sum((item["planned"] for item in planned_items), Decimal(0)) > 0 else None,
                    "估算投入成本": sum((dec(item["investment"]) for item in cost_items), Decimal(0)),
                    "结算毛利润": sum((item["gross_profit"] for item in planned_items), Decimal(0)),
                    "当前回收余额": balance if cost_items else "",
                    "回本状态": status,
                    "稳定回本月数（从上架月算）": recovery_index + 1 if recovery_index is not None else "",
                    "稳定回本自然月": months[recovery_index] if recovery_index is not None else "",
                    "单个ASIN已回本数": sum(item["status"] == "已回本" for item in planned_items),
                    "monthly_net": monthly_net, "monthly_profit": monthly_profit,
                    "monthly_qty": monthly_qty, "balances": balances,
                }
                cohort_batches[name] = value
                q_summary.append({key: item for key, item in value.items() if not key.startswith("monthly_") and key != "balances"})

            total_cumulative_profit = Decimal(0)
            for index, month in enumerate(months, start=1):
                row: dict[str, Any] = {"上架批次": launch_month, "经营第几个月": index, "财务自然月": month}
                q_month_profit = Decimal(0)
                for name in ("Q1", "Q2", "Q3"):
                    batch = cohort_batches[name]
                    profit = batch["monthly_profit"][month]
                    q_month_profit += profit
                    row.update({
                        f"{name}当月销量": batch["monthly_qty"][month],
                        f"{name}当月结算毛利润": profit,
                        f"{name}累计回收余额": batch["balances"][index - 1] if batch["balances"] else "",
                    })
                overflow_profit = sum(
                    (item["allocated"]["overflow"].get(month, {}).get("profit", Decimal(0)) for item in cohort),
                    Decimal(0),
                )
                total_month_profit = q_month_profit + overflow_profit
                total_cumulative_profit += total_month_profit
                row.update({
                    "超出Q3当月结算毛利润": overflow_profit,
                    "当月全部结算毛利润": total_month_profit,
                    "累计全部结算毛利润": total_cumulative_profit,
                })
                timeline.append(row)
            timelines[launch_month] = timeline
            overview_row: dict[str, Any] = {"上架批次": launch_month, "ASIN总数": len(cohort)}
            for name in ("Q1", "Q2", "Q3"):
                value = cohort_batches[name]
                overview_row.update({
                    f"{name}计划ASIN数": value["有该Q计划ASIN数"],
                    f"{name}计划量": value["计划量"],
                    f"{name}已售量": value["已售量"],
                    f"{name}结算毛利润": value["结算毛利润"],
                    f"{name}当前回收余额": value["当前回收余额"],
                    f"{name}回本状态": value["回本状态"],
                    f"{name}稳定回本月数": value["稳定回本月数（从上架月算）"],
                })
            overview_row["截至当前全部结算毛利润"] = sum((item["allocated"]["total_gross_profit"] for item in cohort), Decimal(0))
            overview.append(overview_row)
        result[currency] = {
            "overview": overview,
            "q_summary": q_summary,
            "timelines": timelines,
            "details": [item["detail"] for item in internal],
        }
    return result


def combined_q_explanations(currency: str) -> list[dict[str, str]]:
    return [
        {"字段": "上架批次", "大白话是什么意思": "同一个上架月份的固定ASIN集合。", "具体怎么算": "直接读取基础统一表的模型分析起算月。", "数据从哪里来": "ASIN统一基础表", "容易误解的地方": "不是采购计划创建月。"},
        {"字段": "Q1/Q2/Q3计划量", "大白话是什么意思": "这个ASIN第1、第2、第3次已完成采购计划分别计划买多少件。", "具体怎么算": "同SKU已完成采购计划按创建时间排序；第1个为Q1，第2个为Q2，第3个为Q3。SKU对多个ASIN时等额分摊。", "数据从哪里来": "lingxing_purchase_plan.quantity_plan", "容易误解的地方": "只使用采购数量，不使用采购计划时间切利润周期。"},
        {"字段": "Q1/Q2/Q3已售量", "大白话是什么意思": "累计销量中已经消耗了该Q批多少件。", "具体怎么算": "从上架后财务销量开始，前Q1计划量件归Q1；超过Q1后归Q2；超过Q1+Q2后归Q3。", "数据从哪里来": "lingxing_profit_asin.total_sales_quantity + Q计划量", "容易误解的地方": "这是按数量依次消耗，不再依靠FBA断货判断。"},
        {"字段": "财务单位采购及运输成本", "大白话是什么意思": "财务表确认的每件商品采购成本加采购运输成本。", "具体怎么算": "该Q批已售部分的 abs(cg_price)+abs(cg_transport_costs)，除以该Q批财务销量；没有该批销量时使用该ASIN全期平均单位成本兜底。", "数据从哪里来": "lingxing_profit_asin.cg_price、cg_transport_costs", "容易误解的地方": "这是财务成本估算，不是采购计划表中的报价。"},
        {"字段": "估算投入成本", "大白话是什么意思": "为了采购这一整批货，估算需要先压进去多少钱。", "具体怎么算": "Q批计划量×财务单位采购及运输成本。", "数据从哪里来": "采购计划量+财务单位成本", "容易误解的地方": "没有成本数据时不能判断真实回本。"},
        {"字段": "结算毛利润", "大白话是什么意思": "销售落入该Q批的商品，领星财务最终确认赚了或亏了多少钱。", "具体怎么算": "每天的销量按Q1→Q2→Q3数量边界分配，销售额、成本和gross_profit按当天分配销量比例切到对应Q批。", "数据从哪里来": "lingxing_profit_asin.gross_profit", "容易误解的地方": "它是会计结算利润，不直接等于现金已经收回多少。"},
        {"字段": "每月回收金额", "大白话是什么意思": "这个月通过销售实际收回、可用于覆盖最初采购投入的钱。", "具体怎么算": "该Q批当月gross_profit + 财务已确认的采购成本 + 采购运输成本。因为采购成本已在前面作为整批投入扣除，这里要把随销售确认的成本加回来，避免重复扣成本。", "数据从哪里来": "gross_profit、cg_price、cg_transport_costs", "容易误解的地方": "不是简单使用gross_profit从0累计。"},
        {"字段": "当前回收余额", "大白话是什么意思": "截至2026-06，这一批货收回初始投入后还剩多少；负数表示还差多少回本。", "具体怎么算": "-估算投入成本 + 从该Q批开始售卖至今的每月回收金额累计。", "数据从哪里来": "Q批估算投入成本+逐月回收金额", "容易误解的地方": "负数才是真正尚未回本；正数表示已经收回投入并产生现金盈余。"},
        {"字段": "稳定回本月数", "大白话是什么意思": "从上架月份开始数，第几个月真正收回投入，并且后面没有再次跌回负数。", "具体怎么算": "当前回收余额最终必须>=0，再找回收余额回正后不再跌回负数的第一个月。Q2/Q3未开始售卖前的0余额不算回本。", "数据从哪里来": "逐月Q批回收余额", "容易误解的地方": "首月gross_profit为正，不代表首月回本；必须先覆盖整批估算投入。"},
        {"字段": "超出Q3利润", "大白话是什么意思": "累计销量已经超过Q1+Q2+Q3计划量后，仍发生的利润。", "具体怎么算": "超过三批计划量边界的销量和利润单独放到超出Q3，不强行塞回Q1。", "数据从哪里来": "财务累计销量与Q1-Q3数量边界", "容易误解的地方": "可能意味着存在Q4及以后计划、组合映射误差或采购计划未补齐。"},
        {"字段": "币种", "大白话是什么意思": "当前工作簿所有金额使用的货币。", "具体怎么算": "当前文件固定为{0}。".format(currency), "数据从哪里来": "lingxing_profit_asin.currency_code", "容易误解的地方": "GBP与EUR不换汇、不相加。"},
    ]


def combined_overview_total(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {}
    total: dict[str, Any] = {
        "上架批次": "合计",
        "ASIN总数": sum(int(row["ASIN总数"]) for row in rows),
        "截至当前全部结算毛利润": sum((dec(row["截至当前全部结算毛利润"]) for row in rows), Decimal(0)),
    }
    for name in ("Q1", "Q2", "Q3"):
        total.update({
            f"{name}计划ASIN数": sum(int(row[f"{name}计划ASIN数"]) for row in rows),
            f"{name}计划量": sum((dec(row[f"{name}计划量"]) for row in rows), Decimal(0)),
            f"{name}已售量": sum((dec(row[f"{name}已售量"]) for row in rows), Decimal(0)),
            f"{name}结算毛利润": sum((dec(row[f"{name}结算毛利润"]) for row in rows), Decimal(0)),
            f"{name}当前回收余额": sum((dec(row[f"{name}当前回收余额"]) for row in rows if row[f"{name}当前回收余额"] != ""), Decimal(0)),
            f"{name}回本状态": "见各月份",
            f"{name}稳定回本月数": "",
        })
    return total


def q_summary_totals(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals: list[dict[str, Any]] = []
    for name in ("Q1", "Q2", "Q3"):
        values = [row for row in rows if row["Q批次"] == name]
        if not values:
            continue
        planned = sum((dec(row["计划量"]) for row in values), Decimal(0))
        sold = sum((dec(row["已售量"]) for row in values), Decimal(0))
        totals.append({
            "上架批次": "合计", "Q批次": name,
            "批次ASIN总数": sum(int(row["批次ASIN总数"]) for row in values),
            "有该Q计划ASIN数": sum(int(row["有该Q计划ASIN数"]) for row in values),
            "已开始售卖ASIN数": sum(int(row["已开始售卖ASIN数"]) for row in values),
            "尚未开始售卖ASIN数": sum(int(row["尚未开始售卖ASIN数"]) for row in values),
            "成本可计算ASIN数": sum(int(row["成本可计算ASIN数"]) for row in values),
            "计划量": planned, "已售量": sold,
            "售罄率": sold / planned if planned else None,
            "估算投入成本": sum((dec(row["估算投入成本"]) for row in values), Decimal(0)),
            "结算毛利润": sum((dec(row["结算毛利润"]) for row in values), Decimal(0)),
            "当前回收余额": sum((dec(row["当前回收余额"]) for row in values if row["当前回收余额"] != ""), Decimal(0)),
            "回本状态": "见各月份", "稳定回本月数（从上架月算）": "", "稳定回本自然月": "",
            "单个ASIN已回本数": sum(int(row["单个ASIN已回本数"]) for row in values),
        })
    return totals


def asin_cycle_detail_sheets(rows: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    """Build one sheet per observed FBA cycle, with one ASIN-cycle per row."""
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if str(row["FBA周期类型"]) == "未观察到FBA":
            continue
        cycle_no = int(row["FBA周期序号"])
        payback_days = row["周期财务回本天数"]
        grouped[cycle_no].append({
            "上架批次": row["上架批次"],
            "ASIN": row["ASIN"],
            "开发人": row["开发人"],
            "基准SKU": row["基准SKU"],
            "当前标签状态": row["当前标签状态"],
            "首次FBA可售日期": row["首次FBA可售日期"],
            "上架起点依据": row["上架起点依据"],
            "FBA周期": fba_cycle_name(cycle_no),
            "周期开始": row["周期开始"],
            "周期结束/观察截止": row["周期结束"],
            "前次断货观察周": row["前次断货观察周"],
            "售卖周期天数": row["售卖周期天数"],
            "周期FBA可售周数": row["周期FBA可售周数"],
            "Q1计划量": row["Q1计划量"],
            "Q2计划量": row["Q2计划量"],
            "采购计划关联": row["采购计划关联"],
            "周期产品表现销量": row["周期销量"],
            "周期产品表现销售额": row["周期产品表现销售额"],
            "有财务记录": "是" if row["有财务记录"] else "否",
            "财务销量": row["财务销量"],
            "销售额": row["结算销售额"],
            "广告费支出": row["广告费支出"],
            "财务总成本支出": row["财务总成本支出"],
            "其他结算支出/调整": row["其他结算支出/调整"],
            "领星结算毛利润": row["结算毛利润"],
            "是否周期内回本": "是" if payback_days != "" else "否",
            "周期财务回本天数": payback_days,
            "最新观察FBA可售库存": row["最新观察FBA可售库存"],
            "最新FBA观察周": row["最新FBA观察周"],
        })

    sheets: list[tuple[str, list[dict[str, Any]]]] = []
    for cycle_no in sorted(grouped):
        details = sorted(
            grouped[cycle_no],
            key=lambda item: (item["上架批次"], item["开发人"], item["ASIN"]),
        )
        sheets.append((fba_cycle_sheet_name(cycle_no), details))
    return sheets


def cohort_summary(rows: list[dict[str, Any]], summaries: dict[str, dict[str, Any]], currency: str) -> list[dict[str, Any]]:
    asins = {row["ASIN"] for row in rows}
    grouped: dict[str, list[str]] = defaultdict(list)
    for asin in asins:
        grouped[summaries[asin]["launch_month"]].append(asin)
    output: list[dict[str, Any]] = []
    for month, members in sorted(grouped.items()):
        all_rows = [summaries[asin] for asin in members]
        finance = [row.get(f"finance_{currency}", metric_total([])) for row in all_rows]
        retained = [row for row in all_rows if row["label_status"] != "已淘汰"]
        latest_fba_active = [row for row in all_rows if row["latest_fba"] > 0]
        eliminated = [row for row in all_rows if row["label_status"] == "已淘汰"]
        pending = [row for row in all_rows if row["label_status"] == "待淘汰"]
        total_sales = sum((row.get(f"performance_{currency}", {}).get("volume", Decimal(0)) for row in all_rows), Decimal(0))
        total_finance_sales = sum((row["sales"] for row in finance), Decimal(0))
        total_profit = sum((row["profit"] for row in finance), Decimal(0))
        retained_finance = [row.get(f"finance_{currency}", metric_total([])) for row in retained]
        eliminated_finance = [row.get(f"finance_{currency}", metric_total([])) for row in eliminated]
        payback_weeks = [
            relative_week(row["launch_date"], row[f"payback_{currency}"])
            for row in all_rows if row.get(f"payback_{currency}")
        ]
        output.append({
            "上架批次": month,
            "ASIN总数": len(all_rows),
            "总销售量": total_sales,
            "总销售额": sum((row.get(f"performance_{currency}", {}).get("amount", Decimal(0)) for row in all_rows), Decimal(0)),
            "总结算利润": total_profit,
            "总结算销售额": total_finance_sales,
            "总利润率": (total_profit / total_finance_sales) if total_finance_sales else None,
            "最新观察FBA可售库存": sum((row["latest_fba"] for row in all_rows), Decimal(0)),
            "最新观察FBA可售ASIN数": len(latest_fba_active),
            "最新观察FBA可售率": Decimal(len(latest_fba_active)) / Decimal(len(all_rows)) if all_rows else Decimal(0),
            "留存ASIN销售量": sum((row.get(f"performance_{currency}", {}).get("volume", Decimal(0)) for row in retained), Decimal(0)),
            "留存ASIN销售额": sum((row["sales"] for row in retained_finance), Decimal(0)),
            "留存ASIN数": len(retained),
            "留存ASIN结算利润": sum((row["profit"] for row in retained_finance), Decimal(0)),
            "留存ASIN利润率": (sum((row["profit"] for row in retained_finance), Decimal(0)) / sum((row["sales"] for row in retained_finance), Decimal(0))) if sum((row["sales"] for row in retained_finance), Decimal(0)) else None,
            "淘汰ASIN销售量": sum((row.get(f"performance_{currency}", {}).get("volume", Decimal(0)) for row in eliminated), Decimal(0)),
            "淘汰ASIN销售额": sum((row["sales"] for row in eliminated_finance), Decimal(0)),
            "淘汰ASIN总数": len(eliminated),
            "淘汰率": Decimal(len(eliminated)) / Decimal(len(all_rows)) if all_rows else Decimal(0),
            "淘汰ASIN结算利润": sum((row["profit"] for row in eliminated_finance), Decimal(0)),
            "淘汰ASIN利润率": (sum((row["profit"] for row in eliminated_finance), Decimal(0)) / sum((row["sales"] for row in eliminated_finance), Decimal(0))) if sum((row["sales"] for row in eliminated_finance), Decimal(0)) else None,
            "待淘汰ASIN数": len(pending),
            "财务回正ASIN数": len(payback_weeks),
            "平均财务回正周数": Decimal(str(mean(payback_weeks))) if payback_weeks else None,
        })
    return output


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0]) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=headers)
        writer.writeheader()
        writer.writerows([{key: text(value) for key, value in row.items()} for row in rows])


def write_workbook(path: Path, sheets: list[tuple[str, list[dict[str, Any]]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets:
        sheet = workbook.create_sheet(name[:31])
        if not rows:
            sheet.append(["暂无数据"])
            continue
        headers = list(rows[0])
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in rows:
            values: list[Any] = []
            for header in headers:
                value = row.get(header, "")
                # 保持金额、销量、天数为真正的 Excel 数值，便于后续筛选和求和；
                # 仅将 ASIN、SKU、状态等维度保留为文本。
                if isinstance(value, Decimal):
                    values.append(float(value))
                else:
                    values.append(value)
            sheet.append(values)
        for column, header in enumerate(headers, start=1):
            if "率" in header:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, column).number_format = "0.00%"
            elif any(keyword in header for keyword in ("金额", "销售额", "支出", "利润", "成本", "销量", "库存", "计划量", "天数")):
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, column).number_format = "#,##0.00"
            elif "数" in header or "序号" in header:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, column).number_format = "#,##0"
            elif "日期" in header or "观察周" in header or "周期开始" in header or "周期结束" in header:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, column).number_format = "yyyy-mm-dd"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in range(1, len(headers) + 1):
            width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 80) + 1))
            sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 35)
        if name == "字段说明":
            widths = {"A": 24, "B": 42, "C": 72, "D": 48, "E": 58}
            for column, width in widths.items():
                sheet.column_dimensions[column].width = width
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                sheet.row_dimensions[row[0].row].height = 72
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_relative_week(rows: list[dict[str, Any]], summaries: dict[str, dict[str, Any]], weekly: dict[str, list[dict[str, Any]]], finance: dict[tuple[str, str], list[dict[str, Any]]], currency: str) -> list[dict[str, Any]]:
    bucket: dict[tuple[str, str, int], dict[str, Any]] = {}

    def value_for(cohort: str, developer: str, relative: int) -> dict[str, Any]:
        return bucket.setdefault((cohort, developer, relative), {
            "上架批次": cohort, "开发人": developer, "相对周": relative,
            "成熟ASIN": set(), "FBA可售ASIN": set(),
            "产品表现销量": Decimal(0), "产品表现销售额": Decimal(0),
            "财务销量": Decimal(0), "结算销售额": Decimal(0), "结算毛利润": Decimal(0),
            "广告费支出": Decimal(0), "财务总成本支出": Decimal(0),
        })

    for asin, summary in summaries.items():
        if not any(row["ASIN"] == asin for row in rows):
            continue
        start = summary["launch_date"]
        developer = summary["developer"]
        cohort = summary["launch_month"]
        for item in weekly.get(asin, []):
            if item["start"] < start:
                continue
            rel = relative_week(start, item["start"])
            value = value_for(cohort, developer, rel)
            value["成熟ASIN"].add(asin)
            if item["available"] > 0:
                value["FBA可售ASIN"].add(asin)
            metric = item["currency_metrics"].get(currency)
            if metric is not None:
                value["产品表现销量"] += metric["volume"]
                value["产品表现销售额"] += metric["amount"]
        for item in finance.get((asin, currency), []):
            if item["date"] < start:
                continue
            rel = relative_week(start, item["date"])
            value = value_for(cohort, developer, rel)
            value["成熟ASIN"].add(asin)
            value["财务销量"] += item["quantity"]
            value["结算销售额"] += item["sales"]
            value["结算毛利润"] += item["profit"]
            value["广告费支出"] += abs(item["ads"])
            value["财务总成本支出"] += abs(item["cost"])
        for life in rows:
            if life["ASIN"] != asin:
                continue
            cycle_start = life["周期开始"]
            if not isinstance(cycle_start, date):
                continue
            rel = relative_week(start, cycle_start)
            value = value_for(cohort, developer, rel)
            value["成熟ASIN"].add(asin)
            if life["周期FBA可售周数"]:
                value["FBA可售ASIN"].add(asin)
    output: list[dict[str, Any]] = []
    for value in bucket.values():
        matured = len(value["成熟ASIN"])
        active = len(value["FBA可售ASIN"])
        output.append({
            "上架批次": value["上架批次"], "开发人": value["开发人"], "相对周": value["相对周"],
            "成熟ASIN数": matured, "FBA可售ASIN数": active,
            "FBA留存率": Decimal(active) / Decimal(matured) if matured else Decimal(0),
            "产品表现销量": value["产品表现销量"], "产品表现销售额": value["产品表现销售额"],
            "财务销量": value["财务销量"], "结算销售额": value["结算销售额"],
            "广告费支出": value["广告费支出"], "财务总成本支出": value["财务总成本支出"], "结算毛利润": value["结算毛利润"],
        })
    return sorted(output, key=lambda row: (row["上架批次"], row["开发人"], row["相对周"]))


def developer_month(rows: list[dict[str, Any]], summaries: dict[str, dict[str, Any]], weekly: dict[str, list[dict[str, Any]]], finance: dict[tuple[str, str], list[dict[str, Any]]], currency: str) -> list[dict[str, Any]]:
    bucket: dict[tuple[str, str], dict[str, Any]] = {}
    member_asins = {row["ASIN"] for row in rows}
    for asin in member_asins:
        developer = summaries[asin]["developer"]
        for item in weekly.get(asin, []):
            month = item["start"].strftime("%Y-%m")
            key = (month, developer)
            value = bucket.setdefault(key, {"时间": month, "开发人": developer, "ASIN": set(), "FBA可售ASIN": set(), "产品表现销量": Decimal(0), "产品表现销售额": Decimal(0), "财务销量": Decimal(0), "结算销售额": Decimal(0), "广告费支出": Decimal(0), "财务总成本支出": Decimal(0), "结算毛利润": Decimal(0)})
            value["ASIN"].add(asin)
            if item["available"] > 0:
                value["FBA可售ASIN"].add(asin)
            metric = item["currency_metrics"].get(currency)
            if metric is not None:
                value["产品表现销量"] += metric["volume"]
                value["产品表现销售额"] += metric["amount"]
        for item in finance.get((asin, currency), []):
            month = item["date"].strftime("%Y-%m")
            key = (month, developer)
            value = bucket.setdefault(key, {"时间": month, "开发人": developer, "ASIN": set(), "FBA可售ASIN": set(), "产品表现销量": Decimal(0), "产品表现销售额": Decimal(0), "财务销量": Decimal(0), "结算销售额": Decimal(0), "广告费支出": Decimal(0), "财务总成本支出": Decimal(0), "结算毛利润": Decimal(0)})
            value["ASIN"].add(asin)
            value["财务销量"] += item["quantity"]
            value["结算销售额"] += item["sales"]
            value["广告费支出"] += abs(item["ads"])
            value["财务总成本支出"] += abs(item["cost"])
            value["结算毛利润"] += item["profit"]
    return [{
        "时间": value["时间"], "开发人": value["开发人"], "有财务记录ASIN数": len(value["ASIN"]),
        "FBA可售ASIN数": len(value["FBA可售ASIN"]), "产品表现销量": value["产品表现销量"], "产品表现销售额": value["产品表现销售额"],
        "财务销量": value["财务销量"], "结算销售额": value["结算销售额"],
        "广告费支出": value["广告费支出"], "财务总成本支出": value["财务总成本支出"],
        "结算毛利润": value["结算毛利润"],
        "利润率": value["结算毛利润"] / value["结算销售额"] if value["结算销售额"] else None,
    } for value in sorted(bucket.values(), key=lambda row: (row["时间"], row["开发人"]))]


def main() -> None:
    """按上架月份生成 Q1/Q2/Q3 与逐月盈利回本模型。"""
    asins, sku_to_asins, _owners = load_asin_baseline()
    plans = load_completed_plans()
    finance = load_finance(set(asins))
    fixed_cohorts = load_fixed_currency_cohorts()
    model = build_combined_q_profit_model(asins, sku_to_asins, plans, finance, fixed_cohorts)

    for currency in CURRENCIES:
        currency_model = model[currency]
        overview = currency_model["overview"]
        q_summary = currency_model["q_summary"]
        timelines = currency_model["timelines"]
        details = currency_model["details"]
        detail_by_month: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in details:
            detail_by_month[str(row["上架批次"])].append(row)
        write_workbook(
            OUTPUT_ROOT / currency / f"月度上架批次_Q1Q2Q3盈利回本_{currency}.xlsx",
            [
                ("字段说明", combined_q_explanations(currency)),
                ("全部批次汇总", [*overview, combined_overview_total(overview)]),
                ("Q批次汇总", [*q_summary, *q_summary_totals(q_summary)]),
                *[(month, rows) for month, rows in sorted(timelines.items())],
            ]
            if overview else [("暂无月度批次", [])],
        )
        write_workbook(
            OUTPUT_ROOT / currency / f"月度上架批次_ASIN_Q1Q2Q3明细_{currency}.xlsx",
            [
                ("字段说明", combined_q_explanations(currency)),
                *[(month, rows) for month, rows in sorted(detail_by_month.items())],
            ]
            if details else [("暂无月度批次", [])],
        )
    print(OUTPUT_ROOT)


if __name__ == "__main__":
    main()
