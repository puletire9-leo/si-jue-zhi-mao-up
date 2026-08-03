#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按 ASIN 模型起算月生成 GBP、EUR 分开的审查工作簿。"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path

import pymysql
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lingxing_base_access import dec as decimal, mysql_env
from lingxing_model_paths import (
    ASIN_START_BASELINE,
    FBA_INVENTORY_BASELINE,
    LINGXING_DATA_ROOT,
    MONTHLY_MODEL_DIR,
)


BASELINE = ASIN_START_BASELINE
RAW_FBA_BASELINE = FBA_INVENTORY_BASELINE
DATA_START = "2025-04"
DATA_CUTOFF = "2026-07"
WINDOW = re.compile(r"(20\d{2}-\d{2})-\d{2}~")
CURRENCIES = ("GBP", "EUR")
TEAM_DEVELOPERS = frozenset({"蒋舒", "陈杨", "宋凤莉", "刘淼", "龙梦临", "周沁仪", "张子轩", "黄雨珊"})
MONTHLY_SOURCE_DIR = LINGXING_DATA_ROOT / "领星25年到26年6月所有数据，以每月数据"
MODEL_DIR = MONTHLY_MODEL_DIR

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
NEGATIVE_FILL = PatternFill("solid", fgColor="F4CCCC")
POSITIVE_FILL = PatternFill("solid", fgColor="D9EAD3")


def month_index(month: str) -> int:
    year, value = map(int, month.split("-"))
    return year * 12 + value - 1


def month_range(start: str, end: str) -> list[str]:
    return [f"{point // 12:04d}-{point % 12 + 1:02d}" for point in range(month_index(start), month_index(end) + 1)]


def china_currency(country: object) -> str:
    value = str(country or "")
    if "英国" in value:
        return "GBP"
    if "德国" in value:
        return "EUR"
    return ""


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        return list(csv.DictReader(source))


def team_developer(value: str) -> str | None:
    """只保留指定团队；同一 ASIN 的多人归属不重复摊入个人。"""
    names = {name.strip() for name in re.split(r"[,，|/]+", value or "") if name.strip()}
    matched = names & TEAM_DEVELOPERS
    if len(matched) == 1 and len(names) == 1:
        return next(iter(matched))
    if matched:
        return "多人归属待确认"
    return None


def load_baseline() -> dict[str, dict[str, object]]:
    """从数据库 lingxing_product_unified 统一表读取基准。"""
    # 数据源迁移到统一表：product_create_time→listing_open_date（真实上架日，覆盖率更高）；
    # analysis_status 弃用（淘汰判定本就用 listing_tags）；三个国家字段统一为 country(UK/DE)。
    sql = """
        SELECT asin, base_sku, developer, listing_open_date,
               model_start_month, model_start_basis,
               fba_first_available_month, fba_first_available_basis,
               fba_first_available_month, fba_inventory_first_month,
               listing_tags,
               country
        FROM lingxing_product_unified
    """
    rows: dict[str, dict[str, object]] = {}
    with pymysql.connect(**mysql_env()) as conn:
        with conn.cursor() as cur:
            cur.execute(sql)
            for (asin, sku, dev_raw, created_at, start_month, basis,
                 first_fba_month, first_fba_basis, first_available,
                 first_inventory, label,
                 country) in cur.fetchall():
                start_month = str(start_month or "").strip()
                if not start_month or start_month > DATA_CUTOFF:
                    continue
                developer = team_developer(str(dev_raw or ""))
                if developer is None:
                    continue
                label = str(label or "")
                if "欧洲精铺2025淘汰" in label:
                    status = "淘汰"
                elif "待淘汰" in label:
                    status = "待淘汰（暂计入留存）"
                else:
                    status = "留存"
                site_text = str(country or "")
                fixed_currencies: set[str] = set()
                if "UK" in site_text.upper() or "英国" in site_text:
                    fixed_currencies.add("GBP")
                if "DE" in site_text.upper() or "德国" in site_text:
                    fixed_currencies.add("EUR")
                rows[str(asin)] = {
                    "asin": str(asin),
                    "sku": str(sku or ""),
                    "developer": developer,
                    "created_at": str(created_at or ""),
                    "start_month": start_month,
                    "basis": str(basis or ""),
                    "first_fba_month": str(first_fba_month or ""),
                    "first_fba_basis": str(first_fba_basis or ""),
                    "first_available": str(first_available or ""),
                    "first_inventory": str(first_inventory or ""),
                    "label": label,
                    "status": status,
                    "cutoff_status": "",
                    "performance_quantity": defaultdict(Decimal),
                    "performance_sales": defaultdict(Decimal),
                    "available_stock": defaultdict(Decimal),
                    "cutoff_quantity": defaultdict(Decimal),
                    "finance_sales": defaultdict(Decimal),
                    "finance_profit": defaultdict(Decimal),
                    "currencies": fixed_currencies,
                }
    return rows


def marketplace_currency(marketplace: object) -> str:
    """周表 marketplace(UK/DE) 直接映射币种。"""
    value = str(marketplace or "").upper()
    if "UK" in value:
        return "GBP"
    if "DE" in value:
        return "EUR"
    return ""


def load_monthly_performance(records: dict[str, dict[str, object]]) -> None:
    """从周表 lingxing_sku_weekly_performance 按月聚合月度产品表现。

    数据源迁移说明（2026-08）：原 lingxing_asin_monthly_performance 仅到 2026-06 且无同步维护，
    改用周表按 asin+marketplace+year_month 聚合。已验证 6 月聚合与旧月度表差异 < 0.1%：
    - 销量/销售额 = 月内 SUM(volume)/SUM(amount)
    - 出单率分子（近30天销量）= 截止月整月 SUM(volume)
    - FBA 可售 = 截止月最后一周的 SUM(afn_fulfillable_quantity)（月末快照，非累加）
    """
    # 累计销量、销售额：按 asin+marketplace+月 聚合
    sql_sales = """
        SELECT asin, marketplace, `year_month` AS ym,
               SUM(volume) AS vol, SUM(amount) AS amt
        FROM lingxing_sku_weekly_performance
        WHERE `year_month` BETWEEN %s AND %s
        GROUP BY asin, marketplace, `year_month`
    """
    # 出单率分子：截止月整月销量（近30天销量，按 asin+marketplace 聚合全月）
    sql_cutoff_vol = """
        SELECT asin, marketplace, SUM(volume) AS vol
        FROM lingxing_sku_weekly_performance
        WHERE `year_month` = %s
        GROUP BY asin, marketplace
    """
    # FBA 可售月末快照：取截止月内最大 week_start 那一周
    sql_cutoff_fba = """
        SELECT w.asin, w.marketplace, SUM(w.afn_fulfillable_quantity) AS fba
        FROM lingxing_sku_weekly_performance w
        INNER JOIN (
            SELECT MAX(week_start) AS max_ws
            FROM lingxing_sku_weekly_performance
            WHERE `year_month` = %s
        ) m ON w.week_start = m.max_ws
        WHERE w.`year_month` = %s
        GROUP BY w.asin, w.marketplace
    """
    with pymysql.connect(**mysql_env()) as conn:
        with conn.cursor() as cur:
            cur.execute(sql_sales, (DATA_START, DATA_CUTOFF))
            for asin, marketplace, ym, volume, amount in cur.fetchall():
                record = records.get(str(asin or "").strip())
                if not record:
                    continue
                currency = marketplace_currency(marketplace)
                if not currency or currency not in record["currencies"]:
                    continue
                if str(ym) < str(record["start_month"]):
                    continue
                record["performance_quantity"][currency] += decimal(volume)
                record["performance_sales"][currency] += decimal(amount)
            # 出单率分子：截止月整月销量
            cur.execute(sql_cutoff_vol, (DATA_CUTOFF,))
            for asin, marketplace, volume in cur.fetchall():
                record = records.get(str(asin or "").strip())
                if not record:
                    continue
                currency = marketplace_currency(marketplace)
                if not currency or currency not in record["currencies"]:
                    continue
                if DATA_CUTOFF < str(record["start_month"]):
                    continue
                record["cutoff_quantity"][currency] += decimal(volume)
            # FBA 可售月末快照
            cur.execute(sql_cutoff_fba, (DATA_CUTOFF, DATA_CUTOFF))
            for asin, marketplace, fba in cur.fetchall():
                record = records.get(str(asin or "").strip())
                if not record:
                    continue
                currency = marketplace_currency(marketplace)
                if not currency or currency not in record["currencies"]:
                    continue
                if DATA_CUTOFF < str(record["start_month"]):
                    continue
                record["available_stock"][currency] += decimal(fba)


def load_finance(records: dict[str, dict[str, object]]) -> None:
    """从周表 lingxing_sku_weekly_performance 读取结算销售额/利润。

    数据源迁移说明（2026-08）：原 lingxing_profit_asin 财务日数据仅剩零星两周（历史全量已丢失、无备份），
    改用周表 gross_profit（领星结算毛利，已扣采购/运输/广告成本）+ amount（结算销售额）。
    周表完整覆盖 2025-04~2026-07 全 16 个月。
    """
    sql = """
        SELECT asin, marketplace, `year_month` AS ym,
               SUM(amount) AS sales, SUM(gross_profit) AS profit
        FROM lingxing_sku_weekly_performance
        WHERE `year_month` BETWEEN %s AND %s
        GROUP BY asin, marketplace, `year_month`
    """
    with pymysql.connect(**mysql_env()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(sql, (DATA_START, DATA_CUTOFF))
            for asin, marketplace, ym, sales, profit in cursor.fetchall():
                record = records.get(str(asin or "").strip())
                if not record:
                    continue
                currency = marketplace_currency(marketplace)
                if not currency or currency not in record["currencies"]:
                    continue
                if str(ym) < str(record["start_month"]):
                    continue
                record["finance_sales"][currency] += decimal(sales)
                record["finance_profit"][currency] += decimal(profit)


def rate(numerator: Decimal | int, denominator: Decimal | int) -> Decimal | None:
    return Decimal(numerator) / Decimal(denominator) if denominator else None


def report_row(record: dict[str, object], currency: str) -> list[object]:
    return [
        record["asin"], record["sku"], record["developer"], record["created_at"], record["start_month"],
        record["basis"], record["first_fba_month"], record["first_fba_basis"], record["first_inventory"], record["label"],
        record["performance_quantity"][currency], record["performance_sales"][currency],
        record["finance_sales"][currency], record["finance_profit"][currency], record["available_stock"][currency],
        record["status"], record["cutoff_status"],
    ]


def aggregate_row(label: str, batch: list[dict[str, object]], currency: str, include_order_rate: bool = False) -> list[object]:
    """按一批 ASIN 计算月度行或全期间合计行；比例均由汇总金额/数量重新计算。"""
    retained = [record for record in batch if record["status"] != "淘汰"]
    eliminated = [record for record in batch if record["status"] == "淘汰"]
    performance_quantity = sum((record["performance_quantity"][currency] for record in batch), Decimal(0))
    performance_sales = sum((record["performance_sales"][currency] for record in batch), Decimal(0))
    finance_sales = sum((record["finance_sales"][currency] for record in batch), Decimal(0))
    finance_profit = sum((record["finance_profit"][currency] for record in batch), Decimal(0))
    retained_sales = sum((record["performance_sales"][currency] for record in retained), Decimal(0))
    retained_finance_sales = sum((record["finance_sales"][currency] for record in retained), Decimal(0))
    retained_finance_profit = sum((record["finance_profit"][currency] for record in retained), Decimal(0))
    eliminated_sales = sum((record["performance_sales"][currency] for record in eliminated), Decimal(0))
    eliminated_finance_sales = sum((record["finance_sales"][currency] for record in eliminated), Decimal(0))
    eliminated_finance_profit = sum((record["finance_profit"][currency] for record in eliminated), Decimal(0))
    row = [
        label, len(batch), performance_quantity, performance_sales, finance_profit, finance_sales,
        rate(finance_profit, finance_sales),
        sum((record["available_stock"][currency] for record in batch), Decimal(0)),
        rate(len(retained), len(batch)),
        sum((record["performance_quantity"][currency] for record in retained), Decimal(0)), retained_sales,
        len(retained), retained_finance_profit, rate(retained_finance_profit, retained_finance_sales),
        sum((record["performance_quantity"][currency] for record in eliminated), Decimal(0)), eliminated_sales,
        len(eliminated), rate(len(eliminated), len(batch)), eliminated_finance_profit,
        rate(eliminated_finance_profit, eliminated_finance_sales),
    ]
    if include_order_rate:
        available_records = [record for record in batch if record["available_stock"][currency] > 0]
        cutoff_quantity = sum((record["cutoff_quantity"][currency] for record in available_records), Decimal(0))
        # 出单率：2026-06 的近30天销量 ÷ 30 ÷ FBA可售大于0的 ASIN数。
        row.insert(8, rate(cutoff_quantity, len(available_records) * 30))
    return row


def summary_rows(records: list[dict[str, object]], currency: str) -> list[list[object]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        grouped[str(record["start_month"])].append(record)
    rows = [aggregate_row(month, grouped.get(month, []), currency, include_order_rate=True) for month in month_range(DATA_START, DATA_CUTOFF)]
    rows.append(aggregate_row("合计", records, currency, include_order_rate=True))
    return rows


def developer_summary_rows(records: list[dict[str, object]], currency: str) -> list[list[object]]:
    """按开发人汇总，列口径与时间汇总一致。"""
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        developer = str(record["developer"] or "未填写开发人")
        grouped[developer].append(record)
    rows = [aggregate_row(developer, grouped[developer], currency) for developer in sorted(grouped)]
    rows.append(aggregate_row("合计", records, currency))
    return rows


def style_table(sheet, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"].font = Font(bold=True, size=14, color="0B3D47")
    sheet.append(headers)
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    for row in rows:
        sheet.append([float(value) if isinstance(value, Decimal) else value for value in row])
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(2, sheet.max_row)}"
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 34
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = min(max(len(header) + 4, 13), 32)
        if header in {"最新Listing标签", "基准SKU", "开发人", "模型分析起算依据", "数据截止月分析状态"}:
            sheet.column_dimensions[letter].width = 30
        if any(word in header for word in ("销售额", "利润", "库存")):
            for cell in sheet[letter][2:]:
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
        if header == "出单率":
            for cell in sheet[letter][2:]:
                cell.number_format = '0.00%;[Red]-0.00%'
        elif "率" in header:
            for cell in sheet[letter][2:]:
                cell.number_format = '0.00%;[Red]-0.00%'
        if "利润" in header and sheet.max_row >= 3:
            area = f"{letter}3:{letter}{sheet.max_row}"
            sheet.conditional_formatting.add(area, CellIsRule(operator="lessThan", formula=["0"], fill=NEGATIVE_FILL))
            sheet.conditional_formatting.add(area, CellIsRule(operator="greaterThan", formula=["0"], fill=POSITIVE_FILL))


def build_workbook(currency: str, records: list[dict[str, object]], unassigned: list[dict[str, object]]) -> Path:
    output_dir = MODEL_DIR / "03_审查报表" / currency
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"ASIN_FBA可售优先_商品信息创建时间兜底汇总_{DATA_START}至{DATA_CUTOFF}.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    summary_headers = [
        "时间", "ASIN总数", "总销售量", "总销售额", "总结算利润", "总结算销售额", "总利润率", "总可用库存", "出单率",
        "留存率", "留存ASIN销售量", "留存ASIN销售额", "留存ASIN数", "留存ASIN总利润", "留存ASIN利润率",
        "淘汰ASIN销售量", "淘汰ASIN销售额", "淘汰ASIN总数", "淘汰率", "淘汰ASIN总利润", "淘汰ASIN利润率",
    ]
    style_table(summary, f"{currency} 站点 ASIN 模型起算月汇总（截至 {DATA_CUTOFF}）", summary_headers, summary_rows(records, currency))
    for cell in summary[summary.max_row]:
        cell.fill = NOTE_FILL
        cell.font = Font(bold=True)

    developer_summary = workbook.create_sheet("开发人汇总")
    developer_headers = [
        "开发人", "ASIN总数", "总销售量", "总销售额", "总结算利润", "总结算销售额", "总利润率", "总可用库存",
        "留存率", "留存ASIN销售量", "留存ASIN销售额", "留存ASIN数", "留存ASIN总利润", "留存ASIN利润率",
        "淘汰ASIN销售量", "淘汰ASIN销售额", "淘汰ASIN总数", "淘汰率", "淘汰ASIN总利润", "淘汰ASIN利润率",
    ]
    style_table(developer_summary, f"{currency} 站点开发人汇总（截至 {DATA_CUTOFF}）", developer_headers, developer_summary_rows(records, currency))
    for cell in developer_summary[developer_summary.max_row]:
        cell.fill = NOTE_FILL
        cell.font = Font(bold=True)

    detail_headers = [
        "ASIN", "基准SKU", "开发人", "商品信息创建时间", "模型分析起算月", "模型分析起算依据", "FBA可售首现月",
        "FBA可售首现依据", "首次FBA库存月", "最新Listing标签", "累计销售量", "总销售额", "总结算销售额", "总结算利润", "总可用库存", "当前标签状态", "数据截止月分析状态",
    ]
    for month in month_range(DATA_START, DATA_CUTOFF):
        sheet = workbook.create_sheet(f"{month[2:4]}年{month[5:7]}月")
        cohort = [record for record in records if record["start_month"] == month]
        style_table(sheet, f"模型起算月为 {month} 的 {currency} ASIN 明细", detail_headers, [report_row(record, currency) for record in cohort])

    if unassigned:
        sheet = workbook.create_sheet("未分币种")
        style_table(
            sheet,
            "未在产品表现或财务数据中识别到 GBP/EUR 的 ASIN（不进入本币种汇总）",
            ["ASIN", "基准SKU", "开发人", "商品信息创建时间", "模型分析起算月", "模型分析起算依据", "最新Listing标签"],
            [[record[key] for key in ("asin", "sku", "developer", "created_at", "start_month", "basis", "label")] for record in unassigned],
        )

    notes = workbook.create_sheet("口径说明")
    notes.sheet_view.showGridLines = False
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 110
    note_rows = [
        ("报表范围", f"{currency} 单独汇总；仅限指定的 8 位团队开发人；GBP 与 EUR 从不相加、不换汇。"),
        ("批次起算", "优先取月表首次观察到 FBA-可售 > 0 的月份；全期未观察到时，取原始产品表现表商品信息中的创建时间所在月。FBA库存仅作为明细核查字段，不参与起算。"),
        ("创建时间边界", "商品信息创建时间是领星本地记录创建时间，不等同于 FBA 首次可售或亚马逊真实上架日。"),
        ("纳入条件", f"仅纳入模型起算月在 {DATA_START} 至 {DATA_CUTOFF} 的 ASIN；创建时间晚于截止月的 200 个 ASIN 不进入本报表。"),
        ("销售与库存", "总销售量、总销售额、可用库存取周表 lingxing_sku_weekly_performance 按月聚合；销售从模型起算月起累计，可用库存取 2026-07 月末最后一周 FBA 可售快照。"),
        ("汇总出单率", "取 2026-07 月销量。出单率 = 近30天销量 ÷ 30 ÷ FBA可售大于0的ASIN数；不以标签留存/淘汰作为分母。"),
        ("结算利润", "总结算销售额取周表 amount、总结算利润取周表 gross_profit（领星结算毛利，已扣采购/运输/广告成本），按 ASIN、币种、月累计；仅累计模型起算月及之后。原 lingxing_profit_asin 财务日数据历史全量已丢失，改用周表（完整覆盖 2025-04~2026-07）。"),
        ("留存/淘汰", "仅最新 Listing 标签含“欧洲精铺2025淘汰”时归为淘汰；“待淘汰”仍计入留存，但会在明细中单独标记。该分类是截止月标签状态，不表示每个历史月份的状态。"),
        ("开发人汇总", "按 ASIN 基准中的团队开发人聚合，金额、库存及留存/淘汰口径与“汇总”工作表完全一致。多人归属 ASIN 单列“多人归属待确认”，不重复分摊给个人。"),
        ("未分币种", "没有在产品表现或财务数据中识别到 GBP/EUR 的 ASIN 不进入汇总，列在“未分币种”工作表供核查。"),
    ]
    notes["A1"] = "报表口径"
    notes["B1"] = "说明"
    for cell in notes[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in note_rows:
        notes.append(row)
    for row in notes.iter_rows(min_row=2):
        row[0].fill = NOTE_FILL
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    notes.freeze_panes = "A2"
    workbook.save(output)
    return output


def build_developer_workbook(currency: str, developer: str, records: list[dict[str, object]]) -> Path:
    """在币种目录下为单个开发人输出可按批次审查的 ASIN 工作簿。"""
    output_dir = MODEL_DIR / "03_审查报表" / currency / "开发人维度"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"{developer}_ASIN经营报告_{DATA_START}至{DATA_CUTOFF}.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    summary_headers = [
        "时间", "ASIN总数", "总销售量", "总销售额", "总结算利润", "总结算销售额", "总利润率", "总可用库存",
        "留存率", "留存ASIN销售量", "留存ASIN销售额", "留存ASIN数", "留存ASIN总利润", "留存ASIN利润率",
        "淘汰ASIN销售量", "淘汰ASIN销售额", "淘汰ASIN总数", "淘汰率", "淘汰ASIN总利润", "淘汰ASIN利润率",
    ]
    style_table(summary, f"{developer} {currency} ASIN 模型起算月总览（截至 {DATA_CUTOFF}）", summary_headers, summary_rows(records, currency))
    for cell in summary[summary.max_row]:
        cell.fill = NOTE_FILL
        cell.font = Font(bold=True)

    detail_headers = [
        "ASIN", "基准SKU", "商品信息创建时间", "模型分析起算月", "模型分析起算依据", "FBA可售首现月",
        "FBA可售首现依据", "首次FBA库存月", "最新Listing标签", "累计销售量", "总销售额", "总结算销售额", "总结算利润", "总可用库存", "当前标签状态", "数据截止月分析状态",
    ]
    for month in month_range(DATA_START, DATA_CUTOFF):
        cohort = [record for record in records if record["start_month"] == month]
        if not cohort:
            continue
        sheet = workbook.create_sheet(f"{month[2:4]}年{month[5:7]}月")
        rows = [
            [
                record["asin"], record["sku"], record["created_at"], record["start_month"], record["basis"],
                record["first_fba_month"], record["first_fba_basis"], record["first_inventory"], record["label"],
                record["performance_quantity"][currency], record["performance_sales"][currency], record["finance_sales"][currency],
                record["finance_profit"][currency], record["available_stock"][currency], record["status"], record["cutoff_status"],
            ]
            for record in cohort
        ]
        style_table(sheet, f"{developer}：模型起算月为 {month} 的 {currency} ASIN 明细", detail_headers, rows)

    notes = workbook.create_sheet("口径说明")
    notes.sheet_view.showGridLines = False
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 100
    notes.append(["报告口径", "说明"])
    notes.append(["开发人", developer])
    notes.append(["批次起算", "首次观察到 FBA-可售 > 0 的月份；没有 FBA 可售时，以商品信息创建时间所在月兜底。"])
    notes.append(["金额与库存", "销量、销售额、库存、结算利润均来自周表 lingxing_sku_weekly_performance 按月聚合，累计至 2026-07。"])
    notes.append(["标签状态", "仅“欧洲精铺2025淘汰”算淘汰；待淘汰暂计入留存。"])
    for cell in notes[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in notes.iter_rows(min_row=2):
        row[0].fill = NOTE_FILL
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    notes.freeze_panes = "A2"
    workbook.save(output)
    return output


def main() -> None:
    records = load_baseline()
    load_monthly_performance(records)
    load_finance(records)
    assigned: dict[str, list[dict[str, object]]] = {currency: [] for currency in CURRENCIES}
    unassigned = []
    for record in records.values():
        report_currencies = set(record["currencies"]) & set(CURRENCIES)
        if not report_currencies:
            unassigned.append(record)
            continue
        for currency in report_currencies:
            assigned[currency].append(record)
    for currency in CURRENCIES:
        assigned[currency].sort(key=lambda item: (str(item["start_month"]), str(item["asin"])))
        output = build_workbook(currency, assigned[currency], unassigned)
        developer_outputs = []
        for developer in sorted(TEAM_DEVELOPERS):
            developer_records = [record for record in assigned[currency] if record["developer"] == developer]
            if developer_records:
                try:
                    developer_outputs.append(build_developer_workbook(currency, developer, developer_records))
                except OSError as exc:
                    # 开发人文件可能正被 Excel 打开；不能阻断总报表和其他开发人文件刷新。
                    print(f"{currency}: skip_developer_report={developer} reason={exc}")
        print(f"{currency}: asins={len(assigned[currency])} developer_reports={len(developer_outputs)} output={output}")
    print(f"未分币种 ASIN={len(unassigned)}")


if __name__ == "__main__":
    main()
