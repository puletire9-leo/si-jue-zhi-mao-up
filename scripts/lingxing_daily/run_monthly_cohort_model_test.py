"""Run a monthly cohort test from Lingxing exports without treating stock as Q1."""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from python_calamine import load_workbook as load_calamine_workbook


ROOT = Path(__file__).resolve().parents[2]
MONTHLY_DIR = ROOT / "\u4ea7\u54c1\u6570\u636e" / "\u9886\u661f\u6570\u636eapi" / "\u9886\u661f25\u5e74\u523026\u5e746\u6708\u6240\u6709\u6570\u636e\uff0c\u4ee5\u6bcf\u6708\u6570\u636e"
HISTORY_DIR = MONTHLY_DIR / "\u5386\u53f2SKU\u4e0a\u67b6\u57fa\u7840\u6570\u636e_2025-04\u81f32026-06"
TEAM_DIR = HISTORY_DIR / "03_\u56e2\u961f\u5f00\u53d1SKU\u751f\u547d\u5468\u671f"
OUTPUT_DIR = HISTORY_DIR / "04_\u6708\u5ea6\u6a21\u578b\u6d4b\u8bd5"
LIFECYCLE_INPUT = TEAM_DIR / "\u56e2\u961fSKU_\u751f\u547d\u5468\u671f\u5224\u5b9a_\u6570\u636e\u622a\u6b622026-06.csv"
DETAIL_OUTPUT = OUTPUT_DIR / "\u56e2\u961fSKU_\u6708\u5ea6\u4e0a\u67b6\u6279\u6b21\u9636\u6bb5\u660e\u7ec6_\u6570\u636e\u622a\u6b622026-06.csv"
WORKBOOK_OUTPUT = OUTPUT_DIR / "\u56e2\u961fSKU_\u6708\u5ea6\u4e0a\u67b6\u6279\u6b21\u6a21\u578b\u6d4b\u8bd5_\u6570\u636e\u622a\u6b622026-06.xlsx"
REPORT_OUTPUT = OUTPUT_DIR / "\u56e2\u961fSKU_\u6708\u5ea6\u4e0a\u67b6\u6279\u6b21\u6a21\u578b\u6d4b\u8bd5\u62a5\u544a_\u6570\u636e\u622a\u6b622026-06.md"
DATA_CUTOFF_MONTH = "2026-06"
MONTH_PATTERN = re.compile(r"(20\d{2}-\d{2})-\d{2}~")

COUNTRY = "\u56fd\u5bb6"
STORE = "\u5e97\u94fa"
SKU = "SKU"
VOLUME = "\u9500\u91cf"
GROSS_PROFIT = "\u7ed3\u7b97\u6bdb\u5229\u6da6"
AD_SPEND = "\u5e7f\u544a\u82b1\u8d39"
FBA_AVAILABLE = "FBA-\u53ef\u552e"
FBA_INVENTORY = "FBA\u5e93\u5b58"

DETAIL_LABELS = {
    "marketplace": "站点", "country": "国家", "store_name": "店铺", "sku": "SKU",
    "developer": "开发人", "is_combo": "是否组合产品", "combo_counting_status": "组合计数状态",
    "first_fba_month": "首次FBA可售观察月", "first_full_sales_month": "首个完整销售月份",
    "two_month_maturity_month": "两个月成熟月份", "stage": "月度阶段判定",
    "month0_volume": "首次FBA可售月销量", "month0_fba_available_max": "首次FBA可售月最大FBA可售库存",
    "month1_volume": "首个完整月销量", "month1_gross_profit": "首个完整月结算毛利润",
    "month1_ad_spend": "首个完整月广告花费", "month1_fba_available_max": "首个完整月最大FBA可售库存",
    "month2_volume": "第二完整月销量", "month2_gross_profit": "第二完整月结算毛利润",
    "month2_ad_spend": "第二完整月广告花费", "month2_fba_available_max": "第二完整月最大FBA可售库存",
}
SUMMARY_LABELS = {
    "cohort_month": "上架批次月份", "cohort_object_count": "上架批次店铺SKU数",
    "r1_eligible_count": "首个完整月可观察对象数", "r1_sold_count": "首个完整月有出单对象数",
    "two_month_eligible_count": "两个月可观察对象数", "r2_eligible_r1_sold_count": "首月出单且两个月可观察对象数",
    "r2_continued_count": "第二完整月持续出单对象数", "month1_volume": "首个完整月销量合计",
    "month1_gross_profit": "首个完整月结算毛利润合计", "month1_ad_spend": "首个完整月广告花费合计",
    "month2_volume": "第二完整月销量合计", "month2_gross_profit": "第二完整月结算毛利润合计",
    "month2_ad_spend": "第二完整月广告花费合计", "r1_first_full_month_sale_rate": "首个完整月出单率",
    "r2_second_month_continuation_rate": "第二完整月持续出单率", "two_month_survival_rate": "两个月持续出单率",
}


def add_months(month: str, months: int) -> str:
    year, value = map(int, month.split("-"))
    index = year * 12 + value - 1 + months
    return f"{index // 12:04d}-{index % 12 + 1:02d}"


def number(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def marketplace(country: str) -> str:
    return {"\u82f1\u56fd": "UK", "\u5fb7\u56fd": "DE"}.get(country, "UNKNOWN")


def zero_point() -> dict[str, float]:
    return {"volume": 0.0, "gross_profit": 0.0, "ad_spend": 0.0, "fba_available": 0.0, "fba_inventory": 0.0, "raw_row_count": 0.0}


def load_team_skus() -> dict[str, dict[str, str]]:
    team_by_sku: dict[str, dict[str, str]] = {}
    with LIFECYCLE_INPUT.open(encoding="utf-8-sig", newline="") as source:
        for row in csv.DictReader(source):
            if row["生命周期判定"] != "已观察到FBA可售":
                continue
            sku = row["SKU"].strip()
            existing = team_by_sku.get(sku)
            candidate = {
                "developer": row["开发人"],
                "is_combo": row["是否组合产品"],
                "combo_counting_status": row["组合计数状态"],
            }
            if existing and existing != candidate:
                raise RuntimeError(f"Ambiguous local SKU ownership: {sku}")
            team_by_sku[sku] = candidate
    return team_by_sku


def read_monthly_facts(team_by_sku: dict[str, dict[str, str]]) -> tuple[list[str], dict[str, dict[tuple[str, str, str, str], dict[str, float]]]]:
    files: list[tuple[str, Path]] = []
    for path in MONTHLY_DIR.glob("*.xlsx"):
        match = MONTH_PATTERN.search(path.name)
        if match:
            files.append((match.group(1), path))
    files.sort()
    if len(files) != 15:
        raise RuntimeError(f"Expected 15 monthly workbooks, found {len(files)}")

    facts: dict[str, dict[tuple[str, str, str, str], dict[str, float]]] = {}
    for month, path in files:
        print(f"reading={month} file={path.name}", flush=True)
        workbook = load_calamine_workbook(path)
        rows = workbook.get_sheet_by_index(0).iter_rows()
        headers = [str(value) if value is not None else "" for value in next(rows)]
        positions = {field: headers.index(field) for field in (COUNTRY, STORE, SKU, VOLUME, GROSS_PROFIT, AD_SPEND, FBA_AVAILABLE, FBA_INVENTORY)}
        month_facts: dict[tuple[str, str, str, str], dict[str, float]] = {}
        for row in rows:
            sku = str(row[positions[SKU]] or "").strip()
            if sku not in team_by_sku:
                continue
            country = str(row[positions[COUNTRY]] or "").strip()
            store = str(row[positions[STORE]] or "").strip()
            if not country or not store:
                continue
            key = (marketplace(country), country, store, sku)
            point = month_facts.setdefault(key, zero_point())
            point["volume"] += number(row[positions[VOLUME]])
            point["gross_profit"] += number(row[positions[GROSS_PROFIT]])
            point["ad_spend"] += number(row[positions[AD_SPEND]])
            point["fba_available"] = max(point["fba_available"], number(row[positions[FBA_AVAILABLE]]))
            point["fba_inventory"] = max(point["fba_inventory"], number(row[positions[FBA_INVENTORY]]))
            point["raw_row_count"] += 1
        facts[month] = month_facts
    return [month for month, _ in files], facts


def stage(first_fba_month: str, month_one: str, month_two: str, points: dict[str, dict[str, float]]) -> str:
    if first_fba_month == "2025-04":
        return "数据起点左截断，不纳入转化率"
    if month_one > DATA_CUTOFF_MONTH:
        return "尚未到首个完整销售月"
    if points[month_one]["volume"] <= 0:
        return "首个完整月未出单"
    if month_two > DATA_CUTOFF_MONTH:
        return "首月已出单，第二月待观察"
    return "第二完整月持续出单" if points[month_two]["volume"] > 0 else "第二完整月未持续出单"


def percent(numerator: int, denominator: int) -> str:
    return "" if denominator == 0 else f"{numerator / denominator:.2%}"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    team_by_sku = load_team_skus()
    months, facts = read_monthly_facts(team_by_sku)
    first_month = months[0]

    all_keys = set().union(*(month_facts.keys() for month_facts in facts.values()))
    detail_rows: list[dict[str, Any]] = []
    for key in sorted(all_keys):
        first_fba = next((month for month in months if facts[month].get(key, zero_point())["fba_available"] > 0), "")
        if not first_fba:
            continue
        month_one = add_months(first_fba, 1)
        month_two = add_months(first_fba, 2)
        points = {month: facts.get(month, {}).get(key, zero_point()) for month in (first_fba, month_one, month_two)}
        team = team_by_sku[key[3]]
        detail_rows.append({
            "marketplace": key[0], "country": key[1], "store_name": key[2], "sku": key[3],
            "developer": team["developer"], "is_combo": team["is_combo"],
            "combo_counting_status": team["combo_counting_status"], "first_fba_month": first_fba,
            "first_full_sales_month": month_one, "two_month_maturity_month": month_two,
            "stage": stage(first_fba, month_one, month_two, points),
            "month0_volume": points[first_fba]["volume"], "month0_fba_available_max": points[first_fba]["fba_available"],
            "month1_volume": points[month_one]["volume"], "month1_gross_profit": points[month_one]["gross_profit"], "month1_ad_spend": points[month_one]["ad_spend"], "month1_fba_available_max": points[month_one]["fba_available"],
            "month2_volume": points[month_two]["volume"], "month2_gross_profit": points[month_two]["gross_profit"], "month2_ad_spend": points[month_two]["ad_spend"], "month2_fba_available_max": points[month_two]["fba_available"],
        })

    summary: dict[str, dict[str, Any]] = defaultdict(lambda: {
        "cohort_object_count": 0, "r1_eligible_count": 0, "r1_sold_count": 0,
        "two_month_eligible_count": 0, "r2_eligible_r1_sold_count": 0, "r2_continued_count": 0,
        "month1_volume": 0.0, "month1_gross_profit": 0.0, "month1_ad_spend": 0.0,
        "month2_volume": 0.0, "month2_gross_profit": 0.0, "month2_ad_spend": 0.0,
    })
    for row in detail_rows:
        cohort = row["first_fba_month"]
        if cohort == first_month:
            continue
        item = summary[cohort]
        item["cohort_object_count"] += 1
        if row["first_full_sales_month"] <= DATA_CUTOFF_MONTH:
            item["r1_eligible_count"] += 1
            item["month1_volume"] += row["month1_volume"]
            item["month1_gross_profit"] += row["month1_gross_profit"]
            item["month1_ad_spend"] += row["month1_ad_spend"]
            if row["month1_volume"] > 0:
                item["r1_sold_count"] += 1
        if row["two_month_maturity_month"] <= DATA_CUTOFF_MONTH:
            item["two_month_eligible_count"] += 1
            item["month2_volume"] += row["month2_volume"]
            item["month2_gross_profit"] += row["month2_gross_profit"]
            item["month2_ad_spend"] += row["month2_ad_spend"]
            if row["month1_volume"] > 0:
                item["r2_eligible_r1_sold_count"] += 1
                if row["month2_volume"] > 0:
                    item["r2_continued_count"] += 1

    summary_rows = []
    for cohort, item in sorted(summary.items()):
        summary_rows.append({
            "cohort_month": cohort,
            **item,
            "r1_first_full_month_sale_rate": percent(item["r1_sold_count"], item["r1_eligible_count"]),
            "r2_second_month_continuation_rate": percent(item["r2_continued_count"], item["r2_eligible_r1_sold_count"]),
            "two_month_survival_rate": percent(item["r2_continued_count"], item["two_month_eligible_count"]),
        })

    detail_columns = list(detail_rows[0].keys()) if detail_rows else []
    summary_columns = list(summary_rows[0].keys()) if summary_rows else []
    detail_headers = [DETAIL_LABELS[column] for column in detail_columns]
    summary_headers = [SUMMARY_LABELS[column] for column in summary_columns]
    with DETAIL_OUTPUT.open("w", encoding="utf-8-sig", newline="") as destination:
        writer = csv.DictWriter(destination, fieldnames=detail_headers)
        writer.writeheader()
        writer.writerows([{DETAIL_LABELS[key]: value for key, value in row.items()} for row in detail_rows])

    workbook = Workbook(write_only=True)
    overview = workbook.create_sheet("\u4e0a\u67b6\u6279\u6b21\u6c47\u603b")
    overview.append(summary_headers)
    for row in summary_rows:
        overview.append([row[column] for column in summary_columns])
    detail = workbook.create_sheet("SKU\u9636\u6bb5\u660e\u7ec6")
    detail.append(detail_headers)
    for row in detail_rows:
        detail.append([row[column] for column in detail_columns])
    workbook.save(WORKBOOK_OUTPUT)

    total = {key: sum(row[key] for row in summary_rows) for key in (
        "cohort_object_count", "r1_eligible_count", "r1_sold_count", "two_month_eligible_count",
        "r2_eligible_r1_sold_count", "r2_continued_count")}
    report = f"""# 团队 SKU 月度上架批次模型测试\n\n## 输入和范围\n\n- 月度数据：2025-04 至 2026-06 的 15 份领星产品表现导出。\n- 对象粒度：站点 + 店铺 + SKU；同 SKU 跨店铺不合并。\n- 团队范围：本地团队 SKU 中，月度数据已观察到 FBA 可售的对象。\n- 首月 2025-04 的对象存在数据起点左截断，排除在阶段转化率之外。\n\n## 阶段口径\n\n- 首月阶段：首次 FBA 可售后的首个完整自然月，销量大于 0。\n- 次月阶段：首月已出单，且第二个完整自然月仍有销量。\n- 这不是采购批次的正式 R1/R2，也不是 Q1/Q2 盈亏模型：月度库存快照不能代替首批采购量。\n\n## 汇总结果\n\n| 指标 | 数量 | 比率 |\n|---|---:|---:|\n| 可归入非左截断上架批次的店铺-SKU | {total['cohort_object_count']:,} | - |\n| 具备首个完整月观察条件 | {total['r1_eligible_count']:,} | - |\n| 首个完整月有出单 | {total['r1_sold_count']:,} | {percent(total['r1_sold_count'], total['r1_eligible_count'])} |\n| 具备第二完整月观察条件且首月出单 | {total['r2_eligible_r1_sold_count']:,} | - |\n| 第二完整月持续出单 | {total['r2_continued_count']:,} | {percent(total['r2_continued_count'], total['r2_eligible_r1_sold_count'])} |\n| 两个月持续出单占已成熟上架批次 | {total['r2_continued_count']:,} | {percent(total['r2_continued_count'], total['two_month_eligible_count'])} |\n\n## 成熟度\n\n- 2026-04 及更早首次 FBA 可售的上架批次，已具备两个月窗口。\n- 2026-05 上架批次只具备首个完整月观察；2026-06 上架批次尚未有首个完整月。\n- 组合主 SKU 已保留为一个对象；组合子件关系尚未取得，不能自动剔除。\n\n## 限制\n\n- `结算毛利润` 仅作为观察毛利，费用口径未核实，不能称为真实净利润。\n- 采购事实表虽有 `quantity_real`，但月度导出缺少可直接对应的店铺 `sid`；在完成店铺映射前，不能把采购行严格归到某个店铺-SKU 的 Q1/Q2。\n- 精确首个 FBA 可售日、断货日和补货到货日仍需日/周级事实。\n"""
    report = report.replace(
        f"| 首个完整月有出单 | {total['r1_sold_count']:,} | {percent(total['r1_sold_count'], total['r1_eligible_count'])} |\n| 具备第二完整月观察条件且首月出单 |",
        f"| 首个完整月有出单 | {total['r1_sold_count']:,} | {percent(total['r1_sold_count'], total['r1_eligible_count'])} |\n| 具备两个完整月观察条件 | {total['two_month_eligible_count']:,} | - |\n| 具备第二完整月观察条件且首月出单 |",
    )
    REPORT_OUTPUT.write_text(report, encoding="utf-8-sig")
    print(f"team_store_skus={len(detail_rows)} summary_rows={len(summary_rows)} output={OUTPUT_DIR}")


if __name__ == "__main__":
    main()
