"""Build one filterable workbook from the team lifecycle, cohort, and profit summaries."""

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
BASE_DIR = ROOT / "产品数据" / "领星数据api" / "领星25年到26年6月所有数据，以每月数据" / "历史SKU上架基础数据_2025-04至2026-06"
LIFECYCLE_FILE = BASE_DIR / "03_团队开发SKU生命周期" / "团队SKU_生命周期判定_数据截止2026-06.csv"
COHORT_FILE = BASE_DIR / "04_月度模型测试" / "团队SKU_月度上架批次阶段明细_数据截止2026-06.csv"
FINANCE_DIR = BASE_DIR / "05_财务利润周度回补"
OUTPUT_DIR = BASE_DIR / "00_整合工作簿"
OUTPUT_FILE = OUTPUT_DIR / "团队SKU经营总览_生命周期与财务_2025-04至2026-06.xlsx"
MONTHS = [f"2025年{month:02d}月" for month in range(4, 13)] + [f"2026年{month:02d}月" for month in range(1, 7)]

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAD3")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NEGATIVE_FILL = PatternFill("solid", fgColor="F4CCCC")
POSITIVE_FILL = PatternFill("solid", fgColor="D9EAD3")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as source:
        return list(csv.DictReader(source))


def decimal_value(value: str | None) -> Decimal:
    try:
        return Decimal((value or "0").strip() or "0")
    except InvalidOperation:
        return Decimal(0)


def int_value(value: str | None) -> int:
    return int(decimal_value(value))


def excel_value(value: str | Decimal | int | float | None) -> str | float | int | None:
    if isinstance(value, Decimal):
        return float(value)
    return value


def write_table(sheet, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append([title])
    sheet["A1"].font = Font(bold=True, size=14, color="0B3D47")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    sheet.append(headers)
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    for row in rows:
        sheet.append([excel_value(value) for value in row])
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(2, sheet.max_row)}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 32
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=THIN_GRAY)
    for index, header in enumerate(headers, start=1):
        width = min(max(len(header) + 4, 12), 32)
        sheet.column_dimensions[get_column_letter(index)].width = width
        if header in {"品名", "FBA匹配店铺", "最近Listing标签", "领星业务状态", "标签状态说明", "数据截止时销售成熟度"}:
            sheet.column_dimensions[get_column_letter(index)].width = 28
        if header in {"销售额", "广告费", "总成本", "领星其他结算调整项", "毛利润", "结算毛利润", "GBP销售额", "GBP结算毛利润", "EUR销售额", "EUR结算毛利润"}:
            for cell in sheet[get_column_letter(index)][2:]:
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
        if header in {"汇总毛利率", "毛利率"}:
            for cell in sheet[get_column_letter(index)][2:]:
                cell.number_format = '0.00%;[Red]-0.00%'
        if header in {"毛利润", "结算毛利润", "GBP结算毛利润", "EUR结算毛利润"} and sheet.max_row >= 3:
            range_ref = f"{get_column_letter(index)}3:{get_column_letter(index)}{sheet.max_row}"
            sheet.conditional_formatting.add(range_ref, CellIsRule(operator="lessThan", formula=["0"], fill=NEGATIVE_FILL))
            sheet.conditional_formatting.add(range_ref, CellIsRule(operator="greaterThan", formula=["0"], fill=POSITIVE_FILL))


def build_usage_sheet(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "使用说明"
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "团队 SKU 经营总览"
    sheet["A1"].font = Font(bold=True, size=18, color="0B3D47")
    sheet.merge_cells("A1:F1")
    rows = [
        ("用途", "在一个文件内查看团队 SKU 的生命周期、上架批次表现及 2025-04 至 2026-06 的月度财务利润。"),
        ("阅读顺序", "先看“团队总览”“开发批次经营”和“月度利润总览”，需要定位对象时再筛选“开发批次SKU明细”“生命周期明细”“上架批次阶段”或“月度SKU利润”。"),
        ("上架与状态范围", "2025-04 至 2026-06 月度领星产品表现导出；首次 FBA 可售观察月记录历史上架观察，领星业务状态以最近一次前端 listing 标签为准。"),
        ("财务范围", "2025-04 至 2026-06，15 个团队目标店铺、8 位开发人的有效 SKU 前缀；财务事实按 ASIN、店铺、日期和币种幂等。"),
        ("利润口径", "结算毛利润直接取领星 grossProfit。商品成本 totalCost = 采购成本 + 头程 + 其他商品成本；其他结算调整项用于对平领星毛利润。"),
        ("币种", "GBP 与 EUR 单独汇总，未换汇、未相加。"),
        ("标签状态", "标签包含“欧洲精铺2025淘汰”即为淘汰；待淘汰、侵权下架、季节性断货独立显示。未在前端观察到不等于淘汰。"),
        ("组合产品", "仅保留领星财务实际返回的 localSku；组合子件关系尚未取得，不能从财务表自动排除。"),
        ("数据边界", "这是平台结算毛利润，不包含人工、国内运营、汇兑和资金成本等平台外费用。"),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 1).font = BOLD_FONT
        sheet.cell(row_index, 1).fill = SECTION_FILL
        sheet.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(row_index, 2).fill = NOTE_FILL
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 110
    sheet.freeze_panes = "A3"


def lifecycle_overview(lifecycle_rows: list[dict[str, str]]) -> tuple[list[list[object]], list[list[object]]]:
    developer_rows: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    status_rows: dict[tuple[str, str], int] = defaultdict(int)
    for row in lifecycle_rows:
        developer = row["开发人"]
        stats = developer_rows[developer]
        stats["开发SKU数"] += 1
        stats["组合主产品数"] += int(row["是否组合产品"] == "是")
        business_status = row["领星业务状态"]
        stats["上架在售"] += int(business_status == "上架在售")
        stats["淘汰"] += int(business_status == "淘汰")
        stats["未在前端观察"] += int(business_status == "未在领星前端观察到")
        if row["首次FBA可售观察月"]:
            stats["已观察FBA可售"] += 1
        if row["月度FBA匹配状态"] == "月度数据中未出现FBA可售":
            stats["未观察FBA可售"] += 1
        status_rows[(developer, business_status)] += 1
    developer_table = []
    for developer, stats in sorted(developer_rows.items()):
        developer_table.append([
            developer, stats["开发SKU数"], stats["组合主产品数"], stats["上架在售"], stats["淘汰"], stats["未在前端观察"], stats["已观察FBA可售"],
        ])
    lifecycle_status_table = [[developer, status, count] for (developer, status), count in sorted(status_rows.items())]
    return developer_table, lifecycle_status_table


def profit_tables() -> tuple[list[list[object]], list[list[object]], list[list[object]]]:
    monthly_totals: dict[tuple[str, str], dict[str, object]] = defaultdict(lambda: {
        "facts": 0, "skus": set(), "sales_qty": Decimal(0), "sales_amount": Decimal(0), "ads_cost": Decimal(0),
        "total_cost": Decimal(0), "adjustment": Decimal(0), "gross_profit": Decimal(0),
    })
    owner_totals: dict[tuple[str, str, str], dict[str, object]] = defaultdict(lambda: {
        "skus": set(), "sales_qty": Decimal(0), "sales_amount": Decimal(0), "ads_cost": Decimal(0),
        "total_cost": Decimal(0), "adjustment": Decimal(0), "gross_profit": Decimal(0),
    })
    sku_rows: list[list[object]] = []
    for month in MONTHS:
        path = FINANCE_DIR / f"{month}团队SKU财务利润明细.csv"
        for row in read_csv(path):
            currency = row["币种"]
            developer = row["开发人"]
            values = {
                "facts": int_value(row["财务日记录数"]), "sales_qty": decimal_value(row["销量"]),
                "sales_amount": decimal_value(row["销售额"]), "ads_cost": decimal_value(row["广告费"]),
                "total_cost": decimal_value(row["总成本"]), "adjustment": decimal_value(row["领星其他结算调整项"]),
                "gross_profit": decimal_value(row["毛利润"]),
            }
            monthly = monthly_totals[(month, currency)]
            monthly["facts"] += values["facts"]
            monthly["skus"].add(row["SKU"])
            owner = owner_totals[(month, developer, currency)]
            owner["skus"].add(row["SKU"])
            for key in ("sales_qty", "sales_amount", "ads_cost", "total_cost", "adjustment", "gross_profit"):
                monthly[key] += values[key]
                owner[key] += values[key]
            margin = values["gross_profit"] / values["sales_amount"] if values["sales_amount"] else None
            sku_rows.append([
                month, developer, row["SKU"], currency, values["facts"], int_value(row["ASIN数"]), int_value(row["店铺数"]),
                values["sales_qty"], values["sales_amount"], values["ads_cost"], values["total_cost"], values["adjustment"],
                values["gross_profit"], margin,
            ])
    monthly_rows = []
    for (month, currency), row in sorted(monthly_totals.items()):
        margin = row["gross_profit"] / row["sales_amount"] if row["sales_amount"] else None
        monthly_rows.append([
            month, currency, row["facts"], len(row["skus"]), row["sales_qty"], row["sales_amount"], row["ads_cost"],
            row["total_cost"], row["adjustment"], row["gross_profit"], margin,
        ])
    owner_rows = []
    for (month, developer, currency), row in sorted(owner_totals.items()):
        margin = row["gross_profit"] / row["sales_amount"] if row["sales_amount"] else None
        owner_rows.append([
            month, developer, currency, len(row["skus"]), row["sales_qty"], row["sales_amount"], row["ads_cost"],
            row["total_cost"], row["adjustment"], row["gross_profit"], margin,
        ])
    return monthly_rows, owner_rows, sku_rows


def cohort_financial_tables(
    lifecycle_rows: list[dict[str, str]], cohort_rows: list[dict[str, str]], sku_profit_rows: list[list[object]],
) -> tuple[list[list[object]], list[list[object]]]:
    lifecycle_by_sku = {row["SKU"]: row for row in lifecycle_rows}
    stage_by_sku: dict[str, dict[str, int]] = defaultdict(lambda: {
        "store_sku_count": 0, "first_month_sold": 0, "second_month_continued": 0,
    })
    for row in cohort_rows:
        stage = stage_by_sku[row["SKU"]]
        stage["store_sku_count"] += 1
        first_sales = decimal_value(row["首个完整月销量"])
        second_sales = decimal_value(row["第二完整月销量"])
        stage["first_month_sold"] += int(first_sales > 0)
        stage["second_month_continued"] += int(first_sales > 0 and second_sales > 0)

    finance_by_sku: dict[str, dict[str, Decimal | set[str]]] = defaultdict(lambda: {
        "GBP_sales": Decimal(0), "GBP_profit": Decimal(0), "GBP_months": set(),
        "EUR_sales": Decimal(0), "EUR_profit": Decimal(0), "EUR_months": set(),
    })
    for month, _developer, sku, currency, _facts, _asins, _stores, _qty, sales, _ads, _cost, _adjustment, profit, _margin in sku_profit_rows:
        if currency not in {"GBP", "EUR"}:
            continue
        financial = finance_by_sku[str(sku)]
        financial[f"{currency}_sales"] += Decimal(sales)
        financial[f"{currency}_profit"] += Decimal(profit)
        financial[f"{currency}_months"].add(str(month))

    summary: dict[tuple[str, str], dict[str, object]] = defaultdict(lambda: {
        "skus": set(), "fba_observed": 0, "active": 0, "eliminated": 0, "not_observed": 0, "special": 0, "mature": 0,
        "first_month_sold": 0, "second_month_continued": 0, "finance_skus": set(),
        "GBP_sales": Decimal(0), "GBP_profit": Decimal(0), "EUR_sales": Decimal(0), "EUR_profit": Decimal(0),
    })
    detail_rows: list[list[object]] = []
    matched_finance_skus: set[str] = set()
    for row in lifecycle_rows:
        sku = row["SKU"]
        key = (row["创建月份"], row["开发人"])
        bucket = summary[key]
        bucket["skus"].add(sku)
        status = row["领星业务状态"]
        bucket["fba_observed"] += int(status == "已观察到FBA可售")
        bucket["active"] += int(status == "上架在售")
        bucket["eliminated"] += int(status == "淘汰")
        bucket["not_observed"] += int(status == "未在领星前端观察到")
        bucket["special"] += int(status not in {"上架在售", "淘汰", "未在领星前端观察到"})
        bucket["mature"] += int(row["数据截止时销售成熟度"] == "已满两个月，可作为稳定观察样本")
        stage = stage_by_sku[sku]
        bucket["first_month_sold"] += stage["first_month_sold"]
        bucket["second_month_continued"] += stage["second_month_continued"]
        financial = finance_by_sku.get(sku)
        if financial:
            matched_finance_skus.add(sku)
            bucket["finance_skus"].add(sku)
            for currency in ("GBP", "EUR"):
                bucket[f"{currency}_sales"] += financial[f"{currency}_sales"]
                bucket[f"{currency}_profit"] += financial[f"{currency}_profit"]
        detail_rows.append([
            row["创建月份"], row["开发人"], sku, row["品名"], row["领星业务状态"], row["最近领星前端观察月"], row["最近Listing标签"], row["首次FBA可售观察月"],
            row["数据截止时销售成熟度"], stage["store_sku_count"], stage["first_month_sold"], stage["second_month_continued"],
            len(financial["GBP_months"]) if financial else 0, financial["GBP_sales"] if financial else Decimal(0),
            financial["GBP_profit"] if financial else Decimal(0), len(financial["EUR_months"]) if financial else 0,
            financial["EUR_sales"] if financial else Decimal(0), financial["EUR_profit"] if financial else Decimal(0),
        ])

    unmatched_finance = sorted(set(finance_by_sku) - matched_finance_skus)
    if unmatched_finance:
        bucket = summary[("未匹配创建月份", "财务SKU未匹配生命周期")]
        for sku in unmatched_finance:
            financial = finance_by_sku[sku]
            bucket["skus"].add(sku)
            bucket["finance_skus"].add(sku)
            for currency in ("GBP", "EUR"):
                bucket[f"{currency}_sales"] += financial[f"{currency}_sales"]
                bucket[f"{currency}_profit"] += financial[f"{currency}_profit"]
            detail_rows.append([
                "未匹配创建月份", "财务SKU未匹配生命周期", sku, "", "财务数据存在，标签状态库未匹配", "", "", "", 0, 0, 0,
                len(financial["GBP_months"]), financial["GBP_sales"], financial["GBP_profit"], len(financial["EUR_months"]),
                financial["EUR_sales"], financial["EUR_profit"],
            ])

    summary_rows = []
    for (created_month, developer), row in sorted(summary.items()):
        summary_rows.append([
            created_month, developer, len(row["skus"]), row["active"], row["eliminated"], row["not_observed"], row["special"], row["mature"],
            row["first_month_sold"], row["second_month_continued"], len(row["finance_skus"]), row["GBP_sales"], row["GBP_profit"],
            row["EUR_sales"], row["EUR_profit"],
        ])
    return summary_rows, detail_rows


def main() -> None:
    lifecycle_rows = read_csv(LIFECYCLE_FILE)
    cohort_rows = read_csv(COHORT_FILE)
    developer_rows, lifecycle_status_rows = lifecycle_overview(lifecycle_rows)
    monthly_profit_rows, owner_profit_rows, sku_profit_rows = profit_tables()
    cohort_financial_summary_rows, cohort_financial_detail_rows = cohort_financial_tables(
        lifecycle_rows, cohort_rows, sku_profit_rows,
    )

    workbook = Workbook()
    build_usage_sheet(workbook)
    write_table(
        workbook.create_sheet("团队总览"),
        "按开发人汇总：开发 SKU、领星标签状态与历史 FBA 可售观察情况",
        ["开发人", "开发SKU数", "组合主产品数", "上架在售", "淘汰", "未在前端观察", "已观察FBA可售"], developer_rows,
    )
    write_table(
        workbook.create_sheet("开发批次经营"),
        "按开发人和创建月份复盘：领星标签状态、成熟、首月/次月出单与全期间财务（GBP/EUR 不相加）",
        ["创建月份", "开发人", "开发SKU数", "上架在售SKU数", "淘汰SKU数", "未在前端观察SKU数", "特殊标签SKU数", "已满两月观察SKU数", "首月出单店铺SKU数", "两月持续出单店铺SKU数", "有财务记录SKU数", "GBP销售额", "GBP结算毛利润", "EUR销售额", "EUR结算毛利润"],
        cohort_financial_summary_rows,
    )
    write_table(
        workbook.create_sheet("开发批次SKU明细"),
        "每个开发 SKU 的创建批次、领星标签状态、上架后阶段与 2025-04 至 2026-06 财务",
        ["创建月份", "开发人", "SKU", "品名", "领星业务状态", "最近领星前端观察月", "最近Listing标签", "首次FBA可售观察月", "销售成熟度", "观察店铺SKU数", "首月出单店铺SKU数", "两月持续出单店铺SKU数", "GBP财务月份数", "GBP销售额", "GBP结算毛利润", "EUR财务月份数", "EUR销售额", "EUR结算毛利润"],
        cohort_financial_detail_rows,
    )
    write_table(
        workbook.create_sheet("标签状态分布"),
        "按开发人和领星标签业务状态汇总",
        ["开发人", "领星业务状态", "SKU数"], lifecycle_status_rows,
    )
    lifecycle_headers = [
        "开发人", "SKU", "创建时间", "创建月份", "本地状态", "是否组合产品", "品名", "最近领星前端观察月", "最近Listing标签",
        "领星业务状态", "标签状态说明", "首次FBA可售观察月", "FBA匹配店铺SKU数", "FBA匹配店铺", "月度FBA匹配状态", "预计上架月份",
        "首个完整销售月份", "两个月成熟月份", "数据截止时销售成熟度",
    ]
    write_table(
        workbook.create_sheet("标签状态明细"),
        "团队 SKU 领星标签状态明细（可按开发人、标签状态、最近前端观察月筛选）",
        lifecycle_headers,
        [[row.get(header, "") for header in lifecycle_headers] for row in lifecycle_rows],
    )
    cohort_headers = [
        "站点", "国家", "店铺", "SKU", "开发人", "是否组合产品", "首次FBA可售观察月", "首个完整销售月份",
        "两个月成熟月份", "月度阶段判定", "首次FBA可售月销量", "首个完整月销量", "首个完整月结算毛利润",
        "首个完整月广告花费", "第二完整月销量", "第二完整月结算毛利润", "第二完整月广告花费",
    ]
    write_table(
        workbook.create_sheet("上架批次阶段"),
        "店铺-SKU 上架后首月与次月阶段表现",
        cohort_headers,
        [[row.get(header, "") for header in cohort_headers] for row in cohort_rows],
    )
    write_table(
        workbook.create_sheet("月度利润总览"),
        "2025-04 至 2026-06 完整月度团队结算毛利润（GBP 与 EUR 不相加）",
        ["月份", "币种", "财务日记录数", "SKU数", "销量", "销售额", "广告费", "总成本", "领星其他结算调整项", "结算毛利润", "汇总毛利率"],
        monthly_profit_rows,
    )
    write_table(
        workbook.create_sheet("月度开发人利润"),
        "2025-04 至 2026-06 月度开发人结算毛利润",
        ["月份", "开发人", "币种", "SKU数", "销量", "销售额", "广告费", "总成本", "领星其他结算调整项", "结算毛利润", "汇总毛利率"],
        owner_profit_rows,
    )
    write_table(
        workbook.create_sheet("月度SKU利润"),
        "2025-04 至 2026-06 月度 SKU 财务明细（按月份、开发人、SKU 和币种筛选）",
        ["月份", "开发人", "SKU", "币种", "财务日记录数", "ASIN数", "店铺数", "销量", "销售额", "广告费", "总成本", "领星其他结算调整项", "结算毛利润", "汇总毛利率"],
        sku_profit_rows,
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    print(f"lifecycle_rows={len(lifecycle_rows)} cohort_rows={len(cohort_rows)} sku_profit_rows={len(sku_profit_rows)} cohort_detail_rows={len(cohort_financial_detail_rows)} output={OUTPUT_FILE}")


if __name__ == "__main__":
    main()
