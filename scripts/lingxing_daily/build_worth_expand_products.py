# -*- coding: utf-8 -*-
"""
值得拓的产品筛选 —— 1688店雷达方向

口径（用户定案 2026-08）：
  核心：只取统一表 lingxing_product_unified 已知 ASIN
  按上架时间(model_start_month)分组：不限月份，所有月份的 ASIN 都进（每个上架月一个 sheet）
  条件1：产品详情"7天日均" > 0.5
        —— 7天日均 = 最新一周(week_start=2026-07-27) 该 asin 各 SKU 的 SUM(avg_volume)
           (周表 avg_volume = 周销量/7，即该周的日均销量)
  条件2：该 asin "月tacos" < 15%
        —— 月tacos = 最新完整月(2026-07) SUM(spend)/SUM(amount)

输出：产品数据/值得拓的产品，1688店雷达方向/值得拓的产品_全部上架月_<日期>.xlsx
"""
import os
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---- 参数 ----
LATEST_WEEK = "2026-07-27"   # 最新一周窗口起始（7天日均取此周）
TACOS_MONTH = "2026-07"      # 月tacos口径月（最新完整月）
DAY7_MIN = 0.5               # 7天日均下限
TACOS_MAX = 0.15             # 月tacos上限
# MONTHS 分组月份不再写死，运行时从命中数据里动态收集（降序），不限月份

BASE_DIR = os.path.join("产品数据", "值得拓的产品，1688店雷达方向")
DATE_TAG = "2026-08-03"
# 结果放独立子文件夹，方便观看；同目录里 4 个文件：按时间 + 按标/非标各一套 xlsx+md
OUT_DIR = os.path.join(BASE_DIR, f"值得拓的产品_{DATE_TAG}")
# xlsx 与 md 分开放两个子目录
OUT_DIR_XLSX = os.path.join(OUT_DIR, "xlsx")
OUT_DIR_MD = os.path.join(OUT_DIR, "md")
# 按上架月分组
OUT_FILE_TIME = os.path.join(OUT_DIR_XLSX, f"值得拓的产品_按上架月_{DATE_TAG}.xlsx")
OUT_MD_TIME = os.path.join(OUT_DIR_MD, f"值得拓的产品_按上架月_{DATE_TAG}.md")
# 按标品/非标品分组
OUT_FILE_TYPE = os.path.join(OUT_DIR_XLSX, f"值得拓的产品_按标品非标品_{DATE_TAG}.xlsx")
OUT_MD_TYPE = os.path.join(OUT_DIR_MD, f"值得拓的产品_按标品非标品_{DATE_TAG}.md")


def conn():
    return pymysql.connect(
        host=os.environ.get("MYSQL_HOST_EXTERNAL", "127.0.0.1"),
        port=int(os.environ.get("MYSQL_PORT_EXTERNAL", "3310")),
        user=os.environ.get("MYSQL_USERNAME", "root"),
        password=os.environ.get("MYSQL_PASSWORD", "root123456"),
        database=os.environ.get("MYSQL_DATABASE", "sijuelishi"),
        charset="utf8mb4",
    )


QUERY = """
SELECT
    u.model_start_month                         AS start_month,
    u.country                                   AS country,
    u.asin                                      AS asin,
    COALESCE(w.local_name, '')                   AS local_name,
    COALESCE(NULLIF(u.title,''), w.product_name) AS title,
    u.developer                                 AS developer,
    u.base_sku                                  AS base_sku,
    ROUND(w.day7, 3)                            AS day7_avg,
    ROUND(m.amt, 2)                             AS month_amount,
    m.vol                                        AS month_volume,
    ROUND(m.sp, 2)                              AS month_spend,
    ROUND(m.sp / m.amt, 4)                       AS month_tacos,
    u.listing_tags                              AS listing_tags,
    CASE WHEN u.listing_tags LIKE '%%非标品%%' THEN '非标品' ELSE '标品' END AS product_type,
    u.listing_open_date                         AS open_date
FROM lingxing_product_unified u
JOIN (
    SELECT asin, marketplace,
           SUM(avg_volume) AS day7,
           MAX(product_name) AS product_name,
           MAX(local_name)   AS local_name
    FROM lingxing_sku_weekly_performance
    WHERE week_start = %s
    GROUP BY asin, marketplace
    HAVING day7 > %s
) w ON w.asin = u.asin AND w.marketplace = u.country
JOIN (
    SELECT asin, marketplace,
           SUM(spend) AS sp, SUM(amount) AS amt, SUM(volume) AS vol
    FROM lingxing_sku_weekly_performance
    WHERE `year_month` = %s
    GROUP BY asin, marketplace
    HAVING amt > 0 AND sp / amt < %s
) m ON m.asin = u.asin AND m.marketplace = w.marketplace
WHERE u.model_start_month IS NOT NULL AND u.model_start_month <> ''
ORDER BY u.model_start_month DESC, u.country, w.day7 DESC
"""

ALL_COLUMNS = [
    ("asin", "ASIN", 16),
    ("local_name", "中文品名(本地)", 40),
    ("title", "英文标题", 52),
    ("country", "站点", 8),
    ("developer", "开发人", 12),
    ("base_sku", "基准SKU", 16),
    ("start_month", "上架月(起算月)", 14),
    ("day7_avg", "7天日均(>0.5)", 13),
    ("month_tacos", "月TACOS(<15%)", 13),
    ("month_volume", "7月销量", 10),
    ("month_amount", "7月销售额", 12),
    ("month_spend", "7月广告花费", 12),
    ("product_type", "标品/非标品", 11),
    ("listing_tags", "Listing标签", 22),
    ("open_date", "上架日", 20),
]

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
GROUP_FILL = PatternFill("solid", fgColor="D9E1F2")


def style_header(ws, row_idx, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_sheet(ws, rows):
    ncol = len(ALL_COLUMNS)
    # 说明行
    ws.cell(row=1, column=1,
            value=("值得拓的产品筛选 | 核心=统一表已知ASIN | "
                   "7天日均=最新周(2026-07-27~08-02) SUM(周日均)>0.5 | "
                   "月TACOS=2026-07 SUM(广告花费)/SUM(销售额)<15%"))
    ws.cell(row=1, column=1).font = Font(bold=True, size=10, color="C00000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    # 表头
    for i, (_, label, _) in enumerate(ALL_COLUMNS, start=1):
        ws.cell(row=2, column=i, value=label)
    style_header(ws, 2, ncol)

    # 数据
    r = 3
    for row in rows:
        for i, (key, _, _) in enumerate(ALL_COLUMNS, start=1):
            val = row.get(key)
            if key == "month_tacos" and val is not None:
                cell = ws.cell(row=r, column=i, value=float(val))
                cell.number_format = "0.00%"
            elif key == "open_date" and val is not None:
                ws.cell(row=r, column=i, value=str(val))
            else:
                ws.cell(row=r, column=i, value=val)
        r += 1

    # 列宽
    for i, (_, _, width) in enumerate(ALL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30


def month_label(m):
    """2026-07 -> 2026年07月上架"""
    y, mm = m.split("-")
    return f"{y}年{mm}月上架"


def sheet_label(m):
    """2026-07 -> 202607（sheet 名限长且不能含特殊字符，用紧凑年月）"""
    return m.replace("-", "")


def write_grouped_workbook(out_file, groups, group_keys, group_title_fn,
                           sheet_name_fn, summary_first_col):
    """通用：按 groups(dict: key->rows) 生成一个工作簿（汇总 sheet + 每组一个 sheet）。"""
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "汇总"
    ws0.cell(row=1, column=1, value=f"值得拓的产品 —— 按{summary_first_col}分组汇总").font = Font(bold=True, size=12)
    ws0.cell(row=3, column=1, value=summary_first_col)
    ws0.cell(row=3, column=2, value="UK")
    ws0.cell(row=3, column=3, value="DE")
    ws0.cell(row=3, column=4, value="合计")
    style_header(ws0, 3, 4)
    rr = 4
    grand = tot_uk = tot_de = 0
    for k in group_keys:
        rows = groups.get(k, [])
        uk = sum(1 for x in rows if x["country"] == "UK")
        de = sum(1 for x in rows if x["country"] == "DE")
        ws0.cell(row=rr, column=1, value=group_title_fn(k))
        ws0.cell(row=rr, column=2, value=uk)
        ws0.cell(row=rr, column=3, value=de)
        ws0.cell(row=rr, column=4, value=uk + de)
        grand += uk + de; tot_uk += uk; tot_de += de
        rr += 1
    ws0.cell(row=rr, column=1, value="总计").font = Font(bold=True)
    ws0.cell(row=rr, column=2, value=tot_uk).font = Font(bold=True)
    ws0.cell(row=rr, column=3, value=tot_de).font = Font(bold=True)
    ws0.cell(row=rr, column=4, value=grand).font = Font(bold=True)
    for i, w in enumerate([16, 10, 10, 10], start=1):
        ws0.column_dimensions[get_column_letter(i)].width = w

    for k in group_keys:
        ws = wb.create_sheet(title=sheet_name_fn(k))
        write_sheet(ws, groups.get(k, []))

    wb.save(out_file)
    return grand


def main():
    os.makedirs(OUT_DIR_XLSX, exist_ok=True)
    os.makedirs(OUT_DIR_MD, exist_ok=True)
    c = conn()
    cur = c.cursor()
    cur.execute(QUERY, (LATEST_WEEK, DAY7_MIN, TACOS_MONTH, TACOS_MAX))
    cols = [d[0] for d in cur.description]
    data = [dict(zip(cols, row)) for row in cur.fetchall()]
    cur.close(); c.close()

    # ---- 分组一：按上架月（降序，最新月在前） ----
    by_month = {}
    for row in data:
        by_month.setdefault(row["start_month"], []).append(row)
    months = sorted(by_month.keys(), reverse=True)
    grand = write_grouped_workbook(
        OUT_FILE_TIME, by_month, months,
        group_title_fn=month_label, sheet_name_fn=sheet_label,
        summary_first_col="上架月")
    write_markdown_time(by_month, months, grand)

    # ---- 分组二：按标品/非标品 ----
    by_type = {}
    for row in data:
        by_type.setdefault(row["product_type"], []).append(row)
    type_keys = [k for k in ["标品", "非标品"] if by_type.get(k)]
    write_grouped_workbook(
        OUT_FILE_TYPE, by_type, type_keys,
        group_title_fn=lambda k: k, sheet_name_fn=lambda k: k,
        summary_first_col="标品/非标品")
    write_markdown_type(by_type, type_keys, grand)

    print(f"已生成 4 个文件于: {OUT_DIR}")
    print(f"总计 {grand} 个 ASIN")
    print("[按上架月] 覆盖 %d 个月:" % len(months))
    for m in months:
        rows = by_month.get(m, [])
        print(f"  {m}: {len(rows)} (UK {sum(1 for x in rows if x['country']=='UK')} / "
              f"DE {sum(1 for x in rows if x['country']=='DE')})")
    print("[按标品/非标品]:")
    for k in type_keys:
        rows = by_type.get(k, [])
        print(f"  {k}: {len(rows)} (UK {sum(1 for x in rows if x['country']=='UK')} / "
              f"DE {sum(1 for x in rows if x['country']=='DE')})")


def _fmt(v, pct=False):
    if v is None:
        return ""
    if pct:
        return f"{float(v) * 100:.2f}%"
    s = str(v).replace("|", "/").replace("\n", " ").strip()
    return s


MD_HEADER = [
    f"# 值得拓的产品（1688店雷达方向）—— {DATE_TAG}",
    "",
    "> 核心：只取统一表 `lingxing_product_unified` 已知 ASIN。",
    "> 筛选条件：",
    ">",
    "> - **7天日均 > 0.5**：最新一周（2026-07-27~08-02）该 ASIN 各 SKU 的 SUM(周日均)，周日均 = 周销量÷7",
    "> - **月TACOS < 15%**：最新完整月 2026-07 的 SUM(广告花费) / SUM(销售额)",
    "> - **标品/非标品**：看 Listing 标签，含「欧洲精铺2025非标品」即为非标品，其余为标品",
]

MD_TABLE_HEAD = [
    "| ASIN | 中文品名(本地) | 英文标题 | 站点 | 开发人 | 7天日均 | 月TACOS | 7月销量 | 7月销售额 | 标/非标 | 上架日 |",
    "|---|---|---|---|---|---:|---:|---:|---:|---|---|",
]


def _md_row(r):
    return (
        f"| {_fmt(r.get('asin'))} "
        f"| {_fmt(r.get('local_name'))} "
        f"| {_fmt(r.get('title'))} "
        f"| {_fmt(r.get('country'))} "
        f"| {_fmt(r.get('developer'))} "
        f"| {_fmt(r.get('day7_avg'))} "
        f"| {_fmt(r.get('month_tacos'), pct=True)} "
        f"| {_fmt(r.get('month_volume'))} "
        f"| {_fmt(r.get('month_amount'))} "
        f"| {_fmt(r.get('product_type'))} "
        f"| {_fmt(str(r.get('open_date'))[:10] if r.get('open_date') else '')} |"
    )


def write_markdown_time(by_month, months, grand):
    lines = list(MD_HEADER)
    lines.append("> - 本文件按**上架月**（起算月 `model_start_month`）分组，不限月份，所有上架月都进。")
    lines.append("")
    lines.append("## 汇总")
    lines.append("")
    lines.append("| 上架月 | UK | DE | 合计 |")
    lines.append("|---|---:|---:|---:|")
    tot_uk = tot_de = 0
    for m in months:
        rows = by_month.get(m, [])
        uk = sum(1 for x in rows if x["country"] == "UK")
        de = sum(1 for x in rows if x["country"] == "DE")
        tot_uk += uk; tot_de += de
        lines.append(f"| {month_label(m)} | {uk} | {de} | {uk + de} |")
    lines.append(f"| **总计** | **{tot_uk}** | **{tot_de}** | **{grand}** |")
    lines.append("")
    for m in months:
        rows = by_month.get(m, [])
        lines.append(f"## {month_label(m)}（{len(rows)} 个）")
        lines.append("")
        if not rows:
            lines.append("_无_"); lines.append(""); continue
        lines.extend(MD_TABLE_HEAD)
        for r in rows:
            lines.append(_md_row(r))
        lines.append("")
    with open(OUT_MD_TIME, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_markdown_type(by_type, type_keys, grand):
    lines = list(MD_HEADER)
    lines.append("> - 本文件按**标品/非标品**分组（不分上架月）。")
    lines.append("")
    lines.append("## 汇总")
    lines.append("")
    lines.append("| 标品/非标品 | UK | DE | 合计 |")
    lines.append("|---|---:|---:|---:|")
    tot_uk = tot_de = 0
    for k in type_keys:
        rows = by_type.get(k, [])
        uk = sum(1 for x in rows if x["country"] == "UK")
        de = sum(1 for x in rows if x["country"] == "DE")
        tot_uk += uk; tot_de += de
        lines.append(f"| {k} | {uk} | {de} | {uk + de} |")
    lines.append(f"| **总计** | **{tot_uk}** | **{tot_de}** | **{grand}** |")
    lines.append("")
    for k in type_keys:
        rows = by_type.get(k, [])
        lines.append(f"## {k}（{len(rows)} 个）")
        lines.append("")
        if not rows:
            lines.append("_无_"); lines.append(""); continue
        lines.extend(MD_TABLE_HEAD)
        for r in rows:
            lines.append(_md_row(r))
        lines.append("")
    with open(OUT_MD_TYPE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    main()
