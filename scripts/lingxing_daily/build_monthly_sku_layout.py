"""Build the monthly first-FBA SKU layout from pre-filtered Lingxing workbooks."""

from __future__ import annotations

import csv
import re
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = ROOT / "\u4ea7\u54c1\u6570\u636e" / "\u9886\u661f\u6570\u636eapi" / "\u9886\u661f25\u5e74\u523026\u5e746\u6708\u6240\u6709\u6570\u636e\uff0c\u4ee5\u6bcf\u6708\u6570\u636e"
HISTORY_DIR = SOURCE_DIR / "\u5386\u53f2SKU\u4e0a\u67b6\u57fa\u7840\u6570\u636e_2025-04\u81f32026-06"
OUTPUT_DIR = HISTORY_DIR / "02_\u5386\u53f2SKU\u4e0a\u67b6\u57fa\u7ebf"
OVERVIEW_DIR = HISTORY_DIR / "01_\u5feb\u901f\u603b\u89c8"
WINDOW = re.compile(r"(20\d{2}-\d{2}-\d{2})~(20\d{2}-\d{2}-\d{2})")

DATE_FIELD = "\u521b\u5efa\u65f6\u95f4"
COUNTRY = "\u56fd\u5bb6"
STORE = "\u5e97\u94fa"
SKU = "SKU"
MSKU = "MSKU"
ASIN = "ASIN"
PARENT_ASIN = "\u7236ASIN"
FBA = "FBA-\u53ef\u552e"
TAGS = "listing\u6807\u7b7e"
PRODUCT = "\u54c1\u540d"
TITLE = "\u6807\u9898"

FIELD_LABELS = {
    "marketplace": "站点", "country": "国家", "store_name": "店铺", "sku": "SKU",
    "first_fba_month": "首次FBA可售观察月", "first_fba_window_start": "首次观察窗口开始日",
    "first_fba_window_end": "首次观察窗口结束日", "first_fba_max": "首次观察月最大FBA可售库存",
    "observed_active_months": "观察期内FBA可售月份数", "first_fba_status": "首次FBA可售判定",
    "mskus": "MSKU列表", "asins": "ASIN列表", "parent_asins": "父ASIN列表",
    "listing_tags": "Listing标签", "product_names": "品名", "titles": "商品标题",
}


def as_number(value) -> float:
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def marketplace(country: str) -> str:
    return {"\u82f1\u56fd": "UK", "\u5fb7\u56fd": "DE"}.get(country, "UNKNOWN")


def main() -> None:
    files = sorted(SOURCE_DIR.glob("*.xlsx"), key=lambda path: WINDOW.search(path.name).group(1))
    if len(files) != 15:
        raise RuntimeError(f"Expected 15 monthly files, found {len(files)}")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OVERVIEW_DIR.mkdir(parents=True, exist_ok=True)
    entities: dict[tuple[str, str, str, str], dict] = {}
    monthly_new: Counter[str] = Counter()
    first_window = None

    for path in files:
        match = WINDOW.search(path.name)
        if match is None:
            raise RuntimeError(f"Cannot read window from {path.name}")
        window_start, window_end = match.groups()
        month = window_start[:7]
        first_window = first_window or month
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(value) if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        positions = {field: headers.index(field) for field in (COUNTRY, STORE, SKU, MSKU, ASIN, PARENT_ASIN, FBA, TAGS, PRODUCT, TITLE)}
        monthly_active: dict[tuple[str, str, str, str], dict] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            country = str(row[positions[COUNTRY]] or "").strip()
            store = str(row[positions[STORE]] or "").strip()
            sku = str(row[positions[SKU]] or "").strip()
            if not country or not store or not sku or as_number(row[positions[FBA]]) <= 0:
                continue
            key = (marketplace(country), country, store, sku)
            item = monthly_active.setdefault(key, {"mskus": set(), "asins": set(), "parent_asins": set(), "tags": set(), "product_names": set(), "titles": set(), "max_fba": 0.0})
            for field, target in ((MSKU, "mskus"), (ASIN, "asins"), (PARENT_ASIN, "parent_asins"), (TAGS, "tags"), (PRODUCT, "product_names"), (TITLE, "titles")):
                value = str(row[positions[field]] or "").strip()
                if value:
                    item[target].add(value)
            item["max_fba"] = max(item["max_fba"], as_number(row[positions[FBA]]))
        workbook.close()
        for key, item in monthly_active.items():
            entity = entities.get(key)
            if entity is None:
                entity = {"marketplace": key[0], "country": key[1], "store_name": key[2], "sku": key[3], "first_fba_month": month, "first_fba_window_start": window_start, "first_fba_window_end": window_end, "first_fba_max": item["max_fba"], "observed_active_months": 0, "mskus": set(), "asins": set(), "parent_asins": set(), "tags": set(), "product_names": set(), "titles": set()}
                entities[key] = entity
                monthly_new[month] += 1
            entity["observed_active_months"] += 1
            for field in ("mskus", "asins", "parent_asins", "tags", "product_names", "titles"):
                entity[field].update(item[field])

    columns = ["marketplace", "country", "store_name", "sku", "first_fba_month", "first_fba_window_start", "first_fba_window_end", "first_fba_max", "observed_active_months", "first_fba_status", "mskus", "asins", "parent_asins", "listing_tags", "product_names", "titles"]
    rows = []
    for entity in entities.values():
        row = {field: entity[field] for field in columns if field in entity}
        row["first_fba_status"] = (
            "数据起点已可售，真实首次上架月未知"
            if entity["first_fba_month"] == first_window
            else "月度数据中首次观察到FBA可售"
        )
        for field in ("mskus", "asins", "parent_asins", "tags", "product_names", "titles"):
            row_name = "listing_tags" if field == "tags" else field
            row[row_name] = " | ".join(sorted(entity[field]))
        rows.append(row)
    rows.sort(key=lambda row: (row["first_fba_month"], row["marketplace"], row["store_name"], row["sku"]))

    csv_path = OUTPUT_DIR / "\u5386\u53f2SKU_\u9996\u6b21FBA\u53ef\u552e\u6392\u5e03\u8868_2025-04\u81f32026-06.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=[FIELD_LABELS[column] for column in columns])
        writer.writeheader()
        writer.writerows([{FIELD_LABELS[key]: value for key, value in row.items()} for row in rows])
    summary_path = OVERVIEW_DIR / "\u5386\u53f2SKU_\u6708\u5ea6\u65b0\u589e\u6c47\u603b_2025-04\u81f32026-06.csv"
    with summary_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle); writer.writerow(["首次FBA可售观察月", "新增店铺SKU数"])
        writer.writerows(sorted(monthly_new.items()))
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("SKU首次FBA可售月")
    sheet.append([FIELD_LABELS[column] for column in columns])
    for row in rows: sheet.append([row.get(column, "") for column in columns])
    summary = workbook.create_sheet("月度新增汇总")
    summary.append(["首次FBA可售观察月", "新增店铺SKU数"])
    for month, count in sorted(monthly_new.items()): summary.append([month, count])
    workbook.save(OUTPUT_DIR / "\u5386\u53f2SKU_\u9996\u6b21FBA\u53ef\u552e\u6392\u5e03\u8868_2025-04\u81f32026-06.xlsx")
    print(f"rows={len(rows)} csv={csv_path}")


if __name__ == "__main__":
    main()
