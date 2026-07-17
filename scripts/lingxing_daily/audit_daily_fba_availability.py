#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Stream one raw daily Lingxing workbook and count FBA-available SKU/ASIN."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from prepare_daily_exports import ASIN, DATE, FBA_AVAILABLE, SKU, iter_workbook_rows


ROOT = Path(__file__).resolve().parents[2]
ANNUAL_SOURCE_IDS = ("934061734342180864", "934062080035405824")


def decimal(value: str) -> Decimal:
    try:
        return Decimal(value or "0")
    except InvalidOperation:
        return Decimal(0)


def annual_asin_baseline() -> set[str]:
    files = []
    for source_id in ANNUAL_SOURCE_IDS:
        files.extend(ROOT.rglob(f"*{source_id}.xlsx"))
    asins: set[str] = set()
    for path in files:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        asin_index = headers.index(ASIN)
        for row in rows:
            asin = str(row[asin_index] or "").strip()
            if asin:
                asins.add(asin)
        workbook.close()
    if not asins:
        raise RuntimeError("未找到两份年度 ASIN 基线源")
    return asins


def empty_months():
    return defaultdict(lambda: {"skus": set(), "asins": set(), "rows": 0})


def summarize(months):
    return {
        month: {"FBA可售SKU数": len(value["skus"]), "FBA可售ASIN数": len(value["asins"]), "FBA可售原始记录行数": value["rows"]}
        for month, value in sorted(months.items())
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="原始产品表现 ASIN xlsx")
    args = parser.parse_args()
    baseline = annual_asin_baseline()
    all_months = empty_months()
    baseline_months = empty_months()
    rows_scanned = 0
    for _, row in iter_workbook_rows(args.source):
        rows_scanned += 1
        if decimal(row.get(FBA_AVAILABLE, "")) <= 0:
            continue
        month = row.get(DATE, "")[:7]
        asin = row.get(ASIN, "").strip()
        sku = row.get(SKU, "").strip()
        if not month or not asin:
            continue
        for buckets in (all_months, baseline_months if asin in baseline else None):
            if buckets is None:
                continue
            bucket = buckets[month]
            bucket["rows"] += 1
            bucket["asins"].add(asin)
            if sku:
                bucket["skus"].add(sku)
    result = {
        "source": str(args.source),
        "sourceRowsScanned": rows_scanned,
        "annualAsinBaseline": len(baseline),
        "原始全量": summarize(all_months),
        "年度ASIN基线内": summarize(baseline_months),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
