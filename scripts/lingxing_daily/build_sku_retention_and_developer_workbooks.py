"""Build per-SKU retention/return and per-developer monthly workbooks."""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
BASE_DIR = ROOT / "产品数据" / "领星数据api" / "领星25年到26年6月所有数据，以每月数据" / "历史SKU上架基础数据_2025-04至2026-06"
LIFECYCLE_FILE = BASE_DIR / "03_团队开发SKU生命周期" / "团队SKU_生命周期判定_数据截止2026-06.csv"
FINANCE_DIR = BASE_DIR / "05_财务利润周度回补"
OUTPUT_DIR = BASE_DIR / "00_整合工作簿"
RETENTION_OUTPUT = OUTPUT_DIR / "SKU留存回报与财务回正_2025-04至2026-06.xlsx"
DEVELOPER_OUTPUT = OUTPUT_DIR / "开发人月度与批次经营_2025-04至2026-06.xlsx"
MONTH_PATTERN = re.compile(r"(20\d{2})年(\d{2})月团队SKU财务利润明细\.csv")

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAD3")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NEGATIVE_FILL = PatternFill("solid", fgColor="F4CCCC")
POSITIVE_FILL = PatternFill("solid", fgColor="D9EAD3")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as source:
        return list(csv.DictReader(source))


def amount(value: str | Decimal | None) -> Decimal:
    try:
        return Decimal(str(value or "0").strip() or "0")
    except InvalidOperation:
        return Decimal(0)


def integer(value: str | Decimal | None) -> int:
    return int(amount(value))


def month_index(month: str) -> int:
    year, value = map(int, month.split("-"))
    return year * 12 + value - 1


def month_distance(start: str, end: str) -> int:
    return month_index(end) - month_index(start)


def month_range(start: str, end: str) -> list[str]:
    return [f"{index // 12:04d}-{index % 12 + 1:02d}" for index in range(month_index(start), month_index(end) + 1)]


def finance_files() -> list[tuple[str, Path]]:
    files = []
    for path in FINANCE_DIR.glob("*团队SKU财务利润明细.csv"):
        match = MONTH_PATTERN.fullmatch(path.name)
        if match:
            files.append((f"{match.group(1)}-{match.group(2)}", path))
    return sorted(files)


def style_table(sheet, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append([title])
    sheet["A1"].font = Font(bold=True, size=14, color="0B3D47")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.append(headers)
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    for values in rows:
        sheet.append([float(value) if isinstance(value, Decimal) else value for value in values])
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(2, sheet.max_row)}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 32
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = min(max(len(header) + 4, 12), 28)
        if header in {"品名", "最近Listing标签", "标签状态说明", "财务回正说明"}:
            sheet.column_dimensions[letter].width = 34
        if any(key in header for key in ("销售额", "广告费", "总成本", "调整项", "毛利润", "累计利润", "结算支出")):
            for cell in sheet[letter][2:]:
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
        if any(key in header for key in ("率", "ROI")):
            for cell in sheet[letter][2:]:
                cell.number_format = '0.00%;[Red]-0.00%'
        if "毛利润" in header or "累计利润" in header:
            area = f"{letter}3:{letter}{sheet.max_row}"
            sheet.conditional_formatting.add(area, CellIsRule(operator="lessThan", formula=["0"], fill=NEGATIVE_FILL))
            sheet.conditional_formatting.add(area, CellIsRule(operator="greaterThan", formula=["0"], fill=POSITIVE_FILL))


def usage_sheet(workbook: Workbook, title: str, rows: list[tuple[str, str]]) -> None:
    sheet = workbook.active
    sheet.title = "口径说明"
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=18, color="0B3D47")
    sheet.merge_cells("A1:E1")
    for index, (label, value) in enumerate(rows, start=3):
        sheet.cell(index, 1, label).fill = SECTION_FILL
        sheet.cell(index, 1).font = Font(bold=True)
        sheet.cell(index, 2, value).fill = NOTE_FILL
        sheet.cell(index, 2).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 115


def load_data() -> tuple[dict[str, dict[str, str]], list[str], dict[tuple[str, str], dict[str, dict[str, Decimal | int]]]]:
    lifecycle = {row["SKU"]: row for row in read_csv(LIFECYCLE_FILE)}
    monthly_data: dict[tuple[str, str], dict[str, dict[str, Decimal | int]]] = defaultdict(dict)
    months = []
    for month, path in finance_files():
        months.append(month)
        for row in read_csv(path):
            key = (row["SKU"], row["币种"])
            monthly_data[key][month] = {
                "sales_qty": amount(row["销量"]), "sales_amount": amount(row["销售额"]), "ads_cost": amount(row["广告费"]),
                "total_cost": amount(row["总成本"]), "adjustment": amount(row["领星其他结算调整项"]),
                "gross_profit": amount(row["毛利润"]), "facts": integer(row["财务日记录数"]),
                "developer": row["开发人"],
            }
    return lifecycle, sorted(months), monthly_data


def retention_rows(
    lifecycle: dict[str, dict[str, str]], months: list[str], monthly_data: dict[tuple[str, str], dict[str, dict[str, Decimal | int]]],
) -> list[list[object]]:
    cutoff = months[-1]
    currencies_by_sku: dict[str, set[str]] = defaultdict(set)
    for sku, currency in monthly_data:
        currencies_by_sku[sku].add(currency)
    rows = []
    for sku, life in lifecycle.items():
        for currency in sorted(currencies_by_sku.get(sku, {"无财务币种"})):
            by_month = monthly_data.get((sku, currency), {})
            financial_months = sorted(by_month)
            active_months = [month for month in financial_months if by_month[month]["sales_qty"] > 0]
            first_finance = financial_months[0] if financial_months else ""
            first_sale = active_months[0] if active_months else ""
            last_sale = active_months[-1] if active_months else ""
            observable_months = month_distance(first_sale, cutoff) + 1 if first_sale else 0
            sales_retention = Decimal(len(active_months)) / observable_months if observable_months else None
            continuous_months = 0
            if first_sale:
                for month in month_range(first_sale, cutoff):
                    if by_month.get(month, {}).get("sales_qty", Decimal(0)) <= 0:
                        break
                    continuous_months += 1
            total_sales = sum((point["sales_amount"] for point in by_month.values()), Decimal(0))
            total_ads = sum((point["ads_cost"] for point in by_month.values()), Decimal(0))
            total_cost = sum((point["total_cost"] for point in by_month.values()), Decimal(0))
            total_adjustment = sum((point["adjustment"] for point in by_month.values()), Decimal(0))
            total_profit = sum((point["gross_profit"] for point in by_month.values()), Decimal(0))
            settlement_outflow = max(Decimal(0), -(total_ads + total_cost + total_adjustment))
            gross_margin = total_profit / total_sales if total_sales else None
            settlement_roi = total_profit / settlement_outflow if settlement_outflow else None
            cumulative_profit = Decimal(0)
            payback_month = ""
            if first_finance:
                for month in month_range(first_finance, cutoff):
                    cumulative_profit += by_month.get(month, {}).get("gross_profit", Decimal(0))
                    if not payback_month and cumulative_profit > 0:
                        payback_month = month
            if not first_finance:
                payback_status = "无财务记录"
                payback_cycle = None
            elif payback_month:
                payback_cycle = month_distance(first_finance, payback_month) + 1
                payback_status = "首月财务转正" if payback_cycle == 1 else "累计财务转正"
            else:
                payback_cycle = None
                payback_status = "截至数据截止月未财务转正"
            if not first_sale:
                retention_status = "未销售"
            elif last_sale == cutoff:
                retention_status = "截至截止月仍销售"
            else:
                retention_status = "截止月未销售"
            rows.append([
                life["开发人"], sku, currency, life["创建月份"], life["领星业务状态"], life["最近领星前端观察月"],
                life["最近Listing标签"], life["首次FBA可售观察月"], first_finance, first_sale, last_sale, len(financial_months),
                len(active_months), observable_months, sales_retention, continuous_months, retention_status,
                total_sales, total_ads, total_cost, total_adjustment, settlement_outflow, total_profit, gross_margin,
                settlement_roi, payback_month, payback_cycle, payback_status,
            ])
    return rows


def retention_monthly_overviews(
    lifecycle: dict[str, dict[str, str]], months: list[str], monthly_data: dict[tuple[str, str], dict[str, dict[str, Decimal | int]]],
) -> tuple[list[list[object]], list[list[object]]]:
    developers = sorted({row["开发人"] for row in lifecycle.values()})
    payback_month: dict[tuple[str, str], str] = {}
    for (sku, currency), by_month in monthly_data.items():
        running = Decimal(0)
        for month in months:
            running += amount(by_month.get(month, {}).get("gross_profit"))
            if running > 0 and (sku, currency) not in payback_month:
                payback_month[(sku, currency)] = month
    grouped: dict[tuple[str, str], dict[str, object]] = defaultdict(lambda: {
        "finance_skus": set(), "sales_skus": set(), "GBP_sales": Decimal(0), "GBP_profit": Decimal(0),
        "EUR_sales": Decimal(0), "EUR_profit": Decimal(0), "GBP_payback": set(), "EUR_payback": set(),
    })
    for (sku, currency), by_month in monthly_data.items():
        developer = lifecycle.get(sku, {}).get("开发人")
        if not developer:
            developer = str(next(iter(by_month.values()))["developer"])
        for month, point in by_month.items():
            bucket = grouped[(developer, month)]
            bucket["finance_skus"].add(sku)
            if amount(point["sales_qty"]) > 0:
                bucket["sales_skus"].add(sku)
            if currency in {"GBP", "EUR"}:
                bucket[f"{currency}_sales"] += amount(point["sales_amount"])
                bucket[f"{currency}_profit"] += amount(point["gross_profit"])
                reached = payback_month.get((sku, currency))
                if reached and reached <= month:
                    bucket[f"{currency}_payback"].add(sku)

    developer_rows = []
    team_source: dict[str, dict[str, object]] = defaultdict(lambda: {
        "finance_skus": set(), "sales_skus": set(), "GBP_sales": Decimal(0), "GBP_profit": Decimal(0),
        "EUR_sales": Decimal(0), "EUR_profit": Decimal(0), "GBP_payback": set(), "EUR_payback": set(),
    })
    previous_sales: dict[str, set[str]] = defaultdict(set)
    for developer in developers:
        for month in months:
            row = grouped[(developer, month)]
            retained = previous_sales[developer] & row["sales_skus"]
            rate = Decimal(len(retained)) / len(previous_sales[developer]) if previous_sales[developer] else None
            developer_rows.append([
                developer, month, len(row["finance_skus"]), len(row["sales_skus"]), len(previous_sales[developer]), len(retained), rate,
                row["GBP_sales"], row["GBP_profit"], len(row["GBP_payback"]), row["EUR_sales"], row["EUR_profit"], len(row["EUR_payback"]),
            ])
            previous_sales[developer] = set(row["sales_skus"])
            team = team_source[month]
            team["finance_skus"].update(row["finance_skus"])
            team["sales_skus"].update(row["sales_skus"])
            for key in ("GBP_sales", "GBP_profit", "EUR_sales", "EUR_profit"):
                team[key] += row[key]
            team["GBP_payback"].update(row["GBP_payback"])
            team["EUR_payback"].update(row["EUR_payback"])
    team_rows = []
    previous_team_sales: set[str] = set()
    for month in months:
        row = team_source[month]
        retained = previous_team_sales & row["sales_skus"]
        rate = Decimal(len(retained)) / len(previous_team_sales) if previous_team_sales else None
        team_rows.append([
            month, len(row["finance_skus"]), len(row["sales_skus"]), len(previous_team_sales), len(retained), rate,
            row["GBP_sales"], row["GBP_profit"], len(row["GBP_payback"]), row["EUR_sales"], row["EUR_profit"], len(row["EUR_payback"]),
        ])
        previous_team_sales = set(row["sales_skus"])
    return team_rows, developer_rows


def retention_cohort_overviews(
    lifecycle: dict[str, dict[str, str]], months: list[str], monthly_data: dict[tuple[str, str], dict[str, dict[str, Decimal | int]]],
) -> tuple[list[list[object]], list[list[object]], dict[str, list[list[object]]]]:
    """Summarise each creation cohort once, then make it available by developer."""
    cutoff = months[-1]
    cohort: dict[tuple[str, str], dict[str, object]] = defaultdict(lambda: {
        "skus": set(), "finance_skus": set(), "sold_skus": set(), "cutoff_sales_skus": set(),
        "active": set(), "eliminated": set(), "special": set(), "not_observed": set(),
        "active_months": 0, "observable_months": 0,
        "GBP_sales": Decimal(0), "GBP_profit": Decimal(0), "GBP_payback": set(),
        "EUR_sales": Decimal(0), "EUR_profit": Decimal(0), "EUR_payback": set(),
    })
    sku_cohorts: dict[str, tuple[str, str]] = {}
    for sku, life in lifecycle.items():
        developer, created_month = life["开发人"], life["创建月份"]
        if not created_month or created_month > cutoff:
            continue
        key = (developer, created_month)
        sku_cohorts[sku] = key
        bucket = cohort[key]
        bucket["skus"].add(sku)
        state = life["领星业务状态"]
        if state == "上架在售":
            bucket["active"].add(sku)
        elif state == "淘汰":
            bucket["eliminated"].add(sku)
        elif state == "未在领星前端观察到":
            bucket["not_observed"].add(sku)
        else:
            bucket["special"].add(sku)

    for sku, key in sku_cohorts.items():
        sales_by_month = {
            month for month in months
            if sum((amount(monthly_data.get((sku, currency), {}).get(month, {}).get("sales_qty")) for currency in ("GBP", "EUR")), Decimal(0)) > 0
        }
        if sales_by_month:
            bucket = cohort[key]
            first_sale = min(sales_by_month)
            bucket["sold_skus"].add(sku)
            bucket["active_months"] += len(sales_by_month)
            bucket["observable_months"] += month_distance(first_sale, cutoff) + 1
            if cutoff in sales_by_month:
                bucket["cutoff_sales_skus"].add(sku)

    for (sku, currency), by_month in monthly_data.items():
        key = sku_cohorts.get(sku)
        if not key or currency not in {"GBP", "EUR"}:
            continue
        bucket = cohort[key]
        bucket["finance_skus"].add(sku)
        running_profit = Decimal(0)
        payback = False
        for month in months:
            point = by_month.get(month, {})
            bucket[f"{currency}_sales"] += amount(point.get("sales_amount"))
            profit = amount(point.get("gross_profit"))
            bucket[f"{currency}_profit"] += profit
            running_profit += profit
            if running_profit > 0:
                payback = True
        if payback:
            bucket[f"{currency}_payback"].add(sku)

    developer_rows: list[list[object]] = []
    per_developer: dict[str, list[list[object]]] = defaultdict(list)
    team_cohort: dict[str, dict[str, object]] = defaultdict(lambda: {
        "skus": set(), "finance_skus": set(), "sold_skus": set(), "cutoff_sales_skus": set(),
        "active": set(), "eliminated": set(), "special": set(), "not_observed": set(),
        "active_months": 0, "observable_months": 0,
        "GBP_sales": Decimal(0), "GBP_profit": Decimal(0), "GBP_payback": set(),
        "EUR_sales": Decimal(0), "EUR_profit": Decimal(0), "EUR_payback": set(),
    })
    for (developer, created_month), bucket in sorted(cohort.items()):
        retention_rate = Decimal(bucket["active_months"]) / int(bucket["observable_months"]) if bucket["observable_months"] else None
        row = [
            developer, created_month, len(bucket["skus"]), len(bucket["finance_skus"]), len(bucket["sold_skus"]), len(bucket["cutoff_sales_skus"]),
            bucket["active_months"], bucket["observable_months"], retention_rate, len(bucket["active"]), len(bucket["eliminated"]),
            len(bucket["special"]), len(bucket["not_observed"]), bucket["GBP_sales"], bucket["GBP_profit"],
            bucket["GBP_profit"] / bucket["GBP_sales"] if bucket["GBP_sales"] else None, len(bucket["GBP_payback"]),
            bucket["EUR_sales"], bucket["EUR_profit"], bucket["EUR_profit"] / bucket["EUR_sales"] if bucket["EUR_sales"] else None,
            len(bucket["EUR_payback"]),
        ]
        developer_rows.append(row)
        per_developer[developer].append(row[1:])
        team = team_cohort[created_month]
        for name in ("skus", "finance_skus", "sold_skus", "cutoff_sales_skus", "active", "eliminated", "special", "not_observed", "GBP_payback", "EUR_payback"):
            team[name].update(bucket[name])
        for name in ("active_months", "observable_months", "GBP_sales", "GBP_profit", "EUR_sales", "EUR_profit"):
            team[name] += bucket[name]

    team_rows = []
    for created_month, bucket in sorted(team_cohort.items()):
        retention_rate = Decimal(bucket["active_months"]) / int(bucket["observable_months"]) if bucket["observable_months"] else None
        team_rows.append([
            created_month, len(bucket["skus"]), len(bucket["finance_skus"]), len(bucket["sold_skus"]), len(bucket["cutoff_sales_skus"]),
            bucket["active_months"], bucket["observable_months"], retention_rate, len(bucket["active"]), len(bucket["eliminated"]),
            len(bucket["special"]), len(bucket["not_observed"]), bucket["GBP_sales"], bucket["GBP_profit"],
            bucket["GBP_profit"] / bucket["GBP_sales"] if bucket["GBP_sales"] else None, len(bucket["GBP_payback"]),
            bucket["EUR_sales"], bucket["EUR_profit"], bucket["EUR_profit"] / bucket["EUR_sales"] if bucket["EUR_sales"] else None,
            len(bucket["EUR_payback"]),
        ])
    return team_rows, developer_rows, per_developer


def developer_monthly_rows(
    lifecycle: dict[str, dict[str, str]], months: list[str], monthly_data: dict[tuple[str, str], dict[str, dict[str, Decimal | int]]],
) -> tuple[list[list[object]], dict[str, list[list[object]]]]:
    new_skus: dict[tuple[str, str], int] = defaultdict(int)
    for life in lifecycle.values():
        new_skus[(life["开发人"], life["创建月份"])] += 1
    grouped: dict[tuple[str, str, str], dict[str, Decimal | int | set[str]]] = defaultdict(lambda: {
        "skus": set(), "sales_skus": set(), "facts": 0, "sales_qty": Decimal(0), "sales_amount": Decimal(0),
        "ads_cost": Decimal(0), "total_cost": Decimal(0), "adjustment": Decimal(0), "gross_profit": Decimal(0),
    })
    for (sku, currency), by_month in monthly_data.items():
        for month, point in by_month.items():
            developer = str(point["developer"])
            bucket = grouped[(developer, month, currency)]
            bucket["skus"].add(sku)
            bucket["sales_skus"].update([sku] if point["sales_qty"] > 0 else [])
            bucket["facts"] += int(point["facts"])
            for key in ("sales_qty", "sales_amount", "ads_cost", "total_cost", "adjustment", "gross_profit"):
                bucket[key] += Decimal(point[key])
    rows_by_developer: dict[str, list[list[object]]] = defaultdict(list)
    cumulative: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    all_rows = []
    for (developer, month, currency), row in sorted(grouped.items()):
        cumulative[(developer, currency)] += row["gross_profit"]
        outflow = max(Decimal(0), -(row["ads_cost"] + row["total_cost"] + row["adjustment"]))
        margin = row["gross_profit"] / row["sales_amount"] if row["sales_amount"] else None
        roi = row["gross_profit"] / outflow if outflow else None
        result = [
            developer, month, currency, new_skus[(developer, month)], len(row["skus"]), len(row["sales_skus"]), row["facts"],
            row["sales_qty"], row["sales_amount"], row["ads_cost"], row["total_cost"], row["adjustment"], outflow,
            row["gross_profit"], margin, roi, cumulative[(developer, currency)],
        ]
        all_rows.append(result)
        rows_by_developer[developer].append(result[1:])
    return all_rows, rows_by_developer


def natural_monthly_overviews(
    lifecycle: dict[str, dict[str, str]], months: list[str], monthly_data: dict[tuple[str, str], dict[str, dict[str, Decimal | int]]],
) -> tuple[list[list[object]], list[list[object]]]:
    new_skus: dict[tuple[str, str], int] = defaultdict(int)
    developers = sorted({row["开发人"] for row in lifecycle.values()})
    for row in lifecycle.values():
        if row["创建月份"] in months:
            new_skus[(row["开发人"], row["创建月份"])] += 1
    grouped: dict[tuple[str, str], dict[str, object]] = defaultdict(lambda: {
        "finance_skus": set(), "sales_skus": set(),
        "GBP_sales": Decimal(0), "GBP_profit": Decimal(0), "EUR_sales": Decimal(0), "EUR_profit": Decimal(0),
    })
    for (sku, currency), by_month in monthly_data.items():
        for month, point in by_month.items():
            key = (str(point["developer"]), month)
            bucket = grouped[key]
            bucket["finance_skus"].add(sku)
            if amount(point["sales_qty"]) > 0:
                bucket["sales_skus"].add(sku)
            if currency in {"GBP", "EUR"}:
                bucket[f"{currency}_sales"] += amount(point["sales_amount"])
                bucket[f"{currency}_profit"] += amount(point["gross_profit"])
    cumulative: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    developer_rows = []
    team_buckets: dict[str, dict[str, object]] = defaultdict(lambda: {
        "new_skus": 0, "finance_skus": set(), "sales_skus": set(),
        "GBP_sales": Decimal(0), "GBP_profit": Decimal(0), "EUR_sales": Decimal(0), "EUR_profit": Decimal(0),
    })
    for developer in developers:
        for month in months:
            row = grouped[(developer, month)]
            cumulative[(developer, "GBP")] += row["GBP_profit"]
            cumulative[(developer, "EUR")] += row["EUR_profit"]
            result = [
                developer, month, new_skus[(developer, month)], len(row["finance_skus"]), len(row["sales_skus"]),
                row["GBP_sales"], row["GBP_profit"], row["GBP_profit"] / row["GBP_sales"] if row["GBP_sales"] else None,
                cumulative[(developer, "GBP")], row["EUR_sales"], row["EUR_profit"],
                row["EUR_profit"] / row["EUR_sales"] if row["EUR_sales"] else None, cumulative[(developer, "EUR")],
            ]
            developer_rows.append(result)
            team = team_buckets[month]
            team["new_skus"] += new_skus[(developer, month)]
            team["finance_skus"].update(row["finance_skus"])
            team["sales_skus"].update(row["sales_skus"])
            for key in ("GBP_sales", "GBP_profit", "EUR_sales", "EUR_profit"):
                team[key] += row[key]
    team_cumulative = {"GBP": Decimal(0), "EUR": Decimal(0)}
    team_rows = []
    for month in months:
        row = team_buckets[month]
        team_cumulative["GBP"] += row["GBP_profit"]
        team_cumulative["EUR"] += row["EUR_profit"]
        team_rows.append([
            month, row["new_skus"], len(row["finance_skus"]), len(row["sales_skus"]),
            row["GBP_sales"], row["GBP_profit"], row["GBP_profit"] / row["GBP_sales"] if row["GBP_sales"] else None,
            team_cumulative["GBP"], row["EUR_sales"], row["EUR_profit"],
            row["EUR_profit"] / row["EUR_sales"] if row["EUR_sales"] else None, team_cumulative["EUR"],
        ])
    return team_rows, developer_rows


def cohort_tracker_rows(
    lifecycle: dict[str, dict[str, str]], months: list[str], monthly_data: dict[tuple[str, str], dict[str, dict[str, Decimal | int]]],
) -> tuple[list[list[object]], list[list[object]], dict[str, list[list[object]]]]:
    first_month, cutoff_month = months[0], months[-1]
    cohort_meta: dict[tuple[str, str], dict[str, set[str]]] = defaultdict(lambda: {
        "skus": set(), "active": set(), "eliminated": set(), "not_observed": set(), "special": set(), "tracked": set(),
    })
    finance_skus = {sku for sku, _currency in monthly_data}
    tracked_skus: set[str] = set()
    for sku, life in lifecycle.items():
        created_month = life["创建月份"]
        if not created_month or created_month > cutoff_month:
            continue
        key = (life["开发人"], created_month)
        meta = cohort_meta[key]
        meta["skus"].add(sku)
        state = life["领星业务状态"]
        if state == "上架在售":
            meta["active"].add(sku)
        elif state == "淘汰":
            meta["eliminated"].add(sku)
        elif state == "未在领星前端观察到":
            meta["not_observed"].add(sku)
        else:
            meta["special"].add(sku)
        if life["最近领星前端观察月"] or sku in finance_skus:
            meta["tracked"].add(sku)
            tracked_skus.add(sku)

    cumulative: dict[tuple[str, str, str], Decimal] = {}
    payback_month: dict[tuple[str, str], str] = {}
    for (sku, currency), by_month in monthly_data.items():
        running = Decimal(0)
        for month in months:
            running += amount(by_month.get(month, {}).get("gross_profit"))
            cumulative[(sku, currency, month)] = running
            if running > 0 and (sku, currency) not in payback_month:
                payback_month[(sku, currency)] = month

    summary: dict[tuple[str, str, str], dict[str, object]] = {}
    for (developer, cohort_month), meta in cohort_meta.items():
        for performance_month in month_range(max(cohort_month, first_month), cutoff_month):
            summary[(developer, cohort_month, performance_month)] = {
                **meta, "sales_skus": set(), "GBP_sales": Decimal(0), "GBP_profit": Decimal(0),
                "EUR_sales": Decimal(0), "EUR_profit": Decimal(0), "GBP_cumulative": Decimal(0), "EUR_cumulative": Decimal(0),
                "GBP_payback": set(), "EUR_payback": set(),
            }

    detail_rows: list[list[object]] = []
    for sku in sorted(tracked_skus):
        life = lifecycle[sku]
        developer, cohort_month = life["开发人"], life["创建月份"]
        if not cohort_month or cohort_month > cutoff_month:
            continue
        for performance_month in month_range(max(cohort_month, first_month), cutoff_month):
            gbp = monthly_data.get((sku, "GBP"), {}).get(performance_month, {})
            eur = monthly_data.get((sku, "EUR"), {}).get(performance_month, {})
            gbp_qty, eur_qty = amount(gbp.get("sales_qty")), amount(eur.get("sales_qty"))
            has_sale = gbp_qty + eur_qty > 0
            bucket = summary[(developer, cohort_month, performance_month)]
            if has_sale:
                bucket["sales_skus"].add(sku)
            for currency, point in (("GBP", gbp), ("EUR", eur)):
                bucket[f"{currency}_sales"] += amount(point.get("sales_amount"))
                bucket[f"{currency}_profit"] += amount(point.get("gross_profit"))
                bucket[f"{currency}_cumulative"] += cumulative.get((sku, currency, performance_month), Decimal(0))
                reached = payback_month.get((sku, currency))
                if reached and reached <= performance_month:
                    bucket[f"{currency}_payback"].add(sku)
            detail_rows.append([
                developer, cohort_month, month_distance(cohort_month, performance_month) + 1, performance_month, sku, life["品名"],
                life["领星业务状态"], life["最近Listing标签"], life["首次FBA可售观察月"],
                gbp_qty, amount(gbp.get("sales_amount")), amount(gbp.get("gross_profit")), cumulative.get((sku, "GBP", performance_month), Decimal(0)),
                payback_month.get((sku, "GBP"), ""), eur_qty, amount(eur.get("sales_amount")), amount(eur.get("gross_profit")),
                cumulative.get((sku, "EUR", performance_month), Decimal(0)), payback_month.get((sku, "EUR"), ""),
                "本月有销售" if has_sale else "本月无销售",
            ])

    summary_rows: list[list[object]] = []
    developer_rows: dict[str, list[list[object]]] = defaultdict(list)
    for (developer, cohort_month, performance_month), row in sorted(summary.items()):
        result = [
            developer, cohort_month, month_distance(cohort_month, performance_month) + 1, performance_month,
            len(row["skus"]), len(row["tracked"]), len(row["active"]), len(row["eliminated"]), len(row["not_observed"]),
            len(row["special"]), len(row["sales_skus"]), Decimal(len(row["sales_skus"])) / len(row["skus"]) if row["skus"] else None,
            row["GBP_sales"], row["GBP_profit"], row["GBP_cumulative"], len(row["GBP_payback"]),
            row["EUR_sales"], row["EUR_profit"], row["EUR_cumulative"], len(row["EUR_payback"]),
        ]
        summary_rows.append(result)
        developer_rows[developer].append(result[1:])
    return summary_rows, detail_rows, developer_rows


def main() -> None:
    lifecycle, months, monthly_data = load_data()
    retention = retention_rows(lifecycle, months, monthly_data)
    team_retention_rows, developer_retention_rows = retention_monthly_overviews(lifecycle, months, monthly_data)
    team_cohort_retention_rows, developer_cohort_retention_rows, retention_developer_sheets = retention_cohort_overviews(lifecycle, months, monthly_data)
    team_natural_rows, developer_natural_rows = natural_monthly_overviews(lifecycle, months, monthly_data)
    cohort_summary_rows, cohort_detail_rows, cohort_developer_sheets = cohort_tracker_rows(lifecycle, months, monthly_data)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    retention_book = Workbook()
    usage_sheet(retention_book, "SKU 留存、回报与财务回正", [
        ("分析对象", "每行是一个团队 SKU 加一种结算币种；GBP 与 EUR 绝不相加。无财务记录 SKU 仍保留一行。"),
        ("销售月留存率", "销售活跃月数 / 从首次销售月到 2026-06 的可观察自然月数。它衡量销售持续性，不是客户留存。"),
        ("连续销售月数", "从首次销售月起，连续每月销量大于 0 的月份数；首次断月即停止累计。"),
        ("结算投入回报率", "累计结算毛利润 / 累计结算支出。结算支出 = -(广告费 + 商品成本 + 其他结算调整项)，均来自领星财务事实。"),
        ("财务口径回正", "从首次有财务记录月起，累计结算毛利润首次大于 0 的月份。它不等于真实采购现金回本，因为首批采购投入尚不能严格匹配到店铺-SKU且未处理汇率。"),
        ("业务状态", "以最近一次领星前端 Listing 标签为准；“欧洲精铺2025淘汰”即淘汰。"),
        ("月间销售留存率", "上月有销售的 SKU 中，本月仍有销售的比例；首个财务月没有上月基数，留空而不是填 0。"),
        ("开品批次总览", "每行是一批在同一个创建月份开出的 SKU。先看全员批次，再按开发人筛选或进入开发人专页；这里的留存率是该批所有已销售 SKU 的销售活跃月数 / 可观察月数。"),
    ])
    retention_monthly_headers = ["表现月份", "有财务SKU数", "本月销售SKU数", "上月销售SKU数", "本月仍销售SKU数", "月间销售留存率", "GBP本月销售额", "GBP本月结算毛利润", "GBP截至本月财务回正SKU数", "EUR本月销售额", "EUR本月结算毛利润", "EUR截至本月财务回正SKU数"]
    developer_retention_headers = ["开发人", *retention_monthly_headers]
    style_table(retention_book.create_sheet("全员月度留存总览"), "全员自然月销售留存与财务回正（GBP 与 EUR 横向分列）", retention_monthly_headers, team_retention_rows)
    style_table(retention_book.create_sheet("开发人月度留存总览"), "开发人自然月销售留存与财务回正（每人每月一行）", developer_retention_headers, developer_retention_rows)
    cohort_retention_headers = ["开品批次", "批次开发SKU数", "有财务SKU数", "累计有销售SKU数", "截止月仍销售SKU数", "销售活跃月数合计", "销售可观察月数合计", "销售月留存率", "截止月上架在售SKU数", "截止月淘汰SKU数", "其他标签状态SKU数", "未前端观察SKU数", "GBP累计销售额", "GBP累计结算毛利润", "GBP累计结算毛利率", "GBP财务回正SKU数", "EUR累计销售额", "EUR累计结算毛利润", "EUR累计结算毛利率", "EUR财务回正SKU数"]
    developer_cohort_retention_headers = ["开发人", *cohort_retention_headers]
    style_table(retention_book.create_sheet("全员开品批次回报总览"), "全员各开品批次的留存、标签状态与累计财务回报", cohort_retention_headers, team_cohort_retention_rows)
    style_table(retention_book.create_sheet("开发人开品批次回报"), "开发人各开品批次的留存、标签状态与累计财务回报", developer_cohort_retention_headers, developer_cohort_retention_rows)
    for developer, rows in sorted(retention_developer_sheets.items()):
        style_table(retention_book.create_sheet(f"{developer}-留存"), f"{developer} 各开品批次的留存与累计财务回报", cohort_retention_headers, rows)
    style_table(
        retention_book.create_sheet("SKU留存回报明细"),
        "团队 SKU 留存、回报与财务回正（2025-04 至 2026-06）",
        ["开发人", "SKU", "币种", "创建月份", "领星业务状态", "最近前端观察月", "最近Listing标签", "首次FBA可售月", "首次财务月", "首次销售月", "最近销售月", "财务观察月数", "销售活跃月数", "销售可观察月数", "销售月留存率", "首次销售起连续月数", "销售留存状态", "累计销售额", "累计广告费", "累计商品成本", "累计其他结算调整项", "累计结算支出", "累计结算毛利润", "累计结算毛利率", "结算投入回报率(ROI)", "财务回正月", "财务回正周期(月)", "财务回正说明"],
        retention,
    )
    retention_book.save(RETENTION_OUTPUT)

    developer_book = Workbook()
    usage_sheet(developer_book, "开发人月度与批次经营", [
        ("阅读顺序", "先看“全员自然月总览”，再看“开发人自然月总览”；需要追踪某位开发人的某一批产品时，看“开发批次月度总览”或该开发人批次 Sheet。"),
        ("自然月总览", "一行是一个自然月，GBP/EUR 横向分列；新增开发 SKU 数不按币种重复。"),
        ("批次追踪", "主视图按开发人、创建月份和表现月份组织。批次月龄 = 表现月份相对创建月份的自然月序号。"),
        ("批次状态", "上架在售、淘汰等是截至 2026-06 的最近领星标签状态，不能倒灌为历史月份状态。"),
        ("追踪明细", "逐 SKU 月度明细只包括已出现在领星前端或有财务记录的 SKU；未在前端观察到 SKU 仍保留在批次汇总计数。"),
        ("新增开发SKU数", "本地产品创建月份等于该自然月的 SKU 数；例如刘淼 2026-05 有 1 个 SKU（2610750），2026-06 为 0。"),
        ("利润", "结算毛利润直接来自领星 grossProfit；GBP 与 EUR 分开，不换汇、不相加。"),
        ("结算投入回报率", "当月结算毛利润 / 当月结算支出。该指标用于经营效率比较，不代表采购资金 ROI。"),
        ("累计结算毛利润", "同一开发人、同一币种按月份累计的领星结算毛利润。"),
    ])
    natural_headers = ["表现月份", "当月新增开发SKU数", "有财务SKU数", "本月销售SKU数", "GBP本月销售额", "GBP本月结算毛利润", "GBP本月毛利率", "GBP累计结算毛利润", "EUR本月销售额", "EUR本月结算毛利润", "EUR本月毛利率", "EUR累计结算毛利润"]
    developer_natural_headers = ["开发人", *natural_headers]
    cohort_headers = ["开发人", "开发批次", "批次月龄", "表现月份", "批次开发SKU数", "追踪SKU数", "截至截止月上架在售SKU数", "截至截止月淘汰SKU数", "未前端观察SKU数", "特殊标签SKU数", "本月销售SKU数", "本月销售留存率", "GBP本月销售额", "GBP本月结算毛利润", "GBP累计结算毛利润", "截至本月GBP财务回正SKU数", "EUR本月销售额", "EUR本月结算毛利润", "EUR累计结算毛利润", "截至本月EUR财务回正SKU数"]
    cohort_detail_headers = ["开发人", "开发批次", "批次月龄", "表现月份", "SKU", "品名", "当前领星业务状态", "最近Listing标签", "首次FBA可售月", "GBP本月销量", "GBP本月销售额", "GBP本月结算毛利润", "GBP累计结算毛利润", "GBP财务回正月", "EUR本月销量", "EUR本月销售额", "EUR本月结算毛利润", "EUR累计结算毛利润", "EUR财务回正月", "本月销售状态"]
    style_table(developer_book.create_sheet("全员自然月总览"), "全体开发人自然月经营总览（GBP 与 EUR 横向分列）", natural_headers, team_natural_rows)
    style_table(developer_book.create_sheet("开发人自然月总览"), "开发人自然月经营总览（每人每月仅一行）", developer_natural_headers, developer_natural_rows)
    style_table(developer_book.create_sheet("开发批次月度总览"), "开发批次在各表现月份的经营表现", cohort_headers, cohort_summary_rows)
    style_table(developer_book.create_sheet("开发批次SKU月度明细"), "开发批次 SKU 在各表现月份的经营明细", cohort_detail_headers, cohort_detail_rows)
    for developer, rows in sorted(cohort_developer_sheets.items()):
        style_table(developer_book.create_sheet(f"{developer}-批次"), f"{developer} 开发批次月度表现", cohort_headers[1:], rows)
    developer_book.save(DEVELOPER_OUTPUT)
    print(f"retention_rows={len(retention)} team_retention_rows={len(team_retention_rows)} developer_retention_rows={len(developer_retention_rows)} team_cohort_retention_rows={len(team_cohort_retention_rows)} developer_cohort_retention_rows={len(developer_cohort_retention_rows)} team_natural_rows={len(team_natural_rows)} developer_natural_rows={len(developer_natural_rows)} cohort_summary_rows={len(cohort_summary_rows)} cohort_detail_rows={len(cohort_detail_rows)} developers={len(cohort_developer_sheets)}")
    print(f"retention={RETENTION_OUTPUT}")
    print(f"developer={DEVELOPER_OUTPUT}")


if __name__ == "__main__":
    main()
