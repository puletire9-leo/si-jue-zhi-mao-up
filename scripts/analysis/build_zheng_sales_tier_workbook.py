from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
RUN_ID = "zheng_clean_no_variants_20260707"
MODEL_DIR = ROOT / "产品数据" / "邓总店铺" / "sellersprite_raw" / RUN_ID / "sales_tier_model"
SOURCE_CSV = MODEL_DIR / "zheng_shop_products_sales_tier_all.csv"
OUTPUT_XLSX = MODEL_DIR / "郑总店铺_UK_DE_商品销量分级模型.xlsx"

TIER_ORDER = ["A", "B", "C", "D", "UNKNOWN"]
TIER_SORT = {tier: index for index, tier in enumerate(TIER_ORDER)}


def parse_date_ms(value: Any) -> str:
    if pd.isna(value):
        return ""
    try:
        ts = pd.to_datetime(float(value), unit="ms", errors="coerce")
    except (TypeError, ValueError):
        return ""
    if pd.isna(ts):
        return ""
    return ts.strftime("%Y-%m-%d")


def category_level(value: Any, depth: int) -> str:
    if pd.isna(value):
        return ""
    parts = [part.strip() for part in str(value).split(":") if part.strip()]
    return ":".join(parts[:depth])


def load_products() -> pd.DataFrame:
    df = pd.read_csv(SOURCE_CSV, encoding="utf-8-sig")
    df["salesTier"] = df["salesTier"].fillna("UNKNOWN").where(df["salesTier"].isin(TIER_ORDER), "UNKNOWN")
    df["units"] = pd.to_numeric(df["units"], errors="coerce")
    df["bsr"] = pd.to_numeric(df["bsr"], errors="coerce")
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["availableDateText"] = df["availableDate"].map(parse_date_ms)
    df["categoryL1"] = df["nodeLabelPath"].map(lambda value: category_level(value, 1))
    df["categoryL2"] = df["nodeLabelPath"].map(lambda value: category_level(value, 2))
    df["tierRank"] = df["salesTier"].map(TIER_SORT).fillna(99).astype(int)
    return df.sort_values(["marketplace", "tierRank", "sellerName", "units"], ascending=[True, True, True, False])


def percent(value: float) -> str:
    return f"{value * 100:.1f}%"


def build_overview(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for marketplace, group in df.groupby("marketplace"):
        counts = group["salesTier"].value_counts().reindex(TIER_ORDER, fill_value=0)
        total = int(len(group))
        ab = int(counts["A"] + counts["B"])
        abc = int(ab + counts["C"])
        rows.append(
            {
                "国家": marketplace,
                "商品总数": total,
                "A_利润锚点": int(counts["A"]),
                "A占比": percent(counts["A"] / total if total else 0),
                "B_强候选": int(counts["B"]),
                "B占比": percent(counts["B"] / total if total else 0),
                "C_稳定池": int(counts["C"]),
                "C占比": percent(counts["C"] / total if total else 0),
                "D_测品未起量": int(counts["D"]),
                "D占比": percent(counts["D"] / total if total else 0),
                "UNKNOWN": int(counts["UNKNOWN"]),
                "A+B": ab,
                "A+B占比": percent(ab / total if total else 0),
                "A+B+C": abc,
                "A+B+C占比": percent(abc / total if total else 0),
            }
        )
    return pd.DataFrame(rows)


def top_category(group: pd.DataFrame, tiers: list[str]) -> str:
    subset = group[group["salesTier"].isin(tiers) & group["categoryL2"].astype(bool)]
    if subset.empty:
        return ""
    counts = subset["categoryL2"].value_counts()
    return f"{counts.index[0]} ({int(counts.iloc[0])})"


def build_seller_distribution(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (marketplace, seller_name), group in df.groupby(["marketplace", "sellerName"], dropna=False):
        counts = group["salesTier"].value_counts().reindex(TIER_ORDER, fill_value=0)
        total = int(len(group))
        ab = int(counts["A"] + counts["B"])
        abc = int(ab + counts["C"])
        rows.append(
            {
                "国家": marketplace,
                "店铺名": seller_name,
                "商品总数": total,
                "A": int(counts["A"]),
                "B": int(counts["B"]),
                "C": int(counts["C"]),
                "D": int(counts["D"]),
                "UNKNOWN": int(counts["UNKNOWN"]),
                "A占比": percent(counts["A"] / total if total else 0),
                "A+B": ab,
                "A+B占比": percent(ab / total if total else 0),
                "A+B+C": abc,
                "A+B+C占比": percent(abc / total if total else 0),
                "D占比": percent(counts["D"] / total if total else 0),
                "A主类目": top_category(group, ["A"]),
                "ABC主类目": top_category(group, ["A", "B", "C"]),
                "D主类目": top_category(group, ["D"]),
            }
        )
    return pd.DataFrame(rows).sort_values(["国家", "A+B+C", "A"], ascending=[True, False, False])


def build_category_summary(df: pd.DataFrame) -> pd.DataFrame:
    subset = df[df["categoryL2"].astype(bool)].copy()
    summary = (
        subset.groupby(["marketplace", "salesTier", "categoryL2"], dropna=False)
        .agg(
            商品数=("asin", "count"),
            店铺数=("sellerName", "nunique"),
            月销量合计=("units", "sum"),
            月销量中位数=("units", "median"),
            BSR中位数=("bsr", "median"),
        )
        .reset_index()
    )
    summary["等级排序"] = summary["salesTier"].map(TIER_SORT)
    summary = summary.sort_values(["marketplace", "等级排序", "商品数"], ascending=[True, True, False]).drop(columns=["等级排序"])
    return summary.rename(columns={"marketplace": "国家", "salesTier": "销量等级", "categoryL2": "榜单类目"})


def product_view(df: pd.DataFrame) -> pd.DataFrame:
    columns = {
        "marketplace": "国家",
        "salesTier": "销量等级",
        "sellerName": "店铺名",
        "sellerId": "sellerId",
        "asin": "ASIN",
        "parentAsin": "父ASIN",
        "asinUrl": "ASIN链接",
        "units": "月销量units",
        "title": "标题",
        "brand": "品牌",
        "price": "价格",
        "bsrId": "BSR大类",
        "bsr": "BSR排名",
        "categoryL2": "二级榜单类目",
        "nodeLabelPath": "完整榜单类目",
        "bestSellerSubcategories": "BestSeller子类目",
        "availableDateText": "上架时间",
        "variations": "变体数",
        "imageUrl": "图片链接",
        "sourceResponse": "原始响应文件",
    }
    return df[list(columns)].rename(columns=columns)


def explanation_rows() -> list[list[Any]]:
    return [
        ["郑总店铺 UK/DE 商品销量分级模型", ""],
        ["数据源", f"{RUN_ID}，卖家精灵 competitor-lookup 店铺名查询"],
        ["请求口径", "variation=Y，不含变体；本文件直接使用卖家精灵返回行，不再二次父 ASIN 去重"],
        ["为什么要分级", "把店铺商品拆成利润锚点、强候选、稳定池、测品池，便于看一家精铺店的结构，而不是只看商品总数"],
        ["为什么不用固定好坏判断", "A/B/C/D 是销量阶段信号，不是简单好坏。D 可能是新测品，也可能是未起量品；A 是结果，但不代表完整打法"],
        ["A", "月销量 units >= 100。代表已经跑出来的利润锚点，可看店铺最终靠什么赚钱"],
        ["B", "月销量 50-99。代表强候选盈利池，通常已经有稳定销量但还没达到 A 的强度"],
        ["C", "月销量 15-49。代表稳定小单/潜力池，是判断店铺主打法和可复制方向的重要中间层"],
        ["D", "月销量 0-14。代表测品池或未起量池；在精铺模型里不是垃圾，而是后续转化观察对象"],
        ["UNKNOWN", "units 为空。数据缺失，不参与正负判断"],
        ["看 A", "看利润答案：这家店最后跑出来的赚钱方向"],
        ["看 A+B", "看确定性盈利和强候选池"],
        ["看 A+B+C", "看店铺稳定经营画像和主打法，比只看 A 更完整"],
        ["看 D", "看测品规模和未起量压力。后续如要更细，需要叠加上架时间"],
        ["注意", "这是商品基础特征处理层/数据辅助层，不等于方法卡命中；方法卡仍应保持独立筛选逻辑"],
    ]


def write_dataframe(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False)


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells[:100]]
            width = max([len(value.encode("gbk", errors="ignore")) // 2 + 2 for value in values] + [10])
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width, 45)
        sheet.auto_filter.ref = sheet.dimensions

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        for link_header in ("ASIN链接", "图片链接"):
            if link_header not in headers:
                continue
            link_col = headers.index(link_header) + 1
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=link_col)
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"
        if "ASIN" in headers and "ASIN链接" in headers:
            asin_col = headers.index("ASIN") + 1
            url_col = headers.index("ASIN链接") + 1
            for row in range(2, sheet.max_row + 1):
                url = sheet.cell(row=row, column=url_col).value
                asin_cell = sheet.cell(row=row, column=asin_col)
                if url and asin_cell.value:
                    asin_cell.hyperlink = str(url)
                    asin_cell.style = "Hyperlink"

    intro = workbook["00_分级说明"]
    intro.freeze_panes = None
    intro.column_dimensions["A"].width = 22
    intro.column_dimensions["B"].width = 110
    for row in range(1, intro.max_row + 1):
        intro.cell(row=row, column=1).font = Font(bold=True)
    intro["A1"].font = Font(bold=True, size=15, color="1F4E78")
    intro.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    workbook.save(path)


def main() -> int:
    df = load_products()
    overview = build_overview(df)
    seller_distribution = build_seller_distribution(df)
    category_summary = build_category_summary(df)
    products = product_view(df)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(explanation_rows(), columns=["项目", "说明"]).to_excel(writer, sheet_name="00_分级说明", index=False)
        write_dataframe(writer, overview, "01_两国总览")
        write_dataframe(writer, seller_distribution, "02_店铺分布")
        write_dataframe(writer, category_summary, "03_榜单类目汇总")
        write_dataframe(writer, products, "04_全部商品明细")

        for marketplace in ["UK", "DE"]:
            for tier in TIER_ORDER:
                sheet_df = products[(products["国家"] == marketplace) & (products["销量等级"] == tier)].copy()
                sheet_df = sheet_df.sort_values(["店铺名", "月销量units"], ascending=[True, False])
                write_dataframe(writer, sheet_df, f"{marketplace}_{tier}")

    style_workbook(OUTPUT_XLSX)
    print(f"output={OUTPUT_XLSX}")
    print(f"rows={len(df)}")
    print(overview.to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
