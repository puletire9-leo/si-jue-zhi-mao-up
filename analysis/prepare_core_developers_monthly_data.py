from __future__ import annotations

import argparse
import csv
import json
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font

from analyze_june_2026_developer_directions import (
    HEADER_FILL,
    WHITE_BOLD,
    analyze,
    developer_overview,
    mysql_config,
    write_table,
)


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = (
    ROOT
    / "产品数据"
    / "思考理实团队的开品方向"
    / "第一版"
    / "第一层_核心开发人开品分析_陈杨宋凤莉蒋舒"
)
DEVELOPERS = ("陈杨", "宋凤莉", "蒋舒")
START_MONTH = "2025-04"
END_MONTH = "2026-06"

DETAIL_HEADERS = [
    "创建月份",
    "开发人",
    "SKU",
    "创建时间",
    "产品名称",
    "采购价(CNY)",
    "是否组合主SKU",
    "本地产品状态",
    "开品状态",
    "是否定制",
    "是否多件/套装",
    "主方向",
    "产品族",
    "主题",
    "商品标准化类型",
    "价值主导类型",
    "标准化判定依据",
    "标准化判定置信度",
    "是否需要人工复核",
    "风险标记",
    "第一层判断",
    "判断说明",
    "产品族首次出现月份（开发人）",
    "产品族最近出现月份（开发人）",
    "产品族出现月份数（开发人）",
    "产品族出现月份列表（开发人）",
    "本月产品族状态",
]


def month_sequence(start_month: str, end_month: str) -> list[str]:
    start_year, start_value = map(int, start_month.split("-"))
    end_year, end_value = map(int, end_month.split("-"))
    current = start_year * 12 + start_value - 1
    end = end_year * 12 + end_value - 1
    months = []
    while current <= end:
        months.append(f"{current // 12:04d}-{current % 12 + 1:02d}")
        current += 1
    return months


def month_folder(month: str) -> str:
    year, value = map(int, month.split("-"))
    return f"{str(year)[2:]}年{value}月"


def month_display(month: str) -> str:
    year, value = map(int, month.split("-"))
    return f"{year}年{value:02d}月"


def end_exclusive(month: str) -> str:
    year, value = map(int, month.split("-"))
    index = year * 12 + value
    return f"{index // 12:04d}-{index % 12 + 1:02d}-01"


def load_rows(start_month: str, end_month: str) -> list[dict[str, Any]]:
    placeholders = ", ".join(["%s"] * len(DEVELOPERS))
    sql = f"""
        SELECT sku, product_developer, product_name, cg_price, is_combo,
               status_text, open_status, lx_create_time
        FROM lingxing_local_product
        WHERE product_developer IN ({placeholders})
          AND lx_create_time >= %s
          AND lx_create_time < %s
          AND sku IS NOT NULL AND sku <> ''
        ORDER BY lx_create_time, product_developer, sku
    """
    params = (*DEVELOPERS, f"{start_month}-01", end_exclusive(end_month))
    connection = pymysql.connect(**mysql_config())
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            columns = [item[0] for item in cursor.description]
            return [dict(zip(columns, values)) for values in cursor.fetchall()]
    finally:
        connection.close()


def deduplicate_rows(rows: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    unique: dict[tuple[str, str], dict[str, Any]] = {}
    duplicates: list[dict[str, Any]] = []
    for row in rows:
        key = (str(row.get("product_developer") or ""), str(row.get("sku") or ""))
        if key in unique:
            duplicates.append(row)
            continue
        unique[key] = row
    return list(unique.values()), duplicates


def prepare_details(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    classified = analyze(rows)
    details: list[dict[str, Any]] = []
    for source, item in zip(rows, classified):
        created = source["lx_create_time"]
        month = created.strftime("%Y-%m") if hasattr(created, "strftime") else str(created)[:7]
        detail = dict(item)
        detail["创建月份"] = month
        detail["本地产品状态"] = str(source.get("status_text") or "")
        detail["开品状态"] = str(source.get("open_status") or "")
        details.append(detail)

    family_months: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in details:
        family_months[(row["开发人"], row["产品族"])].add(row["创建月份"])

    for row in details:
        months = sorted(family_months[(row["开发人"], row["产品族"])])
        row["产品族首次出现月份（开发人）"] = months[0]
        row["产品族最近出现月份（开发人）"] = months[-1]
        row["产品族出现月份数（开发人）"] = len(months)
        row["产品族出现月份列表（开发人）"] = "、".join(months)
        row["本月产品族状态"] = "新增品线" if row["创建月份"] == months[0] else "跨月延续"
    return details


def overview_rows(details: list[dict[str, Any]], months: list[str]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for month in months:
        selected_month = [row for row in details if row["创建月份"] == month]
        for developer in DEVELOPERS:
            selected = [row for row in selected_month if row["开发人"] == developer]
            families = {row["产品族"] for row in selected}
            new_families = {row["产品族"] for row in selected if row["本月产品族状态"] == "新增品线"}
            rows.append([
                month,
                developer,
                len(selected),
                sum(row["是否组合主SKU"] == "是" for row in selected),
                sum(row["是否定制"] == "是" for row in selected),
                sum(row["是否多件/套装"] == "是" for row in selected),
                len(families),
                len(new_families),
                len(families - new_families),
                sum(row["商品标准化类型"] == "标品" for row in selected),
                sum(row["商品标准化类型"] == "非标" for row in selected),
                sum(row["商品标准化类型"] == "待复核" for row in selected),
                sum(row["风险标记"] != "无明显硬风险" for row in selected),
            ])
    return rows


def family_registry_rows(details: list[dict[str, Any]]) -> list[list[Any]]:
    buckets: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in details:
        buckets[(row["开发人"], row["产品族"])].append(row)

    output: list[list[Any]] = []
    for (developer, family), items in sorted(buckets.items()):
        months = sorted({row["创建月份"] for row in items})
        direction = Counter(row["主方向"] for row in items).most_common(1)[0][0]
        themes = Counter(
            theme
            for row in items
            for theme in str(row["主题"]).split("、")
            if theme != "通用/无明显主题"
        )
        output.append([
            developer,
            family,
            direction,
            len(items),
            months[0],
            months[-1],
            len(months),
            "、".join(months),
            "、".join(theme for theme, _ in themes.most_common(5)) or "通用/无明显主题",
            sum(row["风险标记"] != "无明显硬风险" for row in items),
        ])
    return output


def monthly_family_rows(details: list[dict[str, Any]], month: str) -> list[list[Any]]:
    buckets: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in details:
        if row["创建月份"] == month:
            buckets[(row["开发人"], row["产品族"])].append(row)

    output: list[list[Any]] = []
    for (developer, family), items in sorted(buckets.items()):
        sample = items[0]
        output.append([
            developer,
            family,
            Counter(row["主方向"] for row in items).most_common(1)[0][0],
            len(items),
            sample["本月产品族状态"],
            sample["产品族首次出现月份（开发人）"],
            sample["产品族最近出现月份（开发人）"],
            sample["产品族出现月份数（开发人）"],
            sample["产品族出现月份列表（开发人）"],
            "、".join(name for name, _ in Counter(row["主题"] for row in items).most_common(3)),
        ])
    return output


def add_notes_sheet(workbook: Workbook, title: str, explanations: list[tuple[str, str]]) -> None:
    sheet = workbook.active
    sheet.title = "00_数据说明"
    sheet.append([title])
    sheet["A1"].font = Font(bold=True, size=17, color="0B3D47")
    sheet.append(["项目", "说明"])
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
    for item in explanations:
        sheet.append(item)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 110
    sheet.freeze_panes = "A3"


def save_workbook_safely(workbook: Workbook, target: Path, attempts: int = 5) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f"{target.stem}.tmp{target.suffix}")
    last_error: OSError | None = None
    try:
        for attempt in range(attempts):
            try:
                workbook.save(temporary)
                break
            except OSError as error:
                last_error = error
                if attempt == attempts - 1:
                    raise
                time.sleep(0.2 * (attempt + 1))

        for attempt in range(attempts):
            try:
                temporary.replace(target)
                return
            except OSError as error:
                last_error = error
                if attempt == attempts - 1:
                    raise
                time.sleep(0.2 * (attempt + 1))
    finally:
        if temporary.exists():
            temporary.unlink()
    if last_error:
        raise last_error


def write_detail_sheet(workbook: Workbook, name: str, title: str, details: list[dict[str, Any]]) -> None:
    write_table(
        workbook.create_sheet(name),
        title,
        DETAIL_HEADERS,
        [[row.get(header, "") for header in DETAIL_HEADERS] for row in details],
    )
    sheet = workbook[name]
    for cell in sheet["F"][2:]:
        cell.number_format = "0.00"


def build_month_workbook(details: list[dict[str, Any]], month: str, target: Path) -> None:
    selected = [row for row in details if row["创建月份"] == month]
    workbook = Workbook()
    add_notes_sheet(
        workbook,
        f"{month_display(month)}三位核心开发人开品基础数据",
        [
            ("时间范围", f"创建月份为{month}。"),
            ("开发人", "陈杨、宋凤莉、蒋舒。"),
            ("当前阶段", "只做第一层数据整理，不评价FBA上架、销量、利润和最终经营成败。"),
            ("产品族状态", "新增品线=该开发人的这个产品族在15个月观察范围内首次出现；跨月延续=此前月份已经出现。2025-04是观察起点，不能据此证明历史上从未开发。"),
            ("季节性说明", "本文件先保留月份、产品族和主题事实，季节性结论在逐月分析阶段单独判断。"),
            ("标品/非标", "定制、玩具、图像审美和娱乐情绪主导判非标；功能、适配、维修、收纳和工具主导判标品；标题证据不足保留为待复核。"),
            ("旧文件", "原陈杨与宋凤莉2026年4—6月分析保持不变，没有删除或覆盖。"),
        ],
    )

    overview_headers = [
        "开发人", "SKU数", "产品族数", "定制SKU数", "定制占比", "多件套装SKU数",
        "多件套装占比", "组合主SKU数", "重复产品名SKU数", "采购价中位数", "采购价均值",
        "有风险标记SKU数", "风险占比", "符合主方向", "普通测试", "需复核", "高风险",
        "标品SKU数", "非标SKU数", "标准化待复核SKU数", "新增品线数", "跨月延续品线数",
    ]
    overview_data = []
    for developer in DEVELOPERS:
        overview = developer_overview(selected, developer)
        developer_rows = [row for row in selected if row["开发人"] == developer]
        families = {row["产品族"] for row in developer_rows}
        new_families = {row["产品族"] for row in developer_rows if row["本月产品族状态"] == "新增品线"}
        overview_data.append([
            overview.get(header, "") for header in overview_headers[:-5]
        ] + [
            sum(row["商品标准化类型"] == "标品" for row in developer_rows),
            sum(row["商品标准化类型"] == "非标" for row in developer_rows),
            sum(row["商品标准化类型"] == "待复核" for row in developer_rows),
            len(new_families),
            len(families - new_families),
        ])
    write_table(workbook.create_sheet("01_三人数量汇总"), "三人本月基础数量汇总", overview_headers, overview_data)
    for column in ("E", "G", "M"):
        for cell in workbook["01_三人数量汇总"][column][2:]:
            cell.number_format = "0.00%"

    write_detail_sheet(workbook, "02_全部SKU基础明细", "三人本月全部SKU基础明细", selected)
    for index, developer in enumerate(DEVELOPERS, start=3):
        write_detail_sheet(
            workbook,
            f"0{index}_{developer}SKU",
            f"{developer}{month_display(month)}全部SKU基础明细",
            [row for row in selected if row["开发人"] == developer],
        )

    family_headers = [
        "开发人", "产品族", "主方向", "本月SKU数", "本月产品族状态", "首次月份",
        "最近月份", "出现月份数", "出现月份列表", "本月主要主题",
    ]
    write_table(
        workbook.create_sheet("06_本月产品族状态"),
        "本月产品族新增与跨月延续基础表",
        family_headers,
        monthly_family_rows(details, month),
    )
    save_workbook_safely(workbook, target)


def build_master_workbook(details: list[dict[str, Any]], months: list[str], target: Path) -> None:
    workbook = Workbook()
    add_notes_sheet(
        workbook,
        "三位核心开发人开品基础总表（2025-04至2026-06）",
        [
            ("数据源", "MySQL lingxing_local_product。"),
            ("开发人", "陈杨、宋凤莉、蒋舒；三人作为团队开发方向的核心样本。"),
            ("时间范围", "2025-04-01至2026-06-30，共15个月。"),
            ("计数口径", "按开发人+SKU去重；组合主SKU仍按一个SKU记录。"),
            ("当前阶段", "该文件是后续逐月分析的统一基础数据，不包含FBA、销量、利润结论。"),
            ("标品/非标", "非标不只等于定制：定制、玩具、图像审美、装饰礼品和娱乐情绪商品均属于非标；功能主导商品属于标品。"),
            ("后续分析", "在现有方向、套装、定制和风险分析上，新增季节性、常驻品线、跨月延续和新增品线。2025-04的新增只表示观察窗口内首次出现。"),
        ],
    )

    overview_headers = [
        "创建月份", "开发人", "SKU数", "组合主SKU数", "定制SKU数", "多件套装SKU数",
        "产品族数", "新增品线数", "跨月延续品线数", "标品SKU数", "非标SKU数",
        "标准化待复核SKU数", "风险SKU数",
    ]
    write_table(
        workbook.create_sheet("01_月份开发人数量"),
        "15个月三位开发人基础数量",
        overview_headers,
        overview_rows(details, months),
    )
    write_detail_sheet(workbook, "02_三人全部SKU", "三人15个月全部SKU基础明细", details)

    family_headers = [
        "开发人", "产品族", "主方向", "SKU总数", "首次月份", "最近月份",
        "出现月份数", "出现月份列表", "主要主题", "风险SKU数",
    ]
    write_table(
        workbook.create_sheet("03_产品族跨月登记"),
        "三位开发人产品族跨月登记",
        family_headers,
        family_registry_rows(details),
    )
    for index, developer in enumerate(DEVELOPERS, start=4):
        write_detail_sheet(
            workbook,
            f"0{index}_{developer}全部SKU",
            f"{developer}15个月全部SKU基础明细",
            [row for row in details if row["开发人"] == developer],
        )
    save_workbook_safely(workbook, target)


def build_review_workbook(details: list[dict[str, Any]], target: Path) -> None:
    review_rows = [row for row in details if row.get("商品标准化类型") == "待复核"]
    workbook = Workbook()
    add_notes_sheet(
        workbook,
        "标品/非标待复核SKU审查表",
        [
            ("用途", "只集中展示标题和产品族证据不足、暂时不能可靠判断标品或非标的SKU。"),
            ("复核问题", "人工确认消费者购买它的核心原因是功能/适配，还是图像/审美/娱乐/情绪价值。"),
            ("处理方法", "确认后把稳定规则补回基础分类脚本，再统一重建全部基础表，不直接手改月度Excel。"),
            ("当前数量", f"共{len(review_rows)}个待复核SKU。"),
        ],
    )

    month_developer_counts = Counter((row["创建月份"], row["开发人"]) for row in review_rows)
    summary_rows = [
        [month, developer, month_developer_counts[(month, developer)]]
        for month in sorted({row["创建月份"] for row in review_rows})
        for developer in DEVELOPERS
        if month_developer_counts[(month, developer)]
    ]
    write_table(
        workbook.create_sheet("01_开发人月份汇总"),
        "待复核SKU按开发人与月份汇总",
        ["创建月份", "开发人", "待复核SKU数"],
        summary_rows,
    )

    family_buckets: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in review_rows:
        family_buckets[(row["开发人"], row["产品族"])].append(row)
    family_rows = []
    for (developer, family), items in sorted(
        family_buckets.items(), key=lambda item: (-len(item[1]), item[0][0], item[0][1])
    ):
        family_rows.append([
            developer,
            family,
            len(items),
            "、".join(sorted({row["创建月份"] for row in items})),
            "；".join(str(row["产品名称"]) for row in items[:3]),
        ])
    write_table(
        workbook.create_sheet("02_待复核产品族"),
        "待复核产品族汇总",
        ["开发人", "产品族", "SKU数", "出现月份", "标题示例"],
        family_rows,
    )

    write_detail_sheet(workbook, "03_全部待复核SKU", "全部待复核SKU", review_rows)
    for index, developer in enumerate(DEVELOPERS, start=4):
        write_detail_sheet(
            workbook,
            f"0{index}_{developer}待复核",
            f"{developer}待复核SKU",
            [row for row in review_rows if row["开发人"] == developer],
        )
    save_workbook_safely(workbook, target)


def write_csv(details: list[dict[str, Any]], target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DETAIL_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(details)


def main() -> None:
    parser = argparse.ArgumentParser(description="整理三位核心开发人15个月的开品基础数据")
    parser.add_argument("--start-month", default=START_MONTH, help="开始月份，格式YYYY-MM")
    parser.add_argument("--end-month", default=END_MONTH, help="结束月份，格式YYYY-MM")
    args = parser.parse_args()

    months = month_sequence(args.start_month, args.end_month)
    raw_rows = load_rows(args.start_month, args.end_month)
    unique_rows, duplicate_rows = deduplicate_rows(raw_rows)
    details = prepare_details(unique_rows)

    master_name = f"00_三人开品基础总表_{args.start_month}至{args.end_month}.xlsx"
    csv_name = f"00_三人开品基础明细_{args.start_month}至{args.end_month}.csv"
    review_name = f"00_标品非标待复核审查表_{args.start_month}至{args.end_month}.xlsx"
    build_master_workbook(details, months, OUTPUT_ROOT / master_name)
    write_csv(details, OUTPUT_ROOT / csv_name)
    build_review_workbook(details, OUTPUT_ROOT / review_name)

    for month in months:
        target = OUTPUT_ROOT / month_folder(month) / f"{month_display(month)}_三人开品基础数据.xlsx"
        build_month_workbook(details, month, target)

    payload = {
        "output_root": str(OUTPUT_ROOT),
        "range": [args.start_month, args.end_month],
        "raw_rows": len(raw_rows),
        "unique_rows": len(unique_rows),
        "duplicate_rows": len(duplicate_rows),
        "developers": Counter(row["开发人"] for row in details),
        "months": {
            month: Counter(row["开发人"] for row in details if row["创建月份"] == month)
            for month in months
        },
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2, default=dict))


if __name__ == "__main__":
    main()
