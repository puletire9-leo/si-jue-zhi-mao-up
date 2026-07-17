#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按首次 FBA 可售月生成团队 SKU 总览和开发人报告。

保留早期报告的阅读形式：总览工作簿是“汇总 + 每月明细”，每位开发人
各有一个“汇总 + 每月明细”工作簿。与旧版不同的是：

* 上架批次只认 ``首次FBA可售观察月``，不再使用本地产品创建时间；
* 一行代表一个团队 SKU，不再按 ASIN 重复归并；
* 标签状态是截至 2026-06 的当前状态，不能倒灌为历史月份状态；
* GBP、EUR 分列，不换汇、不相加。
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook as openpyxl_load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = ROOT / "产品数据" / "领星数据api" / "领星25年到26年6月所有数据，以每月数据" / "历史SKU上架基础数据_2025-04至2026-06"
LIFECYCLE_FILE = DATA_ROOT / "03_团队开发SKU生命周期" / "团队SKU_生命周期判定_数据截止2026-06.csv"
FINANCE_DIR = DATA_ROOT / "05_财务利润周度回补"
TIME_OUTPUT_DIR = ROOT / "analysis" / "时间维度"
DEVELOPER_OUTPUT_DIR = ROOT / "analysis" / "开发人维度"
DATA_START = "2025-04"
DATA_CUTOFF = "2026-06"
CURRENCIES = ("GBP", "EUR")
MONTH_FILE = re.compile(r"(20\d{2})年(\d{2})月团队SKU财务利润明细\.csv")
PERFORMANCE_WINDOW = re.compile(r"（(20\d{2}-\d{2})-\d{2}~")

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E2F3")
NEGATIVE_FILL = PatternFill("solid", fgColor="F4CCCC")
POSITIVE_FILL = PatternFill("solid", fgColor="D9EAD3")


def decimal(value: object | None) -> Decimal:
    try:
        return Decimal(str(value or "").strip() or "0")
    except InvalidOperation:
        return Decimal(0)


def month_index(month: str) -> int:
    year, value = map(int, month.split("-"))
    return year * 12 + value - 1


def month_range(start: str, end: str) -> list[str]:
    return [f"{value // 12:04d}-{value % 12 + 1:02d}" for value in range(month_index(start), month_index(end) + 1)]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as source:
        return list(csv.DictReader(source))


def load_product_performance(records: dict[str, "SkuRecord"]) -> None:
    """Add cumulative product-performance sales and the June available-stock snapshot."""
    try:
        from python_calamine import load_workbook as calamine_load_workbook

        def sheet_rows(path: Path):
            return calamine_load_workbook(path).get_sheet_by_index(0).iter_rows()
    except ImportError:
        def sheet_rows(path: Path):
            workbook = openpyxl_load_workbook(path, read_only=True, data_only=True)
            return workbook.worksheets[0].iter_rows(values_only=True)

    fields = ("SKU", "国家", "销量", "销售额", "FBA-可售", "可用库存")
    for path in sorted((DATA_ROOT.parent).glob("*.xlsx")):
        match = PERFORMANCE_WINDOW.search(path.name)
        if not match:
            continue
        month = match.group(1)
        rows = sheet_rows(path)
        headers = [str(value or "") for value in next(rows)]
        positions = {field: headers.index(field) for field in fields}
        for row in rows:
            sku = str(row[positions["SKU"]] or "").strip()
            record = records.get(sku)
            if not record or month < record.fba_month:
                continue
            currency = "GBP" if str(row[positions["国家"]] or "").strip() == "英国" else "EUR"
            if decimal(row[positions["FBA-可售"]]) > 0:
                record.currency_fba_month.setdefault(currency, month)
            if currency not in record.currency_fba_month:
                continue
            record.performance_quantity[currency] += decimal(row[positions["销量"]])
            record.performance_sales[currency] += decimal(row[positions["销售额"]])
            if month == DATA_CUTOFF:
                record.available_stock[currency] += decimal(row[positions["可用库存"]])


@dataclass
class SkuRecord:
    developer: str
    sku: str
    fba_month: str
    product_name: str
    fba_stores: str
    fba_store_sku_count: int
    business_status: str
    latest_label: str
    latest_observed_month: str
    finance_months: set[str] = field(default_factory=set)
    sales_months: set[str] = field(default_factory=set)
    sales_quantity: Decimal = Decimal(0)
    currency_sales: dict[str, Decimal] = field(default_factory=lambda: defaultdict(Decimal))
    currency_profit: dict[str, Decimal] = field(default_factory=lambda: defaultdict(Decimal))
    currency_fba_month: dict[str, str] = field(default_factory=dict)
    performance_quantity: dict[str, Decimal] = field(default_factory=lambda: defaultdict(Decimal))
    performance_sales: dict[str, Decimal] = field(default_factory=lambda: defaultdict(Decimal))
    available_stock: dict[str, Decimal] = field(default_factory=lambda: defaultdict(Decimal))
    currency_payback_month: dict[str, str] = field(default_factory=dict)

    @property
    def sales_retention_rate(self) -> Decimal | None:
        if not self.sales_months:
            return None
        observable = month_index(DATA_CUTOFF) - month_index(min(self.sales_months)) + 1
        return Decimal(len(self.sales_months)) / observable


def load_records() -> tuple[list[SkuRecord], int]:
    lifecycle_rows = read_csv(LIFECYCLE_FILE)
    records: dict[str, SkuRecord] = {}
    for row in lifecycle_rows:
        fba_month = row["首次FBA可售观察月"]
        if not fba_month or not (DATA_START <= fba_month <= DATA_CUTOFF):
            continue
        records[row["SKU"]] = SkuRecord(
            developer=row["开发人"], sku=row["SKU"], fba_month=fba_month, product_name=row["品名"].strip(),
            fba_stores=row["FBA匹配店铺"], fba_store_sku_count=int(decimal(row["FBA匹配店铺SKU数"])),
            business_status=row["领星业务状态"], latest_label=row["最近Listing标签"],
            latest_observed_month=row["最近领星前端观察月"],
        )

    monthly_points: dict[tuple[str, str], dict[str, dict[str, Decimal]]] = defaultdict(dict)
    for path in FINANCE_DIR.glob("*团队SKU财务利润明细.csv"):
        match = MONTH_FILE.fullmatch(path.name)
        if not match:
            continue
        month = f"{match.group(1)}-{match.group(2)}"
        for row in read_csv(path):
            sku, currency = row["SKU"], row["币种"]
            if sku not in records or currency not in CURRENCIES:
                continue
            monthly_points[(sku, currency)][month] = {
                "quantity": decimal(row["销量"]),
                "sales": decimal(row["销售额"]),
                "profit": decimal(row["毛利润"]),
            }

    for (sku, currency), points in monthly_points.items():
        record = records[sku]
        running_profit = Decimal(0)
        for month in month_range(DATA_START, DATA_CUTOFF):
            point = points.get(month)
            if not point:
                continue
            record.finance_months.add(month)
            record.sales_quantity += point["quantity"]
            record.currency_sales[currency] += point["sales"]
            record.currency_profit[currency] += point["profit"]
            running_profit += point["profit"]
            if point["quantity"] > 0:
                record.sales_months.add(month)
            if running_profit > 0 and currency not in record.currency_payback_month:
                record.currency_payback_month[currency] = month
    load_product_performance(records)
    return sorted(records.values(), key=lambda row: (row.fba_month, row.developer, row.sku)), len(lifecycle_rows)


def detail_row(record: SkuRecord, currency: str) -> list[object]:
    return [
        record.sku, record.developer, record.currency_fba_month[currency], record.fba_stores, record.performance_quantity[currency],
        record.performance_sales[currency], record.currency_sales[currency], record.currency_profit[currency],
        record.available_stock[currency], record.business_status,
    ]


def summary_rows(records: list[SkuRecord], months: list[str], currency: str) -> list[list[object]]:
    grouped: dict[str, list[SkuRecord]] = defaultdict(list)
    for record in records:
        grouped[record.currency_fba_month[currency]].append(record)
    rows = []
    for month in months:
        batch = grouped.get(month, [])
        statuses = defaultdict(list)
        for record in batch:
            statuses[record.business_status].append(record)
        active = statuses["上架在售"]
        eliminated = statuses["淘汰"]
        performance_sales = sum((record.performance_sales[currency] for record in batch), Decimal(0))
        settlement_sales = sum((record.currency_sales[currency] for record in batch), Decimal(0))
        settlement_profit = sum((record.currency_profit[currency] for record in batch), Decimal(0))
        active_performance_sales = sum((record.performance_sales[currency] for record in active), Decimal(0))
        active_settlement_profit = sum((record.currency_profit[currency] for record in active), Decimal(0))
        active_settlement_sales = sum((record.currency_sales[currency] for record in active), Decimal(0))
        eliminated_performance_sales = sum((record.performance_sales[currency] for record in eliminated), Decimal(0))
        eliminated_settlement_profit = sum((record.currency_profit[currency] for record in eliminated), Decimal(0))
        eliminated_settlement_sales = sum((record.currency_sales[currency] for record in eliminated), Decimal(0))
        rows.append([
            month, len(batch), sum((record.performance_quantity[currency] for record in batch), Decimal(0)), performance_sales,
            settlement_profit, settlement_sales, settlement_profit / settlement_sales if settlement_sales else None,
            sum((record.available_stock[currency] for record in batch), Decimal(0)), Decimal(len(active)) / len(batch) if batch else None,
            sum((record.performance_quantity[currency] for record in active), Decimal(0)), active_performance_sales, len(active),
            active_settlement_profit, active_settlement_profit / active_settlement_sales if active_settlement_sales else None,
            sum((record.performance_quantity[currency] for record in eliminated), Decimal(0)), eliminated_performance_sales, len(eliminated),
            Decimal(len(eliminated)) / len(batch) if batch else None, eliminated_settlement_profit,
            eliminated_settlement_profit / eliminated_settlement_sales if eliminated_settlement_sales else None,
        ])
    return rows


def style_sheet(sheet, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append([title])
    sheet["A1"].font = Font(bold=True, size=14, color="0B3D47")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.append(headers)
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    for row in rows:
        sheet.append([float(value) if isinstance(value, Decimal) else value for value in row])
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(2, sheet.max_row)}"
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 34
    for column, header in enumerate(headers, start=1):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = min(max(len(header) + 4, 13), 28)
        if header in {"品名", "FBA匹配店铺", "截至截止月最新Listing标签"}:
            sheet.column_dimensions[letter].width = 38
        if any(name in header for name in ("销售额", "结算毛利润", "总结算利润")):
            for cell in sheet[letter][2:]:
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
        if "率" in header:
            for cell in sheet[letter][2:]:
                cell.number_format = '0.00%;[Red]-0.00%'
        if "毛利润" in header or "总结算利润" in header:
            area = f"{letter}3:{letter}{sheet.max_row}"
            sheet.conditional_formatting.add(area, CellIsRule(operator="lessThan", formula=["0"], fill=NEGATIVE_FILL))
            sheet.conditional_formatting.add(area, CellIsRule(operator="greaterThan", formula=["0"], fill=POSITIVE_FILL))


def append_method_sheet(workbook: Workbook, lifecycle_total: int, fba_total: int) -> None:
    sheet = workbook.create_sheet("口径说明")
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "报告口径"
    sheet["A1"].font = Font(bold=True, size=16, color="0B3D47")
    sheet.merge_cells("A1:B1")
    rows = [
        ("上架批次", "按该币种站点首次观察到 FBA 可售的月份归类，不使用本地产品创建时间。"),
        ("分析对象", f"团队本地 SKU 共 {lifecycle_total:,} 个；其中 {fba_total:,} 个在 2025-04 至 2026-06 的月度数据中首次观察到 FBA 可售，纳入本报告。"),
        ("标签状态", "业务状态和 Listing 标签均为截至 2026-06 的最新观察结果，仅用于当前状态，不代表历史每个月状态。"),
        ("币种范围", "本文件仅包含该文件夹对应币种的站点数据，所有金额均为同一币种的数值。"),
        ("销售和结算", "总销售额来自月度产品表现；总结算销售额和总结算利润来自领星财务事实。"),
        ("库存", "总可用库存取 2026-06 月度产品表现的月末快照；不是 15 个月库存累计。"),
        ("存活和淘汰", "存活 = 截至 2026-06 的领星业务状态为“上架在售”；淘汰 = 当前标签含“欧洲精铺2025淘汰”。"),
    ]
    for index, (label, value) in enumerate(rows, start=3):
        sheet.cell(index, 1, label).fill = NOTE_FILL
        sheet.cell(index, 1).font = Font(bold=True)
        sheet.cell(index, 2, value).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 115


SUMMARY_HEADERS = [
    "时间", "SKU总数", "总销售量", "总销售额", "总结算利润", "总结算销售额", "总利润率", "总可用库存", "存活率",
    "存活sku销售量", "存活sku销售额", "存活SKU数", "留存SKU总利润", "留存SKU利润率",
    "淘汰SKU销售量", "淘汰SKU销售额", "淘汰SKU总数", "淘汰率", "淘汰SKU总利润", "淘汰SKU利润率",
]
DETAIL_HEADERS = ["SKU", "开发人", "首次FBA可售月", "FBA匹配店铺", "累计销售量", "总销售额", "总结算销售额", "总结算利润", "总可用库存", "当前状态"]


def write_report(path: Path, title: str, records: list[SkuRecord], lifecycle_total: int, include_developer: bool, currency: str) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    months = sorted({record.currency_fba_month[currency] for record in records})
    style_sheet(workbook.create_sheet("汇总"), title, SUMMARY_HEADERS, summary_rows(records, months, currency))
    for month in months:
        detail = [detail_row(record, currency) for record in records if record.currency_fba_month[currency] == month]
        headers = DETAIL_HEADERS if include_developer else [header for index, header in enumerate(DETAIL_HEADERS) if index != 1]
        if not include_developer:
            detail = [[value for index, value in enumerate(row) if index != 1] for row in detail]
        style_sheet(workbook.create_sheet(month.replace("2025-", "25年").replace("2026-", "26年") + "月"), f"首次 FBA 可售于 {month} 的 SKU 明细", headers, detail)
    append_method_sheet(workbook, lifecycle_total, len(records))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def generate_report_from_excel() -> None:
    """保留旧函数名，数据源已更新为生命周期和财务事实层。"""
    records, lifecycle_total = load_records()
    developers = sorted({record.developer for record in records})
    for currency in CURRENCIES:
        currency_records = [record for record in records if currency in record.currency_fba_month]
        write_report(
            TIME_OUTPUT_DIR / currency / "sku_summary_from_excel.xlsx",
            f"团队 SKU {currency} 站点首次 FBA 可售月份总览（截至 2026-06）",
            currency_records, lifecycle_total, include_developer=True, currency=currency,
        )
        for developer in sorted({record.developer for record in currency_records}):
            developer_records = [record for record in currency_records if record.developer == developer]
            write_report(
                DEVELOPER_OUTPUT_DIR / currency / f"{developer}_sku_report.xlsx",
                f"{developer} {currency} 站点首次 FBA 可售月份总览（截至 2026-06）",
                developer_records, lifecycle_total, include_developer=False, currency=currency,
            )
    # These reports were generated by the previous mixed-currency version of this script.
    # Keep the currency folders as the only current team-report entry points.
    mixed_reports = [TIME_OUTPUT_DIR / "sku_summary_from_excel.xlsx"]
    mixed_reports.extend(DEVELOPER_OUTPUT_DIR / f"{developer}_sku_report.xlsx" for developer in developers)
    for path in mixed_reports:
        try:
            path.unlink(missing_ok=True)
        except PermissionError:
            print(f"mixed_report_in_use={path}")
    print(f"source_lifecycle_skus={lifecycle_total} fba_available_skus={len(records)} developers={len(developers)}")
    print(f"time_reports={TIME_OUTPUT_DIR}")
    print(f"developer_reports={DEVELOPER_OUTPUT_DIR}")


if __name__ == "__main__":
    generate_report_from_excel()
