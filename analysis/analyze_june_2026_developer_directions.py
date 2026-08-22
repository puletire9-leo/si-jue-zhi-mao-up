from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from decimal import Decimal
from pathlib import Path
from statistics import median
from typing import Any

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = (
    ROOT
    / "产品数据"
    / "思考理实团队的开品方向"
    / "第一版"
    / "第一层_开发人开品分析_陈杨与宋凤莉"
)
DEVELOPERS = ("陈杨", "宋凤莉")

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_BOLD = Font(color="FFFFFF", bold=True)


def mysql_config() -> dict[str, Any]:
    scripts_dir = ROOT / "scripts" / "lingxing_daily"
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    from lingxing_base_access import mysql_env

    config = mysql_env()
    if not os.environ.get("MYSQL_PORT_EXTERNAL"):
        ports = json.loads((ROOT / "config" / "ports.json").read_text(encoding="utf-8"))
        config["port"] = int(ports["mysql"])
    return config


def month_range(month: str) -> tuple[str, str]:
    year, value = map(int, month.split("-"))
    next_index = year * 12 + value
    return f"{year:04d}-{value:02d}-01", f"{next_index // 12:04d}-{next_index % 12 + 1:02d}-01"


def month_names(month: str) -> tuple[str, str, str]:
    year, value = map(int, month.split("-"))
    return f"{year}年{value:02d}月", f"{str(year)[2:]}年{value}月", f"{year}年{value:02d}月"


def output_path(month: str) -> Path:
    file_month, folder_month, _ = month_names(month)
    return OUTPUT_ROOT / folder_month / f"{file_month}_陈杨与宋凤莉_全部开品方向分析.xlsx"


def load_rows(month: str) -> list[dict[str, Any]]:
    start_date, end_date = month_range(month)
    connection = pymysql.connect(**mysql_config())
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT sku, product_developer, product_name, cg_price, is_combo,
                       status_text, open_status, lx_create_time
                FROM lingxing_local_product
                WHERE product_developer IN ('陈杨', '宋凤莉')
                  AND lx_create_time >= %s
                  AND lx_create_time < %s
                  AND sku IS NOT NULL AND sku <> ''
                ORDER BY product_developer, lx_create_time, sku
                """
            , (start_date, end_date))
            columns = [item[0] for item in cursor.description]
            return [dict(zip(columns, values)) for values in cursor.fetchall()]
    finally:
        connection.close()


DIRECTION_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("健康康复与身体护理", ("瑜伽", "太极", "锻炼", "康复", "训练器", "按摩", "淋巴", "去茧", "吸奶器", "护理", "脚锉", "磨脚器")),
    ("汽车骑行户外与功能配件", ("汽车", "自行车", "头盔", "钓鱼", "飞镖", "战术", "烧烤", "燃气", "气罐", "水管", "车标", "后视镜")),
    ("宠物用品", ("宠物", "猫狗", "牵引绳", "舔食", "喂食提醒", "饮水壶", "饮水器", "猫玩具")),
    ("派对贺卡与礼品", ("派对", "气球", "蛋糕", "邀请卡", "贺卡", "礼物", "口袋拥抱", "婚礼", "生日", "退休礼物", "纪念币")),
    ("DIY手工与文具材料", ("diy", "DIY", "手工", "刺绣", "珠子", "颜料", "水彩", "绘画套装", "钻石画", "数字油画", "戳戳画", "涂色", "涂鸦本", "和弦卡", "磨钻", "橡皮", "拼豆", "材料包", "贴纸", "练习册", "圆珠笔", "笔袋", "书签", "绘图笔")),
    ("家纺布艺与日用", ("沙滩巾", "浴巾", "枕套", "毛毯", "毯子", "床单", "抱枕套")),
    ("定制平面载体与轻装饰", ("亚克力", "铁皮", "海报", "挂画", "帆布袋", "化妆包", "束口袋", "午餐包", "鼠标垫", "木质摆件", "木质挂牌", "木制钥匙架", "木质钥匙架", "收纳筐", "厨房毛巾")),
    ("玩具解压与3D打印", ("玩具", "3D打印", "3d打印", "人偶", "积木", "感官石", "拼图", "叠叠乐", "跳跳球", "魔珠", "解压", "不倒翁", "小动物公仔", "造景小摆件")),
    ("宗教纪念与精神礼物", ("耶稣", "圣母", "天主教", "修女", "十字架", "天使", "祈祷")),
    ("服饰饰品与随身配件", ("帽子", "鸭舌帽", "棒球帽", "太阳帽", "假发网帽", "服饰", "衣服", "发带", "发夹", "袜套", "耳套", "手扇", "指环", "裤链", "纽扣", "鞋扣", "首饰", "珠宝", "手链", "项链", "耳环", "发簪", "零钱包", "手机气囊支架")),
    ("厨房日用与收纳", ("厨房", "磨蒜", "冰格", "吸管", "油脂", "沙拉", "水杯", "杯垫", "毛巾", "元宝巾", "围裙", "收纳", "蛋杯", "漏斗", "杯架", "洗碗", "置物架", "沥水", "隔热垫", "空锡盒")),
    ("花园家居与立体装饰", ("花园", "鸟浴", "植物", "喂水器", "地插", "挂钩", "墙面装饰", "装饰摆件", "铁艺", "风车", "园林", "土筛", "花压机", "铁锹", "播种铲", "打草绳", "鸟舍", "花盆", "种植挂袋", "水位控制阀", "浮球阀", "花艺摆件")),
    ("电子与功能设备", ("带电", "电动", "温度计", "电瓶", "3D打印机配件", "3d打印机配件", "喷嘴", "太阳能", "发光")),
    ("图像主题载体与装饰周边", ("钥匙扣", "马克杯", "挂牌", "挂件", "挂饰", "徽章", "立牌", "海报", "油画", "风铃", "硬币", "墙贴", "车贴", "冰箱贴", "背景纸画", "背景贴", "玻璃挂牌")),
    ("玩具手办与解压收藏", ("公仔", "捏捏", "玩偶", "人偶", "手办", "积木", "迷宫", "变形车", "解压", "观影器", "扭动", "龙蛋", "美人鱼", "卡皮巴拉", "摆件", "雕塑", "雕像", "树脂", "飞轮", "指尖陀螺", "水枪", "水袖", "泡泡机", "仿真昆虫", "娃衣", "镖靶", "山海经", "蜘蛛侠", "桌游", "骰子", "彩泥", "麻将", "扑克", "棋", "游戏")),
    ("包袋收纳与出行用品", ("背包", "腰包", "钱包", "旅行包", "护照本", "卡册", "收纳包", "单肩方包", "单肩包", "方包", "竿袋", "鱼竿套")),
    ("阅读文具与办公用品", ("笔记本", "记事本", "台历", "广告笔", "存钱本", "剪贴", "手账", "集卡", "图表", "食物清单", "早教", "学习字母", "学习闪卡", "阅读追踪器", "阅读计数器", "计数摆件")),
    ("美妆美甲与个护工具", ("美甲", "点花笔", "印花笔", "蜜粉刷", "化妆刷", "头发蓬松", "指甲钳", "洗澡刷", "美容", "美妆")),
    ("汽车维修与适配配件", ("奥迪", "宝马", "奔驰", "丰田", "遥控", "汽车钥匙", "汽车配件", "车载", "拖车", "联轴器", "油箱盖", "刹车胶")),
    ("功能工具与维修配件", ("工具", "修理", "替换", "更换", "密封", "兼容", "接头", "适用", "适配", "固定", "清洁", "过滤", "搅拌机配件", "支架", "展示架", "防水鞋套", "保护套", "锁杆", "转轴", "刹车", "除刺器", "起草器", "开孔器", "穿针器", "吸力棒", "吸铁棒", "压脚", "穿线器", "吊钩", "油箱盖", "排水阀", "控制阀", "遮阳板", "桌孔盖", "连接扣", "单车铃", "铃铛", "冰袋", "漏斗", "螺丝", "螺栓", "垫圈", "刮片", "批头", "夹具", "修补", "防撞", "防尘罩", "压缩", "烟盘", "带盖")),
    ("户外运动与钓鱼用品", ("钓鱼", "鱼竿", "鱼钩", "假饵", "串钩", "渔", "高尔夫", "路亚", "游泳", "划水掌", "轮滑", "跑步机", "健身球")),
    ("母婴儿童功能用品", ("婴儿车", "儿童安全座椅", "宝宝", "婴儿", "奶瓶", "儿童防护")),
]

FAMILY_KEYWORDS = (
    "钻石画", "沙滩巾", "数字油画", "戳戳画", "马克杯", "汽车杯套件", "汽车杯",
    "围裙", "枕套", "防水透明收纳包", "塑料杯", "超细纤维毛巾", "鼠标垫",
    "首饰盒", "鸭舌帽", "背包", "零钱包", "单肩包", "餐包", "钥匙扣", "贴纸",
    "冰格", "积木", "挂饰", "记事本", "杯套", "车枕套", "发簪", "项圈",
    "亚克力拼图挂牌", "亚克力吊坠挂牌", "亚克力挂牌", "亚克力立牌", "亚克力摆件",
    "亚克力地插", "铁皮挂牌", "铁皮画", "海报画芯", "帆布袋", "化妆包", "束口袋",
    "侧兜午餐包", "护腕鼠标垫", "木质摆件", "木质挂牌", "木质钥匙架", "金属地插",
    "金属铁艺装饰画", "金属铁艺花艺摆件", "纪念币", "书本收纳袋", "阅读追踪器",
    "阅读计数器", "杯垫", "感官石", "3D打印人偶", "3d打印幸运人偶", "3D打印",
    "蛋糕装饰", "邀请卡", "贺卡", "手链", "鸭舌帽", "棒球帽", "厨房毛巾",
    "水彩画练习册", "DIY珠子套装", "diy珠子套装", "数字油画", "收纳筐",
)

THEME_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("高地牛", ("高地牛", "牛牛", "小牛", "牛款", "呆呆牛")),
    ("鸟类/知更鸟", ("知更鸟", "罗宾", "蜂鸟", "小鸟", "蓝鸟", "红鸟", "海鸥", "鸟")),
    ("猫狗宠物", ("猫", "狗", "腊肠", "小狗", "狗爪")),
    ("龙与幻想", ("龙", "女巫", "仙女", "精灵", "UFO", "独角兽")),
    ("天使宗教", ("天使", "耶稣", "圣母", "十字架", "天主教", "修女")),
    ("花卉蝴蝶", ("花", "蝴蝶", "向日葵", "郁金香", "雏菊", "水仙")),
    ("城市旅行", ("伦敦", "巴黎", "意大利", "土耳其", "苏格兰", "牙买加", "旅行地图")),
    ("书籍阅读", ("书", "阅读", "读者", "第四翼", "4th wing")),
    ("海洋水生", ("海洋", "鱼", "鲸", "章鱼", "海浪", "水母", "海星", "蝾螈")),
    ("老年健康", ("老年", "老人", "椅子瑜伽", "太极", "康复")),
    ("幽默成人", ("fxxk", "FUCK", "fuck", "WTF", "dick", "傻", "老家伙", "OLD FUCKERS")),
]

IP_KEYWORDS = ("第四翼", "4th wing", "APEROL", "Aperol", "大力水手", "不良人", "一梦江湖", "K-pop", "kpop")
SAFETY_KEYWORDS = ("燃气", "气罐", "电瓶断电", "婴儿车", "吸奶器", "电动喷雾", "泳池电动", "水管接头", "3D打印机配件", "3d打印机配件")
ELECTRIC_KEYWORDS = ("带电", "电动", "USB", "usb", "太阳能", "发光", "温度计")
LIQUID_KEYWORDS = ("带液体", "颜料", "喷雾")
CHEMICAL_EFFECT_KEYWORDS = ("牙垢剂", "护油笔", "木器护理", "清洁剂", "除锈", "杀虫", "除草", "驱鼠", "修复剂")
CONTENT_KEYWORDS = ("fxxk", "FUCK", "fuck", "WTF", "dick", "OLD FUCKERS")
COMPATIBILITY_KEYWORDS = ("汽车", "水管", "3D打印机", "3d打印机", "喷嘴", "配件", "吸奶器")


def includes_any(text: str, keywords: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(keyword.lower() in lowered for keyword in keywords)


def classify_direction_with_family(name: str, family: str) -> str:
    for direction, keywords in DIRECTION_RULES:
        if includes_any(name, keywords):
            return direction
    if family in {"挂牌", "钥匙扣", "马克杯", "挂饰", "徽章", "内框油画", "玻璃挂牌"}:
        return "图像主题载体与装饰周边"
    if family in {"背包", "钱包", "旅行包", "收纳包", "卡册", "竿袋"}:
        return "包袋收纳与出行用品"
    return "其他小商品与待细分方向"


def classify_direction(name: str, family: str = "") -> str:
    return classify_direction_with_family(name, family)


NONSTANDARD_HARD_KEYWORDS = (
    "定制", "玩具", "公仔", "捏捏", "玩偶", "人偶", "手办", "积木", "拼图", "迷宫",
    "变形车", "解压", "娃娃", "霸王龙", "龙蛋", "观影器", "扭动", "麻将", "扑克", "棋",
    "游戏", "桌游", "骰子", "彩泥", "泡泡机", "仿真昆虫", "冰箱贴", "背景纸画", "背景贴", "绘画套装", "摆件", "雕塑", "雕像", "树脂", "美人鱼", "刺绣", "无功能", "挂牌", "挂件", "挂饰", "飞轮", "指尖陀螺", "水枪", "水袖", "娃衣", "镖靶", "山海经", "蜘蛛侠", "烟花", "亚克力", "菜谱", "图表", "食物清单", "拉花", "涂鸦本", "和弦卡", "钻石画", "数字油画", "戳戳画", "贺卡", "邀请卡", "纪念币", "徽章", "钥匙扣",
    "项链", "手链", "耳环", "发饰", "派对", "礼物", "礼盒", "装饰", "挂画", "油画",
)

FUNCTION_HARD_KEYWORDS = (
    "配件", "修理", "替换", "更换", "密封", "兼容", "接头", "适用", "适配", "固定", "清洁", "过滤", "工具",
    "支架", "展示架", "保护套", "防水", "锁杆", "转轴", "刹车", "除刺器", "起草器", "开孔器", "穿针器", "吸力棒", "吸铁棒", "脚锉", "磨脚器", "指甲钳", "洗澡刷", "沥水", "隔热垫", "压脚", "穿线器", "吊钩", "打草绳", "排水阀", "控制阀", "浮球阀", "水龙头", "延伸器", "扎带", "毛刷", "超声波", "驱狗", "驱鸟", "鸟浴", "游泳帽", "游泳", "泵", "油箱盖", "钓", "鱼钩", "假饵", "串钩", "园林", "土筛", "花压机", "铁锹", "播种铲", "鸟舍", "花盆", "种植挂袋", "早教", "学习字母", "学习闪卡", "冰袋", "漏斗", "螺丝", "螺栓", "垫圈", "批头", "夹具", "修补", "防撞", "防尘罩", "压缩", "烟盘", "车载", "带盖", "腰包", "防滑", "耳套", "鼻贴", "空锡盒",
    "遮阳板", "桌孔盖", "连接扣", "单车铃", "铃铛", "划水掌", "收纳", "整理", "防护",
)

IMAGE_CARRIER_FAMILIES = {
    "亚克力挂牌", "亚克力拼图挂牌", "亚克力吊坠挂牌", "亚克力立牌", "亚克力摆件",
    "亚克力地插", "铁皮挂牌", "铁皮画", "海报", "海报画芯", "帆布袋", "化妆包",
    "束口袋", "马克杯", "沙滩巾", "钥匙扣", "挂牌", "挂饰", "徽章", "杯垫", "枕套", "杯架", "蛋杯",
}

STRONG_VISUAL_KEYWORDS = (
    "图案", "印花", "主题", "卡通", "复古", "可爱", "高地牛", "蝴蝶", "圣诞", "万圣",
    "复活节", "彩虹", "爱心", "花朵", "动物",
)

CARRIER_THEME_KEYWORDS = STRONG_VISUAL_KEYWORDS + ("花", "龙", "鸟", "猫", "狗")

NONSTANDARD_DIRECTIONS = {
    "定制平面载体与轻装饰", "图像主题载体与装饰周边", "玩具解压与3D打印",
    "玩具手办与解压收藏", "派对贺卡与礼品", "DIY手工与文具材料",
    "宗教纪念与精神礼物", "服饰饰品与随身配件",
    "阅读文具与办公用品",
}

STANDARD_DIRECTIONS = {
    "健康康复与身体护理", "汽车骑行户外与功能配件", "宠物用品", "家纺布艺与日用",
    "厨房日用与收纳", "电子与功能设备", "功能工具与维修配件", "户外运动与钓鱼用品",
    "美妆美甲与个护工具", "母婴儿童功能用品", "包袋收纳与出行用品", "汽车维修与适配配件",
    "花园家居与立体装饰",
}


def _first_keyword(text: str, keywords: tuple[str, ...]) -> str:
    lowered = text.lower()
    for keyword in keywords:
        if keyword.lower() in lowered:
            return keyword
    return ""


def classify_standardization(name: str, direction: str, family: str) -> dict[str, str]:
    custom_keyword = _first_keyword(name, ("定制",))
    if custom_keyword:
        return {
            "商品标准化类型": "非标",
            "价值主导类型": "定制主导",
            "标准化判定依据": "标题明确包含“定制”",
            "标准化判定置信度": "高",
            "是否需要人工复核": "否",
        }

    nonstandard_keyword = _first_keyword(name, NONSTANDARD_HARD_KEYWORDS)
    if nonstandard_keyword:
        return {
            "商品标准化类型": "非标",
            "价值主导类型": "娱乐情绪主导" if nonstandard_keyword in {"玩具", "公仔", "捏捏", "玩偶", "人偶", "手办", "积木", "解压", "游戏"} else "图像审美主导",
            "标准化判定依据": f"标题包含非标特征词“{nonstandard_keyword}”",
            "标准化判定置信度": "高",
            "是否需要人工复核": "否",
        }

    functional_keyword = _first_keyword(name, FUNCTION_HARD_KEYWORDS)
    if functional_keyword:
        return {
            "商品标准化类型": "标品",
            "价值主导类型": "功能主导",
            "标准化判定依据": f"标题包含功能/适配特征词“{functional_keyword}”",
            "标准化判定置信度": "高",
            "是否需要人工复核": "否",
        }

    strong_visual_keyword = _first_keyword(name, STRONG_VISUAL_KEYWORDS)
    carrier_theme_keyword = _first_keyword(name, CARRIER_THEME_KEYWORDS)
    image_family = family in IMAGE_CARRIER_FAMILIES or "杯架" in family
    if image_family and carrier_theme_keyword:
        return {
            "商品标准化类型": "非标",
            "价值主导类型": "图像审美主导",
            "标准化判定依据": f"图像载体产品包含视觉主题词“{carrier_theme_keyword}”",
            "标准化判定置信度": "中",
            "是否需要人工复核": "否",
        }

    if direction in NONSTANDARD_DIRECTIONS:
        return {
            "商品标准化类型": "非标",
            "价值主导类型": "图像审美主导",
            "标准化判定依据": "主方向以图像、装饰、礼品、文创或审美价值为主",
            "标准化判定置信度": "中",
            "是否需要人工复核": "否",
        }

    if direction in STANDARD_DIRECTIONS:
        return {
            "商品标准化类型": "标品",
            "价值主导类型": "功能主导",
            "标准化判定依据": "主方向以功能、适配、收纳或工具属性为主",
            "标准化判定置信度": "中",
            "是否需要人工复核": "否",
        }

    if strong_visual_keyword:
        return {
            "商品标准化类型": "非标",
            "价值主导类型": "图像审美主导",
            "标准化判定依据": f"标题包含明确视觉/主题词“{strong_visual_keyword}”且未发现更强功能证据",
            "标准化判定置信度": "中",
            "是否需要人工复核": "否",
        }

    return {
        "商品标准化类型": "待复核",
        "价值主导类型": "待判断",
        "标准化判定依据": "标题和产品族暂不足以判断功能主导还是视觉/创意主导",
        "标准化判定置信度": "低",
        "是否需要人工复核": "是",
    }


def classify_family(name: str) -> str:
    for keyword in FAMILY_KEYWORDS:
        if keyword.lower() in name.lower():
            return keyword.replace("3d", "3D")
    text = re.sub(r"^【[^】]+】", "", name).strip()
    while text.startswith("【"):
        text = re.sub(r"^【[^】]+】", "", text).strip()
    text = re.split(r"【|（|\(|-|—|–|：|:", text, maxsplit=1)[0].strip("- ：:")
    return text[:24] or "待人工归族"


def classify_theme(name: str) -> str:
    themes = [theme for theme, keywords in THEME_RULES if includes_any(name, keywords)]
    return "、".join(themes) if themes else "通用/无明显主题"


def risk_flags(name: str) -> list[str]:
    flags: list[str] = []
    if includes_any(name, IP_KEYWORDS):
        flags.append("明确IP/品牌词复核")
    if includes_any(name, SAFETY_KEYWORDS):
        flags.append("安全/适配责任复核")
    if includes_any(name, ELECTRIC_KEYWORDS):
        flags.append("带电履约复核")
    if includes_any(name, LIQUID_KEYWORDS):
        flags.append("液体/颜料履约复核")
    if includes_any(name, CHEMICAL_EFFECT_KEYWORDS):
        flags.append("化学/功效复核")
    if includes_any(name, CONTENT_KEYWORDS):
        flags.append("成人文案/平台内容复核")
    if includes_any(name, COMPATIBILITY_KEYWORDS):
        flags.append("规格适配复核")
    return flags


def first_layer_judgement(name: str, direction: str, flags: list[str]) -> tuple[str, str]:
    if "明确IP/品牌词复核" in flags or "安全/适配责任复核" in flags or "化学/功效复核" in flags:
        return "高风险复核", "方向可能有需求，但在进入运营前必须先解决IP、安全或规格适配问题。"
    if flags:
        return "可做但需复核", "商品形态可以测试，但带电、液体、成人文案或兼容性会增加履约与平台风险。"
    if direction in {
        "定制平面载体与轻装饰", "派对贺卡与礼品", "DIY手工与文具材料",
        "玩具解压与3D打印", "花园家居与立体装饰",
    }:
        return "符合团队主方向", "轻小、可定制、可做套装或主题差异化，符合当前团队能力。"
    return "普通测试方向", "没有明显硬风险，但需要结合第二层上架率和第三层利润再决定是否扩大。"


def analyze(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        name = str(row.get("product_name") or "").strip()
        family = classify_family(name)
        direction = classify_direction(name, family)
        flags = risk_flags(name)
        standardization = classify_standardization(name, direction, family)
        if direction == "其他小商品与待细分方向":
            direction = {
                "标品": "功能标品与杂项配件",
                "非标": "创意非标与收藏杂项",
                "待复核": "待人工方向复核",
            }[standardization["商品标准化类型"]]
        judgement, note = first_layer_judgement(name, direction, flags)
        output.append({
            "开发人": str(row["product_developer"]),
            "SKU": str(row["sku"]),
            "创建时间": row["lx_create_time"],
            "产品名称": name,
            "采购价(CNY)": float(row.get("cg_price") or 0),
            "是否组合主SKU": "是" if row.get("is_combo") else "否",
            "是否定制": "是" if "定制" in name else "否",
            "是否多件/套装": "是" if re.search(r"\d+\s*(pcs|sets|set|cps|cps)|套装|\d+件套|\d+个装", name, re.I) else "否",
            "主方向": direction,
            "产品族": family,
            "主题": classify_theme(name),
            "风险标记": "、".join(flags) if flags else "无明显硬风险",
            "第一层判断": judgement,
            "判断说明": note,
            **standardization,
        })
    return output


def direction_summary(rows: list[dict[str, Any]], developer: str) -> list[list[Any]]:
    selected = [row for row in rows if row["开发人"] == developer]
    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in selected:
        buckets[row["主方向"]].append(row)
    result = []
    for direction, items in sorted(buckets.items(), key=lambda item: (-len(item[1]), item[0])):
        result.append([
            direction,
            len(items),
            len(items) / len(selected) if selected else 0,
            sum(row["是否定制"] == "是" for row in items),
            sum(row["是否多件/套装"] == "是" for row in items),
            sum(row["风险标记"] != "无明显硬风险" for row in items),
            len({row["产品族"] for row in items}),
            "、".join(name for name, _ in Counter(row["产品族"] for row in items).most_common(5)),
        ])
    return result


def developer_overview(rows: list[dict[str, Any]], developer: str) -> dict[str, Any]:
    selected = [row for row in rows if row["开发人"] == developer]
    prices = [Decimal(str(row["采购价(CNY)"])) for row in selected]
    judgements = Counter(row["第一层判断"] for row in selected)
    exact_names = Counter(row["产品名称"] for row in selected)
    custom_count = sum(row["是否定制"] == "是" for row in selected)
    bundle_count = sum(row["是否多件/套装"] == "是" for row in selected)
    risk_count = sum(row["风险标记"] != "无明显硬风险" for row in selected)
    return {
        "开发人": developer,
        "SKU数": len(selected),
        "产品族数": len({row["产品族"] for row in selected}),
        "定制SKU数": custom_count,
        "定制占比": custom_count / len(selected) if selected else 0,
        "多件套装SKU数": bundle_count,
        "多件套装占比": bundle_count / len(selected) if selected else 0,
        "组合主SKU数": sum(row["是否组合主SKU"] == "是" for row in selected),
        "重复产品名SKU数": sum(count for count in exact_names.values() if count > 1),
        "采购价中位数": float(median(prices)) if prices else 0,
        "采购价均值": float(sum(prices) / len(prices)) if prices else 0,
        "有风险标记SKU数": risk_count,
        "风险占比": risk_count / len(selected) if selected else 0,
        "符合主方向": judgements["符合团队主方向"],
        "普通测试": judgements["普通测试方向"],
        "需复核": judgements["可做但需复核"],
        "高风险": judgements["高风险复核"],
    }


def write_table(sheet, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"].font = Font(bold=True, size=15, color="0B3D47")
    sheet.append(headers)
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, start=1):
        width = 15
        if header in {"产品名称", "判断说明"}:
            width = 58
        elif header in {"风险标记", "产品族", "主题", "主方向"}:
            width = 26
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook(rows: list[dict[str, Any]], month: str, target: Path) -> None:
    display_month, _, _ = month_names(month)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    notes = workbook.active
    notes.title = "00_阅读说明"
    notes.append([f"{display_month}陈杨与宋凤莉开品方向分析"])
    notes["A1"].font = Font(bold=True, size=17, color="0B3D47")
    notes.append(["项目", "说明"])
    for cell in notes[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
    explanations = [
        ("分析范围", f"仅分析领星本地产品中创建时间位于{month}的陈杨、宋凤莉SKU。"),
        ("分析层级", "这是第一层开发开品方向分析，不使用销量、利润或淘汰标签倒推判断。"),
        ("数据库类目", "两人的category_name均为空；主方向与产品族根据完整产品名称按固定规则补充。"),
        ("第一层判断", "用于决定进入运营前的优先级和复核要求，不代表最终经营成败。"),
        ("金额", "采购价来自领星本地产品cg_price，币种为CNY；不是亚马逊售价。"),
        ("后续验证", "第二层验证FBA上架率，第三层验证销量、利润、留存和回本。"),
    ]
    for item in explanations:
        notes.append(item)
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 100
    notes.freeze_panes = "A3"

    overview_headers = list(developer_overview(rows, DEVELOPERS[0]).keys())
    overview_rows = [[developer_overview(rows, developer)[header] for header in overview_headers] for developer in DEVELOPERS]
    write_table(workbook.create_sheet("01_两人对比"), f"两人{display_month}开品总览", overview_headers, overview_rows)
    overview_sheet = workbook["01_两人对比"]
    for index, header in enumerate(overview_headers, start=1):
        if "占比" in header:
            for cell in overview_sheet[get_column_letter(index)][2:]:
                cell.number_format = "0.00%"

    summary_headers = ["主方向", "SKU数", "占比", "定制SKU数", "多件套装SKU数", "风险SKU数", "产品族数", "主要产品族"]
    for developer in DEVELOPERS:
        write_table(
            workbook.create_sheet(f"02_{developer}方向汇总"),
            f"{developer} {display_month}开品方向汇总",
            summary_headers,
            direction_summary(rows, developer),
        )
        summary_sheet = workbook[f"02_{developer}方向汇总"]
        for cell in summary_sheet["C"][2:]:
            cell.number_format = "0.00%"

    detail_headers = [
        "开发人", "SKU", "创建时间", "产品名称", "采购价(CNY)", "是否组合主SKU",
        "是否定制", "是否多件/套装", "主方向", "产品族", "主题", "风险标记",
        "第一层判断", "判断说明",
    ]
    for developer in DEVELOPERS:
        selected = [row for row in rows if row["开发人"] == developer]
        detail_rows = [[row[header] for header in detail_headers] for row in selected]
        write_table(
            workbook.create_sheet(f"03_{developer}全部SKU"),
            f"{developer} {display_month}全部开品逐条分析",
            detail_headers,
            detail_rows,
        )
        sheet = workbook[f"03_{developer}全部SKU"]
        for cell in sheet["E"][2:]:
            cell.number_format = "0.00"

    workbook.save(target)


def main() -> None:
    parser = argparse.ArgumentParser(description="按月份分析陈杨和宋凤莉的全部开品方向")
    parser.add_argument("--month", default="2026-06", help="月份，格式 YYYY-MM")
    args = parser.parse_args()
    target = output_path(args.month)
    rows = analyze(load_rows(args.month))
    build_workbook(rows, args.month, target)
    payload = {
        "month": args.month,
        "output": str(target),
        "developers": {developer: developer_overview(rows, developer) for developer in DEVELOPERS},
        "directions": {
            developer: [
                {"direction": row[0], "count": row[1], "rate": row[2], "risk": row[5], "families": row[7]}
                for row in direction_summary(rows, developer)
            ]
            for developer in DEVELOPERS
        },
        "top_families": {
            developer: Counter(row["产品族"] for row in rows if row["开发人"] == developer).most_common(20)
            for developer in DEVELOPERS
        },
        "risk_flags": {
            developer: Counter(
                flag
                for row in rows if row["开发人"] == developer
                for flag in str(row["风险标记"]).split("、")
                if flag != "无明显硬风险"
            )
            for developer in DEVELOPERS
        },
        "top_themes": {
            developer: Counter(
                theme
                for row in rows if row["开发人"] == developer
                for theme in str(row["主题"]).split("、")
                if theme != "通用/无明显主题"
            ).most_common(20)
            for developer in DEVELOPERS
        },
        "price_bands": {
            developer: {
                "0-5元": sum(0 <= row["采购价(CNY)"] <= 5 for row in rows if row["开发人"] == developer),
                "5-10元": sum(5 < row["采购价(CNY)"] <= 10 for row in rows if row["开发人"] == developer),
                "10-20元": sum(10 < row["采购价(CNY)"] <= 20 for row in rows if row["开发人"] == developer),
                "20元以上": sum(row["采购价(CNY)"] > 20 for row in rows if row["开发人"] == developer),
            }
            for developer in DEVELOPERS
        },
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
