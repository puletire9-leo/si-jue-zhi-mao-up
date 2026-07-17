from __future__ import annotations

import argparse
import csv
import re
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SOURCE_CSV = (
    ROOT
    / "产品数据"
    / "思考理实团队的开品方向"
    / "第一版"
    / "第一层_核心开发人开品分析_陈杨宋凤莉蒋舒"
    / "00_三人开品基础明细_2025-04至2026-06.csv"
)
OUTPUT_ROOT = (
    ROOT
    / "产品数据"
    / "思考理实团队的开品方向"
    / "第一版"
    / "第一层开发skll提取"
    / "05_逐标题分析基础"
)
DEEP_REVIEW_ROOT = (
    ROOT
    / "产品数据"
    / "思考理实团队的开品方向"
    / "第一版"
    / "第一层开发skll提取"
    / "07_逐月深审"
)
OUTPUT_PATH = OUTPUT_ROOT / "25年7月" / "2025年07月_三位核心开发人_逐标题开发分析.xlsx"
TRIAL_MONTH = "2025-07"
DEVELOPERS = ("陈杨", "蒋舒", "宋凤莉")
REQUIRED_SOURCE_COLUMNS = (
    "创建月份",
    "开发人",
    "SKU",
    "创建时间",
    "产品名称",
    "是否组合主SKU",
)
EXPLICIT_PARENT_COLUMNS = ("关联组合主SKU", "组合主SKU", "父SKU")
TEST_STATUS_VALUES = {"测试", "测试数据", "test", "demo"}
VAGUE_PRODUCT_BODIES = {
    "其他商品",
    "小商品",
    "杂项商品",
    "待确认产品本体",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_BOLD = Font(color="FFFFFF", bold=True)

MASTER_INDEX_HEADERS = [
    "创建月份",
    "开发人",
    "SKU",
    "原始完整标题",
    "产品本体",
    "产品族",
    "核心功能",
    "使用场景",
    "适用人群",
    "适配对象",
    "主题元素",
    "材质与结构",
    "建议售卖时间",
    "商品标准化类型",
    "标准化判断依据",
    "风险提示",
    "方向Skill原子",
    "开发方法原子",
    "组合打法原子",
    "产品方案ID",
    "产品方案键",
    "关联组合主SKU",
    "关联完整产品标题",
    "标题分析依据",
    "信息补充依据",
    "分析置信度",
    "是否待复核",
    "待复核原因",
    "月份文件夹",
    "XLSX文件",
    "CSV文件",
    "搜索关键词",
]


SCENE_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("万圣节", ("万圣节", "万圣", "halloween")),
    ("圣诞节", ("圣诞", "christmas")),
    ("婚礼", ("婚礼", "婚纱", "新娘")),
    ("生日派对", ("生日", "派对", "蛋糕顶饰", "蛋糕插牌")),
    ("汽车使用", ("汽车", "挡风玻璃", "车载", "车门", "车窗")),
    ("自行车骑行", ("自行车", "单车", "骑行")),
    ("维修替换", ("维修", "修理", "替换", "更换", "密封环", "卡扣", "接头")),
    ("DIY手工", ("DIY", "diy", "手工", "钻石画", "刺绣", "绘画")),
    ("户外花园", ("花园", "户外", "园林", "种植", "鸟浴")),
    ("宠物使用", ("宠物", "猫咪", "狗狗", "猫狗", "犬用", "牵引", "喂食", "宠物美容", "狗面具", "宠物套装")),
    ("儿童使用", ("儿童", "宝宝", "婴儿", "早教")),
    ("家居日用", ("家居", "厨房", "浴室", "桌面", "收纳")),
)

AUDIENCE_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("儿童", ("儿童", "男童", "女童", "小孩")),
    ("婴幼儿", ("婴儿", "宝宝", "幼儿")),
    ("女性", ("女士", "女性", "女款", "新娘")),
    ("男性", ("男士", "男性", "男款")),
    ("老人", ("老人", "老年", "长者")),
    ("宠物及宠物主人", ("宠物", "猫咪", "狗狗", "猫狗", "犬用", "牵引", "喂食", "狗面具", "宠物套装")),
    ("车主", ("汽车", "车载", "车门", "车窗", "挡风玻璃")),
    ("骑行人群", ("自行车", "单车", "骑行")),
    ("手工爱好者", ("DIY", "diy", "钻石画", "刺绣", "手工")),
)

FUNCTION_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("密封防漏", ("密封", "防漏", "垫圈", "密封环")),
    ("连接适配", ("适配", "适用", "兼容", "接头", "卡扣", "连接", "转换")),
    ("维修替换", ("维修", "修理", "替换", "更换", "泵", "压脚", "火花塞", "钻头", "锯片", "扳手", "套筒")),
    ("清洁维护", ("清洁", "清洗", "除尘", "毛刷", "刮片")),
    ("收纳整理", ("收纳", "整理", "储存", "置物", "挂钩")),
    ("安全防护", ("保护", "防护", "防撞", "防水", "防滑", "遮阳")),
    ("穿戴造型", ("服装", "长袍", "头饰", "发夹", "发箍", "项链", "手链", "帽子", "生日帽", "女巫帽")),
    ("装饰布置", ("装饰", "吊饰", "摆件", "相框", "挂牌", "挂饰", "蛋糕顶饰")),
    ("娱乐解压", ("玩具", "解压", "游戏", "公仔", "跳舞", "感官石")),
    ("手工创作", ("DIY", "diy", "钻石画", "刺绣", "绘画", "贴纸")),
    ("园艺种植", ("种植", "花园", "播种", "园林", "花盆")),
    ("提示记录", ("提醒", "记录", "整理本", "卡片", "标签")),
    ("温度测量", ("温度计", "水温测试仪", "水温卡")),
    ("打磨去角质", ("脚锉", "磨脚器", "去死皮")),
    ("防滑固定", ("防滑垫", "防滑扣", "固定器", "保护罩", "保护套")),
    ("切割钻孔", ("刮刀", "裁纸器", "线锯", "钻头", "扳手", "塞尺")),
    ("美甲磁吸造型", ("猫眼磁铁", "指甲磁铁", "美甲磁铁")),
    ("宠物训练或装扮", ("吠叫阻扰", "驱狗", "宠物套装", "狗面具", "宠物翅膀")),
)

MATERIAL_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("金属", ("金属", "合金", "铁艺", "不锈钢", "铜", "铝")),
    ("木制", ("木制", "木质", "木头", "榉木")),
    ("亚克力", ("亚克力",)),
    ("塑料", ("塑料", "PVC", "pvc")),
    ("硅胶", ("硅胶",)),
    ("橡胶", ("橡胶", "胶圈")),
    ("树脂", ("树脂",)),
    ("毛绒", ("毛绒",)),
    ("纸质", ("纸", "卡纸", "贴纸")),
    ("布艺", ("布", "绒", "纱", "蕾丝", "毛毡")),
    ("尼龙", ("尼龙", "nylon")),
    ("玻璃", ("玻璃",)),
)

THEME_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("万圣节", ("万圣节", "万圣", "halloween", "死神", "幽灵", "鬼头")),
    ("圣诞节", ("圣诞", "christmas")),
    ("花卉蝴蝶", ("花卉", "花朵", "蝴蝶", "珍珠蝴蝶")),
    ("动物", ("猫咪", "狗狗", "猫狗", "章鱼", "小鸟", "龙虾", "螃蟹")),
    ("幻想角色", ("龙", "女巫", "死神", "怪物", "恶魔")),
    ("婚礼", ("婚礼", "婚纱", "新娘")),
)

BRAND_AND_DEVICE_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("凯驰高压清洗机", ("凯驰", "Karcher", "karcher")),
    ("奥迪汽车", ("奥迪", "Audi", "audi")),
    ("宝马汽车", ("宝马", "BMW", "bmw")),
    ("奔驰汽车", ("奔驰", "Mercedes", "mercedes")),
    ("丰田汽车", ("丰田", "Toyota", "toyota")),
    ("高压清洗机", ("高压清洗机",)),
    ("挡风玻璃系统", ("挡风玻璃", "清洗泵")),
    ("自行车", ("自行车", "单车")),
    ("缝纫机", ("压脚", "缝纫机")),
)

PRODUCT_BODY_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("车载LED电子表", ("车载LED电子表",)),
    ("黑曜石猫咪水晶摆件", ("黑曜石", "猫咪", "摆件")),
    ("宠物生日装扮套装", ("宠物生日套装", "帽子", "三角巾")),
    ("身体贴钻", ("身体贴钻",)),
    ("专用地网盘", ("专用地网盘",)),
    ("高压清洗机密封环与接头卡扣", ("高压清洗机", "密封环", "接头卡扣")),
    ("儿童死神服装套装", ("儿童", "死神", "服装")),
    ("汽车挡风挡雨板金属卡扣", ("汽车挡风挡雨板", "金属卡扣")),
    ("自行车合金脚柱", ("自行车", "合金脚柱")),
    ("挡风玻璃清洗泵", ("挡风玻璃", "清洗泵")),
    ("万圣节服装套装", ("万圣节", "服装", "套装")),
    ("万圣节装饰吊饰", ("万圣节", "吊饰")),
    ("婚纱压脚", ("婚纱", "压脚")),
    ("蛋糕顶饰", ("蛋糕顶饰",)),
    ("电动跳舞章鱼玩具", ("电动", "跳舞", "章鱼", "玩具")),
)

COLOUR_WORDS = (
    "黑色", "白色", "红色", "蓝色", "绿色", "黄色", "粉色", "紫色", "橙色",
    "银色", "金色", "灰色", "棕色", "黑", "白", "红", "蓝", "绿", "黄", "粉",
)

FAMILY_KEYWORDS = tuple(
    sorted(
        {
            "高压清洗机密封环",
            "汽车挡风挡雨板金属卡扣",
            "挡风玻璃清洗泵",
            "自行车合金脚柱",
            "电动跳舞章鱼玩具",
            "万圣节儿童死神套装",
            "万圣节儿童服装",
            "婚纱压脚",
            "蛋糕顶饰",
            "火花塞",
            "耳骨夹",
            "耳钉",
            "肚脐钉",
            "腰链",
            "胸针",
            "领结",
            "钻石画",
            "数字油画",
            "杯垫",
            "贴纸",
            "手链",
            "感官石",
            "发夹",
            "发箍",
            "项链",
            "钥匙扣",
            "相框",
            "卡扣",
            "清洗泵",
            "密封环",
            "接头",
            "压脚",
            "服装",
            "长袍",
            "帽子",
            "生日帽",
            "女巫帽",
            "挂牌",
            "贺卡",
            "邀请卡",
        },
        key=len,
        reverse=True,
    )
)

NONSTANDARD_KEYWORDS = (
    "定制",
    "玩具",
    "公仔",
    "解压",
    "万圣节",
    "万圣",
    "钻石画",
    "数字油画",
    "贴纸",
    "贺卡",
    "邀请卡",
    "装饰",
    "吊饰",
    "挂饰",
    "项链",
    "手链",
    "耳骨夹",
    "耳钉",
    "肚脐钉",
    "肚脐扣",
    "腰链",
    "腰饰",
    "胸针",
    "领结",
    "发夹",
    "发箍",
    "服装",
    "长袍",
    "摆件",
    "相框",
    "3D打印",
    "3d打印",
    "配饰",
    "手环",
    "手镯",
    "装饰扣",
    "车贴",
    "挂件",
    "点钻",
    "幽灵",
    "蝙蝠",
    "纹身贴",
    "花朵",
    "水晶",
    "插旗",
    "拉旗",
    "蛋糕插排",
    "戒指",
    "耳环",
    "耳线",
    "身体链",
    "乳钉",
    "眉钉",
    "唇钉",
    "袖扣",
    "精灵耳朵",
    "假胡子",
    "塔罗牌",
    "立体书",
    "迷你恐龙",
    "陀螺",
    "马头棒",
    "放屁垫",
    "史莱姆耳朵",
    "纸花",
    "彩绘模板",
    "彩绘套装",
    "眼影睫毛贴",
    "情趣卡片",
    "石雕",
    "圣经学习本",
    "Bible学习计划本",
    "百科全书",
    "串珠",
    "猫毛收纳钥匙扣",
    "宠物毛发纪念",
    "减压神器",
    "吹球机",
    "溜溜球",
    "修女服饰",
    "气球",
)

FUNCTIONAL_KEYWORDS = (
    "密封",
    "接头",
    "卡扣",
    "适配",
    "适用",
    "兼容",
    "维修",
    "修理",
    "替换",
    "清洗泵",
    "压脚",
    "工具",
    "收纳",
    "保护",
    "防水",
    "防滑",
    "固定",
    "支架",
    "自行车",
    "火花塞",
    "钻头",
    "锯片",
    "扳手",
    "套筒",
    "塞尺",
    "脚锉",
    "磨脚器",
    "温度计",
    "水温测试仪",
    "防滑垫",
    "收纳袋",
    "滤水",
    "线锯",
    "液位指示器",
    "雨刷",
    "遮阳板",
    "固定器",
    "保护罩",
    "清洗机",
    "挂钩",
    "切割器",
    "切割片",
    "缝纫套装",
    "推剪梳",
    "限位梳",
    "偏光镜",
    "耳塞",
    "拖把",
    "工作灯",
    "护照夹",
    "测量尺",
    "手机挂绳",
    "口红包",
    "分趾器",
    "按钮开关",
    "空气炸锅烤架",
    "经络刷",
    "滚轮",
    "铰链",
    "收集袋",
    "握把管",
    "灯座",
    "饼干压模",
    "书页支撑夹",
    "管夹",
    "倒车灯",
    "滑轮",
    "排档套",
    "调平垫片",
    "药盒",
    "洗眼杯",
    "笔夹",
    "口部贴",
    "脚趾矫正器",
    "安全带调节器",
    "碟刹锁",
    "电缆",
    "插座垫片",
    "腰带收紧夹",
    "背带锁扣",
    "假发头架",
    "遮纹身袖套",
    "自拍镜",
    "拍摄神器",
    "喷嘴",
    "剥线钳",
    "分装药盒",
    "药丸收纳盒",
    "睡眠帽",
    "粉扑",
    "粘毛器",
    "去蒂器",
    "织带梳",
    "记分器",
    "鱼类休息室",
    "队长袖标",
    "眼镜挂绳",
    "工作证手机挂绳",
    "假发头架",
)

IP_RISK_KEYWORDS = ("第四翼", "4th wing", "APEROL", "大力水手", "不良人", "一梦江湖", "K-pop", "kpop")
ELECTRIC_RISK_KEYWORDS = ("带电", "电动", "USB", "usb", "太阳能", "发光")
SAFETY_RISK_KEYWORDS = ("燃气", "气罐", "婴儿车", "吸奶器", "高压清洗机")
EFFECT_RISK_KEYWORDS = ("牙垢剂", "护理液", "清洁剂", "除锈", "杀虫", "除草", "修复剂")

SOURCE_DIRECTION_ATOMS: dict[str, tuple[str, ...]] = {
    "DIY手工与文具材料": ("DIY手工商品开发",),
    "功能工具与维修配件": ("功能维修配件开发",),
    "厨房日用与收纳": ("厨房餐饮功能商品开发",),
    "派对贺卡与礼品": ("派对庆典商品开发",),
    "图像主题载体与装饰周边": ("家居装饰商品开发",),
    "玩具手办与解压收藏": ("玩具娱乐商品开发",),
    "服饰饰品与随身配件": ("服饰饰品开发",),
    "创意非标与收藏杂项": ("玩具娱乐商品开发",),
    "电子与功能设备": ("电子与手机功能配件开发",),
    "包袋收纳与出行用品": ("包袋与出行收纳开发",),
    "定制平面载体与轻装饰": ("家居装饰商品开发",),
    "阅读文具与办公用品": ("阅读文具与学习用品开发",),
    "花园家居与立体装饰": ("家居装饰商品开发",),
    "宗教纪念与精神礼物": ("阅读文具与学习用品开发",),
    "功能标品与杂项配件": ("功能维修配件开发",),
    "健康康复与身体护理": ("健康护理与康复辅助开发",),
    "玩具解压与3D打印": ("玩具娱乐商品开发",),
    "汽车骑行户外与功能配件": ("汽车功能配件开发",),
    "母婴儿童功能用品": ("健康护理与康复辅助开发",),
    "汽车维修与适配配件": ("汽车功能配件开发",),
    "家纺布艺与日用": ("家居收纳与防护商品开发",),
}

# 2025-10 至 2026-06 全量逐标题审核中确认的商品语义。
# 这些规则只使用可复用的商品本体/用途词，不使用 SKU 或月份硬编码；
# 纯编号、供应商标签和无法识别本体的标题仍必须留在待复核区。
AUDITED_DIRECTION_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    (
        "健康护理与康复辅助开发",
        (
            "面部提升贴", "耳穴贴", "牙缝刷", "齿间刷", "脚趾分离器",
            "分趾器", "v脸绷带", "V脸绷带", "颈部皱纹贴", "护趾套",
            "重力眼罩", "遮光眼罩", "防蓝光眼镜", "老年人口水巾",
        ),
    ),
    (
        "美妆个护功能商品开发",
        (
            "防脱梳", "造型梳", "画眉辅助尺", "磨甲修剪器", "烫睫毛",
            "沐浴球", "银器抛光布", "面部提拉", "身体贴钻",
        ),
    ),
    (
        "厨房餐饮功能商品开发",
        (
            "滤奶酪布", "奶酪布", "研磨器搅蒜器", "搅蒜器", "樱桃去核器",
            "碗罩", "面包发酵罩", "饺子器", "烤箱架拉拔器", "威士忌杯",
            "海浪玻璃杯", "冰淇淋杯", "玉米饼模具", "汉堡碟子",
            "咖啡豆碟", "咖啡豆盘", "酒瓶", "玻璃弯管",
        ),
    ),
    (
        "DIY手工商品开发",
        (
            "木珠", "圆珠", "描绘模版", "描绘模板", "布料画笔", "丙烯高光笔",
            "邮票压花器", "邮票框压花器", "切割垫板", "切割垫", "缝纫尺模板",
            "针法样本书", "绒皮绳", "沙盘模型树", "模型树仿真", "打结训练板",
            "画架支撑杆", "马克笔", "记号笔", "塑料彩色实色小圆片",
            "创意涂鸦盖章", "粘贴画", "马赛克贴画", "微缩模型绘画手柄",
            "柔光胶带",
        ),
    ),
    (
        "阅读文具与学习用品开发",
        (
            "日历", "圣经工作簿", "男性圣经", "本子款式", "文具盒", "铅笔盒",
            "书籍计数器", "阅读器", "书页夹", "记分牌", "可擦中性笔",
            "便利贴", "四合一印章", "针法样本书", "愿望清单",
            "应援照片", "学生铅笔文具袋", "笔架", "企鹅卡片", "乌龟卡片",
            "拥抱卡片",
        ),
    ),
    (
        "派对庆典商品开发",
        (
            "春节红包", "马年红包", "情人节刮刮卡", "情人节鲜花花束包装",
            "花束包装", "佩戴式花束夹", "胸花手腕花", "毕业典礼拍照用品",
            "GRAD毕业", "订婚快乐横幅", "生日肩带", "MOM TO BE肩带",
            "BIRTHDAY BADDIE", "感恩祝福告别", "中东节礼品纸袋",
            "世界杯球迷助威围巾", "世界杯比赛矩阵图", "世界杯矩阵图",
            "西部牛仔拍照道具", "牙买加旗帜", "牙买加串旗", "腊肠犬横幅",
            "充气篝火", "派对拍照", "粉色绶带",
        ),
    ),
    (
        "包袋与出行收纳开发",
        (
            "手提包", "斜挎包", "书籍手提袋", "运动腰带水壶", "托特包",
            "折叠购物袋", "麂皮绒女包", "背心包", "卡包", "名片盒",
            "尼龙网袋", "礼品纸袋", "飞机盒", "铅笔盒", "文具袋",
        ),
    ),
    (
        "服饰功能配件开发",
        (
            "睡帽", "长款手臂套", "遮阳口面罩", "裤脚防拖地", "隐形增高后垫袜",
            "收腰夹", "披肩夹", "收裤腰别针", "带垫肩", "运动腕带", "排球护肘",
            "防晒手掌", "窗帘遮光扣", "窗帘磁吸扣",
        ),
    ),
    (
        "服饰饰品开发",
        (
            "护士怀表", "弹簧扣8字扣", "S925弹簧扣", "蕾丝花边", "毛线帽",
            "发圈", "新娘大腿环", "连体泳衣", "豹纹泳衣", "芭蕾舞裙",
            "动漫假发", "cos长卷", "太阳眼镜", "太阳镜", "墨镜", "袜子",
            "折扇", "荆棘皇冠",
            "手串", "围巾", "花束夹", "牛头", "五角星", "链条",
        ),
    ),
    (
        "宠物用品开发",
        (
            "仓鼠", "猫舔食碗", "狗舔食碗", "舔食碗", "狗训练项圈",
            "项圈牵引绳", "攀爬爬架", "爬宠喂食夹", "海龟带卡片",
        ),
    ),
    (
        "户外运动商品开发",
        (
            "欧鲤钓线组", "鱼钩", "饵笼", "打窝器", "成人游泳圈", "双节棍",
            "运动腰带水壶", "排球护肘", "世界杯球迷", "助威围巾",
        ),
    ),
    (
        "汽车功能配件开发",
        (
            "轮廓标反光片", "夜光气门嘴帽", "氧传感器", "汽车除雪",
            "除雪仪", "防冻仪", "跑车四联画",
        ),
    ),
    (
        "电子与手机功能配件开发",
        (
            "非接触式支付卡夹", "万用表鳄鱼夹", "鳄鱼夹线", "DC电源",
            "断电开关",
        ),
    ),
    (
        "花园园艺功能商品开发",
        (
            "仿真睡莲", "多肉花插", "饮水鸟", "花插夜光马灯",
        ),
    ),
    (
        "家居收纳与防护商品开发",
        (
            "地漏片盖", "排风口盖", "通风管", "窗帘磁吸扣", "窗帘遮光扣",
            "磁吸式窗帘", "卷帘安装架", "罗马卷帘", "美边线", "美缝贴",
            "自粘拉手", "抽屉橱柜门把手", "柜门把手", "强力吸盘粘钩",
            "马桶盖脚垫", "桌布", "纸巾架", "卷纸架", "大容量折叠购物袋",
            "挂墙式单格名片盒", "玻璃瓶保护盖", "碗罩", "美缝条自粘纸",
            "牛皮纸气泡袋",
        ),
    ),
    (
        "功能维修配件开发",
        (
            "地网盘", "反光片", "弹簧扣", "磁性解扣器", "磁铁性解扣器",
            "解扣器磁扣取钉器", "86弹片", "圆环磁铁", "气门嘴帽",
            "吉他擦弦器", "银器抛光布", "保护盖", "磁吸扣",
        ),
    ),
    (
        "通用工具开发",
        (
            "旋转切割刀", "布料旋转切割刀", "卷尺PVC软皮尺", "木工打磨塑形抛光器",
            "羊毛百叶轮抛光盘", "邮票压花器", "解扣器", "画架支撑杆",
            "推夹器", "烤箱架拉拔器", "强力磁铁性解扣器",
        ),
    ),
    (
        "家居装饰商品开发",
        (
            "景观墙艺画", "墙艺画", "水晶玛瑙", "水晶石", "疗愈石", "紫水晶簇",
            "窗景摄影拍摄道具", "仿真睡莲", "多肉花插", "模型树", "海螺贝壳",
            "木质墙饰", "金属工艺品", "金属标志", "书挡", "书档", "金属书架",
            "钥匙挂架", "纸巾架", "摆件", "纪念物", "水晶包子", "水晶小笼包",
            "玻璃杯木质垫", "威士忌杯", "迷你造景热带鱼", "纪念章鱼酒瓶",
            "骷髅头酒瓶", "龙冰淇淋杯", "鲸鲨仿羊绒地垫", "相框",
            "杯木质垫",
        ),
    ),
    (
        "玩具娱乐商品开发",
        (
            "EDC推牌", "磁性旋转啪啪币", "鬼灭之刃", "女团卷卷贴", "女团拼贴换脸贴",
            "企鹅疗愈石", "马疗愈石", "水豚独角兽机甲", "磁吸迷你指尖板",
            "手指滑板", "滚珠魔方", "3D旋转环", "猜猜我是谁", "食人花鸭",
            "夜光亮片小乌龟", "迷你夜光海洋生物", "1/75彩色人物模型",
            "香蕉双节棍", "饮水鸟", "高地编织正能量", "正能量编织鲨鱼",
            "老毛俱乐部", "KATSEYE", "KPopDemonHunters", "卡皮治郎", "卡皮高达",
            "彩色表情小笼包", "傻鹅纸巾",
        ),
    ),
)

# 深度逐条审核后的主方向优先级。仅放入标题语义明确、容易被通用关键词
# （例如“猫”“贴纸”“支架”“电子”）误导的产品，先确定真实商品用途。
PRIMARY_DIRECTION_OVERRIDE_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("家居装饰商品开发", (
        "黑曜石小猫咪水晶", "爱心水晶礼物",
    )),
    ("派对庆典商品开发", (
        "贺卡", "刮刮乐", "情人节吸管", "爱心纸拉花", "爱心螺旋吊饰",
        "激光亮片情人节", "蛋糕装饰摆件", "农场动物蛋糕装饰",
    )),
    ("美妆装饰商品开发", (
        "身体贴钻", "脸部纹身贴", "纹身贴", "嘴唇A款", "面部装饰",
        "美甲水钻", "美甲钻套装",
    )),
    ("健康护理与康复辅助开发", (
        "呼吸解压项链", "义齿盒", "红外线鼻炎器",
    )),
    ("宠物用品开发", (
        "鸟笼清洁器", "鸟笼刷", "鱼缸外置过滤器", "宠物生日套装",
    )),
    ("户外运动商品开发", (
        "钓鱼钳", "网球拍记分器", "高尔夫记分卡",
    )),
    ("阅读文具与学习用品开发", (
        "图钉", "信纸套装", "阅读障碍辅助", "书签卡片", "便签收纳盒",
        "线圈本", "笔记本本子", "快捷键鼠标垫",
    )),
    ("DIY手工商品开发", (
        "钩针套装", "针具磁力收纳器", "按扣金属子母扣", "免缝纽扣",
        "透明树脂配件", "弹簧扣8字扣",
    )),
    ("包袋与出行收纳开发", (
        "卫生巾收纳袋", "零钱包", "钱包", "女团礼品袋",
    )),
    ("服饰功能配件开发", (
        "假发网睡帽", "护士怀表",
    )),
    ("服饰饰品开发", (
        "石头串珠钥匙链", "串珠手链", "发饰套装", "亚克力钥匙扣",
        "佩戴式花束夹", "蜥蜴钥匙扣",
    )),
    ("通用工具开发", (
        "六角冲击套筒", "蜡烛灯芯中心定位工具", "购物车钥匙扣",
    )),
    ("汽车功能配件开发", (
        "车载LED电子表", "轮胎除锈清洁套件",
    )),
    ("玩具娱乐商品开发", (
        "儿童电动高铁仿真模型", "情人节心形积木", "骷髅头乌鸦硬币",
    )),
    ("家居收纳与防护商品开发", (
        "悬挂吊钩", "窗帘固定器", "钥匙收纳铁艺挂架", "窗帘磁吸扣",
        "沙发床挡板", "墙上置物架", "淋浴喷头支架", "衣柜杆支架",
        "家用壁挂钩", "垃圾袋收纳架", "橱柜抽屉拉手", "银器抛光布",
    )),
    ("家居装饰商品开发", (
        "汽车挂件", "汽车吊饰", "装饰车挂", "汽车挂饰", "猫咪情侣2D亚克力摆件",
        "黑曜石小猫咪水晶", "高尔夫口袋拥抱", "骷髅头乌鸦硬币",
    )),
)

STANDARDIZATION_OVERRIDE_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("非标", (
        "宠物生日套装", "汽车挂件", "汽车吊饰", "装饰车挂", "汽车挂饰",
        "身体贴钻", "纹身贴", "高尔夫口袋拥抱", "情人节吸管",
        "女团礼品袋",
    )),
    ("标品", (
        "六角冲击套筒", "鸟笼清洁器", "鸟笼刷", "悬挂吊钩", "窗帘固定器",
        "购物车钥匙扣", "阅读障碍辅助", "针具磁力收纳器", "车载LED电子表",
        "网球拍记分器", "红外线鼻炎器", "鱼缸外置过滤器", "义齿盒",
        "银器抛光布", "卫生巾收纳袋", "双向导向梳", "钩针套装",
        "墙上置物架", "淋浴喷头支架", "衣柜杆支架", "家用壁挂钩",
        "垃圾袋收纳架", "橱柜抽屉拉手", "呼吸解压项链",
        "美甲旋转式贴钻粘钻笔", "快捷键鼠标垫", "弹簧扣8字扣",
        "按扣金属子母扣", "免缝纽扣", "护士怀表",
    )),
)

FORCE_REVIEW_TITLES = {
    "花朵手柄套装--9pcs",
}

TITLE_ANALYSIS_HEADERS = [
    "创建月份",
    "开发人",
    "SKU",
    "原始完整标题",
    "规范产品名称",
    "产品本体",
    "产品族",
    "核心功能",
    "使用场景",
    "适用人群",
    "适配对象",
    "主题元素",
    "材质与结构",
    "数量与组合",
    "季节属性",
    "建议售卖时间",
    "商品标准化类型",
    "标准化判断依据",
    "差异化方式",
    "风险提示",
    "方向Skill原子",
    "开发方法原子",
    "组合打法原子",
    "产品方案ID",
    "产品方案键",
    "关联组合主SKU",
    "关联完整产品标题",
    "标题分析依据",
    "信息补充依据",
    "分析置信度",
    "是否待复核",
    "待复核原因",
]


def contains_all(text: str, keywords: Iterable[str]) -> bool:
    lowered = text.lower()
    return all(keyword.lower() in lowered for keyword in keywords)


def contains_any(text: str, keywords: Iterable[str]) -> bool:
    lowered = text.lower()
    return any(keyword.lower() in lowered for keyword in keywords)


def classify_family(title: str) -> str:
    lowered = title.lower()
    for keyword in FAMILY_KEYWORDS:
        if keyword.lower() in lowered:
            return keyword
    text = normalize_product_name(title)
    text = re.split(r"【|（|\(|-|—|–|：|:", text, maxsplit=1)[0]
    return text[:24].strip() or "待人工归族"


def classify_title_standardization(title: str) -> dict[str, str]:
    if "定制" in title:
        return {
            "商品标准化类型": "非标",
            "标准化判断依据": "标题明确包含“定制”",
            "标准化置信度": "高",
        }
    for standardization_type, keywords in STANDARDIZATION_OVERRIDE_RULES:
        if contains_any(title, keywords):
            return {
                "商品标准化类型": standardization_type,
                "标准化判断依据": (
                    f"逐标题深审确认该商品以"
                    f"{'功能/适配' if standardization_type == '标品' else '主题/造型'}价值为主"
                ),
                "标准化置信度": "高",
            }
    if contains_any(title, NONSTANDARD_KEYWORDS):
        matched = next(keyword for keyword in NONSTANDARD_KEYWORDS if keyword.lower() in title.lower())
        return {
            "商品标准化类型": "非标",
            "标准化判断依据": f"标题包含非标/主题化特征“{matched}”",
            "标准化置信度": "高",
        }
    if contains_any(title, FUNCTIONAL_KEYWORDS):
        matched = next(keyword for keyword in FUNCTIONAL_KEYWORDS if keyword.lower() in title.lower())
        return {
            "商品标准化类型": "标品",
            "标准化判断依据": f"标题包含功能/适配特征“{matched}”",
            "标准化置信度": "高",
        }
    return {
        "商品标准化类型": "待复核",
        "标准化判断依据": "标题尚不足以确认功能主导或主题审美主导",
        "标准化置信度": "低",
    }


def title_risk_flags(title: str) -> list[str]:
    flags: list[str] = []
    if contains_any(title, IP_RISK_KEYWORDS):
        flags.append("明确IP/品牌词复核")
    if contains_any(title, ELECTRIC_RISK_KEYWORDS):
        flags.append("带电履约复核")
    if contains_any(title, SAFETY_RISK_KEYWORDS):
        flags.append("安全/适配责任复核")
    if contains_any(title, EFFECT_RISK_KEYWORDS):
        flags.append("化学/功效复核")
    if contains_any(title, ("汽车", "适配", "适用", "兼容", "凯驰")):
        flags.append("规格适配复核")
    return flags


def matched_labels(
    text: str,
    rules: Iterable[tuple[str, tuple[str, ...]]],
) -> list[str]:
    lowered = text.lower()
    return [
        label
        for label, keywords in rules
        if any(keyword.lower() in lowered for keyword in keywords)
    ]


def join_labels(labels: Iterable[str], fallback: str) -> str:
    unique = list(dict.fromkeys(label for label in labels if label))
    return "、".join(unique) if unique else fallback


def normalize_product_name(title: str) -> str:
    text = re.sub(r"^【[^】]+】", "", title).strip()
    text = re.sub(r"^3d打印\s*[-—–_:：]*", "", text, flags=re.I)
    text = re.sub(r"\bL-\d+(?:-\d+)?\b", "", text, flags=re.I)
    text = re.sub(r"建议身高\s*\d+\s*[-~—–]\s*\d+", "", text)
    text = re.sub(r"\s+", "", text)
    return text.strip("-—–，,；;：:")


def extract_product_body(title: str) -> tuple[str, bool]:
    for label, required_keywords in PRODUCT_BODY_RULES:
        if contains_all(title, required_keywords):
            return label, True

    family = classify_family(title)
    if family and family not in {"待人工归族", title[:24]} and len(family) <= 16:
        return family, True

    cleaned = normalize_variant_key(title)
    if cleaned and not re.fullmatch(r"[A-Za-z0-9款新品神秘]+", cleaned):
        return cleaned[:32], False
    return "待确认产品本体", False


def extract_quantity_and_bundle(title: str) -> str:
    quantities = re.findall(
        r"\d+(?:\.\d+)?\s*(?:pcs|sets?|件套装|件套|件|个装|个|只装|只|枚|套|张|片|包|组)",
        title,
        flags=re.I,
    )
    details = ""
    if any(token in title for token in ("+", "套装", "件套")):
        left_positions = [position for position in (title.find("（"), title.find("(")) if position >= 0]
        if left_positions:
            details = title[min(left_positions) :].strip()
        elif "+" in title:
            details = title
    parts = list(dict.fromkeys(quantities))
    if details:
        parts.append(details)
    return "；".join(parts) if parts else "单件/标题未写明组合数量"


def extract_season(title: str) -> str:
    seasons = []
    if re.search(r"万圣|halloween", title, re.I):
        seasons.append("节日季节性：万圣节")
    if re.search(r"圣诞|christmas", title, re.I):
        seasons.append("节日季节性：圣诞节")
    if any(keyword in title for keyword in ("夏季", "沙滩", "游泳")):
        seasons.append("夏季场景")
    if any(keyword in title for keyword in ("冬季", "保暖", "耳套")):
        seasons.append("冬季场景")
    return join_labels(seasons, "标题未见明确季节词")


def extract_differentiation(
    title: str,
    themes: list[str],
    audiences: list[str],
    fitments: list[str],
    materials: list[str],
    quantity: str,
) -> list[str]:
    methods: list[str] = []
    if quantity != "单件/标题未写明组合数量":
        methods.append("套装/数量配置")
    if themes:
        methods.append("主题元素差异化")
    if fitments:
        methods.append("设备/车型适配")
    if audiences:
        methods.append("适用人群细分")
    if materials:
        methods.append("材质/结构差异化")
    if any(word in title for word in COLOUR_WORDS) or re.search(r"\b(?:XS|S|M|L|XL|XXL|\d+XL)\b", title, re.I):
        methods.append("颜色/尺寸规格差异化")
    if "+" in title:
        methods.append("跨组件组合")
    return list(dict.fromkeys(methods))


def title_is_insufficient(
    title: str,
    product_body: str,
    body_confirmed: bool,
    semantic_dimension_count: int,
) -> bool:
    if product_body == "待确认产品本体":
        return True
    if body_confirmed or semantic_dimension_count > 0:
        return False
    reduced = normalize_product_name(title)
    reduced = re.sub(
        r"(?:宝蓝|深蓝|浅蓝|藏蓝|天蓝|湖蓝|玫红|粉橙|咖啡|米白|墨绿)色?",
        "",
        reduced,
    )
    for colour in COLOUR_WORDS:
        reduced = reduced.replace(colour, "")
    reduced = re.sub(
        r"\d+(?:\.\d+)?\s*(?:号|款|pcs|sets?|件|个|只|枚|套|张|片|包|组|cm|mm|米|条)?",
        "",
        reduced,
        flags=re.I,
    )
    reduced = re.sub(r"[A-Za-z0-9+*×/，,、（）()\[\]【】\s_-]", "", reduced)
    reduced = re.sub(r"^(?:一条|一件|一套|一组|一包)+$", "", reduced)
    return len(reduced) < 2


def select_primary_direction_atom(
    title: str,
    product_body: str,
    atoms: Iterable[str],
) -> str:
    candidates = list(dict.fromkeys(atom for atom in atoms if atom))
    if not candidates:
        return "待确认开发方向"
    if len(candidates) == 1:
        return candidates[0]

    text = f"{title}|{product_body}"
    selection_rules: tuple[tuple[str, tuple[str, ...]], ...] = (
        ("宠物用品开发", ("宠物", "猫咪", "狗狗", "猫狗", "犬用", "宠物美容", "猫毛收纳", "鱼类休息室")),
        ("汽车功能配件开发", ("汽车", "车载", "挡风玻璃", "宝马", "奔驰", "奥迪", "摩托车", "换挡", "排档")),
        ("骑行功能配件开发", ("自行车", "单车", "骑行")),
        ("健康护理与康复辅助开发", ("药盒", "洗眼杯", "脚趾矫正", "口部贴", "康复", "经络刷")),
        ("美妆个护功能商品开发", ("美甲", "指甲", "脚锉", "磨脚", "推剪", "理发器", "化妆", "粉扑")),
        ("美妆装饰商品开发", ("眼影", "睫毛", "面部装饰", "雀斑", "纹身遮盖")),
        ("阅读文具与学习用品开发", ("学习本", "计划本", "百科全书", "早教卡片", "立体书", "阅读")),
        ("包袋与出行收纳开发", ("护照夹", "口红包", "化妆包", "手机挂绳", "工作证", "眼镜挂绳")),
        ("DIY手工商品开发", ("DIY", "diy", "钻石画", "刺绣", "手工", "串珠", "贴纸书")),
        ("玩具娱乐商品开发", ("玩具", "公仔", "解压", "减压", "感官石", "3D打印", "3d打印", "毛绒")),
        ("服饰饰品开发", ("服装", "长袍", "发夹", "发箍", "项链", "手链", "耳环", "耳钉", "戒指", "腰链", "腰饰", "胸针", "头饰", "面具")),
        ("服饰功能配件开发", ("隐形肩带", "腰带收紧夹", "遮纹身袖套", "睡眠帽", "防滑耳套", "内衣防滑扣")),
        ("派对庆典商品开发", ("生日", "派对", "周岁", "蛋糕顶饰", "蛋糕插排", "纸花", "拉旗")),
        ("家居装饰商品开发", ("摆件", "挂饰", "太阳捕手", "水晶鸟", "圣诞树装饰", "乔迁礼物")),
        ("户外运动商品开发", ("露营", "高尔夫", "足球", "台球", "鱼竿", "指南针", "求生口哨")),
        ("花园园艺功能商品开发", ("花园", "园艺", "浇花", "除草", "打草", "种植")),
        ("厨房餐饮功能商品开发", ("厨房", "空气炸锅", "咖啡机", "餐具", "过滤网", "筷子", "去蒂器")),
        ("电子与手机功能配件开发", ("手机摄影", "自拍镜", "拍摄神器", "LED工作灯", "按钮开关", "灯座")),
        ("电子测量功能商品开发", ("温度计", "水温测试仪", "水温卡")),
        ("家居收纳与防护商品开发", ("收纳", "防滑垫", "扶手垫", "钥匙盒", "粘毛器", "拖把")),
        ("功能维修配件开发", ("密封", "替换", "更换", "连接器", "接头", "卡扣", "清洗泵", "喷嘴", "配件")),
        ("通用工具开发", ("工具", "扳手", "刮刀", "套筒", "塞尺", "钻头", "锯片", "裁纸器", "拆表")),
    )
    for atom, keywords in selection_rules:
        if atom in candidates and contains_any(text, keywords):
            return atom

    priority = tuple(atom for atom, _ in selection_rules)
    for atom in priority:
        if atom in candidates:
            return atom
    return candidates[0]


def direction_atoms(
    title: str,
    product_body: str,
    functions: list[str],
    scenes: list[str],
) -> list[str]:
    for atom, keywords in PRIMARY_DIRECTION_OVERRIDE_RULES:
        if contains_any(title, keywords):
            return [atom]
    atoms: list[str] = []
    if any(value in functions for value in ("密封防漏", "连接适配", "维修替换", "清洁维护")):
        atoms.append("功能维修配件开发")
    if any(keyword in title for keyword in ("汽车", "挡风玻璃", "车载", "水箱盖", "行车记录仪", "安全带套")):
        atoms.append("汽车功能配件开发")
    if "火花塞" in title and not any(keyword in title for keyword in ("割草机", "油锯", "园林")):
        atoms.append("汽车功能配件开发")
    if any(keyword in title for keyword in ("自行车", "单车", "骑行")):
        atoms.append("骑行功能配件开发")
    if any(keyword in title for keyword in ("DIY", "diy", "钻石画", "刺绣", "手工")):
        atoms.append("DIY手工商品开发")
    if any(
        keyword in title
        for keyword in (
            "服装", "发夹", "发箍", "项链", "手链", "帽子", "生日帽", "女巫帽", "耳骨夹", "耳钉",
            "肚脐钉", "肚脐扣", "乳钉", "眉钉", "唇钉", "腰链", "腰饰",
            "胸针", "领结", "首饰", "戒指", "耳环", "耳线", "身体链", "袖扣",
            "精灵耳朵", "头巾", "围脖", "修女服饰", "头巾别针",
            "隐形肩带", "头带", "装扮道具", "面具",
        )
    ):
        atoms.append("服饰饰品开发")
    if any(keyword in title for keyword in ("玩具", "公仔", "解压", "减压", "感官石", "吹球机", "溜溜球")):
        atoms.append("玩具娱乐商品开发")
    if any(keyword in title for keyword in ("迷你恐龙", "磁力片", "陀螺", "放屁垫", "史莱姆耳朵", "马头棒")):
        atoms.append("玩具娱乐商品开发")
    if any(keyword in title for keyword in ("宠物", "猫咪", "猫毛收纳", "宠物毛发", "狗狗", "猫狗", "犬用", "狗面具", "宠物套装", "吠叫阻扰", "鱼类休息室")):
        atoms.append("宠物用品开发")
    if any(keyword in title for keyword in ("搓澡", "洗澡", "梳子", "去黑头", "化妆", "美甲", "指甲", "猫眼磁铁", "敷贴", "涂抹器")):
        atoms.append("美妆个护功能商品开发")
    if any(keyword in title for keyword in ("眼影睫毛贴", "彩绘模板", "彩绘模版", "彩绘套装", "纹身遮盖贴")):
        atoms.append("美妆装饰商品开发")
    if any(keyword in title for keyword in ("推剪梳", "限位梳", "理发器", "电推剪", "粉扑", "织带梳", "假发头架")):
        atoms.append("美妆个护功能商品开发")
    if any(keyword in title for keyword in ("垫发", "蓬松隐形造型", "发网", "U型夹")):
        atoms.append("服饰饰品开发")
    if any(keyword in title for keyword in ("药盒", "药丸收纳盒", "分装药盒", "洗眼杯", "口部贴", "脚趾矫正器", "耳塞", "分趾器", "经络刷", "去角质")):
        atoms.append("健康护理与康复辅助开发")
    if any(keyword in title for keyword in ("脚锉", "磨脚器", "去死皮")):
        atoms.append("美妆个护功能商品开发")
    if any(keyword in title for keyword in ("花园", "浇花", "施肥", "除草", "打草", "种植", "园艺")):
        atoms.append("花园园艺功能商品开发")
    if any(keyword in title for keyword in ("咖啡机", "咖啡过滤", "咖啡滤纸", "咖啡豆计量勺", "计量勺", "餐具", "厨房", "过滤网", "空锡盒", "空气炸锅", "饼干压模", "筷子", "去蒂器", "瓜果螺旋切割器")):
        atoms.append("厨房餐饮功能商品开发")
    if any(keyword in title for keyword in ("露营", "鱼竿", "求生口哨", "指南针", "毽子")) or (
        "高尔夫" in title
        and not any(keyword in title for keyword in ("汽车", "奥迪", "大众", "VW", "适用"))
    ):
        atoms.append("户外运动商品开发")
    if any(keyword in title for keyword in ("足球", "队长袖标", "台球", "斯洛克")):
        atoms.append("户外运动商品开发")
    if any(keyword in title for keyword in ("刮刀", "拆表", "打草头", "钻夹头", "裁纸器", "吊钩", "支架", "调音叉", "工具")):
        atoms.append("通用工具开发")
    if any(keyword in title for keyword in ("切割器", "切割片", "缝纫套装", "滚轮", "铰链", "管夹", "调平垫片", "插座垫片", "打孔针套件", "盒子切割器")):
        atoms.append("通用工具开发")
    if any(keyword in title for keyword in ("握把管", "滑轮", "背带锁扣")):
        atoms.append("功能维修配件开发")
    if any(keyword in title for keyword in ("喷嘴", "剥线钳")):
        atoms.append("功能维修配件开发")
    if any(keyword in title for keyword in ("周岁", "生日", "派对", "拉旗", "插旗", "蛋糕顶饰", "蛋糕插排")):
        atoms.append("派对庆典商品开发")
    if any(keyword in title for keyword in ("纸花", "假胡子", "马头棒", "放屁垫", "夜店", "情趣卡片", "塔罗牌", "装扮道具")):
        atoms.append("派对庆典商品开发")
    if any(keyword in title for keyword in ("针笔", "起针铲", "贴纸书", "闪光纸", "绘画纸", "点钻笔", "托盘", "胶泥")):
        atoms.append("DIY手工商品开发")
    if any(keyword in title for keyword in ("串珠", "缝纫套装", "墙面绘画模板")):
        atoms.append("DIY手工商品开发")
    if any(keyword in title for keyword in ("3D打印", "3d打印", "接球游戏")):
        atoms.append("玩具娱乐商品开发")
    if any(keyword in title for keyword in ("温度计", "水温测试仪", "水温卡")):
        atoms.append("电子测量功能商品开发")
    if any(keyword in title for keyword in ("LED", "led", "按钮开关", "灯座", "偏光镜手机摄影", "自拍镜", "拍摄神器")):
        atoms.append("电子与手机功能配件开发")
    if any(keyword in title for keyword in ("收纳袋", "隔热垫", "扶手垫", "防滑垫", "种子储存整理本")):
        atoms.append("家居收纳与防护商品开发")
    if any(keyword in title for keyword in ("护照夹", "口红包", "树叶收集袋", "药盒", "钥匙盒隐藏器", "书页支撑夹", "粘毛器", "拖把")):
        atoms.append("包袋与出行收纳开发" if any(word in title for word in ("护照夹", "口红包")) else "家居收纳与防护商品开发")
    if any(keyword in title for keyword in ("手机挂绳", "工作证", "眼镜挂绳", "笔夹")):
        atoms.append("包袋与出行收纳开发")
    if "化妆包" in title:
        atoms.append("包袋与出行收纳开发")
    if any(keyword in title for keyword in ("钥匙盒", "滑轮", "家具调平")):
        atoms.append("家居收纳与防护商品开发")
    if any(keyword in title for keyword in ("车贴", "雨刷", "遮阳板", "宝马", "进气通风管")):
        atoms.append("汽车功能配件开发")
    if any(keyword in title for keyword in ("安全带调节器", "摩托车", "碟刹锁", "排档套", "倒车灯")):
        atoms.append("汽车功能配件开发")
    if "换挡" in title:
        atoms.append("汽车功能配件开发")
    if any(keyword in title for keyword in ("手环", "手镯", "配饰", "鞋子装饰扣", "纹身贴", "虎眼石")):
        atoms.append("服饰饰品开发")
    if any(keyword in title for keyword in ("面部装饰珍珠贴纸", "雀斑纹身贴")):
        atoms.append("美妆装饰商品开发")
    if any(keyword in title for keyword in ("眼镜硅防滑耳套", "内衣防滑扣")):
        atoms.append("服饰功能配件开发")
    if any(keyword in title for keyword in ("腰带收紧夹", "遮纹身袖套")):
        atoms.append("服饰功能配件开发")
    if "睡眠帽" in title:
        atoms.append("服饰功能配件开发")
    if "吊篮挂钩" in title:
        atoms.append("花园园艺功能商品开发")
    if any(keyword in title for keyword in ("太阳捕手", "水晶鸟", "亚克力挂饰", "水晶挂件")):
        atoms.append("家居装饰商品开发")
    if any(keyword in title for keyword in ("石雕", "纸花", "墙面绘画模板", "树脂摆件", "装饰摆件", "荡秋千情侣幽灵", "圣诞树装饰", "乔迁礼物", "礼物挂饰")):
        atoms.append("家居装饰商品开发")
    if any(keyword in title for keyword in ("圣经学习本", "Bible学习计划本", "百科全书", "记分器", "学习本", "计划本")):
        atoms.append("阅读文具与学习用品开发")
    if any(keyword in title for keyword in ("镰刀", "蝙蝠翅膀")):
        atoms.append("服饰饰品开发")
    if any(keyword in title for keyword in ("塞尺", "线锯", "窗帘固定器", "液位指示器")):
        atoms.append("通用工具开发")
    supplemental_direction_rules: tuple[tuple[str, tuple[str, ...]], ...] = (
        (
            "家居收纳与防护商品开发",
            (
                "坐垫", "门挡", "柜门抽屉", "衣柜门把手", "家具可调节支脚",
                "移门拉手", "窗帘塑料调节钩", "封边条", "布线扣", "线夹",
                "出线孔面板", "肥皂泡沫分配器",
            ),
        ),
        (
            "厨房餐饮功能商品开发",
            (
                "切条器", "抹茶搅拌器", "蛋白粉容器", "多功能叉勺", "开瓶器",
                "微波炉", "解冻菜板", "刻度量杯", "零食容器", "洋葱切",
                "可加热蒸笼",
            ),
        ),
        (
            "包袋与出行收纳开发",
            (
                "礼品袋", "钱夹", "手机包", "证件卡套", "卡套", "薰衣草布袋",
            ),
        ),
        (
            "电子与手机功能配件开发",
            (
                "开关吊坠拉线", "手机夹片", "继电器模块", "电池隔离器",
                "光伏隔离开关", "遥控器磁吸贴", "收音机天线", "保护面板膜",
                "AAC设备", "AAC自闭症设备", "平板电脑舒适手带",
            ),
        ),
        (
            "美妆个护功能商品开发",
            (
                "唇刷架", "胡子填充笔", "针孔眼镜", "刮痧板", "足部量脚器",
                "牙科抛光条", "眼镜腿套",
            ),
        ),
        (
            "通用工具开发",
            (
                "缝纫枪", "圆角定位模具", "折叠木尺", "勾线笔", "干墙画架",
                "疏通器", "针线套装", "断路器安全锁", "缝纫机多边形强磁定规",
                "缝纫机定规",
            ),
        ),
        (
            "玩具娱乐商品开发",
            (
                "充气野人钉锤", "放屁机", "智力解扣", "玉米扭扭乐", "LOMO小卡",
                "猎魔女团", "磁性分数学习", "立体降临节日历", "北极熊",
            ),
        ),
        (
            "DIY手工商品开发",
            (
                "EVA海棉贴片", "EVA海绵贴片", "秋季工艺套件", "缝纫线",
                "红果三角叶", "实心浮雕贴",
            ),
        ),
        (
            "服饰饰品开发",
            (
                "流苏链条", "腰带夹", "领带夹", "拉链头", "皮革拉链",
                "麋鹿头饰", "励志蜜蜂不锈钢钥匙圈",
            ),
        ),
        (
            "家居装饰商品开发",
            (
                "纪念日旗帜", "纪念旗帜", "Lest We Forget flag", "洗手间门牌", "门牌",
                "倒数日历", "英国日历", "看书猫日历", "圣诞车枕套", "相框架",
            ),
        ),
        (
            "服饰功能配件开发",
            ("收腰神器松紧带",),
        ),
        (
            "汽车功能配件开发",
            (
                "油管单向止回阀", "车轮螺母", "MINI Cooper", "Mazda CX3",
                "房车窗户", "奥迪迎宾灯", "方向盘套", "前拖车盖", "滤清器进气管",
                "雪挡", "防风绑带",
            ),
        ),
        (
            "户外运动商品开发",
            (
                "口哨", "登山杖", "花样滑冰鞋套", "滑冰鞋套",
            ),
        ),
        (
            "花园园艺功能商品开发",
            (
                "蔬菜防冻罩", "保温裹树布", "雨桶水箱网罩",
            ),
        ),
        (
            "宠物用品开发",
            ("激光逗猫棒",),
        ),
        (
            "阅读文具与学习用品开发",
            ("课堂良好行为激励卡", "Holy Quran", "古兰经", "信封"),
        ),
        (
            "派对庆典商品开发",
            ("新年黑金纸质彩色3D眼镜", "新年", "洗礼礼物"),
        ),
    )
    for atom, keywords in supplemental_direction_rules:
        if contains_any(title, keywords):
            atoms.append(atom)
    for atom, keywords in AUDITED_DIRECTION_RULES:
        if contains_any(title, keywords):
            atoms.append(atom)
    atoms = list(dict.fromkeys(atoms))
    if "玩具娱乐商品开发" in atoms and any(
        keyword in title for keyword in ("玩具", "减压神器", "3D打印", "3d打印")
    ):
        atoms = [
            atom
            for atom in atoms
            if atom not in {"通用工具开发", "服饰饰品开发"}
        ]
    if "美妆个护功能商品开发" in atoms and any(
        keyword in title for keyword in ("美甲工具", "造型工具", "推剪梳", "限位梳")
    ):
        atoms = [atom for atom in atoms if atom != "通用工具开发"]
    if "服饰饰品开发" in atoms and any(
        keyword in title for keyword in ("造型工具", "垫发", "发网", "U型夹")
    ):
        atoms = [atom for atom in atoms if atom != "通用工具开发"]
    if "厨房餐饮功能商品开发" in atoms and any(
        keyword in title for keyword in ("瓜果", "草莓", "饼干", "空气炸锅", "水槽")
    ):
        atoms = [atom for atom in atoms if atom != "通用工具开发"]
    if "化妆包" in title:
        atoms = [
            atom
            for atom in atoms
            if atom not in {"美妆个护功能商品开发", "玩具娱乐商品开发"}
        ]
        atoms.append("包袋与出行收纳开发")
    if "汽车功能配件开发" in atoms and "手机支架" in title:
        atoms = [atom for atom in atoms if atom != "通用工具开发"]
    return [select_primary_direction_atom(title, product_body, atoms)]


def method_atoms(
    title: str,
    methods: list[str],
    functions: list[str],
    season: str,
) -> list[str]:
    atoms: list[str] = []
    if "套装/数量配置" in methods or "跨组件组合" in methods:
        atoms.append("套装组合开发")
    if "设备/车型适配" in methods:
        atoms.append("设备适配开发")
    if "主题元素差异化" in methods:
        atoms.append("主题元素差异化开发")
    if season != "标题未见明确季节词":
        atoms.append("季节主题提前布局")
    if "适用人群细分" in methods:
        atoms.append("人群场景细分开发")
    if "颜色/尺寸规格差异化" in methods:
        atoms.append("规格变体开发")
    if any(value in functions for value in ("密封防漏", "连接适配", "维修替换")):
        atoms.append("维修替换需求开发")
    return atoms or ["单品基础开发"]


def standardization_for_title(title: str, product_family: str) -> dict[str, str]:
    del product_family
    return classify_title_standardization(title)


def infer_sales_window(
    title: str,
    directions: list[str],
    season: str,
) -> str:
    if "万圣节" in season or "万圣节场景商品开发" in directions:
        return "8—10月主售，9—10月重点；建议7—8月完成上架和备货"
    if "圣诞节" in season:
        return "10—12月主售，11—12月重点；建议9—10月完成上架和备货"
    if "夏季场景" in season or any(word in title for word in ("游泳", "沙滩", "水枪")):
        return "4—8月主售，5—7月重点；其余月份作为常规补充"
    if "冬季场景" in season:
        return "9月至次年2月主售，10—12月重点"
    if "花园园艺功能商品开发" in directions:
        return "3—8月主售，4—7月重点；温暖地区可全年销售"
    if "派对庆典商品开发" in directions:
        return "全年可售；生日、婚礼和节日前2—6周重点推广"
    if "玩具娱乐商品开发" in directions:
        return "全年可售；10—12月礼品季及学校假期需求更强"
    if "DIY手工商品开发" in directions:
        return "全年可售；节日前、学校假期和礼品季可加强"
    if "服饰饰品开发" in directions or "美妆装饰商品开发" in directions:
        return "全年可售；节日、派对、婚礼和礼品季可加强"
    return "全年常驻销售；结合具体使用场景做阶段性推广"


def infer_missing_dimensions(
    title: str,
    product_body: str,
    functions: list[str],
    scenes: list[str],
    audiences: list[str],
    fitments: list[str],
    themes: list[str],
    materials: list[str],
    directions: list[str],
    standardization_type: str,
    season: str,
) -> dict[str, str]:
    inferred_fields: list[str] = []

    if functions:
        function_value = join_labels(functions, "")
    else:
        function_rules = (
            ("万圣节场景商品开发", "节日装扮、角色造型或场景布置"),
            ("服饰功能配件开发", "穿戴固定、防滑或舒适性辅助"),
            ("服饰饰品开发", "装饰穿戴与造型搭配"),
            ("美妆装饰商品开发", "面部或身体装饰与造型表达"),
            ("美妆个护功能商品开发", "个人清洁、护理或美容辅助"),
            ("电子测量功能商品开发", "温度检测与使用状态提示"),
            ("功能维修配件开发", "安装、替换、维修或功能恢复"),
            ("汽车功能配件开发", "车辆维护、替换或功能补充"),
            ("骑行功能配件开发", "骑行安装、维护或使用辅助"),
            ("通用工具开发", "安装、拆卸、切割、测量或维修辅助"),
            ("DIY手工商品开发", "手工制作、点钻、绘画或材料整理"),
            ("派对庆典商品开发", "庆典布置、气氛营造或活动装扮"),
            ("玩具娱乐商品开发", "娱乐互动、解压或收藏展示"),
            ("花园园艺功能商品开发", "园艺维护、种植或户外整理"),
            ("厨房餐饮功能商品开发", "食品处理、饮品制作或厨房整理"),
            ("家居收纳与防护商品开发", "家庭收纳、保护或防滑固定"),
            ("家居装饰商品开发", "家居空间装饰与视觉点缀"),
            ("户外运动商品开发", "户外活动、运动训练或应急辅助"),
            ("宠物用品开发", "宠物训练、日常使用或造型装扮"),
            ("健康护理与康复辅助开发", "健康管理、身体护理或轻康复辅助"),
            ("电子与手机功能配件开发", "照明、拍摄、开关连接或移动设备使用辅助"),
            ("包袋与出行收纳开发", "随身物品携带、证件保护和出行整理"),
            ("阅读文具与学习用品开发", "学习、阅读、记录或知识内容展示"),
        )
        function_value = next(
            (f"推断：{value}" for atom, value in function_rules if atom in directions),
            "无法可靠推断：标题及关联方向均不足以判断核心功能",
        )
        inferred_fields.append("核心功能")

    if scenes:
        scene_value = join_labels(scenes, "")
    else:
        scene_rules = (
            ("万圣节场景商品开发", "万圣节派对、角色扮演和节日布置"),
            ("服饰功能配件开发", "日常穿戴、服装固定和出行使用"),
            ("服饰饰品开发", "日常穿搭、派对造型和礼品赠送"),
            ("美妆装饰商品开发", "化妆造型、派对装扮和拍照场景"),
            ("美妆个护功能商品开发", "家庭个人护理、美容美甲或足部护理"),
            ("电子测量功能商品开发", "浴室、洗浴、水温检测或家庭测量"),
            ("汽车功能配件开发", "车辆日常使用、保养和维修"),
            ("骑行功能配件开发", "自行车安装、骑行和维护"),
            ("功能维修配件开发", "家庭、车库或工作间维修"),
            ("通用工具开发", "家庭DIY、车库、工作间或专业维修"),
            ("DIY手工商品开发", "家庭手工、点钻、绘画和创作桌面"),
            ("派对庆典商品开发", "生日、婚礼、周年和聚会布置"),
            ("玩具娱乐商品开发", "家庭娱乐、亲子互动、派对和礼品"),
            ("花园园艺功能商品开发", "花园、庭院、草坪和户外种植"),
            ("厨房餐饮功能商品开发", "家庭厨房、咖啡和餐饮准备"),
            ("家居收纳与防护商品开发", "家庭收纳、浴室、衣橱或日常防护"),
            ("家居装饰商品开发", "室内、窗边、花园和礼品装饰"),
            ("户外运动商品开发", "露营、球类、钓鱼或户外活动"),
            ("宠物用品开发", "宠物日常使用、训练、护理或节日装扮"),
            ("健康护理与康复辅助开发", "家庭健康管理、身体护理、用药整理或康复辅助"),
            ("电子与手机功能配件开发", "移动拍摄、工作照明、电子连接或设备使用"),
            ("包袋与出行收纳开发", "旅行、通勤、证件携带和日常随身收纳"),
            ("阅读文具与学习用品开发", "家庭学习、阅读、办公记录和知识展示"),
        )
        scene_value = next(
            (f"推断：{value}" for atom, value in scene_rules if atom in directions),
            "无法可靠推断：标题缺少产品本体或使用环境",
        )
        inferred_fields.append("使用场景")

    if audiences:
        audience_value = join_labels(audiences, "")
    else:
        if any(word in title for word in ("脚锉", "磨脚", "去死皮")):
            audience_value = "推断：需要足部去角质、去硬皮护理的成人用户"
        else:
            audience_rules = (
                ("万圣节场景商品开发", "参加万圣节活动、派对或角色扮演的人群及家长"),
                ("服饰功能配件开发", "有穿戴固定、防滑或舒适性需求的用户"),
                ("服饰饰品开发", "关注穿搭、饰品和礼品需求的消费者"),
                ("美妆装饰商品开发", "化妆、派对造型和个性装饰用户"),
                ("美妆个护功能商品开发", "有个人清洁、护理或美容需求的成人用户"),
                ("电子测量功能商品开发", "家庭用户、家长及关注水温安全的人群"),
                ("汽车功能配件开发", "对应车型车主、汽修和车辆维护人员"),
                ("骑行功能配件开发", "自行车用户、骑行爱好者和维修人员"),
                ("功能维修配件开发", "家庭DIY用户、维修人员和设备使用者"),
                ("通用工具开发", "家庭DIY用户、手工人员和专业维修人员"),
                ("DIY手工商品开发", "手工爱好者、点钻用户、亲子创作人群"),
                ("派对庆典商品开发", "派对组织者、家庭用户和礼品购买者"),
                ("玩具娱乐商品开发", "儿童、家庭用户、礼品购买者或收藏人群"),
                ("花园园艺功能商品开发", "家庭园艺用户、庭院业主和种植爱好者"),
                ("厨房餐饮功能商品开发", "家庭烹饪、咖啡和餐饮使用者"),
                ("家居收纳与防护商品开发", "家庭用户、租住用户和收纳整理人群"),
                ("家居装饰商品开发", "家居装饰用户、园艺用户和礼品购买者"),
                ("户外运动商品开发", "户外运动、露营、球类或钓鱼爱好者"),
                ("宠物用品开发", "宠物主人、宠物训练者和宠物装扮消费者"),
                ("健康护理与康复辅助开发", "有健康管理、身体护理或康复辅助需求的用户"),
                ("电子与手机功能配件开发", "手机用户、电子设备使用者和工作照明用户"),
                ("包袋与出行收纳开发", "旅行者、通勤用户和日常随身携带人群"),
                ("阅读文具与学习用品开发", "学生、家庭学习者、阅读者和办公记录用户"),
            )
            audience_value = next(
                (f"推断：{value}" for atom, value in audience_rules if atom in directions),
                "无法可靠推断：标题缺少产品本体和购买者线索",
            )
        inferred_fields.append("适用人群")

    if fitments:
        fitment_value = join_labels(fitments, "")
    elif (height_match := re.search(r"建议身高\s*(\d+\s*[-~—–]\s*\d+)", title)):
        fitment_value = f"标题明确：建议身高{height_match.group(1).replace(' ', '')}cm的人群穿戴"
    elif "花洒" in title and "温度" in title:
        fitment_value = "推断：花洒出水口、淋浴水路或同类水温检测位置"
        inferred_fields.append("适配对象")
    elif any(word in title for word in ("水温测试仪", "水温卡")):
        fitment_value = "推断：浴缸、洗澡水、水盆或同类水温检测场景"
        inferred_fields.append("适配对象")
    elif any(word in title for word in ("脚锉", "磨脚", "去死皮")):
        fitment_value = "推断：脚部、脚底、脚跟及足部硬皮区域"
        inferred_fields.append("适配对象")
    elif any(word in title for word in ("耳骨夹", "耳钉", "耳环")):
        fitment_value = "推断：耳廓、耳骨或耳垂位置"
        inferred_fields.append("适配对象")
    elif "肚脐" in title:
        fitment_value = "推断：肚脐穿刺饰品位置"
        inferred_fields.append("适配对象")
    elif "腰" in title:
        fitment_value = "推断：腰部穿戴和舞蹈服饰搭配"
        inferred_fields.append("适配对象")
    elif any(atom in directions for atom in ("服饰饰品开发", "美妆装饰商品开发", "玩具娱乐商品开发", "派对庆典商品开发", "家居装饰商品开发")):
        fitment_value = "不适用：非设备或型号适配类商品"
    elif directions == ["待确认开发方向"]:
        fitment_value = "无法可靠推断：标题缺少适配对象信息"
        inferred_fields.append("适配对象")
    else:
        fitment_value = f"推断：适用于“{product_body}”对应的使用对象或作业位置"
        inferred_fields.append("适配对象")

    if themes:
        theme_value = join_labels(themes, "")
    elif standardization_type == "标品":
        theme_value = "不适用：功能型商品，以功能和适配为主"
    elif directions == ["待确认开发方向"]:
        theme_value = "无法可靠推断：标题缺少主题或图像元素"
        inferred_fields.append("主题元素")
    else:
        theme_value = "推断：通用装饰、造型或审美主题"
        inferred_fields.append("主题元素")

    if materials:
        material_value = join_labels(materials, "")
        if any(word in title for word in ("脚锉", "磨脚")) and "榉木" in title:
            material_value = f"{material_value}（标题明确榉木）、磨砂锉面"
    else:
        material_rules = (
            (("脚锉", "磨脚"), "榉木/塑料手柄与磨砂锉面，具体以实物为准"),
            (("温度计", "测试仪", "水温卡"), "电子测温组件及塑料/金属外壳，具体以实物为准"),
            (("服装", "长袍", "布", "条纹", "格子"), "纺织布料或服装辅料，具体材质待核实"),
            (("耳钉", "耳骨夹", "项链", "手链", "肚脐", "腰饰", "胸针"), "常见饰品金属/合金及装饰件，具体材质待核实"),
            (("贴纸", "绘画纸", "贺卡", "插旗", "拉旗"), "纸张、胶粘或印刷材料，具体规格待核实"),
            (("玩具", "公仔", "3D打印", "3d打印"), "塑料、树脂或3D打印材料，具体材质待核实"),
            (("工具", "扳手", "锯", "钻头", "塞尺", "卡扣", "挂钩"), "金属主体，部分含塑料/橡胶辅助件，具体材质待核实"),
        )
        material_value = next(
            (
                f"推断：{value}"
                for keywords, value in material_rules
                if any(keyword in title for keyword in keywords)
            ),
            "",
        )
        if not material_value:
            direction_material_rules = (
                ("万圣节场景商品开发", "织物、塑料、金属、纸质或复合装饰材料，具体材质需核实"),
                ("功能维修配件开发", "常见金属、塑料或橡胶功能结构，具体材质需核实"),
                ("汽车功能配件开发", "汽车级金属、塑料或橡胶结构，具体材质需核实"),
                ("骑行功能配件开发", "金属、塑料或橡胶骑行配件结构，具体材质需核实"),
                ("通用工具开发", "金属主体及塑料/橡胶辅助件，具体材质需核实"),
                ("服饰功能配件开发", "织物、硅胶、塑料或金属辅件，具体材质需核实"),
                ("服饰饰品开发", "饰品金属、合金、皮革或装饰材料，具体材质需核实"),
                ("美妆装饰商品开发", "胶粘、纸张、塑料或饰品材料，具体材质需核实"),
                ("美妆个护功能商品开发", "塑料、硅胶、金属或电子组件，具体材质需核实"),
                ("电子测量功能商品开发", "电子元件及塑料/金属外壳，具体材质需核实"),
                ("DIY手工商品开发", "纸张、塑料、金属或手工耗材，具体材质需核实"),
                ("派对庆典商品开发", "纸张、塑料、织物或装饰材料，具体材质需核实"),
                ("玩具娱乐商品开发", "塑料、树脂、织物或复合材料，具体材质需核实"),
                ("花园园艺功能商品开发", "金属、塑料或耐候复合材料，具体材质需核实"),
                ("厨房餐饮功能商品开发", "食品接触级金属、塑料、硅胶或织物，具体材质需核实"),
                ("家居收纳与防护商品开发", "织物、塑料、硅胶或金属结构，具体材质需核实"),
                ("家居装饰商品开发", "玻璃、亚克力、金属、木材或树脂，具体材质需核实"),
                ("户外运动商品开发", "金属、塑料、织物或耐候复合材料，具体材质需核实"),
                ("宠物用品开发", "织物、塑料、硅胶或电子组件，具体材质需核实"),
                ("健康护理与康复辅助开发", "塑料、硅胶、织物或金属辅助结构，具体材质需核实"),
                ("电子与手机功能配件开发", "电子元件、金属、塑料或硅胶结构，具体材质需核实"),
                ("包袋与出行收纳开发", "皮革、织物、塑料或金属连接件，具体材质需核实"),
                ("阅读文具与学习用品开发", "纸张、印刷材料、塑料或金属辅件，具体规格需核实"),
            )
            material_value = next(
                (
                    f"推断：{value}"
                    for atom, value in direction_material_rules
                    if atom in directions
                ),
                "无法可靠推断：标题未提供材质且产品类型不足以安全推断",
            )
        inferred_fields.append("材质与结构")

    sales_window = infer_sales_window(title, directions, season)
    inferred_fields.append("建议售卖时间")
    basis = (
        f"根据产品本体“{product_body}”、方向“{'、'.join(directions)}”"
        f"及常见购买和使用方式，补充推断字段：{'、'.join(dict.fromkeys(inferred_fields))}。"
        "推断内容用于开发分析，具体材质、适配和人群仍应以实物及页面资料复核。"
    )
    return {
        "核心功能": function_value,
        "使用场景": scene_value,
        "适用人群": audience_value,
        "适配对象": fitment_value,
        "主题元素": theme_value,
        "材质与结构": material_value,
        "建议售卖时间": sales_window,
        "信息补充依据": basis,
    }


def analyze_title(
    title: str,
    source_direction: str = "",
    source_standardization: str = "",
) -> dict[str, str]:
    normalized = normalize_product_name(title)
    product_body, body_confirmed = extract_product_body(title)
    family = classify_family(title)
    functions = matched_labels(title, FUNCTION_RULES)
    scenes = matched_labels(title, SCENE_RULES)
    audiences = matched_labels(title, AUDIENCE_RULES)
    fitments = matched_labels(title, BRAND_AND_DEVICE_RULES)
    themes = matched_labels(title, THEME_RULES)
    materials = matched_labels(title, MATERIAL_RULES)
    quantity = extract_quantity_and_bundle(title)
    season = extract_season(title)
    differences = extract_differentiation(
        title,
        themes,
        audiences,
        fitments,
        materials,
        quantity,
    )
    semantic_dimension_count = sum(
        bool(values)
        for values in (functions, scenes, audiences, fitments, themes, materials)
    )
    insufficient_title = title_is_insufficient(
        title,
        product_body,
        body_confirmed,
        semantic_dimension_count,
    )
    if title in FORCE_REVIEW_TITLES:
        insufficient_title = True
    directions = (
        ["待确认开发方向"]
        if insufficient_title
        else direction_atoms(title, product_body, functions, scenes)
    )
    source_atoms = SOURCE_DIRECTION_ATOMS.get(source_direction, ())
    if directions == ["待确认开发方向"] and source_atoms and not insufficient_title:
        directions = [
            select_primary_direction_atom(title, product_body, source_atoms)
        ]
    methods = method_atoms(title, differences, functions, season)
    standardization = standardization_for_title(title, family)
    if insufficient_title:
        standardization = {
            "商品标准化类型": "待复核",
            "标准化判断依据": "标题不足以分析，不能依赖宽泛主方向猜测商品本体",
            "标准化置信度": "低",
        }
    if standardization["商品标准化类型"] == "待复核":
        nonstandard_directions = {
            "万圣节场景商品开发",
            "服饰饰品开发",
            "玩具娱乐商品开发",
            "DIY手工商品开发",
            "派对庆典商品开发",
            "家居装饰商品开发",
            "美妆装饰商品开发",
            "阅读文具与学习用品开发",
        }
        standard_directions = {
            "功能维修配件开发",
            "汽车功能配件开发",
            "骑行功能配件开发",
            "宠物用品开发",
            "美妆个护功能商品开发",
            "花园园艺功能商品开发",
            "厨房餐饮功能商品开发",
            "户外运动商品开发",
            "通用工具开发",
            "电子测量功能商品开发",
            "家居收纳与防护商品开发",
            "服饰功能配件开发",
            "健康护理与康复辅助开发",
            "电子与手机功能配件开发",
            "包袋与出行收纳开发",
        }
        if any(atom in nonstandard_directions for atom in directions):
            standardization = {
                "商品标准化类型": "非标",
                "标准化判断依据": f"完整标题语义归入“{directions[0]}”，以主题、审美或娱乐价值为主",
                "标准化置信度": "中",
            }
        elif any(atom in standard_directions for atom in directions):
            standardization = {
                "商品标准化类型": "标品",
                "标准化判断依据": f"完整标题语义归入“{directions[0]}”，以功能或适配价值为主",
                "标准化置信度": "中",
            }
    if (
        standardization["商品标准化类型"] == "待复核"
        and source_standardization in {"标品", "非标"}
        and not insufficient_title
    ):
        standardization = {
            "商品标准化类型": source_standardization,
            "标准化判断依据": (
                f"沿用基础明细中的“{source_direction or '已审查方向'}”判断，"
                "并由本月逐标题分析继续复核"
            ),
            "标准化置信度": "中",
        }
    completed_dimensions = infer_missing_dimensions(
        title,
        product_body,
        functions,
        scenes,
        audiences,
        fitments,
        themes,
        materials,
        directions,
        standardization["商品标准化类型"],
        season,
    )
    risks = title_risk_flags(title)

    evidence_parts = [f"产品本体按标题识别为“{product_body}”"]
    if functions:
        evidence_parts.append(f"功能词指向“{'、'.join(functions)}”")
    if scenes:
        evidence_parts.append(f"场景词指向“{'、'.join(scenes)}”")
    if fitments:
        evidence_parts.append(f"适配对象为“{'、'.join(fitments)}”")
    if differences:
        evidence_parts.append(f"差异化来自“{'、'.join(differences)}”")

    low_information = (
        insufficient_title
        or product_body == "待确认产品本体"
        or directions == ["待确认开发方向"]
        or (
            not body_confirmed
            and semantic_dimension_count == 0
            and standardization["商品标准化类型"] == "待复核"
        )
    )
    if low_information:
        confidence = "低"
    elif body_confirmed and semantic_dimension_count >= 3:
        confidence = "高"
    else:
        confidence = "中"
    needs_review = "是" if low_information or standardization["商品标准化类型"] == "待复核" else "否"
    review_reason = ""
    if insufficient_title:
        review_reason = "标题不足以分析"
    elif needs_review == "是":
        review_reason = standardization["标准化判断依据"]

    combination = f"{directions[0]}×{methods[0]}"
    return {
        "原始完整标题": title,
        "规范产品名称": normalized,
        "产品本体": product_body,
        "产品族": family,
        "核心功能": completed_dimensions["核心功能"],
        "使用场景": completed_dimensions["使用场景"],
        "适用人群": completed_dimensions["适用人群"],
        "适配对象": completed_dimensions["适配对象"],
        "主题元素": completed_dimensions["主题元素"],
        "材质与结构": completed_dimensions["材质与结构"],
        "数量与组合": quantity,
        "季节属性": season,
        "建议售卖时间": completed_dimensions["建议售卖时间"],
        "商品标准化类型": standardization["商品标准化类型"],
        "标准化判断依据": standardization["标准化判断依据"],
        "差异化方式": join_labels(differences, "标题未明确差异化方式"),
        "风险提示": join_labels(risks, "无明显硬风险"),
        "方向Skill原子": join_labels(directions, "待确认开发方向"),
        "开发方法原子": join_labels(methods, "单品基础开发"),
        "组合打法原子": combination,
        "关联组合主SKU": "",
        "关联完整产品标题": "",
        "标题分析依据": "；".join(evidence_parts),
        "信息补充依据": completed_dimensions["信息补充依据"],
        "分析置信度": confidence,
        "是否待复核": needs_review,
        "待复核原因": review_reason,
    }


def normalize_variant_key(title: str) -> str:
    text = normalize_product_name(title)
    text = re.sub(r"（.*", "", text)
    text = re.sub(r"\(.*", "", text)
    text = re.sub(
        r"\d+(?:\.\d+)?\s*(?:pcs|sets?|件套装|件套|件|个装|个|只装|只|枚|套|张|片|包|组|cm|mm)",
        "",
        text,
        flags=re.I,
    )
    text = re.sub(r"\b(?:XS|S|M|L|XL|XXL|\d+XL)\b", "", text, flags=re.I)
    for word in COLOUR_WORDS:
        text = text.replace(word, "")
    text = re.sub(r"[-—–_/+*×（）()，,；;：:\s]", "", text)
    return text or "待确认产品方案"


def normalize_context_key(title: str) -> str:
    text = normalize_product_name(title).lower()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]", "", text)


def numeric_sku(value: str) -> int | None:
    return int(value) if value.isdigit() else None


def find_combo_parent_index(
    index: int,
    rows: list[dict[str, Any]],
    analyses: list[dict[str, Any]],
) -> int | None:
    row = rows[index]
    if str(row.get("是否组合主SKU") or "") == "是":
        return None
    explicit_parent = next(
        (
            str(row.get(column) or "").strip()
            for column in EXPLICIT_PARENT_COLUMNS
            if str(row.get(column) or "").strip()
        ),
        "",
    )
    if explicit_parent:
        for parent_index, parent_row in enumerate(rows):
            if (
                str(parent_row.get("SKU") or "").strip() == explicit_parent
                and str(parent_row.get("开发人") or "").strip()
                == str(row.get("开发人") or "").strip()
                and str(parent_row.get("创建月份") or "").strip()
                == str(row.get("创建月份") or "").strip()
            ):
                return parent_index
        return None
    month = analyses[index]["创建月份"]
    developer = analyses[index]["开发人"]
    component_key = normalize_context_key(analyses[index]["原始完整标题"])
    component_variant_key = normalize_variant_key(
        analyses[index]["原始完整标题"]
    )
    component_created_at = str(row.get("创建时间") or "").strip()
    component_sku = numeric_sku(analyses[index]["SKU"])
    candidates: list[tuple[int, int, int, int]] = []
    for parent_index, parent_row in enumerate(rows):
        if str(parent_row.get("是否组合主SKU") or "") != "是":
            continue
        parent = analyses[parent_index]
        if parent["创建月份"] != month or parent["开发人"] != developer:
            continue
        parent_key = normalize_context_key(parent["原始完整标题"])
        parent_variant_key = normalize_variant_key(parent["原始完整标题"])
        parent_created_at = str(parent_row.get("创建时间") or "").strip()
        same_creation_batch = bool(
            component_created_at
            and parent_created_at
            and component_created_at == parent_created_at
        )
        parent_sku = numeric_sku(parent["SKU"])
        distance = (
            abs(parent_sku - component_sku)
            if parent_sku is not None and component_sku is not None
            else 999999
        )
        exact_component = len(component_key) >= 4 and component_key in parent_key
        batch_semantic_component = (
            same_creation_batch
            and distance <= 4
            and len(component_variant_key) >= 2
            and component_variant_key in parent_variant_key
        )
        if exact_component or batch_semantic_component:
            candidates.append(
                (
                    0 if same_creation_batch else 1,
                    0 if exact_component else 1,
                    distance,
                    parent_index,
                )
            )
    return min(candidates)[3] if candidates else None


def apply_combo_parent_context(
    component: dict[str, Any],
    parent: dict[str, Any],
) -> None:
    component_name = normalize_product_name(component["原始完整标题"])
    component["关联组合主SKU"] = parent["SKU"]
    component["关联完整产品标题"] = parent["原始完整标题"]
    component["_产品方案键"] = parent["_产品方案键"]
    component["产品本体"] = f"{parent['产品本体']}的组合组件：{component_name[:24]}"
    component["产品族"] = parent["产品族"]
    component["核心功能"] = f"推断：作为“{parent['产品本体']}”的组合组件，承担配套或规格补充作用"
    component["使用场景"] = parent["使用场景"]
    component["适用人群"] = parent["适用人群"]
    component["适配对象"] = parent["适配对象"]
    if component["主题元素"].startswith(("无法可靠推断", "推断：通用")):
        component["主题元素"] = parent["主题元素"]
    if component["材质与结构"].startswith("无法可靠推断"):
        component["材质与结构"] = f"关联推断：{parent['材质与结构']}"
    component["建议售卖时间"] = parent["建议售卖时间"]
    if component["商品标准化类型"] == "待复核":
        component["商品标准化类型"] = parent["商品标准化类型"]
        component["标准化判断依据"] = f"根据关联组合主SKU {parent['SKU']} 的完整商品语义判断"
    component["方向Skill原子"] = parent["方向Skill原子"]
    methods = component["开发方法原子"].split("、")
    methods.append("组合组件复用")
    component["开发方法原子"] = join_labels(methods, "组合组件复用")
    component["组合打法原子"] = (
        f"{component['方向Skill原子'].split('、')[0]}×"
        f"{component['开发方法原子'].split('、')[0]}"
    )
    component["标题分析依据"] += f"；通过组合主SKU {parent['SKU']} 的完整标题补足产品上下文"
    component["信息补充依据"] += (
        f"；该标题是组合主SKU {parent['SKU']} 的组件，"
        f"结合完整标题“{parent['原始完整标题']}”补充使用场景、人群、适配和售卖时间。"
    )
    component["分析置信度"] = "中" if parent["分析置信度"] != "低" else "低"
    component["是否待复核"] = "否" if parent["是否待复核"] == "否" else "是"
    component["待复核原因"] = parent.get("待复核原因", "")


def deep_review_override_path(
    month: str,
    review_root: Path = DEEP_REVIEW_ROOT,
) -> Path:
    year, month_value = month.split("-")
    folder = f"{year[-2:]}年{int(month_value)}月"
    prefix = f"{year}年{int(month_value):02d}月"
    return review_root / folder / f"{prefix}_逐标题深审修正.csv"


def load_deep_review_overrides(
    month: str,
    review_root: Path = DEEP_REVIEW_ROOT,
) -> dict[str, dict[str, str]]:
    path = deep_review_override_path(month, review_root)
    if not path.exists():
        return {}
    required_columns = {
        "创建月份",
        "SKU",
        "建议产品本体",
        "建议方向Skill原子",
        "建议商品标准化类型",
        "建议是否待复核",
        "建议待复核原因",
        "风险提示补充",
    }
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = set(reader.fieldnames or ())
        missing = required_columns - headers
        if missing:
            raise ValueError(f"逐条深审修正表缺少字段：{path} {sorted(missing)}")
        overrides: dict[str, dict[str, str]] = {}
        for row in reader:
            if row.get("创建月份") != month:
                raise ValueError(f"逐条深审修正表月份不一致：{path} {row}")
            sku = str(row.get("SKU") or "").strip()
            if not sku or sku in overrides:
                raise ValueError(f"逐条深审修正表SKU为空或重复：{path} {sku}")
            overrides[sku] = row
    return overrides


def apply_deep_review_override(
    analysis: dict[str, str],
    override: dict[str, str],
) -> None:
    product_body = str(override.get("建议产品本体") or "").strip()
    direction = str(override.get("建议方向Skill原子") or "").strip()
    standardization = str(override.get("建议商品标准化类型") or "").strip()
    needs_review = str(override.get("建议是否待复核") or "").strip()
    review_reason = str(override.get("建议待复核原因") or "").strip()
    risk = str(override.get("风险提示补充") or "").strip()

    if product_body:
        analysis["产品本体"] = product_body
    if direction:
        analysis["方向Skill原子"] = direction
    if standardization:
        analysis["商品标准化类型"] = standardization
        analysis["标准化判断依据"] = "经当月逐标题深审核定"
    if needs_review in {"是", "否"}:
        analysis["是否待复核"] = needs_review
        analysis["待复核原因"] = review_reason if needs_review == "是" else ""
        analysis["分析置信度"] = "低" if needs_review == "是" else "中"
    if risk:
        existing_risks = []
        if analysis["风险提示"] != "无明显硬风险":
            existing_risks = analysis["风险提示"].split("、")
        analysis["风险提示"] = join_labels([*existing_risks, risk], "无明显硬风险")
    analysis["组合打法原子"] = (
        f"{analysis['方向Skill原子']}×"
        f"{analysis['开发方法原子'].split('、')[0]}"
    )
    analysis["标题分析依据"] += "；经当月逐标题深审核定产品本体、主方向或标准化属性"


def analyze_source_rows(
    rows: list[dict[str, Any]],
    deep_review_overrides: dict[str, dict[str, str]] | None = None,
) -> list[dict[str, Any]]:
    deep_review_overrides = deep_review_overrides or {}
    output: list[dict[str, Any]] = []
    for row in rows:
        title = str(row.get("产品名称") or row.get("原始完整标题") or "").strip()
        month = str(row.get("创建月份") or "")
        developer = str(row.get("开发人") or "")
        sku = str(row.get("SKU") or "")
        analysis = analyze_title(
            title,
            str(row.get("主方向") or ""),
            str(row.get("商品标准化类型") or ""),
        )
        deep_review_override = deep_review_overrides.get(sku)
        if deep_review_override:
            apply_deep_review_override(analysis, deep_review_override)
        scheme_key = (
            normalize_context_key(analysis["产品本体"])
            if deep_review_override and analysis["是否待复核"] == "否"
            else normalize_variant_key(title)
        )
        if analysis["是否待复核"] == "是":
            scheme_key = f"待复核-{month}-{developer}-{sku}"
        output.append(
            {
                "创建月份": month,
                "开发人": developer,
                "SKU": sku,
                **analysis,
                "产品方案键": scheme_key,
                "_产品方案键": scheme_key,
                "_是否组合主SKU": str(row.get("是否组合主SKU") or ""),
            }
        )
    for index, component in enumerate(output):
        parent_index = find_combo_parent_index(index, rows, output)
        if parent_index is not None:
            apply_combo_parent_context(component, output[parent_index])
    return output


def assign_product_scheme_ids(
    analyses: list[dict[str, Any]],
    historical_ids: dict[tuple[str, str], str] | None = None,
) -> None:
    historical_ids = historical_ids or {}
    keys = sorted(
        {
            (row["创建月份"], row["开发人"], row["_产品方案键"])
            for row in analyses
        }
    )
    key_to_id: dict[tuple[str, str, str], str] = {}
    counters: defaultdict[tuple[str, str], int] = defaultdict(int)
    for month, developer, scheme_key in keys:
        counters[(month, developer)] += 1
        key_to_id[(month, developer, scheme_key)] = historical_ids.get(
            (developer, scheme_key),
            f"{month.replace('-', '')}-{developer}-P{counters[(month, developer)]:03d}",
        )
    for row in analyses:
        row["产品方案ID"] = key_to_id[
            (row["创建月份"], row["开发人"], row["_产品方案键"])
        ]
        row["产品方案键"] = row["_产品方案键"]


def enforce_product_scheme_direction(
    analyses: list[dict[str, Any]],
) -> None:
    buckets: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in analyses:
        buckets[row["产品方案ID"]].append(row)

    for items in buckets.values():
        parent = next(
            (item for item in items if item.get("_是否组合主SKU") == "是"),
            None,
        )
        eligible_directions = [
            item["方向Skill原子"]
            for item in items
            if item["方向Skill原子"] != "待确认开发方向"
        ]
        selected = ""
        if parent and parent["方向Skill原子"] != "待确认开发方向":
            selected = parent["方向Skill原子"]
        elif eligible_directions:
            counts = Counter(eligible_directions)
            ranked = counts.most_common()
            if len(ranked) == 1 or ranked[0][1] > ranked[1][1]:
                selected = ranked[0][0]
            elif len(set(eligible_directions)) == 1:
                selected = eligible_directions[0]

        if not selected:
            has_conflicting_directions = len(set(eligible_directions)) > 1
            for item in items:
                item["方向Skill原子"] = "待确认开发方向"
                item["组合打法原子"] = (
                    f"待确认开发方向×{item['开发方法原子'].split('、')[0]}"
                )
                item["分析置信度"] = "低"
                item["是否待复核"] = "是"
                if has_conflicting_directions:
                    item["待复核原因"] = "同一产品方案的商品方向冲突，无法确定唯一主导方向"
                elif not str(item.get("待复核原因") or "").strip():
                    item["待复核原因"] = "标题不足以确定唯一商品方向"
            continue

        for item in items:
            item["方向Skill原子"] = selected
            item["组合打法原子"] = (
                f"{selected}×{item['开发方法原子'].split('、')[0]}"
            )


def build_product_scheme_rows(
    rows: list[dict[str, Any]],
    analyses: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    analyses = analyses or analyze_source_rows(rows)
    if not all(str(row.get("产品方案ID") or "").strip() for row in analyses):
        assign_product_scheme_ids(analyses)
    enforce_product_scheme_direction(analyses)
    buckets: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in analyses:
        buckets[row["产品方案ID"]].append(row)

    output: list[dict[str, Any]] = []
    for scheme_id, items in sorted(buckets.items()):
        first = next(
            (item for item in items if item.get("_是否组合主SKU") == "是"),
            items[0],
        )
        sku_list = sorted(item["SKU"] for item in items)
        has_combo_components = any(item["关联组合主SKU"] for item in items)
        output.append(
            {
                "产品方案ID": scheme_id,
                "创建月份": first["创建月份"],
                "开发人": first["开发人"],
                "产品本体": first["产品本体"],
                "产品族": first["产品族"],
                "SKU数": len(items),
                "SKU列表": "、".join(sku_list),
                "标题示例": "；".join(item["原始完整标题"] for item in items[:3]),
                "变体关系": (
                    "组合主SKU及组件"
                    if has_combo_components
                    else "颜色/规格/数量变体"
                    if len(items) > 1
                    else "单一标题方案"
                ),
                "方向Skill原子": first["方向Skill原子"],
                "开发方法原子": join_labels(
                    (
                        atom
                        for item in items
                        for atom in item["开发方法原子"].split("、")
                    ),
                    "单品基础开发",
                ),
                "待复核SKU数": sum(item["是否待复核"] == "是" for item in items),
            }
        )
    return output


def build_skill_evidence_rows(
    analyses: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    evidence: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in analyses:
        for atom in row["方向Skill原子"].split("、"):
            evidence[("方向型", atom)].append(row)
        for atom in row["开发方法原子"].split("、"):
            evidence[("方法型", atom)].append(row)
        evidence[("组合打法", row["组合打法原子"])].append(row)

    output: list[dict[str, Any]] = []
    for (skill_type, atom), items in sorted(evidence.items()):
        scheme_ids = sorted({item["产品方案ID"] for item in items})
        developers = sorted({item["开发人"] for item in items})
        output.append(
            {
                "Skill类型": skill_type,
                "Skill原子": atom,
                "SKU证据数": len(items),
                "产品方案证据数": len(scheme_ids),
                "开发人数": len(developers),
                "开发人": "、".join(developers),
                "产品方案ID示例": "、".join(scheme_ids[:10]),
                "标题证据示例": "；".join(item["原始完整标题"] for item in items[:3]),
                "当前判断": "候选证据，需跨月复现" if atom not in {"待确认开发方向", "单品基础开发"} else "待复核",
            }
        )
    return output


def build_developer_overview_rows(
    analyses: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    analysis_month = analyses[0]["创建月份"] if analyses else TRIAL_MONTH
    developer_groups: list[tuple[str, list[dict[str, Any]]]] = [
        (developer, [row for row in analyses if row["开发人"] == developer])
        for developer in DEVELOPERS
    ]
    developer_groups.append(("合计", analyses))
    for developer, items in developer_groups:
        direction_counter = Counter(
            atom
            for item in items
            for atom in item["方向Skill原子"].split("、")
            if atom != "待确认开发方向"
        )
        method_counter = Counter(
            atom
            for item in items
            for atom in item["开发方法原子"].split("、")
            if atom != "单品基础开发"
        )
        output.append(
            {
                "月份": analysis_month,
                "开发人": developer,
                "SKU数": len(items),
                "产品方案数": len({item["产品方案ID"] for item in items}),
                "产品本体数": len({item["产品本体"] for item in items}),
                "标品数": sum(item["商品标准化类型"] == "标品" for item in items),
                "非标数": sum(item["商品标准化类型"] == "非标" for item in items),
                "待复核数": sum(item["是否待复核"] == "是" for item in items),
                "明确套装/数量配置数": sum("套装/数量配置" in item["差异化方式"] for item in items),
                "明确季节性数": sum(item["季节属性"] != "标题未见明确季节词" for item in items),
                "主要方向Skill原子": "、".join(f"{name}({count})" for name, count in direction_counter.most_common(5)),
                "主要方法Skill原子": "、".join(f"{name}({count})" for name, count in method_counter.most_common(5)),
            }
        )
    return output


def month_folder_name(month: str) -> str:
    year, month_value = map(int, month.split("-"))
    return f"{str(year)[2:]}年{month_value}月"


def month_display_text(month: str) -> str:
    year, month_value = map(int, month.split("-"))
    return f"{year}年{month_value}月"


def month_file_prefix(month: str) -> str:
    year, month_value = map(int, month.split("-"))
    return f"{year}年{month_value:02d}月_三位核心开发人_逐标题开发分析"


def validate_source_columns(fieldnames: Iterable[str] | None) -> None:
    available = {str(name or "").strip() for name in (fieldnames or ())}
    missing = [name for name in REQUIRED_SOURCE_COLUMNS if name not in available]
    if missing:
        raise ValueError(
            "第一层开品基础数据缺少必需字段：" + "、".join(missing)
        )


def is_test_source_row(row: dict[str, Any]) -> bool:
    sku = str(row.get("SKU") or "").strip().lower()
    if sku.startswith(("test", "demo", "测试")):
        return True
    title = str(row.get("产品名称") or row.get("原始完整标题") or "").strip()
    if re.search(r"(?:^|_)测试数据|(?:^|_)测试产品|初级测试产品", title, re.I):
        return True
    for field in ("数据类型", "记录类型", "本地产品状态", "开品状态"):
        value = str(row.get(field) or "").strip().lower()
        if value in TEST_STATUS_VALUES:
            return True
    return False


def source_filter_stats(
    source_path: Path | None,
    month: str,
) -> dict[str, int]:
    if source_path is None or not source_path.exists():
        return {"筛选前范围": 0, "测试数据排除": 0, "空SKU或标题排除": 0}
    with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        validate_source_columns(reader.fieldnames)
        scoped = [
            row
            for row in reader
            if row.get("创建月份") == month
            and row.get("开发人") in DEVELOPERS
        ]
    return {
        "筛选前范围": len(scoped),
        "测试数据排除": sum(is_test_source_row(row) for row in scoped),
        "空SKU或标题排除": sum(
            not str(row.get("SKU") or "").strip()
            or not str(row.get("产品名称") or "").strip()
            for row in scoped
            if not is_test_source_row(row)
        ),
    }


def load_historical_scheme_ids(
    current_month: str,
    output_root: Path = OUTPUT_ROOT,
) -> dict[tuple[str, str], str]:
    candidates: dict[tuple[str, str], tuple[str, str]] = {}
    for path in sorted(output_root.glob("*年*月/*_逐标题开发分析.csv")):
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                month = str(row.get("创建月份") or "").strip()
                if not month or month >= current_month:
                    continue
                developer = str(row.get("开发人") or "").strip()
                scheme_id = str(row.get("产品方案ID") or "").strip()
                if str(row.get("是否待复核") or "").strip() == "是":
                    continue
                scheme_key = str(row.get("产品方案键") or "").strip()
                if not scheme_key:
                    scheme_key = normalize_variant_key(
                        str(row.get("关联完整产品标题") or row.get("原始完整标题") or "")
                    )
                if not developer or not scheme_id or not scheme_key:
                    continue
                key = (developer, scheme_key)
                previous = candidates.get(key)
                if previous is None or month < previous[0]:
                    candidates[key] = (month, scheme_id)
    return {key: value[1] for key, value in candidates.items()}


def month_self_checks(
    source_rows: list[dict[str, Any]],
    analyses: list[dict[str, Any]],
    historical_ids: dict[tuple[str, str], str] | None = None,
) -> dict[str, bool]:
    historical_ids = historical_ids or {}
    direction_one_to_one = all(
        len([atom for atom in str(row.get("方向Skill原子") or "").split("、") if atom]) == 1
        for row in analyses
    )
    nonreview_vague_body = any(
        str(row.get("是否待复核") or "") != "是"
        and str(row.get("产品本体") or "").strip() in VAGUE_PRODUCT_BODIES
        for row in analyses
    )
    historical_reuse = all(
        historical_ids.get((row["开发人"], row["产品方案键"]), row["产品方案ID"])
        == row["产品方案ID"]
        for row in analyses
    )
    scheme_direction_one_to_one = all(
        len(
            {
                str(item.get("方向Skill原子") or "")
                for item in analyses
                if item.get("产品方案ID") == row.get("产品方案ID")
            }
        )
        == 1
        for row in analyses
    )
    return {
        "标题100%覆盖": len(source_rows) == len(analyses),
        "产品方案ID100%覆盖": all(
            str(row.get("产品方案ID") or "").strip() for row in analyses
        ),
        "产品方案键100%覆盖": all(
            str(row.get("产品方案键") or "").strip() for row in analyses
        ),
        "方向原子1对1": direction_one_to_one,
        "产品方案方向1对1": scheme_direction_one_to_one,
        "非待复核空泛本体为0": not nonreview_vague_body,
        "跨月产品方案ID复用": historical_reuse,
    }


def assert_month_self_checks(checks: dict[str, bool]) -> None:
    failed = [name for name, passed in checks.items() if not passed]
    if failed:
        raise ValueError("第一层强制自检未通过：" + "、".join(failed))


def load_month_rows(
    month: str,
    source_path: Path = SOURCE_CSV,
) -> list[dict[str, Any]]:
    with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        validate_source_columns(reader.fieldnames)
        rows = list(reader)
    return [
        row
        for row in rows
        if row.get("创建月份") == month
        and row.get("开发人") in DEVELOPERS
        and str(row.get("SKU") or "").strip()
        and str(row.get("产品名称") or "").strip()
        and not is_test_source_row(row)
    ]


def load_trial_rows(source_path: Path = SOURCE_CSV) -> list[dict[str, Any]]:
    return load_month_rows(TRIAL_MONTH, source_path)


def write_table_sheet(
    workbook: Workbook,
    title: str,
    sheet_name: str,
    headers: list[str],
    rows: Iterable[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = worksheet.cell(row=1, column=1, value=title)
    cell.fill = HEADER_FILL
    cell.font = WHITE_BOLD
    cell.alignment = Alignment(horizontal="center")
    for column, header in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=2, column=column, value=header)
        header_cell.fill = SUBHEADER_FILL
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(rows, start=3):
        for column, header in enumerate(headers, start=1):
            value = row.get(header, "")
            data_cell = worksheet.cell(row=row_index, column=column, value=value)
            data_cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{worksheet.max_row}"
    for column, header in enumerate(headers, start=1):
        if header in {"原始完整标题", "标题分析依据", "标题证据示例", "标题示例"}:
            width = 48
        elif header in {"差异化方式", "方向Skill原子", "开发方法原子", "组合打法原子", "数量与组合"}:
            width = 28
        elif header in {"SKU列表", "产品方案ID示例"}:
            width = 32
        else:
            width = min(max(len(header) + 4, 12), 22)
        worksheet.column_dimensions[get_column_letter(column)].width = width


def build_notes_sheet(workbook: Workbook, row_count: int, month: str) -> None:
    worksheet = workbook.active
    worksheet.title = "00_口径说明"
    notes = [
        ("文件用途", f"{month}逐标题开发Skill分析；每个SKU标题都必须形成一条分析记录。"),
        ("标题覆盖", f"本次共处理{row_count}条SKU标题，不抽样、不省略。"),
        ("三层颗粒度", "SKU标题保留工作量；产品方案合并颜色/尺寸/数量变体；Skill按不同产品方案证据统计。"),
        ("事实与推断", "原始标题、数量、品牌、场景等为标题事实；产品本体、Skill原子为结构化分析结论。"),
        ("缺失信息补充", "标题没有直接写明功能、场景、人群、适配、主题或材质时，根据产品本体和常见使用方式补充合理推断，并在信息补充依据中标明。"),
        ("组合上下文", "单个组件标题信息不足时，优先关联同月同开发人的组合主SKU完整标题；组合组件并回主产品方案，不重复计算开发能力。"),
        ("售卖时间", "根据明确节日词、季节场景和常见需求周期给出建议售卖窗口；全年常驻商品也会注明节日或旺季加强时间。"),
        ("待复核", "只有标题和组合上下文都不足以可靠判断时才进入04_待复核；不能为了填满字段凭空编造具体材质或适配型号。"),
        ("Skill状态", "本文件只生成Skill原子和候选证据，不直接认定形成中、稳定或核心Skill。"),
        ("当前限制", "这是首轮可审查试验，需先由用户检查逐标题拆解和产品方案合并效果，再决定是否扩展到后续月份。"),
    ]
    worksheet.append(["逐标题开发Skill分析试验口径"])
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    worksheet["A1"].fill = HEADER_FILL
    worksheet["A1"].font = WHITE_BOLD
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet.append(["项目", "说明"])
    for cell in worksheet[2]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)
    for note in notes:
        worksheet.append(note)
    for row in worksheet.iter_rows(min_row=3):
        row[0].font = Font(bold=True)
        row[0].fill = NOTE_FILL
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 92
    worksheet.freeze_panes = "A3"


def save_workbook_safely(workbook: Workbook, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f"{target.stem}.tmp{target.suffix}")
    workbook.save(temporary)
    shutil.move(str(temporary), str(target))


def build_trial_workbook(
    rows: list[dict[str, Any]],
    target: Path,
    analyses: list[dict[str, Any]] | None = None,
) -> None:
    analyses = analyses or analyze_source_rows(rows)
    if not all(str(row.get("产品方案ID") or "").strip() for row in analyses):
        assign_product_scheme_ids(analyses)
    enforce_product_scheme_direction(analyses)
    scheme_rows = build_product_scheme_rows(rows, analyses)
    skill_rows = build_skill_evidence_rows(analyses)
    review_rows = [row for row in analyses if row["是否待复核"] == "是"]
    overview_rows = build_developer_overview_rows(analyses)
    analysis_month = analyses[0]["创建月份"] if analyses else TRIAL_MONTH

    workbook = Workbook()
    build_notes_sheet(workbook, len(analyses), analysis_month)
    write_table_sheet(
        workbook,
        "逐标题产品分析：每个SKU一行",
        "01_逐标题产品分析",
        TITLE_ANALYSIS_HEADERS,
        analyses,
    )
    write_table_sheet(
        workbook,
        "产品方案分组：颜色、尺寸和数量变体不重复计算开发能力",
        "02_产品方案分组",
        [
            "产品方案ID",
            "创建月份",
            "开发人",
            "产品本体",
            "产品族",
            "SKU数",
            "SKU列表",
            "标题示例",
            "变体关系",
            "方向Skill原子",
            "开发方法原子",
            "待复核SKU数",
        ],
        scheme_rows,
    )
    write_table_sheet(
        workbook,
        "Skill证据映射：同时显示SKU证据和去变体后的产品方案证据",
        "03_Skill证据映射",
        [
            "Skill类型",
            "Skill原子",
            "SKU证据数",
            "产品方案证据数",
            "开发人数",
            "开发人",
            "产品方案ID示例",
            "标题证据示例",
            "当前判断",
        ],
        skill_rows,
    )
    write_table_sheet(
        workbook,
        "待复核标题：必须解决后才能进入正式Skill证据",
        "04_待复核",
        TITLE_ANALYSIS_HEADERS,
        review_rows,
    )
    write_table_sheet(
        workbook,
        "开发人月度总览：同时展示SKU数和产品方案数",
        "05_开发人月度总览",
        [
            "月份",
            "开发人",
            "SKU数",
            "产品方案数",
            "产品本体数",
            "标品数",
            "非标数",
            "待复核数",
            "明确套装/数量配置数",
            "明确季节性数",
            "主要方向Skill原子",
            "主要方法Skill原子",
        ],
        overview_rows,
    )
    save_workbook_safely(workbook, target)


def write_csv_safely(
    target: Path,
    headers: list[str],
    rows: Iterable[dict[str, Any]],
) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f"{target.stem}.tmp{target.suffix}")
    with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(target)


def monthly_output_paths(month: str, output_root: Path) -> tuple[Path, Path]:
    folder = output_root / month_folder_name(month)
    prefix = month_file_prefix(month)
    return folder / f"{prefix}.xlsx", folder / f"{prefix}.csv"


def write_month_quality_audit(
    source_rows: list[dict[str, Any]],
    analyses: list[dict[str, Any]],
    month: str,
    output_root: Path,
    historical_ids: dict[tuple[str, str], str] | None = None,
    source_path: Path | None = None,
) -> Path:
    folder = output_root / month_folder_name(month)
    target = folder / f"{month_file_prefix(month).replace('_三位核心开发人_逐标题开发分析', '')}_开发Skill提取质量审核.md"
    manual_review_path = folder / f"{month_file_prefix(month).replace('_三位核心开发人_逐标题开发分析', '')}_人工质量复核记录.md"
    developer_counts = Counter(row["开发人"] for row in analyses)
    confidence_counts = Counter(row["分析置信度"] for row in analyses)
    standard_counts = Counter(row["商品标准化类型"] for row in analyses)
    pending = [
        row
        for row in analyses
        if row["分析置信度"] == "低" or row["是否待复核"] == "是"
    ]
    scheme_count = len({row["产品方案ID"] for row in analyses if row["产品方案ID"]})
    combo_count = sum(bool(row["关联组合主SKU"]) for row in analyses)
    required_fields = (
        "创建月份",
        "开发人",
        "SKU",
        "原始完整标题",
        "产品本体",
        "产品族",
        "核心功能",
        "使用场景",
        "适用人群",
        "商品标准化类型",
        "产品方案ID",
        "分析置信度",
        "是否待复核",
    )
    blank_counts = {
        field: sum(not str(row.get(field) or "").strip() for row in analyses)
        for field in required_fields
    }
    coverage_ok = len(source_rows) == len(analyses)
    required_ok = not any(blank_counts.values())
    scheme_ok = all(str(row.get("产品方案ID") or "").strip() for row in analyses)
    checks = month_self_checks(source_rows, analyses, historical_ids)
    filter_stats = source_filter_stats(source_path, month)
    status = "结构与追溯审核通过"
    if not (coverage_ok and required_ok and scheme_ok and all(checks.values())):
        status = "审核不通过，必须修复后重新生成"

    lines = [
        f"# {month_display_text(month)}开发Skill提取质量审核",
        "",
        f"> 质量审核结果：{status}。低置信度或待复核记录不进入正式Skill证据。",
        "",
        "## 一、数据范围",
        "",
        f"- 输入源：`{source_path or SOURCE_CSV}`。",
        f"- 月份与开发人筛选前：{filter_stats['筛选前范围']}条。",
        f"- 排除测试数据：{filter_stats['测试数据排除']}条。",
        f"- 排除空SKU或空标题：{filter_stats['空SKU或标题排除']}条。",
        f"- 原始标题：{len(source_rows)}条。",
        f"- 逐标题分析：{len(analyses)}条。",
        f"- 标题覆盖率：{(len(analyses) / len(source_rows) * 100) if source_rows else 0:.2f}%。",
        f"- 产品方案：{scheme_count}个。",
        f"- 组合组件关联：{combo_count}条。",
        f"- 开发人：陈杨{developer_counts.get('陈杨', 0)}、蒋舒{developer_counts.get('蒋舒', 0)}、宋凤莉{developer_counts.get('宋凤莉', 0)}。",
        "",
        "## 二、置信度与标品非标",
        "",
        f"- 高置信度：{confidence_counts.get('高', 0)}条。",
        f"- 中置信度：{confidence_counts.get('中', 0)}条。",
        f"- 低置信度：{confidence_counts.get('低', 0)}条。",
        f"- 标品：{standard_counts.get('标品', 0)}条；非标：{standard_counts.get('非标', 0)}条；待复核：{standard_counts.get('待复核', 0)}条。",
        "",
        "## 三、自动检查",
        "",
        f"- 标题数量完全覆盖：{'通过' if coverage_ok else '未通过'}。",
        f"- 必填字段完整：{'通过' if required_ok else '未通过'}。",
        f"- 产品方案ID覆盖：{'通过' if scheme_ok else '未通过'}。",
        *[
            f"- {name}：{'通过' if passed else '未通过'}。"
            for name, passed in checks.items()
        ],
        "- 组合组件只在标题语义能够包含于组合主SKU时关联，禁止使用相邻SKU强行归并。",
        "- 低置信度和待复核记录由正式Skill证据生成器自动排除。",
        "",
        "## 四、待复核记录",
        "",
    ]
    if not pending:
        lines.append("- 无。")
    else:
        lines.extend([
            "| 开发人 | SKU | 原始完整标题 | 待复核原因 |",
            "|---|---|---|---|",
        ])
        for row in pending:
            reason = str(row.get("待复核原因") or "").strip()
            if not reason:
                reason = (
                    row["风险提示"]
                    if row["风险提示"] != "无明显硬风险"
                    else row["标准化判断依据"]
                )
            lines.append(
                f"| {row['开发人']} | {row['SKU']} | {row['原始完整标题']} | {reason} |"
            )
    lines.extend([
        "",
        "## 五、验收结论",
        "",
        f"- 本月结构审核：{status}。",
        f"- 待复核标题：{len(pending)}条，保留在逐标题底表，不进入正式Skill证据。",
        "- 第一层结论只说明开发方向与方法，尚未接入商品图片、ASIN上架和利润验证。",
    ])
    if manual_review_path.exists():
        lines[3:3] = [
            "",
            f"- [查看人工语义复核记录]({manual_review_path.name})",
        ]
    folder.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f"{target.stem}.tmp{target.suffix}")
    temporary.write_text("\n".join(lines) + "\n", encoding="utf-8")
    temporary.replace(target)
    return target


def search_keyword_text(row: dict[str, Any]) -> str:
    fields = (
        "开发人",
        "SKU",
        "原始完整标题",
        "产品本体",
        "产品族",
        "核心功能",
        "使用场景",
        "适用人群",
        "适配对象",
        "主题元素",
        "材质与结构",
        "建议售卖时间",
        "商品标准化类型",
        "方向Skill原子",
        "开发方法原子",
        "组合打法原子",
    )
    return " | ".join(
        dict.fromkeys(str(row.get(field) or "").strip() for field in fields if row.get(field))
    )


def discover_month_csv_files(output_root: Path) -> list[Path]:
    return sorted(output_root.glob("*年*月/*_逐标题开发分析.csv"))


def rebuild_master_index(output_root: Path) -> tuple[Path, list[dict[str, Any]]]:
    index_rows: list[dict[str, Any]] = []
    for csv_path in discover_month_csv_files(output_root):
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            monthly_rows = list(csv.DictReader(handle))
        xlsx_path = csv_path.with_suffix(".xlsx")
        for row in monthly_rows:
            index_rows.append(
                {
                    **row,
                    "月份文件夹": csv_path.parent.name,
                    "XLSX文件": xlsx_path.relative_to(output_root).as_posix(),
                    "CSV文件": csv_path.relative_to(output_root).as_posix(),
                    "搜索关键词": search_keyword_text(row),
                }
            )
    index_rows.sort(
        key=lambda row: (
            row.get("创建月份", ""),
            row.get("开发人", ""),
            row.get("SKU", ""),
        )
    )
    index_path = output_root / "00_商品标题分析总索引.csv"
    write_csv_safely(index_path, MASTER_INDEX_HEADERS, index_rows)
    return index_path, index_rows


def write_navigation(
    output_root: Path,
    index_rows: list[dict[str, Any]],
) -> Path:
    month_buckets: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in index_rows:
        month_buckets[row["创建月份"]].append(row)

    lines = [
        "# 逐标题商品分析导航",
        "",
        "> 用途：长期保存三位核心开发人的逐标题分析。以后可通过商品标题、SKU、产品本体、使用场景、适用人群或 Skill 关键词检索以前开过的商品，并复用已经完成的分析。",
        "",
        "## 一、最快搜索入口",
        "",
        "- [商品标题分析总索引](00_商品标题分析总索引.csv)：汇总所有月份，一行一个 SKU 标题。",
        "- 用 Excel 打开总索引后，可筛选“原始完整标题”或“搜索关键词”列。",
        "- 在项目终端搜索示例：`rg -i \"万圣节|清洗机|耳骨夹\" \"00_商品标题分析总索引.csv\"`。",
        "- 搜索结果中的“XLSX文件”和“CSV文件”可以定位到具体月份的完整分析。",
        "",
        "## 二、月份入口",
        "",
        "| 月份 | 标题数 | 产品方案数 | 开发人 | XLSX | CSV | 质量审核 |",
        "|---|---:|---:|---|---|---|---|",
    ]
    for month in sorted(month_buckets):
        rows = month_buckets[month]
        folder = month_folder_name(month)
        prefix = month_file_prefix(month)
        developers = "、".join(sorted({row["开发人"] for row in rows}))
        scheme_count = len({row["产品方案ID"] for row in rows})
        xlsx_relative = f"{folder}/{prefix}.xlsx"
        csv_relative = f"{folder}/{prefix}.csv"
        audit_name = prefix.replace(
            "_三位核心开发人_逐标题开发分析",
            "_开发Skill提取质量审核",
        )
        audit_relative = f"{folder}/{audit_name}.md"
        lines.append(
            f"| {month} | {len(rows)} | {scheme_count} | {developers} | "
            f"[打开XLSX]({xlsx_relative}) | [打开CSV]({csv_relative}) | "
            f"[打开审核]({audit_relative}) |"
        )

    lines.extend(
        [
            "",
            "## 三、文件维护规则",
            "",
            "1. 每个月单独建立一个 `YY年M月` 文件夹。",
            "2. 月份文件夹内同时保存XLSX、CSV和质量审核Markdown。",
            "3. XLSX 用于逐层审查；CSV 用于程序读取；质量审核记录本月验收结果和待复核标题。",
            "4. 根目录总索引由生成脚本自动重建，不手工复制月份数据。",
            "5. 相同分析再次出现时，优先复用既有产品本体、场景、人群和 Skill 定义，再补充新证据。",
            "",
            "## 四、检索建议",
            "",
            "可搜索的常用关键词包括：商品标题、SKU、开发人、产品本体、功能、场景、人群、适配对象、主题、材质、售卖时间、方向 Skill 和开发方法 Skill。",
            "",
        ]
    )
    navigation_path = output_root / "00_逐标题分析导航.md"
    navigation_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = navigation_path.with_name(f"{navigation_path.stem}.tmp{navigation_path.suffix}")
    temporary.write_text("\n".join(lines), encoding="utf-8")
    temporary.replace(navigation_path)
    return navigation_path


def publish_month_analysis(
    rows: list[dict[str, Any]],
    month: str,
    output_root: Path = OUTPUT_ROOT,
    source_path: Path | None = None,
) -> dict[str, Path]:
    if not rows:
        raise ValueError(f"{month}没有可发布的逐标题数据")
    analyses = analyze_source_rows(
        rows,
        load_deep_review_overrides(month),
    )
    historical_ids = load_historical_scheme_ids(month, output_root)
    assign_product_scheme_ids(analyses, historical_ids)
    enforce_product_scheme_direction(analyses)
    checks = month_self_checks(rows, analyses, historical_ids)
    assert_month_self_checks(checks)
    xlsx_path, csv_path = monthly_output_paths(month, output_root)
    build_trial_workbook(rows, xlsx_path, analyses)
    write_csv_safely(csv_path, TITLE_ANALYSIS_HEADERS, analyses)
    audit_path = write_month_quality_audit(
        rows,
        analyses,
        month,
        output_root,
        historical_ids,
        source_path,
    )
    index_path, index_rows = rebuild_master_index(output_root)
    navigation_path = write_navigation(output_root, index_rows)
    return {
        "xlsx": xlsx_path,
        "csv": csv_path,
        "audit": audit_path,
        "index": index_path,
        "navigation": navigation_path,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按月份发布逐标题开发Skill分析、CSV索引和导航")
    parser.add_argument("--source", type=Path, default=SOURCE_CSV)
    parser.add_argument("--month", default=TRIAL_MONTH)
    parser.add_argument("--output-root", type=Path, default=OUTPUT_ROOT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows = load_month_rows(args.month, args.source)
    if not rows:
        raise ValueError(f"没有读取到{args.month}三位开发人的标题数据：{args.source}")
    result = publish_month_analysis(
        rows,
        args.month,
        args.output_root,
        args.source,
    )
    print(f"已生成XLSX：{result['xlsx']}")
    print(f"已生成CSV：{result['csv']}")
    print(f"已生成质量审核：{result['audit']}")
    print(f"已更新导航：{result['navigation']}")
    print(f"已更新总索引：{result['index']}")
    print(f"标题记录：{len(rows)}")


if __name__ == "__main__":
    main()
