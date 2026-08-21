# -*- coding: utf-8 -*-
"""把图片锚定嵌入到 openpyxl 工作表的指定单元格。

只负责"往单元格放图"这一件事：锚点定位、等比缩放、行高列宽联动。
图片来源可以是本地路径或字节流。不依赖 DB。
"""
from __future__ import annotations

import io
from pathlib import Path

from openpyxl.drawing.image import Image as _XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# 经验值：图片列宽约 28(字符) 对应 ~200px，图片显示 190px 留白不贴边；
# 行高 150pt 容纳 190px 图（1px ≈ 0.75pt，190px ≈ 142pt，留余量）。
DEFAULT_CELL_PX = 190
DEFAULT_ROW_PT = 150.0
DEFAULT_COL_WIDTH = 28.0


def embed_image(
    ws: Worksheet,
    row: int,
    col: int,
    source: str | Path | bytes,
    *,
    size_px: int = DEFAULT_CELL_PX,
    set_row_height: bool = True,
    row_height_pt: float = DEFAULT_ROW_PT,
) -> None:
    """在 (row, col) 单元格锚定一张图（均为 1 基坐标）。

    Args:
        source: 本地图片路径或图片字节。
        size_px: 图片显示边长（正方形缩放）。
        set_row_height: 是否同步把该行高度撑到 row_height_pt。
    """
    img = _XLImage(io.BytesIO(source) if isinstance(source, bytes) else source)
    img.width = size_px
    img.height = size_px
    ws.add_image(img, f"{get_column_letter(col)}{row}")
    if set_row_height:
        ws.row_dimensions[row].height = row_height_pt


def set_image_column_width(
    ws: Worksheet, col: int, width: float = DEFAULT_COL_WIDTH
) -> None:
    """把图片所在列的宽度设为容纳图片的固定值。"""
    ws.column_dimensions[get_column_letter(col)].width = width
