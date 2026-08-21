# -*- coding: utf-8 -*-
"""按 SKU 列给 xlsx 填图片列。

典型场景：一张表每行一个 SKU，要在"图片"列放上该 SKU 对应的产品图。
SKU 到图片的映射（走领星 listing、走 COS、走本地目录……）因业务而异，
所以本模块把"SKU -> 图片"抽象成一个注入的 resolver 回调，能力库本身
不碰数据库、不写死数据源。

resolver 约定：
    resolve(sku: str, market: str | None) -> str | Path | bytes | None
    返回图片路径 / 字节，或 None（找不到）。market 由 sheet 站点推断。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional, Union

import openpyxl

from .image_embed import (
    DEFAULT_CELL_PX,
    DEFAULT_COL_WIDTH,
    DEFAULT_ROW_PT,
    embed_image,
    set_image_column_width,
)

ImageSource = Union[str, Path, bytes]
ImageResolver = Callable[[str, Optional[str]], Optional[ImageSource]]
MarketResolver = Callable[[str], Optional[str]]


@dataclass
class SheetFillResult:
    sheet: str
    filled: int = 0        # 成功嵌图
    no_image: int = 0      # 有 SKU 但 resolver 返回 None
    blank_sku: int = 0     # 行内无 SKU，跳过


@dataclass
class FillReport:
    output_path: str
    sheets: list[SheetFillResult] = field(default_factory=list)

    @property
    def total_filled(self) -> int:
        return sum(s.filled for s in self.sheets)


def _default_market(sheet_title: str) -> Optional[str]:
    """从 sheet 名推断站点：含 UK->英国, DE->德国。

    WPS 导出的 sheet 名可能是 GBK 被 latin-1 误读的乱码，先尝试还原。
    """
    name = sheet_title
    try:
        name = sheet_title.encode("latin-1").decode("gbk")
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    upper = name.upper()
    if "UK" in upper or "英国" in name:
        return "英国"
    if "DE" in upper or "德国" in name:
        return "德国"
    return None


def fill_images_by_sku(
    xlsx_path: str | Path,
    output_path: str | Path,
    resolver: ImageResolver,
    *,
    sku_col: int,
    img_col: int,
    data_start_row: int = 2,
    market_resolver: MarketResolver = _default_market,
    clear_existing: bool = True,
    size_px: int = DEFAULT_CELL_PX,
    row_height_pt: float = DEFAULT_ROW_PT,
    col_width: float = DEFAULT_COL_WIDTH,
) -> FillReport:
    """遍历所有 sheet，按 SKU 列填图片列，另存到 output_path。

    Args:
        sku_col / img_col: 1 基列号（A=1）。
        data_start_row: 数据起始行（跳过表头），1 基。
        market_resolver: sheet 名 -> 站点，传给 resolver 做站点优先选图。
        clear_existing: 填图前清空图片列原有内容（如 WPS DISPIMG 占位公式）。
        size_px / row_height_pt / col_width: 图片与单元格尺寸。

    Returns:
        FillReport，含每个 sheet 的 filled/no_image/blank_sku 统计。
    """
    wb = openpyxl.load_workbook(xlsx_path)
    report = FillReport(output_path=str(output_path))
    try:
        for ws in wb.worksheets:
            market = market_resolver(ws.title)
            set_image_column_width(ws, img_col, col_width)
            result = SheetFillResult(sheet=_default_market_label(ws.title))
            for row in range(data_start_row, ws.max_row + 1):
                raw = ws.cell(row, sku_col).value
                if raw is None or not str(raw).strip():
                    result.blank_sku += 1
                    continue
                sku = str(raw).strip()
                if clear_existing:
                    ws.cell(row, img_col).value = None
                source = resolver(sku, market)
                if source is None:
                    result.no_image += 1
                    continue
                embed_image(
                    ws, row, img_col, source,
                    size_px=size_px, row_height_pt=row_height_pt,
                )
                result.filled += 1
            report.sheets.append(result)
        wb.save(output_path)
        return report
    finally:
        wb.close()


def _default_market_label(sheet_title: str) -> str:
    """给报告用的可读 sheet 名（还原 WPS 乱码）。"""
    try:
        return sheet_title.encode("latin-1").decode("gbk")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return sheet_title
