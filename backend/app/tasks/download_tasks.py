"""
Celery 下载任务模块

功能：
- 在后台 Celery worker 中执行批量下载和 ZIP 打包
- 支持失败重试（最多 3 次，间隔 60s）
- 任务去重：同一用户同源任务只允许一个 pending/processing
- xlsx 异步导出：按 ASIN 列表查询数据库生成带图标的 Excel 文件
"""

import asyncio
import io
import logging
import os
from datetime import datetime
from typing import List, Dict, Any, Optional

from celery import shared_task
import aiohttp
import openpyxl
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from .celery_app import celery_app
from ..config import settings
from ..repositories.mysql_repo import MySQLRepository
from ..services.download_task_service import DownloadTaskService, DownloadTaskStatus

logger = logging.getLogger(__name__)

# xlsx 导出字段列表（顺序 = Excel 列序，中文表头→数据库列名）
XLSX_COLUMNS: list[dict[str, str]] = [
    ("图片", "image_url_col"),
    ("销量", "units"),
    ("价格", "price"),
    ("站点", "marketplace"),
    ("ASIN", "asin"),
    ("标题", "title"),
    ("品牌", "brand"),
    ("图片链接", "image_url"),
    ("分类ID路径", "node_id_path"),
    ("分类路径", "node_label_path"),
    ("卖家名称", "seller_name"),
    ("卖家国家", "seller_nation"),
    ("配送方式", "fulfillment"),
    ("重量", "weight"),
    ("尺寸", "dimension"),
    ("包装尺寸", "pkg_dimensions"),
    ("包装重量", "pkg_weight"),
    ("上架日期", "available_date"),
]
# 没有"图片"列的普通数据列
DATA_COLUMNS = [col for col in XLSX_COLUMNS if col[1] != "image_url_col"]

# SQL 查询语句（按 source 映射表名）
TABLE_MAP: dict[str, str] = {
    "premium_products": "premium_products",
    "competitor_clean": "competitor_products",
    "competitor_raw": "competitor_products",
    "shop_products": "shop_products",
}

# xlsx 列宽（近似值）
COL_WIDTHS: dict[str, int] = {
    "图片": 14,
    "销量": 10,
    "价格": 10,
    "站点": 8,
    "ASIN": 16,
    "标题": 50,
    "品牌": 18,
    "图片链接": 40,
    "分类ID路径": 35,
    "分类路径": 40,
    "卖家名称": 22,
    "卖家国家": 12,
    "配送方式": 10,
    "重量": 10,
    "尺寸": 15,
    "包装尺寸": 18,
    "包装重量": 10,
    "上架日期": 14,
}

# 图片尺寸（嵌入 Excel 单元格的大小）
IMAGE_SIZE = 80  # px


async def _fetch_image(session: aiohttp.ClientSession, url: str) -> Optional[bytes]:
    """直接抓取亚马逊图片（后端无 CORS 限制，仅需 UA）。"""
    if not url:
        return None
    try:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=15),
                               headers={"User-Agent": "Mozilla/5.0 (compatible; Bot)"}) as resp:
            if resp.status == 200:
                return await resp.read()
    except Exception:
        pass
    return None


def _format_date(val: Any) -> str:
    """毫秒时间戳 → zh-CN 日期字符串。"""
    if val is None or val == "" or val == 0:
        return ""
    try:
        epoch = int(val)
        dt = datetime.fromtimestamp(epoch / 1000)
        return dt.strftime("%Y/%-m/%-d")
    except (ValueError, OSError, OverflowError):
        return str(val) if val else ""


def _build_xlsx(rows: List[Dict[str, Any]], image_buffers: Dict[int, Optional[bytes]]) -> bytes:
    """使用 openpyxl 构建 xlsx 并返回 bytes。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "选品导出"

    # 写表头
    col_idx_map: dict[str, int] = {}  # column_key → 1-based index
    for i, (header, col_key) in enumerate(XLSX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        col_idx_map[col_key] = i
        if header in COL_WIDTHS:
            ws.column_dimensions[chr(64 + i)].width = COL_WIDTHS[header]

    # 写数据行
    for row_idx, row_data in enumerate(rows, start=2):
        # 销量（数字不转字符串，保留数值格式）
        ws.cell(row=row_idx, column=col_idx_map["units"],
                value=row_data.get("units") or 0)
        # 价格（数字）
        ws.cell(row=row_idx, column=col_idx_map["price"],
                value=float(row_data["price"]) if row_data.get("price") else "")

        for col_key in ("marketplace", "asin", "title", "brand", "image_url",
                        "node_id_path", "node_label_path", "seller_name",
                        "seller_nation", "fulfillment", "weight", "dimension",
                        "pkg_dimensions", "pkg_weight"):
            idx = col_idx_map[col_key]
            ws.cell(row=row_idx, column=idx, value=str(row_data.get(col_key) or ""))

        # 上架日期特殊格式化
        avail = row_data.get("available_date")
        ws.cell(row=row_idx, column=col_idx_map["available_date"],
                value=_format_date(avail))

        # 行高
        ws.row_dimensions[row_idx].height = IMAGE_SIZE

    # 嵌入图片（必须在数据写入之后）
    for row_idx, buf in image_buffers.items():
        if buf is None:
            continue
        try:
            pil_img = PILImage.open(io.BytesIO(buf))
            # 等比例缩放到 IMAGE_SIZE
            w, h = pil_img.size
            if w > h:
                nw, nh = IMAGE_SIZE, int(IMAGE_SIZE * h / w)
            else:
                nw, nh = int(IMAGE_SIZE * w / h), IMAGE_SIZE
            pil_img = pil_img.resize((nw, nh), PILImage.LANCZOS)
            img_bytes = io.BytesIO()
            pil_img.save(img_bytes, format="PNG")
            img_bytes.seek(0)

            xl_img = XLImage(img_bytes)
            xl_img.width = nw
            xl_img.height = nh
            # 锚定到该行第一列（图片列），+2 因为数据行从 row 2 开始
            cell_addr = f"A{row_idx + 2}"
            ws.add_image(xl_img, cell_addr)
        except Exception as e:
            logger.warning(f"嵌入图片失败 row={row_idx}: {e}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@shared_task(bind=True, name="app.tasks.download_tasks.execute_xlsx_export")
def execute_xlsx_export(
    self,
    task_id: str,
    asins: List[str],
    source: str,
    marketplace: str,
):
    """
    Celery 任务：按 ASIN 列表生成 xlsx 文件并更新下载任务状态。

    Args:
        task_id: 下载任务 ID
        asins:  选中商品 ASIN 列表
        source: 数据源标识（premium_products / competitor_clean / shop_products）
        marketplace: 站点
    """
    return asyncio.run(_execute_xlsx_export_async(self, task_id, asins, source, marketplace))


async def _execute_xlsx_export_async(
    self,
    task_id: str,
    asins: List[str],
    source: str,
    marketplace: str,
):
    mysql = MySQLRepository(
        host=settings.MYSQL_HOST,
        port=settings.MYSQL_PORT,
        user=settings.MYSQL_USER,
        password=settings.MYSQL_PASSWORD,
        database=settings.MYSQL_DATABASE,
        min_size=settings.MYSQL_WORKER_POOL_MIN_SIZE,
        pool_size=settings.MYSQL_WORKER_POOL_SIZE,
        max_overflow=settings.MYSQL_WORKER_MAX_OVERFLOW,
        pool_timeout=settings.MYSQL_POOL_TIMEOUT,
    )
    await mysql.connect()

    try:
        svc = DownloadTaskService()
        svc.set_mysql_repo(mysql)

        # 1. 标记进行中
        await svc._update_task_status(task_id, DownloadTaskStatus.PROCESSING, progress=0)

        table = TABLE_MAP.get(source)
        if table is None:
            raise ValueError(f"不支持的 XLSX 导出数据源: {source}")
        total = len(asins)

        # 2. 分页查询数据库（避免单次 IN 过长）
        # 需要的主要字段
        fields = ("asin", "marketplace", "title", "brand", "image_url",
                  "node_id_path", "node_label_path", "seller_name", "seller_nation",
                  "fulfillment", "weight", "dimension", "pkg_dimensions",
                  "pkg_weight", "available_date", "units", "price")

        rows_data: List[Dict[str, Any]] = []
        BATCH = 200
        for i in range(0, total, BATCH):
            batch_asins = asins[i:i + BATCH]
            placeholders = ",".join("%s" for _ in batch_asins)
            sql = f"""
                SELECT {','.join(fields)} FROM {table}
                WHERE asin IN ({placeholders}) AND marketplace = %s
                ORDER BY FIELD(asin, {placeholders})
            """
            params = list(batch_asins) + [marketplace] + list(batch_asins)
            results = await mysql.execute_query(sql, tuple(params), fetch_one=False)
            rows_data.extend(results or [])

        # 3. 并发抓取图片
        image_buffers: Dict[int, Optional[bytes]] = {}
        conn = aiohttp.TCPConnector(limit=8)
        async with aiohttp.ClientSession(connector=conn) as session:
            async def fetch_one(idx: int, row: dict):
                url = row.get("image_url") or ""
                buf = await _fetch_image(session, url)
                image_buffers[idx] = buf

            tasks = []
            for idx, row in enumerate(rows_data):
                tasks.append(fetch_one(idx, row))
                # 每批 20 个并发
                if len(tasks) >= 20:
                    await asyncio.gather(*tasks)
                    progress = int((idx + 1) / total * 100) if total else 0
                    await svc._update_task_status(
                        task_id, DownloadTaskStatus.PROCESSING,
                        progress=progress,
                        completed_files=idx + 1,
                    )
                    tasks = []
            if tasks:
                await asyncio.gather(*tasks)

        # 4. 生成 xlsx
        xlsx_bytes = _build_xlsx(rows_data, image_buffers)

        # 5. 保存文件 + 更新任务完成
        from pathlib import Path
        task_dir = Path(__file__).resolve().parent.parent.parent / "下载缓存" / task_id
        task_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = task_dir / f"选品导出_{marketplace}_{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        with open(xlsx_path, "wb") as f:
            f.write(xlsx_bytes)
        file_size = xlsx_path.stat().st_size

        await svc._update_task_status(
            task_id, DownloadTaskStatus.COMPLETED,
            progress=100,
            completed_files=total,
            total_size=file_size,
            local_path=str(xlsx_path),
        )

        logger.info(f"[XLSX] 导出任务完成: {task_id}, ASIN数: {total}, 大小: {file_size / 1024:.1f} KB")
        return {"success": True, "task_id": task_id, "count": total}

    except Exception as e:
        logger.error(f"[XLSX] 导出任务失败: {task_id}, 错误: {e}")
        if self.request.retries < self.max_retries:
            raise self.retry(exc=e, countdown=60)
        try:
            svc = DownloadTaskService()
            svc.set_mysql_repo(mysql)
            await svc._update_task_status(task_id, DownloadTaskStatus.FAILED, error_message=str(e))
        except Exception:
            pass
        raise

    finally:
        await mysql.disconnect()


# ──────────────────────────────────────────────
# 原有的 execute_download（定稿 ZIP 下载）保留不动
# ──────────────────────────────────────────────

@shared_task(bind=True, name="app.tasks.download_tasks.execute_download")
def execute_download(self, task_id: str, files: List[Dict[str, Any]]):
    """
    执行下载任务（定稿 ZIP 打包）

    Args:
        task_id: 下载任务ID
        files: 文件列表 [{"url": "...", "filename": "..."}, ...]

    Returns:
        dict: 执行结果
    """
    return asyncio.run(_execute_download_async(self, task_id, files))


async def _execute_download_async(self, task_id: str, files: List[Dict[str, Any]]):
    """
    异步执行下载任务的内部实现

    Args:
        self: Celery task binding（用于 retry）
        task_id: 下载任务ID
        files: 文件列表
    """
    mysql = MySQLRepository(
        host=settings.MYSQL_HOST,
        port=settings.MYSQL_PORT,
        user=settings.MYSQL_USER,
        password=settings.MYSQL_PASSWORD,
        database=settings.MYSQL_DATABASE,
        min_size=settings.MYSQL_WORKER_POOL_MIN_SIZE,
        pool_size=settings.MYSQL_WORKER_POOL_SIZE,
        max_overflow=settings.MYSQL_WORKER_MAX_OVERFLOW,
        pool_timeout=settings.MYSQL_POOL_TIMEOUT,
    )

    await mysql.connect()

    try:
        svc = DownloadTaskService()
        svc.set_mysql_repo(mysql)
        await svc.execute_task(task_id, files)

        logger.info(f"[OK] 下载任务执行完成: {task_id}")
        return {"success": True, "task_id": task_id}

    except Exception as e:
        logger.error(f"[FAIL] 下载任务执行失败: {task_id}, 错误: {e}")

        if self.request.retries < self.max_retries:
            logger.info(f"重试下载任务: {task_id}, 第 {self.request.retries + 1} 次")
            raise self.retry(exc=e, countdown=60)

        raise

    finally:
        await mysql.disconnect()
